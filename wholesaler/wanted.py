"""
Retail Shop Portal + Smart "Wanted List" Engine
===============================================

Two things live here.

1. SHOP PORTAL (/shop/*) — the retailer's own login. Dashboard, place order,
   my bills, acknowledge bill, credit notes, outstanding. A shop sees ONLY
   its own rows; every query is filtered by the shop_id in their session.

2. WANTED-LIST ENGINE — the reason this exists.

   The old way: a shop sends a list of what they want, written in THEIR
   names ("dolo650", "PAN-40", "augmentin 625 tab"). Someone at the
   distributor reads each line, hunts for it in the item master, and types
   an order. 1 to 1.5 hours.

   The new way:
     * FIRST upload from a shop — we auto-match what we can, and ask a human
       to resolve only the uncertain ones. Every decision is remembered as
       an ALIAS against that shop.
     * EVERY LATER upload — those aliases hit instantly. A shop that always
       writes "dolo650" never has to be asked again. Minutes, not hours.

   Matching ladder, cheapest first:
     1. alias      — this shop has used this exact text before  → instant
     2. global alias — another shop taught us this text          → instant
     3. exact      — normalised text equals an item's name/code  → instant
     4. fuzzy      — close enough; top 3 shown for one-click confirm
     5. none       — we don't stock it → goes to the DEMAND REPORT

   The demand report is the second half of the customer's ask: the lines we
   could NOT match are usually not typos, they're products the distributor
   simply doesn't carry. Aggregated across every shop, that becomes a
   ranked "what should we start stocking" list — real demand, with the
   number of shops asking and the quantity they wanted.
"""
from __future__ import annotations

import io
import re
import csv
import json
import secrets
import functools
from difflib import SequenceMatcher
from datetime import date, datetime

from flask import (request, redirect, url_for, render_template, session,
                   flash, abort, jsonify, Response)

from app import app, conn, login_required, audit, compute_order_totals, next_order_no, rate_for_shop

try:
    from openpyxl import load_workbook
    HAVE_XLSX = True
except ImportError:
    HAVE_XLSX = False


# ══════════════════════════════════════════════════════════════════════
#  SCHEMA
# ══════════════════════════════════════════════════════════════════════
WANTED_SCHEMA = r"""
CREATE TABLE IF NOT EXISTS item_aliases (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    shop_id     INTEGER,                 -- NULL = learned globally
    alias_norm  TEXT NOT NULL,
    alias_raw   TEXT DEFAULT '',
    item_id     INTEGER NOT NULL,
    hits        INTEGER DEFAULT 1,
    created_at  TEXT DEFAULT (datetime('now'))
);
CREATE UNIQUE INDEX IF NOT EXISTS ix_alias ON item_aliases(IFNULL(shop_id,0), alias_norm);

CREATE TABLE IF NOT EXISTS wanted_uploads (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    shop_id      INTEGER NOT NULL,
    filename     TEXT DEFAULT '',
    total_lines  INTEGER DEFAULT 0,
    auto_matched INTEGER DEFAULT 0,
    needs_review INTEGER DEFAULT 0,
    not_stocked  INTEGER DEFAULT 0,
    status       TEXT DEFAULT 'review',   -- review | ordered
    order_id     INTEGER,
    source       TEXT DEFAULT 'admin',    -- admin | shop
    created_at   TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS wanted_lines (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    upload_id   INTEGER NOT NULL,
    raw_name    TEXT DEFAULT '',
    norm_name   TEXT DEFAULT '',
    qty         INTEGER DEFAULT 0,
    item_id     INTEGER,
    match_type  TEXT DEFAULT 'none',      -- alias | global | exact | fuzzy | none
    confidence  REAL DEFAULT 0,
    suggestions TEXT DEFAULT '[]'
);

CREATE TABLE IF NOT EXISTS demand_log (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    norm_name      TEXT UNIQUE NOT NULL,
    raw_name       TEXT DEFAULT '',
    times_asked    INTEGER DEFAULT 0,
    total_qty      INTEGER DEFAULT 0,
    shops_json     TEXT DEFAULT '[]',
    last_asked     TEXT DEFAULT (date('now')),
    status         TEXT DEFAULT 'open',   -- open | sourcing | added | ignored
    note           TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS shop_logins (
    shop_id     INTEGER PRIMARY KEY,
    pin         TEXT DEFAULT '',      -- legacy plaintext; emptied on first login after upgrade
    pin_salt    TEXT DEFAULT '',
    pin_hash    TEXT DEFAULT '',
    last_login  TEXT DEFAULT ''
);
"""

with conn() as _c:
    _c.executescript(WANTED_SCHEMA)


# ══════════════════════════════════════════════════════════════════════
#  SHOP CREDENTIALS
#
#  Retailer secrets used to be stored in shop_logins.pin as PLAIN TEXT and
#  compared with `!=`. Anyone with a copy of a tenant DB - a backup, a
#  support dump, the platform owner - could read every retailer's PIN, and
#  people reuse PINs. Now salted-SHA256, same scheme as staff_users and
#  company logins so there's one hashing story in the codebase.
#
#  Migration is lazy and non-breaking: a row that still has plaintext is
#  verified against it once, then immediately rewritten as a hash (see
#  shop_login). No retailer has to be reset, and the plaintext is gone
#  after their next sign-in. ensure_shop_login_columns() adds the new
#  columns to tenant DBs created before this change.
# ══════════════════════════════════════════════════════════════════════
def _hash_pin(pin: str, salt: str | None = None) -> tuple[str, str]:
    import hashlib
    salt = salt or secrets.token_hex(8)
    return salt, hashlib.sha256((salt + pin).encode()).hexdigest()


def _has_credential(row) -> bool:
    """True if this shop has ever set a secret (hashed or legacy plaintext)."""
    if not row:
        return False
    try:
        if row["pin_hash"]:
            return True
    except (IndexError, KeyError):
        pass
    return bool(row["pin"])


def verify_shop_pin(row, pin: str) -> bool:
    import hmac
    try:
        stored_hash, stored_salt = row["pin_hash"], row["pin_salt"]
    except (IndexError, KeyError):
        stored_hash, stored_salt = "", ""
    if stored_hash:
        return hmac.compare_digest(stored_hash, _hash_pin(pin, stored_salt)[1])
    # legacy plaintext row
    return bool(row["pin"]) and hmac.compare_digest(row["pin"], pin)


def ensure_shop_login_columns():
    with conn() as c:
        cols = {r["name"] for r in c.execute("PRAGMA table_info(shop_logins)").fetchall()}
        for col in ("pin_salt", "pin_hash"):
            if col not in cols:
                c.execute(f"ALTER TABLE shop_logins ADD COLUMN {col} TEXT DEFAULT ''")


def set_shop_pin(shop_id: int, pin: str):
    """Set/reset a shop's password. Never stores it readable."""
    ensure_shop_login_columns()
    salt, h = _hash_pin(pin)
    with conn() as c:
        c.execute("""INSERT INTO shop_logins (shop_id, pin, pin_salt, pin_hash)
                     VALUES (?,'',?,?)
                     ON CONFLICT(shop_id) DO UPDATE SET pin='',
                       pin_salt=excluded.pin_salt, pin_hash=excluded.pin_hash""",
                  (shop_id, salt, h))


# ══════════════════════════════════════════════════════════════════════
#  NORMALISATION + MATCHING
# ══════════════════════════════════════════════════════════════════════
# Pack/format noise that shops write but which never distinguishes a product.
_NOISE = re.compile(
    r"\b(tab|tabs|tablet|tablets|cap|caps|capsule|capsules|syp|syrup|susp|suspension|"
    r"inj|injection|oint|ointment|cream|gel|drops?|sachet|strip|strips|bottle|btl|"
    r"pcs|pc|nos|no|box|pkt|packet|pack)\b", re.I)
_UNIT = re.compile(r"\b\d+\s*(mg|ml|gm|g|mcg|iu|%)\b", re.I)


def norm(s: str) -> str:
    """
    Normalise a product name so 'Dolo 650 Tab', 'dolo-650', 'DOLO650 tablets'
    all collapse to the same key. Strength (650) is KEPT — it distinguishes
    real products. Dosage form (tab/syrup) is dropped — it doesn't.
    """
    s = (s or "").lower().strip()
    s = re.sub(r"[^\w\s%]+", " ", s)     # punctuation → space
    s = _NOISE.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _sim(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


# Fuzzy-matching used to score every unmatched line against the ENTIRE
# catalog (O(catalog_size) SequenceMatcher calls per line). Fine at 50
# items; at a real distributor's 5,000-15,000 SKUs, a 100-line upload could
# mean 1M+ comparisons in one request, monopolising a gunicorn worker for
# tens of seconds on a shared 1-vCPU box. Below, candidates are narrowed to
# a prefix bucket first, capped at MAX_FUZZY_CANDIDATES — cost per line is
# now bounded and independent of catalog size, not just "usually small."
MAX_FUZZY_CANDIDATES = 300
MIN_BUCKET_SIZE = 8   # widen to the 1-char bucket if the 2-char one is this sparse


def load_catalog(c):
    rows = c.execute("""SELECT id, code, name, generic, manufacturer, pack_size, ptr, ptr_b, ptr_c,
                               gst_rate, scheme, moq, stock, mrp
                        FROM wholesale_items WHERE status='active'""").fetchall()
    cat = []
    for r in rows:
        d = dict(r)
        d["_norm"] = norm(r["name"])
        d["_gnorm"] = norm(r["generic"] or "")
        cat.append(d)
    return cat


def build_catalog_index(catalog):
    """
    Built ONCE per upload, not per line.
      exact_norm/exact_code — O(1) lookup for exact matches (was an
        O(catalog_size) scan before).
      idx2/idx1 — items bucketed by the first 2 (then 1, as a fallback)
        characters of their normalised name AND generic name, so a fuzzy
        match only scores plausible candidates instead of everything.
    """
    exact_norm, exact_code = {}, {}
    idx2, idx1 = {}, {}
    for it in catalog:
        if it["_norm"]:
            exact_norm.setdefault(it["_norm"], it)
            idx2.setdefault(it["_norm"][:2], []).append(it)
            idx1.setdefault(it["_norm"][:1], []).append(it)
        if it["_gnorm"]:
            idx2.setdefault(it["_gnorm"][:2], []).append(it)
            idx1.setdefault(it["_gnorm"][:1], []).append(it)
        code = (it["code"] or "").lower().replace(" ", "")
        if code:
            exact_code.setdefault(code, it)
    return {"exact_norm": exact_norm, "exact_code": exact_code, "idx2": idx2, "idx1": idx1}


def _fuzzy_candidates(n, index):
    seen = {}
    for it in index["idx2"].get(n[:2], []):
        seen[it["id"]] = it
    if len(seen) < MIN_BUCKET_SIZE:
        for it in index["idx1"].get(n[:1], []):
            seen[it["id"]] = it
    # dict preserves insertion order; cap bounds worst case regardless of
    # catalog size (e.g. a catalog skewed toward one starting letter).
    return list(seen.values())[:MAX_FUZZY_CANDIDATES]


def match_line(c, index, shop_id, raw_name):
    """Return (item_id, match_type, confidence, suggestions[])."""
    n = norm(raw_name)
    if not n:
        return None, "none", 0.0, []

    # 1 & 2 — learned aliases (this shop first, then anything learned globally)
    row = c.execute("""SELECT item_id FROM item_aliases
                       WHERE alias_norm=? AND shop_id=? LIMIT 1""", (n, shop_id)).fetchone()
    if row:
        return row["item_id"], "alias", 1.0, []
    row = c.execute("""SELECT item_id FROM item_aliases
                       WHERE alias_norm=? AND shop_id IS NULL LIMIT 1""", (n,)).fetchone()
    if row:
        return row["item_id"], "global", 1.0, []

    # 3 — exact on normalised name, or on SKU code — O(1) dict lookup
    it = index["exact_norm"].get(n) or index["exact_code"].get(n.replace(" ", ""))
    if it:
        return it["id"], "exact", 1.0, []

    # 4 — fuzzy, over a bounded candidate set instead of the whole catalog
    scored = []
    for cand in _fuzzy_candidates(n, index):
        s = _sim(n, cand["_norm"])
        if cand["_gnorm"]:
            s = max(s, _sim(n, cand["_gnorm"]) * 0.95)   # generic match is slightly weaker evidence
        # a shared leading token is a strong hint ("pan 40" vs "pantop 40")
        if n.split()[:1] and cand["_norm"].split()[:1] and n.split()[0] == cand["_norm"].split()[0]:
            s = min(1.0, s + 0.12)
        if s > 0.55:
            scored.append((s, cand))
    scored.sort(key=lambda x: -x[0])

    if scored and scored[0][0] >= 0.88:
        best = scored[0]
        return best[1]["id"], "fuzzy", round(best[0], 3), [
            {"id": i["id"], "name": i["name"], "pack": i["pack_size"], "score": round(s, 3)}
            for s, i in scored[:3]]
    if scored:
        return None, "fuzzy", round(scored[0][0], 3), [
            {"id": i["id"], "name": i["name"], "pack": i["pack_size"], "score": round(s, 3)}
            for s, i in scored[:3]]

    # 5 — genuinely not in our catalogue
    return None, "none", 0.0, []


def parse_wanted_file(fs):
    """Accept .xlsx/.csv with (name, qty). Very forgiving about column names."""
    fname = (fs.filename or "").lower()
    raw = fs.read()
    rows = []

    if fname.endswith(".csv"):
        rdr = csv.reader(io.StringIO(raw.decode("utf-8-sig", errors="replace")))
        table = list(rdr)
    elif fname.endswith((".xlsx", ".xlsm")):
        if not HAVE_XLSX:
            return [], "openpyxl not installed — save the file as CSV and retry."
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        table = [[("" if v is None else str(v).strip()) for v in r]
                 for r in wb.active.iter_rows(values_only=True)]
    else:
        return [], "Please upload a .xlsx or .csv file."

    if not table:
        return [], "That file is empty."

    # Find the name/qty columns from a header row if there is one; otherwise
    # assume col0 = name and the first numeric-looking column = qty.
    head = [str(x).lower().strip() for x in table[0]]
    name_i = qty_i = None
    for i, h in enumerate(head):
        if name_i is None and any(k in h for k in ("item", "name", "product", "particular", "description", "medicine")):
            name_i = i
        if qty_i is None and any(k in h for k in ("qty", "quantity", "nos", "count", "req")):
            qty_i = i
    body = table[1:] if (name_i is not None or qty_i is not None) else table
    if name_i is None:
        name_i = 0
    if qty_i is None:
        qty_i = 1 if (body and len(body[0]) > 1) else None

    for r in body:
        if not r or name_i >= len(r):
            continue
        nm = str(r[name_i]).strip()
        if not nm:
            continue
        q = 1
        if qty_i is not None and qty_i < len(r):
            m = re.search(r"\d+", str(r[qty_i]))
            if m:
                q = int(m.group())
        rows.append({"name": nm, "qty": max(1, q)})
    return rows, None


def record_demand(c, line_rows, shop_id):
    """Fold unmatched lines into the aggregate 'we should stock this' report."""
    for ln in line_rows:
        n, rawname, q = ln["norm"], ln["raw"], ln["qty"]
        if not n:
            continue
        row = c.execute("SELECT * FROM demand_log WHERE norm_name=?", (n,)).fetchone()
        if row:
            shops = set(json.loads(row["shops_json"] or "[]"))
            shops.add(shop_id)
            c.execute("""UPDATE demand_log SET times_asked=times_asked+1, total_qty=total_qty+?,
                         shops_json=?, last_asked=date('now'), raw_name=?
                         WHERE norm_name=?""",
                      (q, json.dumps(sorted(shops)), rawname, n))
        else:
            c.execute("""INSERT INTO demand_log (norm_name, raw_name, times_asked, total_qty, shops_json)
                         VALUES (?,?,1,?,?)""", (n, rawname, q, json.dumps([shop_id])))


# ══════════════════════════════════════════════════════════════════════
#  ADMIN: upload a shop's wanted list
# ══════════════════════════════════════════════════════════════════════
@app.route("/wanted", methods=["GET", "POST"])
@login_required
def wanted():
    with conn() as c:
        shops = c.execute("SELECT id, code, name FROM retail_shops WHERE status='active' ORDER BY name").fetchall()
        recent = c.execute("""SELECT u.*, s.name shop_name, s.code shop_code
                              FROM wanted_uploads u JOIN retail_shops s ON s.id=u.shop_id
                              ORDER BY u.id DESC LIMIT 15""").fetchall()

    if request.method == "POST":
        shop_id = int(request.form.get("shop_id") or 0)
        f = request.files.get("file")
        if not shop_id or not f or not f.filename:
            flash("Pick a shop and a file.", "err")
            return redirect(url_for("wanted"))

        rows, err = parse_wanted_file(f)
        if err:
            flash(err, "err")
            return redirect(url_for("wanted"))
        if not rows:
            flash("Couldn't find any item rows in that file.", "err")
            return redirect(url_for("wanted"))

        upload_id = process_wanted(shop_id, rows, f.filename, source="admin")
        return redirect(url_for("wanted_review", upload_id=upload_id))

    return render_template("wanted.html", shops=shops, recent=recent, have_xlsx=HAVE_XLSX)


def process_wanted(shop_id, rows, filename, source="admin"):
    """Shared by the admin upload and the shop-portal upload."""
    with conn() as c:
        catalog = load_catalog(c)
        index = build_catalog_index(catalog)   # once per upload, not per line
        cur = c.execute("""INSERT INTO wanted_uploads (shop_id, filename, total_lines, source)
                           VALUES (?,?,?,?)""", (shop_id, filename, len(rows), source))
        upload_id = cur.lastrowid

        auto = review = missing = 0
        demand_rows = []
        for r in rows:
            item_id, mtype, conf, sugg = match_line(c, index, shop_id, r["name"])
            c.execute("""INSERT INTO wanted_lines
                (upload_id, raw_name, norm_name, qty, item_id, match_type, confidence, suggestions)
                VALUES (?,?,?,?,?,?,?,?)""",
                (upload_id, r["name"], norm(r["name"]), r["qty"], item_id, mtype, conf, json.dumps(sugg)))
            if item_id and mtype in ("alias", "global", "exact"):
                auto += 1
            elif sugg:
                review += 1
            else:
                missing += 1
                demand_rows.append({"norm": norm(r["name"]), "raw": r["name"], "qty": r["qty"]})

        if demand_rows:
            record_demand(c, demand_rows, shop_id)

        c.execute("""UPDATE wanted_uploads SET auto_matched=?, needs_review=?, not_stocked=?
                     WHERE id=?""", (auto, review, missing, upload_id))
    return upload_id


@app.route("/wanted/<int:upload_id>")
@login_required
def wanted_review(upload_id):
    with conn() as c:
        up = c.execute("""SELECT u.*, s.name shop_name, s.code shop_code, s.price_tier
                          FROM wanted_uploads u JOIN retail_shops s ON s.id=u.shop_id
                          WHERE u.id=?""", (upload_id,)).fetchone()
        if not up:
            abort(404)
        lines = c.execute("SELECT * FROM wanted_lines WHERE upload_id=? ORDER BY id", (upload_id,)).fetchall()
        items = {r["id"]: dict(r) for r in c.execute(
            "SELECT id, name, pack_size, ptr, ptr_b, ptr_c, stock FROM wholesale_items").fetchall()}

    parsed = []
    for ln in lines:
        d = dict(ln)
        d["suggestions"] = json.loads(ln["suggestions"] or "[]")
        d["item"] = items.get(ln["item_id"])
        parsed.append(d)
    return render_template("wanted_review.html", up=up, lines=parsed, items=items)


@app.route("/wanted/<int:upload_id>/resolve", methods=["POST"])
@login_required
def wanted_resolve(upload_id):
    """
    Human confirms the uncertain lines. Each confirmation is REMEMBERED as an
    alias so this shop never gets asked about that spelling again — this is
    what turns the second upload from an hour into a couple of minutes.
    """
    line_ids = request.form.getlist("line_id[]")
    chosen = request.form.getlist("choice[]")
    learned = 0
    with conn() as c:
        up = c.execute("SELECT * FROM wanted_uploads WHERE id=?", (upload_id,)).fetchone()
        if not up:
            abort(404)
        for lid, ch in zip(line_ids, chosen):
            if not ch:
                continue
            ln = c.execute("SELECT * FROM wanted_lines WHERE id=?", (int(lid),)).fetchone()
            if not ln:
                continue
            if ch == "skip":
                c.execute("UPDATE wanted_lines SET item_id=NULL, match_type='skipped' WHERE id=?", (ln["id"],))
                continue
            item_id = int(ch)
            c.execute("UPDATE wanted_lines SET item_id=?, match_type='confirmed', confidence=1.0 WHERE id=?",
                      (item_id, ln["id"]))
            try:
                c.execute("""INSERT INTO item_aliases (shop_id, alias_norm, alias_raw, item_id)
                             VALUES (?,?,?,?)""",
                          (up["shop_id"], ln["norm_name"], ln["raw_name"], item_id))
                learned += 1
            except Exception:
                c.execute("""UPDATE item_aliases SET item_id=?, hits=hits+1
                             WHERE shop_id=? AND alias_norm=?""",
                          (item_id, up["shop_id"], ln["norm_name"]))
    flash(f"Saved. Learned {learned} new name(s) for this shop — next upload they'll match instantly.", "ok")
    return redirect(url_for("wanted_review", upload_id=upload_id))


@app.route("/wanted/<int:upload_id>/create-order", methods=["POST"])
@login_required
def wanted_create_order(upload_id):
    with conn() as c:
        up = c.execute("SELECT * FROM wanted_uploads WHERE id=?", (upload_id,)).fetchone()
        if not up:
            abort(404)
        shop = c.execute("SELECT * FROM retail_shops WHERE id=?", (up["shop_id"],)).fetchone()
        lines = c.execute("""SELECT w.*, i.* FROM wanted_lines w
                             JOIN wholesale_items i ON i.id=w.item_id
                             WHERE w.upload_id=? AND w.item_id IS NOT NULL""", (upload_id,)).fetchall()
        if not lines:
            flash("Nothing matched yet — resolve some lines first.", "err")
            return redirect(url_for("wanted_review", upload_id=upload_id))

        order_no = next_order_no()
        cur = c.execute("""INSERT INTO sales_orders (order_no, shop_id, notes, source, created_by)
                           VALUES (?,?,?,?,?)""",
                        (order_no, up["shop_id"], f"From wanted list #{upload_id} ({up['filename']})",
                         "wanted", session.get("ws_user") or "admin"))
        order_id = cur.lastrowid

        for ln in lines:
            rate = rate_for_shop(ln, shop["price_tier"])
            free = 0
            if ln["scheme"] and "+" in (ln["scheme"] or ""):
                try:
                    b, bs = ln["scheme"].split("+"); b, bs = int(b), int(bs)
                    if b > 0:
                        free = (ln["qty"] // b) * bs
                except Exception:
                    pass
            c.execute("""INSERT INTO sales_order_items
                (order_id, item_id, item_name, pack_size, qty, free_qty, rate, gst_rate, amount)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (order_id, ln["item_id"], ln["name"], ln["pack_size"], ln["qty"], free,
                 rate, ln["gst_rate"], ln["qty"] * rate))

        c.execute("UPDATE wanted_uploads SET status='ordered', order_id=? WHERE id=?", (order_id, upload_id))

    compute_order_totals(order_id)
    audit(session.get("ws_user"), "wanted_to_order", "order", order_id, order_no)
    flash(f"Created order {order_no} from the wanted list.", "ok")
    return redirect(url_for("order_detail", order_id=order_id))


# ══════════════════════════════════════════════════════════════════════
#  DEMAND REPORT — what shops want that we don't stock
# ══════════════════════════════════════════════════════════════════════
@app.route("/demand")
@login_required
def demand():
    status = request.args.get("status") or "open"
    with conn() as c:
        if status == "all":
            rows = c.execute("SELECT * FROM demand_log ORDER BY times_asked DESC, total_qty DESC").fetchall()
        else:
            rows = c.execute("SELECT * FROM demand_log WHERE status=? ORDER BY times_asked DESC, total_qty DESC",
                             (status,)).fetchall()
        shop_names = {r["id"]: r["name"] for r in c.execute("SELECT id, name FROM retail_shops").fetchall()}

    out = []
    for r in rows:
        d = dict(r)
        ids = json.loads(r["shops_json"] or "[]")
        d["shop_count"] = len(ids)
        d["shop_names"] = ", ".join(shop_names.get(i, f"#{i}") for i in ids[:6])
        out.append(d)
    totals = {"lines": len(out), "shops": len({s for r in rows for s in json.loads(r["shops_json"] or "[]")})}
    return render_template("demand.html", rows=out, status=status, totals=totals)


@app.route("/demand/<int:did>/<action>", methods=["POST"])
@login_required
def demand_action(did, action):
    if action not in ("sourcing", "added", "ignored", "open"):
        abort(400)
    with conn() as c:
        c.execute("UPDATE demand_log SET status=? WHERE id=?", (action, did))
    flash("Updated.", "ok")
    return redirect(url_for("demand"))


@app.route("/demand/export.csv")
@login_required
def demand_export():
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["product_asked_for", "times_asked", "total_qty", "shops_asking", "last_asked", "status"])
    with conn() as c:
        for r in c.execute("SELECT * FROM demand_log ORDER BY times_asked DESC").fetchall():
            w.writerow([r["raw_name"], r["times_asked"], r["total_qty"],
                        len(json.loads(r["shops_json"] or "[]")), r["last_asked"], r["status"]])
    return Response(out.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=demand-not-stocked.csv"})


@app.route("/aliases")
@login_required
def aliases():
    with conn() as c:
        rows = c.execute("""SELECT a.*, i.name item_name, s.name shop_name
                            FROM item_aliases a
                            LEFT JOIN wholesale_items i ON i.id=a.item_id
                            LEFT JOIN retail_shops s ON s.id=a.shop_id
                            ORDER BY a.hits DESC, a.id DESC LIMIT 400""").fetchall()
    return render_template("aliases.html", rows=rows)


@app.route("/aliases/<int:aid>/delete", methods=["POST"])
@login_required
def alias_delete(aid):
    with conn() as c:
        c.execute("DELETE FROM item_aliases WHERE id=?", (aid,))
    flash("Alias removed — that name will be asked about again next time.", "ok")
    return redirect(url_for("aliases"))


# ══════════════════════════════════════════════════════════════════════
#  SHOP PORTAL — the retailer's own login
# ══════════════════════════════════════════════════════════════════════
def shop_required(f):
    @functools.wraps(f)
    def wrap(*a, **k):
        if not session.get("shop_id") or not session.get("tenant"):
            return redirect(url_for("shop_login"))
        return f(*a, **k)
    return wrap


def _shop_ctx():
    with conn() as c:
        shop = c.execute("SELECT * FROM retail_shops WHERE id=?", (session["shop_id"],)).fetchone()
    return shop


@app.route("/shop/login", methods=["GET", "POST"])
def shop_login():
    error = ""
    company = request.args.get("c") or session.get("tenant") or ""
    if request.method == "POST":
        company = (request.form.get("company") or "").strip().lower()
        code = (request.form.get("shop_code") or "").strip().upper()
        pin = (request.form.get("pin") or "").strip()

        import tenancy as T
        with T.platform_conn() as pc:
            comp = pc.execute("SELECT * FROM companies WHERE slug=? AND status='active'", (company,)).fetchone()
        if not comp:
            error = "Unknown or inactive company code."
        else:
            session["tenant"] = company           # bind DB before querying shops
            from flask import g
            g.tenant_slug, g.tenant_db = company, T.tenant_db_path(company)
            ensure_shop_login_columns()           # self-heal pre-hash tenant DBs
            with conn() as c:
                # status='active' is a SECURITY filter, not a tidiness one.
                # Rows exist here that must never be able to sign in:
                #   'pending'  — self-registered, not yet approved
                #   'rejected' — turned down
                #   'hold'     — auto-created by a public-catalog order
                # Matching on phone as well as code means an unapproved
                # applicant would otherwise log in with their own phone
                # number and set a PIN, skipping approval entirely.
                shop = c.execute(
                    "SELECT * FROM retail_shops WHERE (code=? OR phone=?) AND status='active'",
                    (code, code)).fetchone()
                row = c.execute("SELECT * FROM shop_logins WHERE shop_id=?",
                                (shop["id"],)).fetchone() if shop else None
            if not shop:
                session.pop("tenant", None)
                # Deliberately identical whether the shop is unknown or merely
                # not yet approved - don't confirm to a stranger that a given
                # phone number is registered with this distributor.
                error = "No active shop with that code. If you've just applied, wait for approval."
            elif _has_credential(row):
                if not verify_shop_pin(row, pin):
                    session.pop("tenant", None)
                    error = "Wrong password."
                else:
                    with conn() as c:
                        # Upgrade a legacy plaintext row to a hash the moment
                        # its owner proves they know the secret. Existing shops
                        # keep working; the plaintext disappears on first login
                        # rather than needing every retailer reset at once.
                        if row["pin"]:
                            salt, h = _hash_pin(pin)
                            c.execute("UPDATE shop_logins SET pin='', pin_salt=?, pin_hash=? WHERE shop_id=?",
                                      (salt, h, shop["id"]))
                        c.execute("UPDATE shop_logins SET last_login=datetime('now') WHERE shop_id=?",
                                  (shop["id"],))
                    session["shop_id"], session["shop_name"] = shop["id"], shop["name"]
                    return redirect(url_for("shop_home"))
            else:
                # First sign-in sets the password. Stored hashed from the
                # start - there is no code path that writes a new secret in
                # readable form.
                if len(pin) < 4:
                    session.pop("tenant", None)
                    error = "Choose a password of at least 4 characters for your first sign-in."
                else:
                    salt, h = _hash_pin(pin)
                    with conn() as c:
                        c.execute("""INSERT INTO shop_logins (shop_id, pin, pin_salt, pin_hash, last_login)
                                     VALUES (?,'',?,?,datetime('now'))
                                     ON CONFLICT(shop_id) DO UPDATE SET pin='',
                                       pin_salt=excluded.pin_salt, pin_hash=excluded.pin_hash,
                                       last_login=datetime('now')""", (shop["id"], salt, h))
                    session["shop_id"], session["shop_name"] = shop["id"], shop["name"]
                    return redirect(url_for("shop_home"))

    return render_template("shop_login.html", error=error, company=company)


@app.route("/shop/logout", methods=["POST"])
def shop_logout():
    session.clear()
    return redirect(url_for("shop_login"))


@app.route("/shop")
@shop_required
def shop_home():
    sid = session["shop_id"]
    with conn() as c:
        shop = c.execute("SELECT * FROM retail_shops WHERE id=?", (sid,)).fetchone()
        recent = c.execute("""SELECT invoice_no, invoice_date, total, status
                              FROM invoices WHERE shop_id=? ORDER BY id DESC LIMIT 8""", (sid,)).fetchall()
        outstanding = c.execute("""SELECT COALESCE(SUM(total-paid),0) o FROM invoices
                                   WHERE shop_id=? AND status!='paid'""", (sid,)).fetchone()["o"]
        overdue = c.execute("""SELECT COALESCE(SUM(total-paid),0) o FROM invoices
                               WHERE shop_id=? AND status!='paid' AND due_date<date('now')""", (sid,)).fetchone()["o"]
        open_orders = c.execute("""SELECT COUNT(*) n FROM sales_orders
                                   WHERE shop_id=? AND status IN ('draft','confirmed','dispatched')""",
                                (sid,)).fetchone()["n"]
        month_spend = c.execute("""SELECT COALESCE(SUM(total),0) t FROM invoices
                                   WHERE shop_id=? AND substr(invoice_date,1,7)=strftime('%Y-%m','now')""",
                                (sid,)).fetchone()["t"]
    return render_template("shop_home.html", shop=shop, recent=recent, outstanding=outstanding,
                           overdue=overdue, open_orders=open_orders, month_spend=month_spend)


@app.route("/shop/bills")
@shop_required
def shop_bills():
    with conn() as c:
        rows = c.execute("""SELECT * FROM invoices WHERE shop_id=? ORDER BY id DESC LIMIT 200""",
                         (session["shop_id"],)).fetchall()
    return render_template("shop_bills.html", rows=rows)


@app.route("/shop/bills/<int:inv_id>")
@shop_required
def shop_bill_detail(inv_id):
    with conn() as c:
        # shop_id=? in the WHERE, not just the lookup - a shop must never be
        # able to view another shop's invoice by guessing an id in the URL.
        inv = c.execute("SELECT * FROM invoices WHERE id=? AND shop_id=?",
                        (inv_id, session["shop_id"])).fetchone()
        if not inv:
            abort(404)
        lines = c.execute("SELECT * FROM sales_order_items WHERE order_id=?",
                          (inv["order_id"],)).fetchall() if inv["order_id"] else []

    upi_configured, upi_link = False, ""
    try:
        import upi
        upi_configured = upi.upi_configured()
        if upi_configured:
            upi_link = upi.build_upi_link(round((inv["total"] or 0) - (inv["paid"] or 0), 2), inv["invoice_no"])
    except ImportError:
        pass
    return render_template("shop_bill_detail.html", inv=inv, lines=lines,
                           upi_configured=upi_configured, upi_link=upi_link)


@app.route("/shop/orders")
@shop_required
def shop_orders():
    with conn() as c:
        rows = c.execute("""SELECT * FROM sales_orders WHERE shop_id=? ORDER BY id DESC LIMIT 200""",
                         (session["shop_id"],)).fetchall()
    return render_template("shop_orders.html", rows=rows)


@app.route("/shop/wanted", methods=["GET", "POST"])
@shop_required
def shop_wanted():
    """The shop uploads its own wanted list — same engine, self-service."""
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Choose a file first.", "err")
            return redirect(url_for("shop_wanted"))
        rows, err = parse_wanted_file(f)
        if err:
            flash(err, "err")
            return redirect(url_for("shop_wanted"))
        upload_id = process_wanted(session["shop_id"], rows, f.filename, source="shop")
        with conn() as c:
            up = c.execute("SELECT * FROM wanted_uploads WHERE id=?", (upload_id,)).fetchone()
        flash(f"Received {up['total_lines']} lines — {up['auto_matched']} matched automatically. "
              f"Review and place your order below.", "ok")
        return redirect(url_for("shop_wanted_review", upload_id=upload_id))

    with conn() as c:
        ups = c.execute("""SELECT * FROM wanted_uploads WHERE shop_id=? ORDER BY id DESC LIMIT 20""",
                        (session["shop_id"],)).fetchall()
    return render_template("shop_wanted.html", uploads=ups, have_xlsx=HAVE_XLSX)


@app.route("/shop/wanted/<int:upload_id>")
@shop_required
def shop_wanted_review(upload_id):
    """
    Self-serve review of a shop's own upload. The admin equivalent
    (wanted_review) exists for staff to resolve on a shop's behalf; this is
    the same matched-lines view, scoped to the shop's own session so they
    can confirm quantities and place the order themselves instead of
    waiting on a callback.
    """
    with conn() as c:
        up = c.execute("SELECT * FROM wanted_uploads WHERE id=? AND shop_id=?",
                       (upload_id, session["shop_id"])).fetchone()
        if not up:
            abort(404)
        lines = c.execute("SELECT * FROM wanted_lines WHERE upload_id=? ORDER BY id", (upload_id,)).fetchall()
        items = {r["id"]: dict(r) for r in c.execute(
            "SELECT id, name, pack_size, ptr, ptr_b, ptr_c, scheme, stock FROM wholesale_items").fetchall()}
    shop = _shop_ctx()

    parsed = []
    for ln in lines:
        d = dict(ln)
        d["suggestions"] = json.loads(ln["suggestions"] or "[]")
        d["item"] = items.get(ln["item_id"])
        if d["item"]:
            d["rate"] = rate_for_shop(d["item"], shop["price_tier"])
        for s in d["suggestions"]:
            si = items.get(s["id"])
            s["rate"] = rate_for_shop(si, shop["price_tier"]) if si else 0
        parsed.append(d)
    return render_template("shop_wanted_review.html", up=up, lines=parsed, shop=shop)


@app.route("/shop/wanted/<int:upload_id>/order", methods=["POST"])
@shop_required
def shop_wanted_order(upload_id):
    """
    Turn confirmed lines straight into a placed order - no staff round trip.
    Same order-building logic as wanted_create_order(), but scoped to the
    shop's own upload (re-checked here, not just trusted from the review
    page) and driven by whatever the shop actually ticked/edited rather
    than blindly re-using the original matched qty.
    """
    line_ids = request.form.getlist("line_id[]")
    item_choices = request.form.getlist("item_choice[]")
    qtys = request.form.getlist("qty[]")

    with conn() as c:
        up = c.execute("SELECT * FROM wanted_uploads WHERE id=? AND shop_id=?",
                       (upload_id, session["shop_id"])).fetchone()
        if not up:
            abort(404)
        shop = c.execute("SELECT * FROM retail_shops WHERE id=?", (session["shop_id"],)).fetchone()

        order_lines = []
        for lid, choice, q_str in zip(line_ids, item_choices, qtys):
            if not choice or choice == "skip":
                continue
            qty = int(q_str or 0)
            if qty <= 0:
                continue
            item = c.execute("SELECT * FROM wholesale_items WHERE id=?", (int(choice),)).fetchone()
            if not item:
                continue
            order_lines.append((int(lid), item, qty))

        if not order_lines:
            flash("Nothing selected — tick at least one item to order.", "err")
            return redirect(url_for("shop_wanted_review", upload_id=upload_id))

        order_no = next_order_no()
        cur = c.execute("""INSERT INTO sales_orders (order_no, shop_id, notes, source, created_by)
                           VALUES (?,?,?,?,?)""",
                        (order_no, session["shop_id"], f"From wanted list #{upload_id} ({up['filename']})",
                         "wanted", shop["name"]))
        order_id = cur.lastrowid

        for lid, item, qty in order_lines:
            rate = rate_for_shop(item, shop["price_tier"])
            free = 0
            if item["scheme"] and "+" in (item["scheme"] or ""):
                try:
                    b, bs = item["scheme"].split("+")
                    b, bs = int(b), int(bs)
                    if b > 0:
                        free = (qty // b) * bs
                except Exception:
                    pass
            c.execute("""INSERT INTO sales_order_items
                (order_id, item_id, item_name, pack_size, qty, free_qty, rate, gst_rate, amount)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (order_id, item["id"], item["name"], item["pack_size"], qty, free,
                 rate, item["gst_rate"], qty * rate))

            # Remember the shop's choice as an alias too, same as the staff
            # resolve flow - their next upload matches this line instantly.
            ln = c.execute("SELECT * FROM wanted_lines WHERE id=?", (lid,)).fetchone()
            if ln:
                c.execute("UPDATE wanted_lines SET item_id=?, match_type='confirmed', qty=?, confidence=1.0 WHERE id=?",
                          (item["id"], qty, lid))
                try:
                    c.execute("""INSERT INTO item_aliases (shop_id, alias_norm, alias_raw, item_id)
                                 VALUES (?,?,?,?)""",
                              (session["shop_id"], ln["norm_name"], ln["raw_name"], item["id"]))
                except Exception:
                    c.execute("""UPDATE item_aliases SET item_id=?, hits=hits+1
                                 WHERE shop_id=? AND alias_norm=?""",
                              (item["id"], session["shop_id"], ln["norm_name"]))

        c.execute("UPDATE wanted_uploads SET status='ordered', order_id=? WHERE id=?", (order_id, upload_id))

    compute_order_totals(order_id)
    audit(shop["name"], "shop_wanted_to_order", "order", order_id, order_no)
    flash(f"Order {order_no} placed — {len(order_lines)} item(s). We'll confirm shortly.", "ok")
    return redirect(url_for("shop_orders"))


# ══════════════════════════════════════════════════════════════════════
#  SHOP PORTAL — cart + checkout
#  The cart lives in the browser's localStorage (keyed per shop), not
#  server-side session state: it needs to survive across searches/category
#  changes without a round trip, and localStorage is naturally scoped per
#  browser so two shops on the same distributor never share a cart. The
#  server only ever sees the cart once, at checkout submit.
# ══════════════════════════════════════════════════════════════════════
@app.route("/shop/catalog")
@shop_required
def shop_catalog():
    q = (request.args.get("q") or "").strip()
    category = (request.args.get("category") or "").strip()
    shop = _shop_ctx()
    with conn() as c:
        cats = [r["category"] for r in c.execute(
            "SELECT DISTINCT category FROM wholesale_items WHERE category!='' ORDER BY category").fetchall()]
        if q:
            like = f"%{q}%"
            rows = c.execute(
                "SELECT * FROM wholesale_items WHERE status='active' AND (name LIKE ? OR generic LIKE ?) ORDER BY name LIMIT 200",
                (like, like)).fetchall()
        elif category:
            rows = c.execute(
                "SELECT * FROM wholesale_items WHERE status='active' AND category=? ORDER BY name LIMIT 200",
                (category,)).fetchall()
        else:
            rows = c.execute("SELECT * FROM wholesale_items WHERE status='active' ORDER BY name LIMIT 150").fetchall()

    items = []
    for r in rows:
        d = dict(r)
        d["rate"] = rate_for_shop(r, shop["price_tier"])
        items.append(d)

    return render_template("shop_catalog.html", items=items, categories=cats, q=q, category=category, shop=shop)


@app.route("/shop/cart")
@shop_required
def shop_cart():
    return render_template("shop_cart.html")


@app.route("/shop/cart/checkout", methods=["POST"])
@shop_required
def shop_cart_checkout():
    shop_id = session["shop_id"]
    item_ids = request.form.getlist("item_id[]")
    qtys = request.form.getlist("qty[]")
    notes = (request.form.get("notes") or "").strip()

    if not item_ids:
        flash("Your cart is empty.", "err")
        return redirect(url_for("shop_cart"))

    with conn() as c:
        shop = c.execute("SELECT * FROM retail_shops WHERE id=?", (shop_id,)).fetchone()
        order_no = next_order_no()
        cur = c.execute(
            "INSERT INTO sales_orders (order_no, shop_id, notes, source, created_by) VALUES (?,?,?,?,?)",
            (order_no, shop_id, notes, "shop_portal", session.get("shop_name") or "shop"))
        order_id = cur.lastrowid
        added = 0
        for iid, q_str in zip(item_ids, qtys):
            if not iid or not q_str:
                continue
            qty = int(float(q_str) or 0)
            if qty <= 0:
                continue
            item = c.execute("SELECT * FROM wholesale_items WHERE id=? AND status='active'", (int(iid),)).fetchone()
            if not item:
                continue
            rate = rate_for_shop(item, shop["price_tier"])
            free = 0
            if item["scheme"] and "+" in (item["scheme"] or ""):
                try:
                    b, bs = item["scheme"].split("+"); b, bs = int(b), int(bs)
                    if b > 0:
                        free = (qty // b) * bs
                except Exception:
                    pass
            c.execute("""INSERT INTO sales_order_items
                (order_id, item_id, item_name, pack_size, qty, free_qty, rate, gst_rate, amount)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (order_id, item["id"], item["name"], item["pack_size"], qty, free, rate, item["gst_rate"], qty * rate))
            added += 1

    if added == 0:
        with conn() as c:
            c.execute("DELETE FROM sales_orders WHERE id=?", (order_id,))
        flash("No valid items in your cart.", "err")
        return redirect(url_for("shop_cart"))

    compute_order_totals(order_id)
    audit(session.get("shop_name"), "shop_portal_order", "order", order_id, order_no)
    return render_template("shop_cart_thanks.html", order_no=order_no)


# ══════════════════════════════════════════════════════════════════════
#  RETAILER SELF-REGISTRATION
#
#  Until now a new retail shop could only exist if the distributor typed
#  it in, or if it stumbled through the public catalog (which auto-creates
#  a bare "[Pending] CODE" row with no contact details). Neither turns a
#  shop that finds the public site into a real customer.
#
#  This is the growth loop: the shop applies with its own details, the
#  distributor reviews and approves, the shop can then sign in and order.
#  Applications land as status='pending' - NOT 'active' - so an unapproved
#  shop can never log in or see trade prices. shop_login() already filters
#  logins to real shops; approval is what makes the row usable.
#
#  Tenant comes from the form's company code (validated against
#  platform.db), same as shop_login - a public page has no session to
#  read it from.
# ══════════════════════════════════════════════════════════════════════
@app.route("/shop/register", methods=["GET", "POST"])
def shop_register():
    company = (request.args.get("c") or "").strip().lower()
    error = ""
    if request.method == "POST":
        import tenancy as T
        d = request.form
        company = (d.get("company") or "").strip().lower()
        name = (d.get("name") or "").strip()
        phone = (d.get("phone") or "").strip()

        with T.platform_conn() as pc:
            comp = pc.execute("SELECT * FROM companies WHERE slug=? AND status='active'",
                              (company,)).fetchone()
        if not comp:
            error = "Unknown distributor code — check with your supplier."
        elif not name or not phone:
            error = "Shop name and phone number are required."
        else:
            from flask import g
            g.tenant_slug, g.tenant_db = company, T.tenant_db_path(company)
            with conn() as c:
                dupe = c.execute(
                    "SELECT id, status FROM retail_shops WHERE phone=? OR (name=? AND phone=?)",
                    (phone, name, phone)).fetchone()
                if dupe:
                    # Don't leak whether the existing row is active vs pending -
                    # just stop them creating a second one for the same shop.
                    return render_template("shop_register_done.html",
                                           company=company, existing=True,
                                           comp_name=comp["name"])
                cur = c.execute("""INSERT INTO retail_shops
                    (code, name, contact_person, phone, whatsapp, email, address, city,
                     pincode, gstin, drug_license, status, notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,'pending',?)""",
                    ("", name, (d.get("contact_person") or "").strip(), phone,
                     (d.get("whatsapp") or phone).strip(), (d.get("email") or "").strip(),
                     (d.get("address") or "").strip(), (d.get("city") or "").strip(),
                     (d.get("pincode") or "").strip(), (d.get("gstin") or "").strip(),
                     (d.get("drug_license") or "").strip(),
                     "Self-registered from the public site"))
                audit(name, "self_register", "shop", cur.lastrowid, phone)
            return render_template("shop_register_done.html",
                                   company=company, existing=False, comp_name=comp["name"])

    return render_template("shop_register.html", company=company, error=error)


@app.route("/shops/pending")
@login_required
def shops_pending():
    with conn() as c:
        # 'hold' as well as 'pending': a shop that ordered through the public
        # catalog without an account is auto-created as 'hold', and since
        # shop_login now (correctly) refuses anything but 'active', those
        # rows would be stranded with no way to be activated if this page
        # only listed self-registrations.
        rows = c.execute("""SELECT * FROM retail_shops WHERE status IN ('pending','hold')
                            ORDER BY id DESC""").fetchall()
        routes = c.execute("SELECT * FROM delivery_routes ORDER BY name").fetchall()
    return render_template("shops_pending.html", shops=rows, routes=routes)


@app.route("/shops/<int:shop_id>/approve", methods=["POST"])
@login_required
def shop_approve(shop_id):
    """
    Approve a self-registered shop: assign a real shop code (the login
    identifier), set commercial terms, activate. The code is what the
    retailer types at /shop/login, so it must exist and be unique - a
    pending row is created with an empty code precisely so an unapproved
    application can't be used to sign in.
    """
    d = request.form
    with conn() as c:
        shop = c.execute("SELECT * FROM retail_shops WHERE id=?", (shop_id,)).fetchone()
        if not shop:
            abort(404)
        if shop["status"] not in ("pending", "hold"):
            flash("That application has already been handled.", "err")
            return redirect(url_for("shops_pending"))

        code = (d.get("code") or "").strip().upper()
        if not code:
            n = c.execute("SELECT COUNT(*) FROM retail_shops").fetchone()[0]
            code = f"S{n + 1:04d}"
        if c.execute("SELECT 1 FROM retail_shops WHERE code=? AND id!=?", (code, shop_id)).fetchone():
            flash(f"Shop code '{code}' is already taken — pick another.", "err")
            return redirect(url_for("shops_pending"))

        route_id = d.get("route_id")
        c.execute("""UPDATE retail_shops SET code=?, price_tier=?, credit_limit=?, credit_days=?,
                     route_id=?, status='active' WHERE id=?""",
                  (code, d.get("price_tier") or "A", float(d.get("credit_limit") or 0),
                   int(d.get("credit_days") or 30), int(route_id) if route_id else None, shop_id))
    audit(session.get("ws_user"), "approve_shop", "shop", shop_id, f"{shop['name']} → {code}")

    # Tell them they're live and how to sign in. Best-effort: a missing
    # Twilio config must never block an approval that already succeeded.
    try:
        import whatsapp
        whatsapp.send_whatsapp(shop["whatsapp"] or shop["phone"],
                         f"Your account with us is approved. Sign in at "
                         f"{request.url_root.rstrip('/')}/shop/login with shop code {code} "
                         f"and set your own PIN on first sign-in.")
    except Exception:
        pass

    flash(f"{shop['name']} approved — shop code {code}. They can now sign in and order.", "ok")
    return redirect(url_for("shops_pending"))


@app.route("/shops/<int:shop_id>/reject", methods=["POST"])
@login_required
def shop_reject(shop_id):
    with conn() as c:
        shop = c.execute("SELECT * FROM retail_shops WHERE id=?", (shop_id,)).fetchone()
        if not shop:
            abort(404)
        # Keep the row (status='rejected') rather than deleting: the duplicate
        # check on re-registration relies on it, so a rejected applicant can't
        # simply re-apply in a loop.
        c.execute("UPDATE retail_shops SET status='rejected' WHERE id=?", (shop_id,))
    audit(session.get("ws_user"), "reject_shop", "shop", shop_id, shop["name"])
    flash(f"{shop['name']} rejected.", "ok")
    return redirect(url_for("shops_pending"))


@app.route("/shops/<int:shop_id>/resetpin", methods=["POST"])
@login_required
def shop_reset_pin(shop_id):
    """
    Generate a fresh password for a retailer and show it ONCE.

    Shown once and never again because it is only stored hashed - there is
    no query that can recover it later, for the distributor or for us. If
    the shop loses it, it gets reset again; that's the intended trade.
    """
    with conn() as c:
        shop = c.execute("SELECT id, name FROM retail_shops WHERE id=?", (shop_id,)).fetchone()
    if not shop:
        abort(404)
    new_pin = secrets.token_hex(3).upper()   # 6 hex chars — readable over the phone
    set_shop_pin(shop_id, new_pin)
    audit(session.get("ws_user"), "reset_shop_pin", "shop", shop_id, shop["name"])
    flash(f"New password for {shop['name']}: {new_pin} — send it to them now, "
          f"it is stored hashed and cannot be shown again.", "ok")
    return redirect(request.referrer or url_for("shops"))
