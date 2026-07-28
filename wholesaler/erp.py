"""
MediVision Wholesale — ERP extensions
=====================================

Bolts on the organisation-side modules the wholesaler needs to run
their own business (not just sell to shops):

    - Staff master
    - Daily attendance
    - Monthly payroll (base + OT + advances − deductions)
    - Task assignments (recurring templates → daily task instances)
    - Supplier master
    - Purchase orders (goods received from manufacturers) → stock inward
    - Expense categories + entries + monthly summary
    - Custom fields on shops and items (owner-defined, key/value)
    - Feature toggles (owner can enable/disable modules from Settings)
    - Missing JSON API endpoints used by the front-end

Imported by wsgi.py AFTER `from app import app`, so side-effect
registration works. Idempotent: safe to run schema init many times.
"""
from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from flask import request, redirect, url_for, render_template, session, jsonify, flash, abort

from app import (app, conn, login_required, audit, money,
                 rate_for_shop, session as _s)  # noqa: F401


# ═════════════════════════════════════════════════════════════════════
#  ADDITIONAL SCHEMA
# ═════════════════════════════════════════════════════════════════════
EXTRA_SCHEMA = r"""
-- ── Staff / HR ────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS staff (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT NOT NULL,
    role            TEXT DEFAULT '',
    phone           TEXT DEFAULT '',
    aadhar          TEXT DEFAULT '',
    join_date       TEXT DEFAULT (date('now')),
    base_salary     REAL DEFAULT 0,
    ot_rate_per_hr  REAL DEFAULT 0,
    status          TEXT DEFAULT 'active',
    notes           TEXT DEFAULT '',
    created_at      TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS attendance (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    staff_id        INTEGER NOT NULL,
    day             TEXT NOT NULL,
    status          TEXT DEFAULT 'present',   -- present / absent / half / leave
    hours           REAL DEFAULT 8,
    ot_hours        REAL DEFAULT 0,
    notes           TEXT DEFAULT '',
    UNIQUE(staff_id, day),
    FOREIGN KEY(staff_id) REFERENCES staff(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS advances (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    staff_id        INTEGER NOT NULL,
    amount          REAL DEFAULT 0,
    given_on        TEXT DEFAULT (date('now')),
    period_ym       TEXT,                     -- month it should deduct from
    notes           TEXT DEFAULT '',
    FOREIGN KEY(staff_id) REFERENCES staff(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS payslips (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    staff_id        INTEGER NOT NULL,
    period_ym       TEXT NOT NULL,             -- 2026-07
    days_present    REAL DEFAULT 0,
    days_absent     REAL DEFAULT 0,
    ot_hours        REAL DEFAULT 0,
    base            REAL DEFAULT 0,
    ot_amount       REAL DEFAULT 0,
    advances        REAL DEFAULT 0,
    deductions      REAL DEFAULT 0,
    net             REAL DEFAULT 0,
    generated_at    TEXT DEFAULT (datetime('now')),
    paid            INTEGER DEFAULT 0,
    UNIQUE(staff_id, period_ym),
    FOREIGN KEY(staff_id) REFERENCES staff(id) ON DELETE CASCADE
);

-- ── Purchases from manufacturers ──────────────────────────────────
CREATE TABLE IF NOT EXISTS suppliers (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    code            TEXT UNIQUE,
    name            TEXT NOT NULL,
    contact_person  TEXT DEFAULT '',
    phone           TEXT DEFAULT '',
    email           TEXT DEFAULT '',
    address         TEXT DEFAULT '',
    gstin           TEXT DEFAULT '',
    dl              TEXT DEFAULT '',
    payment_terms   TEXT DEFAULT '30 days',
    notes           TEXT DEFAULT '',
    created_at      TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS purchase_orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    po_no           TEXT UNIQUE,
    supplier_id     INTEGER NOT NULL,
    po_date         TEXT DEFAULT (date('now')),
    status          TEXT DEFAULT 'draft',       -- draft / sent / received / partial / cancelled
    subtotal        REAL DEFAULT 0,
    gst_amount      REAL DEFAULT 0,
    total           REAL DEFAULT 0,
    notes           TEXT DEFAULT '',
    created_at      TEXT DEFAULT (datetime('now')),
    received_at     TEXT,
    FOREIGN KEY(supplier_id) REFERENCES suppliers(id)
);

CREATE TABLE IF NOT EXISTS purchase_order_items (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    po_id           INTEGER NOT NULL,
    item_id         INTEGER,
    item_name       TEXT,
    qty             INTEGER DEFAULT 0,
    received_qty    INTEGER DEFAULT 0,
    rate            REAL DEFAULT 0,
    gst_rate        REAL DEFAULT 0,
    amount          REAL DEFAULT 0,
    batch           TEXT DEFAULT '',
    expiry          TEXT DEFAULT '',
    FOREIGN KEY(po_id) REFERENCES purchase_orders(id) ON DELETE CASCADE
);

-- ── Expenses (opex) ───────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS expenses (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    spent_on        TEXT DEFAULT (date('now')),
    category        TEXT DEFAULT 'other',       -- rent / salary / transport / utilities / fuel / repair / other
    description     TEXT DEFAULT '',
    amount          REAL DEFAULT 0,
    payment_method  TEXT DEFAULT 'cash',
    reference       TEXT DEFAULT '',
    entered_by      TEXT DEFAULT '',
    created_at      TEXT DEFAULT (datetime('now'))
);

-- ── Customization: settings & feature flags & custom fields ──────
CREATE TABLE IF NOT EXISTS settings (
    key             TEXT PRIMARY KEY,
    value           TEXT DEFAULT '',
    updated_at      TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS custom_fields (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    entity          TEXT NOT NULL,              -- 'shop' | 'item'
    field_key       TEXT NOT NULL,              -- e.g. "chief_deity"
    field_label     TEXT NOT NULL,              -- shown in UI
    field_type      TEXT DEFAULT 'text',        -- text / number / date
    sort_order      INTEGER DEFAULT 0,
    UNIQUE(entity, field_key)
);

CREATE TABLE IF NOT EXISTS custom_field_values (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    entity          TEXT NOT NULL,
    entity_id       INTEGER NOT NULL,
    field_key       TEXT NOT NULL,
    value           TEXT DEFAULT '',
    UNIQUE(entity, entity_id, field_key)
);
"""

with conn() as _c:
    _c.executescript(EXTRA_SCHEMA)


# ═════════════════════════════════════════════════════════════════════
#  SETTINGS + FEATURE TOGGLES
# ═════════════════════════════════════════════════════════════════════
DEFAULT_FEATURES = {
    "staff":     True,
    "purchases": True,
    "expenses":  True,
    "custom":    True,
    "reports":   True,
    "catalog":   True,
    "whatsapp":  False,      # off until Twilio wired
}


def setting_get(key, default=""):
    with conn() as c:
        r = c.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return r["value"] if r else default


def setting_set(key, value):
    with conn() as c:
        c.execute("""INSERT INTO settings (key, value) VALUES (?, ?)
                     ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now')""",
                  (key, value))


def features():
    """Return the current feature-flag dict, merging DB overrides on top of defaults."""
    out = dict(DEFAULT_FEATURES)
    with conn() as c:
        rows = c.execute("SELECT key, value FROM settings WHERE key LIKE 'feature.%'").fetchall()
    for r in rows:
        name = r["key"].split(".", 1)[1]
        if name in out:
            out[name] = r["value"] in ("1", "true", "on", "yes")
    return out


@app.context_processor
def _inject_features():
    return {"features": features(), "wholesale_company": setting_get("company.name", "MediVision Wholesale")}


# ═════════════════════════════════════════════════════════════════════
#  MISSING API ENDPOINTS
# ═════════════════════════════════════════════════════════════════════
@app.route("/api/items")
def api_items():
    limit = min(int(request.args.get("limit") or 500), 2000)
    q = (request.args.get("q") or "").strip()
    with conn() as c:
        if q:
            like = f"%{q}%"
            rows = c.execute(
                "SELECT id, code, name, generic, manufacturer, pack_size, mrp, ptr, ptr_b, ptr_c, gst_rate, scheme, stock FROM wholesale_items WHERE status='active' AND (name LIKE ? OR generic LIKE ? OR code LIKE ?) ORDER BY name LIMIT ?",
                (like, like, like, limit),
            ).fetchall()
        else:
            rows = c.execute(
                "SELECT id, code, name, generic, manufacturer, pack_size, mrp, ptr, ptr_b, ptr_c, gst_rate, scheme, stock FROM wholesale_items WHERE status='active' ORDER BY name LIMIT ?",
                (limit,),
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/shops")
def api_shops():
    q = (request.args.get("q") or "").strip()
    with conn() as c:
        if q:
            like = f"%{q}%"
            rows = c.execute(
                "SELECT id, code, name, phone, city, price_tier FROM retail_shops WHERE status='active' AND (name LIKE ? OR code LIKE ? OR phone LIKE ?) ORDER BY name LIMIT 200",
                (like, like, like),
            ).fetchall()
        else:
            rows = c.execute("SELECT id, code, name, phone, city, price_tier FROM retail_shops WHERE status='active' ORDER BY name LIMIT 200").fetchall()
    return jsonify([dict(r) for r in rows])


# ═════════════════════════════════════════════════════════════════════
#  STAFF
# ═════════════════════════════════════════════════════════════════════
@app.route("/staff")
@login_required
def staff_list():
    with conn() as c:
        rows = c.execute("SELECT * FROM staff ORDER BY status, name").fetchall()
    return render_template("staff.html", staff=rows)


@app.route("/staff/save", methods=["POST"])
@login_required
def staff_save():
    d = request.form
    sid = d.get("id")
    vals = {
        "name": (d.get("name") or "").strip(),
        "role": (d.get("role") or "").strip(),
        "phone": (d.get("phone") or "").strip(),
        "aadhar": (d.get("aadhar") or "").strip(),
        "join_date": (d.get("join_date") or date.today().isoformat()).strip(),
        "base_salary": float(d.get("base_salary") or 0),
        "ot_rate_per_hr": float(d.get("ot_rate_per_hr") or 0),
        "status": (d.get("status") or "active").strip(),
        "notes": (d.get("notes") or "").strip(),
    }
    if not vals["name"]:
        flash("Staff name required.", "err")
        return redirect(url_for("staff_list"))
    with conn() as c:
        if sid:
            sets = ", ".join([f"{k}=?" for k in vals.keys()])
            c.execute(f"UPDATE staff SET {sets} WHERE id=?", (*vals.values(), sid))
        else:
            cols = ", ".join(vals.keys())
            qm = ", ".join("?" * len(vals))
            c.execute(f"INSERT INTO staff ({cols}) VALUES ({qm})", tuple(vals.values()))
    audit(session.get("ws_user"), "save", "staff", None, vals["name"])
    flash("Staff saved.", "ok")
    return redirect(url_for("staff_list"))


@app.route("/attendance", methods=["GET", "POST"])
@login_required
def attendance():
    day = request.args.get("day") or date.today().isoformat()

    if request.method == "POST":
        day = request.form.get("day") or date.today().isoformat()
        staff_ids = request.form.getlist("staff_id[]")
        statuses = request.form.getlist("status[]")
        ots = request.form.getlist("ot_hours[]")
        with conn() as c:
            for sid, st, ot in zip(staff_ids, statuses, ots):
                hours = {"present": 8, "half": 4}.get(st, 0)
                c.execute("""INSERT INTO attendance (staff_id, day, status, hours, ot_hours)
                             VALUES (?,?,?,?,?)
                             ON CONFLICT(staff_id, day) DO UPDATE SET
                               status=excluded.status, hours=excluded.hours, ot_hours=excluded.ot_hours""",
                          (int(sid), day, st, hours, float(ot or 0)))
        flash(f"Attendance saved for {day}.", "ok")
        return redirect(url_for("attendance", day=day))

    with conn() as c:
        staff = c.execute("SELECT * FROM staff WHERE status='active' ORDER BY name").fetchall()
        att = {r["staff_id"]: dict(r) for r in c.execute("SELECT * FROM attendance WHERE day=?", (day,)).fetchall()}
    return render_template("attendance.html", staff=staff, att=att, day=day)


@app.route("/payroll")
@login_required
def payroll():
    period = request.args.get("period") or date.today().strftime("%Y-%m")
    year, month = period.split("-")
    # count working days in month = days that are not sundays (simple heuristic; can be overridden)
    from calendar import monthrange
    dim = monthrange(int(year), int(month))[1]
    working_days = sum(1 for d in range(1, dim + 1) if date(int(year), int(month), d).weekday() != 6)

    rows = []
    with conn() as c:
        staff = c.execute("SELECT * FROM staff WHERE status='active'").fetchall()
        for s in staff:
            att = c.execute("SELECT status, ot_hours FROM attendance WHERE staff_id=? AND substr(day,1,7)=?",
                            (s["id"], period)).fetchall()
            present = sum(1 for a in att if a["status"] == "present")
            half = sum(0.5 for a in att if a["status"] == "half")
            absent = sum(1 for a in att if a["status"] == "absent")
            ot_hours = sum(a["ot_hours"] or 0 for a in att)
            per_day = (s["base_salary"] or 0) / max(working_days, 1)
            base = round(per_day * (present + half), 2)
            ot_amount = round(ot_hours * (s["ot_rate_per_hr"] or 0), 2)
            adv_row = c.execute("SELECT COALESCE(SUM(amount),0) a FROM advances WHERE staff_id=? AND (period_ym=? OR (period_ym IS NULL AND substr(given_on,1,7)=?))",
                                (s["id"], period, period)).fetchone()
            advances = adv_row["a"]
            deductions = 0
            net = round(base + ot_amount - advances - deductions, 2)
            rows.append({
                "staff": dict(s),
                "days_present": present + half,
                "days_absent": absent,
                "ot_hours": ot_hours,
                "base": base,
                "ot_amount": ot_amount,
                "advances": advances,
                "deductions": deductions,
                "net": net,
            })

    totals = {
        "base": sum(r["base"] for r in rows),
        "ot_amount": sum(r["ot_amount"] for r in rows),
        "advances": sum(r["advances"] for r in rows),
        "net": sum(r["net"] for r in rows),
    }
    return render_template("payroll.html", rows=rows, period=period, working_days=working_days, totals=totals)


@app.route("/payroll/advance", methods=["POST"])
@login_required
def payroll_advance():
    staff_id = int(request.form.get("staff_id") or 0)
    amount = float(request.form.get("amount") or 0)
    period = (request.form.get("period") or "").strip() or None
    notes = (request.form.get("notes") or "").strip()
    if not staff_id or amount <= 0:
        flash("Staff and amount required.", "err")
    else:
        with conn() as c:
            c.execute("INSERT INTO advances (staff_id, amount, period_ym, notes) VALUES (?,?,?,?)",
                      (staff_id, amount, period, notes))
        flash("Advance recorded.", "ok")
    return redirect(url_for("payroll", period=period or date.today().strftime("%Y-%m")))


# ═════════════════════════════════════════════════════════════════════
#  SUPPLIERS + PURCHASES
# ═════════════════════════════════════════════════════════════════════
@app.route("/suppliers")
@login_required
def suppliers():
    with conn() as c:
        rows = c.execute("SELECT * FROM suppliers ORDER BY name").fetchall()
    return render_template("suppliers.html", suppliers=rows)


@app.route("/suppliers/save", methods=["POST"])
@login_required
def supplier_save():
    d = request.form
    sid = d.get("id")
    vals = {k: (d.get(k) or "").strip() for k in
            ("code", "name", "contact_person", "phone", "email", "address", "gstin", "dl", "payment_terms", "notes")}
    if not vals["name"]:
        flash("Supplier name required.", "err")
        return redirect(url_for("suppliers"))
    if not vals["code"]:
        with conn() as c:
            n = c.execute("SELECT COUNT(*) FROM suppliers").fetchone()[0]
        vals["code"] = f"SUP{n + 1:03d}"
    with conn() as c:
        if sid:
            sets = ", ".join([f"{k}=?" for k in vals.keys()])
            c.execute(f"UPDATE suppliers SET {sets} WHERE id=?", (*vals.values(), sid))
        else:
            cols = ", ".join(vals.keys())
            qm = ", ".join("?" * len(vals))
            c.execute(f"INSERT INTO suppliers ({cols}) VALUES ({qm})", tuple(vals.values()))
    flash("Supplier saved.", "ok")
    return redirect(url_for("suppliers"))


@app.route("/purchases")
@login_required
def purchases():
    with conn() as c:
        rows = c.execute("""
            SELECT p.*, s.name supplier_name, s.code supplier_code
            FROM purchase_orders p JOIN suppliers s ON s.id=p.supplier_id
            ORDER BY p.id DESC LIMIT 100
        """).fetchall()
    return render_template("purchases.html", purchases=rows)


@app.route("/purchases/new", methods=["GET", "POST"])
@login_required
def purchase_new():
    with conn() as c:
        suppliers_rows = c.execute("SELECT id, code, name FROM suppliers ORDER BY name").fetchall()

    if request.method == "POST":
        supplier_id = int(request.form.get("supplier_id") or 0)
        names = request.form.getlist("item_name[]")
        qtys = request.form.getlist("qty[]")
        rates = request.form.getlist("rate[]")
        gsts = request.form.getlist("gst[]")
        item_ids = request.form.getlist("item_id[]")
        notes = request.form.get("notes", "").strip()

        if not supplier_id or not names:
            flash("Supplier and at least one item required.", "err")
            return redirect(url_for("purchase_new"))

        with conn() as c:
            n = c.execute("SELECT COUNT(*) FROM purchase_orders").fetchone()[0]
            po_no = f"PO/{date.today().strftime('%y%m')}/{n + 1:04d}"
            cur = c.execute(
                "INSERT INTO purchase_orders (po_no, supplier_id, notes) VALUES (?,?,?)",
                (po_no, supplier_id, notes),
            )
            po_id = cur.lastrowid
            subtotal = 0
            gst_amount = 0
            for nm, q_str, r_str, g_str, iid in zip(names, qtys, rates, gsts, item_ids):
                nm = (nm or "").strip()
                if not nm:
                    continue
                q = int(q_str or 0)
                r = float(r_str or 0)
                g = float(g_str or 12)
                if q <= 0:
                    continue
                amount = q * r
                subtotal += amount
                gst_amount += amount * g / 100
                c.execute("""INSERT INTO purchase_order_items
                    (po_id, item_id, item_name, qty, rate, gst_rate, amount)
                    VALUES (?,?,?,?,?,?,?)""",
                    (po_id, int(iid) if iid else None, nm, q, r, g, amount))
            c.execute("UPDATE purchase_orders SET subtotal=?, gst_amount=?, total=? WHERE id=?",
                      (subtotal, gst_amount, subtotal + gst_amount, po_id))
        audit(session.get("ws_user"), "create", "purchase_order", po_id, po_no)
        flash(f"Purchase order {po_no} created.", "ok")
        return redirect(url_for("purchase_detail", po_id=po_id))

    return render_template("purchase_new.html", suppliers=suppliers_rows)


@app.route("/purchases/<int:po_id>")
@login_required
def purchase_detail(po_id):
    with conn() as c:
        po = c.execute("""SELECT p.*, s.name supplier_name, s.code supplier_code, s.gstin supplier_gstin
                          FROM purchase_orders p JOIN suppliers s ON s.id=p.supplier_id
                          WHERE p.id=?""", (po_id,)).fetchone()
        if not po:
            abort(404)
        lines = c.execute("SELECT * FROM purchase_order_items WHERE po_id=?", (po_id,)).fetchall()
    return render_template("purchase_detail.html", po=po, lines=lines)


@app.route("/purchases/<int:po_id>/receive", methods=["POST"])
@login_required
def purchase_receive(po_id):
    with conn() as c:
        po = c.execute("SELECT * FROM purchase_orders WHERE id=?", (po_id,)).fetchone()
        if not po:
            abort(404)
        # simple: mark all received & bump stock
        lines = c.execute("SELECT * FROM purchase_order_items WHERE po_id=?", (po_id,)).fetchall()
        for ln in lines:
            recv = ln["qty"] or 0
            c.execute("UPDATE purchase_order_items SET received_qty=? WHERE id=?", (recv, ln["id"]))
            if ln["item_id"]:
                c.execute("UPDATE wholesale_items SET stock=stock+? WHERE id=?", (recv, ln["item_id"]))
        c.execute("UPDATE purchase_orders SET status='received', received_at=? WHERE id=?",
                  (datetime.now().isoformat(timespec="seconds"), po_id))
    audit(session.get("ws_user"), "receive", "purchase_order", po_id, po["po_no"])
    flash("Goods received, stock updated.", "ok")
    return redirect(url_for("purchase_detail", po_id=po_id))


# ═════════════════════════════════════════════════════════════════════
#  EXPENSES
# ═════════════════════════════════════════════════════════════════════
@app.route("/expenses")
@login_required
def expenses():
    period = request.args.get("period") or date.today().strftime("%Y-%m")
    with conn() as c:
        rows = c.execute(
            "SELECT * FROM expenses WHERE substr(spent_on,1,7)=? ORDER BY spent_on DESC, id DESC",
            (period,),
        ).fetchall()
        by_cat = c.execute("""SELECT category, COALESCE(SUM(amount),0) total
                              FROM expenses WHERE substr(spent_on,1,7)=? GROUP BY category ORDER BY total DESC""",
                           (period,)).fetchall()
    total = sum(r["amount"] for r in rows)
    return render_template("expenses.html", expenses=rows, by_cat=by_cat, total=total, period=period)


@app.route("/expenses/save", methods=["POST"])
@login_required
def expense_save():
    d = request.form
    with conn() as c:
        c.execute("""INSERT INTO expenses
            (spent_on, category, description, amount, payment_method, reference, entered_by)
            VALUES (?,?,?,?,?,?,?)""",
            (d.get("spent_on") or date.today().isoformat(),
             (d.get("category") or "other").strip(),
             (d.get("description") or "").strip(),
             float(d.get("amount") or 0),
             (d.get("payment_method") or "cash").strip(),
             (d.get("reference") or "").strip(),
             session.get("ws_user") or "admin"))
    flash("Expense recorded.", "ok")
    return redirect(url_for("expenses"))


# ═════════════════════════════════════════════════════════════════════
#  REPORTS
# ═════════════════════════════════════════════════════════════════════
@app.route("/reports")
@login_required
def reports():
    period = request.args.get("period") or date.today().strftime("%Y-%m")
    with conn() as c:
        sales_by_item = c.execute("""
            SELECT soi.item_name, SUM(soi.qty) qty, SUM(soi.amount) revenue
            FROM sales_order_items soi
            JOIN sales_orders o ON o.id=soi.order_id
            WHERE o.status IN ('invoiced','paid') AND substr(o.order_date,1,7)=?
            GROUP BY soi.item_name ORDER BY revenue DESC LIMIT 30
        """, (period,)).fetchall()

        sales_by_shop = c.execute("""
            SELECT s.name, SUM(i.total) revenue, COUNT(i.id) invoices
            FROM invoices i JOIN retail_shops s ON s.id=i.shop_id
            WHERE substr(i.invoice_date,1,7)=?
            GROUP BY s.id ORDER BY revenue DESC LIMIT 30
        """, (period,)).fetchall()

        # Aging: how much outstanding falls into 0-30 / 31-60 / 61-90 / 90+ days
        today = date.today().isoformat()
        aging = c.execute(f"""
            SELECT
              SUM(CASE WHEN julianday('{today}') - julianday(due_date) <= 0 THEN total-paid ELSE 0 END) current,
              SUM(CASE WHEN julianday('{today}') - julianday(due_date) BETWEEN 1 AND 30 THEN total-paid ELSE 0 END) d30,
              SUM(CASE WHEN julianday('{today}') - julianday(due_date) BETWEEN 31 AND 60 THEN total-paid ELSE 0 END) d60,
              SUM(CASE WHEN julianday('{today}') - julianday(due_date) BETWEEN 61 AND 90 THEN total-paid ELSE 0 END) d90,
              SUM(CASE WHEN julianday('{today}') - julianday(due_date) > 90 THEN total-paid ELSE 0 END) d90plus
            FROM invoices WHERE status!='paid'
        """).fetchone()

        gst_out = c.execute("""SELECT COALESCE(SUM(gst_amount),0) g FROM sales_orders
                               WHERE status IN ('invoiced','paid') AND substr(order_date,1,7)=?""",
                            (period,)).fetchone()["g"]
        gst_in = c.execute("""SELECT COALESCE(SUM(gst_amount),0) g FROM purchase_orders
                              WHERE status='received' AND substr(po_date,1,7)=?""",
                           (period,)).fetchone()["g"]

        exp_total = c.execute("""SELECT COALESCE(SUM(amount),0) t FROM expenses WHERE substr(spent_on,1,7)=?""",
                              (period,)).fetchone()["t"]
        rev_total = c.execute("""SELECT COALESCE(SUM(total),0) t FROM invoices WHERE substr(invoice_date,1,7)=?""",
                              (period,)).fetchone()["t"]

    return render_template("reports.html",
                           period=period,
                           sales_by_item=sales_by_item,
                           sales_by_shop=sales_by_shop,
                           aging=dict(aging) if aging else {},
                           gst_out=gst_out, gst_in=gst_in, gst_net=gst_out - gst_in,
                           exp_total=exp_total, rev_total=rev_total,
                           profit=rev_total - exp_total)


# ═════════════════════════════════════════════════════════════════════
#  CUSTOMIZATION (settings v2)
# ═════════════════════════════════════════════════════════════════════
@app.route("/customize", methods=["GET", "POST"])
@login_required
def customize():
    if request.method == "POST":
        # Feature toggles
        for name in DEFAULT_FEATURES.keys():
            setting_set(f"feature.{name}", "1" if request.form.get(f"feature.{name}") else "0")
        # Company details (persisted in DB in addition to .env; DB wins if set)
        for k in ("company.name", "company.gstin", "company.address", "company.phone",
                  "company.upi", "invoice.terms", "invoice.bank"):
            v = request.form.get(k)
            if v is not None:
                setting_set(k, v.strip())
        flash("Customization saved.", "ok")
        return redirect(url_for("customize"))

    with conn() as c:
        cf_shop = c.execute("SELECT * FROM custom_fields WHERE entity='shop' ORDER BY sort_order, id").fetchall()
        cf_item = c.execute("SELECT * FROM custom_fields WHERE entity='item' ORDER BY sort_order, id").fetchall()
        all_settings = {r["key"]: r["value"] for r in c.execute("SELECT key, value FROM settings").fetchall()}
    try:
        import whatsapp
        wa_configured = whatsapp.whatsapp_configured()
    except ImportError:
        wa_configured = False
    return render_template("customize.html",
                           wa_configured=wa_configured,
                           features=features(),
                           feature_defaults=DEFAULT_FEATURES,
                           cf_shop=cf_shop, cf_item=cf_item,
                           settings=all_settings)


@app.route("/customize/field/add", methods=["POST"])
@login_required
def custom_field_add():
    entity = request.form.get("entity", "shop")
    key = (request.form.get("key") or "").strip().lower().replace(" ", "_")
    label = (request.form.get("label") or "").strip()
    ftype = request.form.get("type") or "text"
    if not key or not label:
        flash("Key and label required.", "err")
        return redirect(url_for("customize"))
    with conn() as c:
        try:
            c.execute("INSERT INTO custom_fields (entity, field_key, field_label, field_type) VALUES (?,?,?,?)",
                      (entity, key, label, ftype))
        except Exception as e:
            flash(f"Could not add: {e}", "err")
            return redirect(url_for("customize"))
    flash("Custom field added.", "ok")
    return redirect(url_for("customize"))


@app.route("/customize/field/<int:fid>/delete", methods=["POST"])
@login_required
def custom_field_del(fid):
    with conn() as c:
        c.execute("DELETE FROM custom_fields WHERE id=?", (fid,))
    flash("Field removed.", "ok")
    return redirect(url_for("customize"))


# ═════════════════════════════════════════════════════════════════════
#  ACTIVITY LOG — "who did what"
#
#  audit() has been writing to audit_log since day one (every order
#  confirm/dispatch/invoice, every payment, every staff/user change), but
#  there was no way to READ it. For a distributor with a salesman and an
#  accountant sharing the system, "who confirmed this order?" is a daily
#  question. Owner-only: it shows every user's actions, which is
#  management information, not something a salesman should browse.
# ═════════════════════════════════════════════════════════════════════
@app.route("/activity")
@login_required
def activity_log():
    actor = (request.args.get("actor") or "").strip()
    entity = (request.args.get("entity") or "").strip()
    with conn() as c:
        q = "SELECT * FROM audit_log"
        where, params = [], []
        if actor:
            where.append("actor=?")
            params.append(actor)
        if entity:
            where.append("entity=?")
            params.append(entity)
        if where:
            q += " WHERE " + " AND ".join(where)
        q += " ORDER BY id DESC LIMIT 300"
        rows = c.execute(q, params).fetchall()
        actors = [r["actor"] for r in c.execute(
            "SELECT DISTINCT actor FROM audit_log WHERE actor IS NOT NULL AND actor!='' ORDER BY actor").fetchall()]
        entities = [r["entity"] for r in c.execute(
            "SELECT DISTINCT entity FROM audit_log WHERE entity IS NOT NULL AND entity!='' ORDER BY entity").fetchall()]
    return render_template("activity.html", rows=rows, actors=actors, entities=entities,
                           actor=actor, entity=entity)


# ═════════════════════════════════════════════════════════════════════
#  BULK STOCK IMPORT
#
#  Adding a catalog one item at a time is fine for a demo and unusable
#  for a real distributor with thousands of SKUs. Reuses the same
#  forgiving column-detection idea as the wanted-list parser, but writes
#  to wholesale_items. Matches on `code` when present, else on exact
#  name, so re-uploading a supplier's updated price list UPDATES rows
#  rather than creating duplicates.
# ═════════════════════════════════════════════════════════════════════
_IMPORT_FIELDS = {
    "code":         ("code", "sku", "item code", "product code"),
    "name":         ("name", "product", "item", "medicine", "description", "particular"),
    "generic":      ("generic", "molecule", "composition", "salt"),
    "manufacturer": ("manufacturer", "company", "mfr", "make", "brand"),
    "pack_size":    ("pack", "packing", "pack size", "packsize"),
    "mrp":          ("mrp", "m.r.p"),
    "ptr":          ("ptr", "rate", "price", "purchase rate", "net rate"),
    "ptr_b":        ("ptr_b", "ptr b", "rate b", "tier b"),
    "ptr_c":        ("ptr_c", "ptr c", "rate c", "tier c"),
    "gst_rate":     ("gst", "gst%", "gst rate", "tax"),
    "hsn":          ("hsn", "hsn code"),
    "scheme":       ("scheme", "offer", "deal"),
    "stock":        ("stock", "qty", "quantity", "closing", "balance"),
    "reorder_level":("reorder", "reorder level", "min qty", "minimum"),
    "category":     ("category", "type", "group"),
    "batch":        ("batch", "batch no", "lot"),
    "expiry":       ("expiry", "exp", "expdt", "expiry date"),
}
_NUMERIC_FIELDS = {"mrp", "ptr", "ptr_b", "ptr_c", "gst_rate", "stock", "reorder_level"}


def _parse_stock_file(fs):
    """Read an .xlsx/.csv into a list of dicts keyed by wholesale_items columns."""
    import io, csv as _csv, re as _re
    fname = (fs.filename or "").lower()
    raw = fs.read()

    if fname.endswith(".csv"):
        table = list(_csv.reader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))))
    elif fname.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return [], "openpyxl isn't installed — save the file as CSV and retry."
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        table = [[("" if v is None else str(v).strip()) for v in r]
                 for r in wb.active.iter_rows(values_only=True)]
    else:
        return [], "Please upload a .xlsx or .csv file."

    table = [r for r in table if any(str(x).strip() for x in r)]
    if len(table) < 2:
        return [], "That file has no data rows under its header."

    header = [str(h).lower().strip() for h in table[0]]
    colmap = {}
    for idx, h in enumerate(header):
        for field, aliases in _IMPORT_FIELDS.items():
            if field in colmap:
                continue
            if any(h == a or h.replace(".", "").replace("_", " ") == a for a in aliases):
                colmap[field] = idx
                break
    if "name" not in colmap:
        return [], ("Couldn't find a product-name column. Name one column "
                    "'Name' (or Product / Item / Medicine) and retry.")

    rows = []
    for r in table[1:]:
        rec = {}
        for field, idx in colmap.items():
            val = str(r[idx]).strip() if idx < len(r) else ""
            if field in _NUMERIC_FIELDS:
                m = _re.search(r"-?\d+(\.\d+)?", val.replace(",", ""))
                rec[field] = float(m.group()) if m else 0
            else:
                rec[field] = val
        if rec.get("name"):
            rows.append(rec)
    return rows, None


@app.route("/items/import", methods=["GET", "POST"])
@login_required
def items_import():
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Choose a file first.", "err")
            return redirect(url_for("items_import"))
        rows, err = _parse_stock_file(f)
        if err:
            flash(err, "err")
            return redirect(url_for("items_import"))
        if not rows:
            flash("No product rows found in that file.", "err")
            return redirect(url_for("items_import"))

        created = updated = 0
        with conn() as c:
            for rec in rows:
                existing = None
                if rec.get("code"):
                    existing = c.execute("SELECT id FROM wholesale_items WHERE code=?", (rec["code"],)).fetchone()
                if not existing:
                    existing = c.execute("SELECT id FROM wholesale_items WHERE name=?", (rec["name"],)).fetchone()

                if existing:
                    fields = [k for k in rec if k != "code"]
                    c.execute(
                        f"UPDATE wholesale_items SET {', '.join(k + '=?' for k in fields)} WHERE id=?",
                        [rec[k] for k in fields] + [existing["id"]],
                    )
                    updated += 1
                else:
                    cols = list(rec.keys())
                    c.execute(
                        f"INSERT INTO wholesale_items ({', '.join(cols)}) VALUES ({', '.join('?' * len(cols))})",
                        [rec[k] for k in cols],
                    )
                    created += 1

        audit(session.get("ws_user"), "import_items", "item", None,
              f"{f.filename}: +{created} new, {updated} updated")
        flash(f"Imported {f.filename} — {created} new item(s), {updated} updated.", "ok")
        return redirect(url_for("items"))

    return render_template("items_import.html")
