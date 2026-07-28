"""
Auto-Reorder Bot
================

Upload an Excel/CSV of current stock → the bot works out what to buy,
how much, and from whom, then turns it into draft purchase orders with
one click.

How it decides (transparent, no black box):

  1. **Sales velocity** — average units sold per day over the lookback
     window, computed from this system's own invoiced order history.
  2. **Days of cover** — current_stock ÷ velocity. How long you'll last.
  3. **Reorder trigger** — flag the item if days-of-cover is below the
     lead time + safety buffer, OR stock is under the item's own
     reorder_level.
  4. **Suggested qty** — enough to reach the target cover window, rounded
     up to the item's MOQ (and to the scheme break-point where a free-goods
     scheme like 10+1 makes a slightly larger order strictly better).

Everything is tunable from the UI: lookback days, lead time, target cover,
safety buffer.

Excel format expected (header row, case-insensitive, extra columns ignored):

    code | name | stock          ← minimum
    code | name | stock | supplier

If `code` matches an item's SKU we use it; otherwise we fuzzy-match on name.
Rows that match nothing are reported back so nothing silently disappears.
"""
from __future__ import annotations

import io
import math
import csv
from datetime import date, timedelta
from flask import render_template, request, redirect, url_for, flash, session, jsonify
from app import app, conn, login_required, audit

try:
    from openpyxl import load_workbook
    HAVE_XLSX = True
except ImportError:
    HAVE_XLSX = False


# ── Defaults for the planning knobs ────────────────────────────────────
DEFAULT_PARAMS = {
    "lookback_days":  60,   # how far back to measure sales velocity
    "lead_time_days": 7,    # how long the supplier takes to deliver
    "target_cover":   30,   # we want this many days of stock after ordering
    "safety_days":    5,    # extra buffer on top of lead time before we panic
}


def parse_upload(file_storage):
    """Return (rows, error). rows = [{code, name, stock, supplier}]"""
    fname = (file_storage.filename or "").lower()
    raw = file_storage.read()
    rows = []

    if fname.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = [h.lower().strip() for h in (reader.fieldnames or [])]
        for r in reader:
            rows.append({(k or "").lower().strip(): (v or "").strip() for k, v in r.items()})
    elif fname.endswith((".xlsx", ".xlsm")):
        if not HAVE_XLSX:
            return [], "openpyxl is not installed on the server — cannot read .xlsx. Save as CSV instead."
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h or "").lower().strip() for h in next(it)]
        except StopIteration:
            return [], "The sheet is empty."
        for vals in it:
            if vals is None:
                continue
            row = {}
            for h, v in zip(header, vals):
                row[h] = "" if v is None else str(v).strip()
            if any(row.values()):
                rows.append(row)
    else:
        return [], "Unsupported file type. Upload .xlsx or .csv."

    # Normalise column names — accept several common spellings
    norm = []
    for r in rows:
        def pick(*names):
            for n in names:
                if n in r and r[n] not in ("", None):
                    return r[n]
            return ""
        code = pick("code", "sku", "item code", "itemcode", "product code")
        name = pick("name", "item", "item name", "product", "description", "particulars")
        stock = pick("stock", "qty", "quantity", "current stock", "closing", "closing stock", "balance")
        supplier = pick("supplier", "company", "manufacturer", "mfr", "vendor")
        if not (code or name):
            continue
        try:
            stock_n = float(str(stock).replace(",", "") or 0)
        except ValueError:
            stock_n = 0
        norm.append({"code": code, "name": name, "stock": stock_n, "supplier": supplier})
    return norm, None


def velocity_map(lookback_days: int) -> dict:
    """item_id → units sold per day over the lookback window."""
    since = (date.today() - timedelta(days=lookback_days)).isoformat()
    out = {}
    with conn() as c:
        rows = c.execute("""
            SELECT soi.item_id, SUM(soi.qty + COALESCE(soi.free_qty,0)) units
            FROM sales_order_items soi
            JOIN sales_orders o ON o.id = soi.order_id
            WHERE o.status IN ('confirmed','dispatched','invoiced','paid')
              AND o.order_date >= ?
              AND soi.item_id IS NOT NULL
            GROUP BY soi.item_id
        """, (since,)).fetchall()
    for r in rows:
        out[r["item_id"]] = (r["units"] or 0) / max(lookback_days, 1)
    return out


def match_item(c, code: str, name: str):
    """Find the item master row for an uploaded line."""
    if code:
        r = c.execute("SELECT * FROM wholesale_items WHERE code = ? COLLATE NOCASE", (code,)).fetchone()
        if r:
            return r
    if name:
        r = c.execute("SELECT * FROM wholesale_items WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
        if r:
            return r
        # loose match — first token of the name (e.g. "Dolo 650 10tab" → "Dolo 650%")
        r = c.execute("SELECT * FROM wholesale_items WHERE name LIKE ? ORDER BY LENGTH(name) LIMIT 1",
                      (f"{name.split()[0]}%",)).fetchone()
        if r:
            return r
    return None


def scheme_round(qty: int, scheme: str) -> int:
    """
    If a scheme like '10+1' exists, rounding UP to the next multiple of the
    base gets free goods — strictly better value for a marginally bigger buy.
    Only rounds up when we're already within 40% of the next break.
    """
    if not scheme or "+" not in scheme:
        return qty
    try:
        base, _bonus = scheme.split("+")
        base = int(base)
    except (ValueError, AttributeError):
        return qty
    if base <= 0 or qty <= 0:
        return qty
    remainder = qty % base
    if remainder == 0:
        return qty
    gap = base - remainder
    if gap <= base * 0.4:          # close enough to the break — round up
        return qty + gap
    return qty


def build_plan(rows, params):
    """Turn uploaded stock rows into a reorder plan."""
    vel = velocity_map(params["lookback_days"])
    trigger_days = params["lead_time_days"] + params["safety_days"]

    plan, unmatched = [], []
    with conn() as c:
        for r in rows:
            item = match_item(c, r["code"], r["name"])
            if not item:
                unmatched.append(r)
                continue

            v = vel.get(item["id"], 0.0)
            stock = r["stock"]
            cover = (stock / v) if v > 0 else (999 if stock > 0 else 0)

            below_reorder = (item["reorder_level"] or 0) > 0 and stock <= item["reorder_level"]
            below_cover = v > 0 and cover < trigger_days

            if not (below_reorder or below_cover):
                continue

            # order enough to reach the target cover window
            if v > 0:
                need = math.ceil(v * params["target_cover"] - stock)
            else:
                need = max((item["reorder_level"] or 0) * 2 - int(stock), item["moq"] or 1)

            need = max(need, item["moq"] or 1)
            need = scheme_round(int(need), item["scheme"])

            plan.append({
                "item_id":    item["id"],
                "code":       item["code"],
                "name":       item["name"],
                "manufacturer": item["manufacturer"] or r["supplier"] or "",
                "pack_size":  item["pack_size"],
                "stock":      stock,
                "velocity":   round(v, 2),
                "cover_days": round(cover, 1) if cover < 999 else None,
                "reorder_level": item["reorder_level"],
                "moq":        item["moq"],
                "scheme":     item["scheme"],
                "suggested":  int(need),
                "ptr":        item["ptr"],
                "value":      round(int(need) * (item["ptr"] or 0), 2),
                "reason":     "below reorder level" if below_reorder else f"only {round(cover,1)}d cover",
                "urgent":     cover < params["lead_time_days"],
            })

    plan.sort(key=lambda p: (not p["urgent"], p["cover_days"] if p["cover_days"] is not None else 999))
    return plan, unmatched


@app.route("/reorder", methods=["GET", "POST"])
@login_required
def reorder():
    params = dict(DEFAULT_PARAMS)
    for k in params:
        try:
            params[k] = int(request.form.get(k) or request.args.get(k) or params[k])
        except (TypeError, ValueError):
            pass

    plan, unmatched, uploaded = [], [], False

    if request.method == "POST" and request.files.get("file"):
        f = request.files["file"]
        if f and f.filename:
            rows, err = parse_upload(f)
            if err:
                flash(err, "err")
            else:
                uploaded = True
                plan, unmatched = build_plan(rows, params)
                session["reorder_plan"] = plan[:500]
                if not plan:
                    flash(f"Read {len(rows)} rows — nothing needs reordering right now. 👍", "ok")
                else:
                    flash(f"Read {len(rows)} rows → {len(plan)} items need reordering.", "ok")
    elif request.method == "POST":
        # "re-run with new numbers" using the previously uploaded plan is not
        # possible without the original stock, so ask for the file again.
        flash("Choose a file to analyse.", "err")

    total_value = sum(p["value"] for p in plan)
    urgent = sum(1 for p in plan if p["urgent"])

    return render_template("reorder.html",
                           plan=plan, unmatched=unmatched, params=params,
                           uploaded=uploaded, total_value=total_value, urgent=urgent,
                           have_xlsx=HAVE_XLSX)


@app.route("/reorder/create-pos", methods=["POST"])
@login_required
def reorder_create_pos():
    """Turn the selected plan lines into draft POs, grouped by manufacturer."""
    item_ids = request.form.getlist("item_id[]")
    qtys = request.form.getlist("qty[]")
    if not item_ids:
        flash("Nothing selected.", "err")
        return redirect(url_for("reorder"))

    groups: dict[str, list] = {}
    with conn() as c:
        for iid, q in zip(item_ids, qtys):
            try:
                iid_i, q_i = int(iid), int(float(q))
            except (TypeError, ValueError):
                continue
            if q_i <= 0:
                continue
            item = c.execute("SELECT * FROM wholesale_items WHERE id=?", (iid_i,)).fetchone()
            if not item:
                continue
            key = (item["manufacturer"] or "Unassigned").strip()
            groups.setdefault(key, []).append((item, q_i))

        created = []
        for mfr, lines in groups.items():
            # find or create a supplier matching the manufacturer name
            sup = c.execute("SELECT * FROM suppliers WHERE name = ? COLLATE NOCASE", (mfr,)).fetchone()
            if not sup:
                n = c.execute("SELECT COUNT(*) FROM suppliers").fetchone()[0]
                cur = c.execute("INSERT INTO suppliers (code, name, notes) VALUES (?,?,?)",
                                (f"SUP{n+1:03d}", mfr, "auto-created by reorder bot"))
                sup_id = cur.lastrowid
            else:
                sup_id = sup["id"]

            n = c.execute("SELECT COUNT(*) FROM purchase_orders").fetchone()[0]
            po_no = f"PO/{date.today().strftime('%y%m')}/{n+1:04d}"
            cur = c.execute("INSERT INTO purchase_orders (po_no, supplier_id, notes) VALUES (?,?,?)",
                            (po_no, sup_id, "Generated by auto-reorder bot"))
            po_id = cur.lastrowid

            subtotal = gst_amt = 0.0
            for item, q_i in lines:
                rate = item["ptr"] or 0
                amount = q_i * rate
                subtotal += amount
                gst_amt += amount * (item["gst_rate"] or 0) / 100
                c.execute("""INSERT INTO purchase_order_items
                    (po_id, item_id, item_name, qty, rate, gst_rate, amount)
                    VALUES (?,?,?,?,?,?,?)""",
                    (po_id, item["id"], item["name"], q_i, rate, item["gst_rate"], amount))
            c.execute("UPDATE purchase_orders SET subtotal=?, gst_amount=?, total=? WHERE id=?",
                      (subtotal, gst_amt, subtotal + gst_amt, po_id))
            created.append((po_no, mfr, len(lines)))

    audit(session.get("ws_user"), "reorder_bot", "purchase_order", None,
          f"{len(created)} POs created")
    flash(f"Created {len(created)} draft PO(s): " +
          ", ".join(f"{po} ({mfr}, {n} items)" for po, mfr, n in created), "ok")
    return redirect(url_for("purchases"))


@app.route("/reorder/template.csv")
@login_required
def reorder_template():
    """Download a starter CSV pre-filled with the current item master."""
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["code", "name", "stock", "supplier"])
    with conn() as c:
        for r in c.execute("SELECT code, name, stock, manufacturer FROM wholesale_items WHERE status='active' ORDER BY name").fetchall():
            w.writerow([r["code"], r["name"], r["stock"], r["manufacturer"] or ""])
    from flask import Response
    return Response(out.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=stock-template.csv"})
