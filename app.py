import json
import os
import sqlite3
import csv
import io
import hashlib
import time
import uuid as _uuid_mod
from datetime import datetime, timedelta, timezone
from typing import Any

from flask import Flask, jsonify, request, send_file, render_template, session, redirect, url_for, Response
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash



BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("PHARMACY_DB_PATH", os.path.join(BASE_DIR, "database.db"))
TEMPLATES_DIR = os.path.join(BASE_DIR, "frontend", "templates")
STATIC_DIR = os.path.join(BASE_DIR, "frontend", "static")


app = Flask(
    __name__,
    template_folder=TEMPLATES_DIR,
    static_folder=STATIC_DIR,
    static_url_path="/static",
)
app.secret_key = os.environ.get("SECRET_KEY", "medivision-secret-key-2026")
CORS(app)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return {row["name"] for row in rows}


def safe_json_loads(raw: Any, fallback: Any) -> Any:
    if raw is None:
        return fallback
    if isinstance(raw, (list, dict)):
        return raw
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return fallback


def json_error(message: str, status_code: int = 400, details: Any = None):
    payload = {"status": "error", "message": message}
    if details is not None:
        payload["details"] = details
    return jsonify(payload), status_code


import urllib.request
import urllib.parse
import urllib.error
import base64
import math


def send_whatsapp(phone: str, message: str) -> dict:
    """Send WhatsApp message via Twilio. Gracefully mocks when credentials are missing."""
    sid   = os.environ.get("TWILIO_ACCOUNT_SID", "")
    token = os.environ.get("TWILIO_AUTH_TOKEN", "")
    from_ = os.environ.get("TWILIO_WHATSAPP_NUMBER", "")

    phone = phone.strip()
    if not phone.startswith("+"):
        phone = f"+91{phone}" if len(phone) == 10 else f"+{phone}"

    if not all([sid, token, from_]):
        return {"status": "mocked", "sid": f"mock_{int(datetime.now().timestamp())}",
                "note": "Twilio credentials not configured"}

    url  = f"https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json"
    body = urllib.parse.urlencode({"To": f"whatsapp:{phone}", "From": from_, "Body": message}).encode()
    auth = base64.b64encode(f"{sid}:{token}".encode()).decode()

    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Authorization", f"Basic {auth}")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")

    try:
        with urllib.request.urlopen(req) as r:
            resp = json.loads(r.read().decode())
            return {"status": "sent", "sid": resp.get("sid", "")}
    except urllib.error.HTTPError as e:
        return {"status": "failed", "error": e.read().decode()}
    except Exception as e:
        return {"status": "failed", "error": str(e)}


def build_bill_receipt_message(bill: dict) -> str:
    """Build a WhatsApp-friendly bill receipt text."""
    lines = [
        f"🏥 *Selvam Medicals & Co*",
        f"Bill #{bill.get('id', '')}  |  {bill.get('date', '')}",
        "──────────────────",
    ]
    for item in bill.get("items", []):
        lines.append(f"• {item.get('n', item.get('name', 'Item'))}  ×{item.get('qty', 1)}  ₹{item.get('p', 0):.0f}")
    lines += [
        "──────────────────",
        f"Subtotal : ₹{bill.get('sub', 0):.2f}",
        f"Discount : ₹{bill.get('disc', 0):.2f}",
        f"GST      : ₹{bill.get('tax', 0):.2f}",
        f"*TOTAL   : ₹{bill.get('total', 0):.2f}*",
        f"Payment  : {str(bill.get('pay', 'Cash')).upper()}",
        "",
        "Thank you for choosing Selvam Medicals! 💊",
        "For queries: +91 XXXXXXXXXX",
    ]
    return "\n".join(lines)


def face_match_customer(query_vector: list, threshold: float = 0.50) -> dict | None:
    """Euclidean distance match against all stored customer face vectors."""
    if not query_vector or not isinstance(query_vector, list):
        return None

    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, name, phone, face_vector FROM customers "
            "WHERE face_vector IS NOT NULL AND face_vector != ''"
        ).fetchall()

    best, best_dist = None, float("inf")

    for row in rows:
        stored = safe_json_loads(row["face_vector"], None)
        if not isinstance(stored, list) or len(stored) != len(query_vector):
            continue
        dist = math.sqrt(sum((a - b) ** 2 for a, b in zip(query_vector, stored)))
        if dist < best_dist:
            best_dist = dist
            best       = dict(row)

    if best and best_dist <= threshold:
        best.pop("face_vector", None)   # don't send the raw vector back
        best["distance"]   = round(best_dist, 4)
        best["confidence"] = round(max(0.0, 1.0 - (best_dist / threshold)) * 100, 1)
        return best
    return None


def required_fields(payload: dict[str, Any], fields: list[str]) -> list[str]:
    missing: list[str] = []
    for field in fields:
        if field not in payload:
            missing.append(field)
            continue
        value = payload[field]
        if value is None:
            missing.append(field)
            continue
        if isinstance(value, str) and not value.strip():
            missing.append(field)
    return missing


def init_db() -> None:
    with get_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bills (
                id TEXT PRIMARY KEY,
                ts INTEGER,
                date TEXT,
                cust TEXT,
                phone TEXT,
                pay TEXT,
                sub REAL,
                disc REAL,
                tax REAL,
                total REAL,
                items TEXT,
                doctor TEXT,
                rx TEXT DEFAULT '',
                prescription TEXT DEFAULT ''
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS medicines (
                id TEXT PRIMARY KEY,
                n TEXT,
                g TEXT,
                c TEXT,
                p REAL,
                s INTEGER,
                batch TEXT DEFAULT '',
                expiry TEXT DEFAULT '',
                p_rate REAL DEFAULT 0,
                p_packing TEXT DEFAULT '',
                s_packing TEXT DEFAULT '',
                p_gst REAL DEFAULT 0,
                s_gst REAL DEFAULT 0,
                disc REAL DEFAULT 0,
                offer TEXT DEFAULT '',
                reorder INTEGER DEFAULT 0,
                max_qty INTEGER DEFAULT 0
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS purchases (
                id TEXT PRIMARY KEY,
                supplier TEXT,
                items TEXT,
                amount REAL,
                date TEXT,
                status TEXT,
                batch TEXT DEFAULT '',
                expiry TEXT DEFAULT '',
                photo TEXT DEFAULT ''
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                phone TEXT,
                visits INTEGER DEFAULT 1,
                total REAL DEFAULT 0,
                address TEXT DEFAULT '',
                email TEXT DEFAULT '',
                face_vector TEXT DEFAULT ''
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS doctors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                specialty TEXT,
                hospital TEXT,
                phone TEXT,
                email TEXT DEFAULT ''
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                phone TEXT,
                gst TEXT DEFAULT '',
                last_order TEXT DEFAULT '-',
                status TEXT DEFAULT 'Active'
            )
            """
        )

        med_count = conn.execute("SELECT COUNT(*) AS c FROM medicines").fetchone()["c"]
        if med_count == 0:
            initial_meds = [
                (
                    "1",
                    "Dolo 650mg",
                    "Paracetamol",
                    "Tablet",
                    28,
                    3,
                    "B101",
                    "2026-12-31",
                    20.5,
                    "1x10",
                    "1x10",
                    12.0,
                    12.0,
                    0.0,
                    "None",
                    10,
                    100,
                ),
                (
                    "2",
                    "Augmentin 625",
                    "Amoxicillin",
                    "Tablet",
                    142,
                    24,
                    "B202",
                    "2026-12-31",
                    110.0,
                    "1x10",
                    "1x10",
                    12.0,
                    12.0,
                    0.0,
                    "None",
                    20,
                    200,
                ),
            ]
            conn.executemany(
                """
                INSERT INTO medicines
                (id, n, g, c, p, s, batch, expiry, p_rate, p_packing, s_packing, p_gst, s_gst, disc, offer, reorder, max_qty)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                initial_meds,
            )

        # ── Platform registrations (wholesale + retail pending approval) ──
        conn.execute("""
            CREATE TABLE IF NOT EXISTS platform_registrations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reg_type TEXT NOT NULL,
                business_name TEXT NOT NULL,
                owner_name TEXT NOT NULL,
                owner_phone TEXT NOT NULL,
                owner_email TEXT,
                alt_phone TEXT,
                address TEXT,
                city TEXT,
                state TEXT,
                pincode TEXT,
                gstin TEXT,
                drug_license TEXT,
                pan TEXT,
                bank_name TEXT,
                bank_account TEXT,
                bank_ifsc TEXT,
                monthly_turnover TEXT,
                no_of_staff TEXT,
                no_of_shops TEXT,
                software_used TEXT,
                hear_from TEXT,
                status TEXT DEFAULT 'pending',
                submitted_at TEXT,
                reviewed_at TEXT,
                reviewed_by TEXT,
                review_notes TEXT,
                username TEXT,
                password_hash TEXT
            )
        """)

        # ── Retail shops (approved) ──────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS retail_shops (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shop_name TEXT NOT NULL,
                owner_name TEXT NOT NULL,
                phone TEXT,
                address TEXT,
                city TEXT,
                gstin TEXT,
                drug_license TEXT,
                username TEXT UNIQUE,
                password_hash TEXT,
                is_active INTEGER DEFAULT 1,
                approved_at TEXT,
                reg_id INTEGER
            )
        """)

        # ── Wholesale accounts (approved) ────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS wholesale_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_name TEXT NOT NULL,
                owner_name TEXT NOT NULL,
                phone TEXT,
                address TEXT,
                city TEXT,
                gstin TEXT,
                drug_license TEXT,
                username TEXT UNIQUE,
                password_hash TEXT,
                is_active INTEGER DEFAULT 1,
                approved_at TEXT,
                reg_id INTEGER
            )
        """)

        # ── Per-shop medicine min/max reorder rules ──────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS shop_reorder_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shop_id INTEGER,
                medicine_id TEXT,
                medicine_name TEXT,
                min_qty INTEGER DEFAULT 0,
                max_qty INTEGER DEFAULT 0,
                reorder_qty INTEGER DEFAULT 0,
                ai_min INTEGER,
                ai_max INTEGER,
                avg_daily_sales REAL DEFAULT 0,
                last_updated TEXT
            )
        """)

        # ── Wholesaler catalog (for AI comparison) ───────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS wholesaler_catalog (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                wholesaler_id INTEGER,
                wholesaler_name TEXT,
                medicine_name TEXT,
                generic_name TEXT,
                category TEXT,
                price REAL,
                mrp REAL,
                discount_pct REAL DEFAULT 0,
                free_offer TEXT,
                min_order_qty INTEGER DEFAULT 1,
                stock_available INTEGER DEFAULT 0,
                last_updated TEXT
            )
        """)

        # ── Daily stock import logs ──────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS stock_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shop_id INTEGER,
                shop_type TEXT DEFAULT 'retail',
                import_date TEXT,
                file_name TEXT,
                item_count INTEGER DEFAULT 0,
                imported_at TEXT,
                imported_by TEXT,
                raw_data TEXT
            )
        """)

        # ── Platform admins ──────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS platform_admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                name TEXT,
                is_super INTEGER DEFAULT 0,
                created_at TEXT
            )
        """)

        # Create default super admin if none exists
        admin_count = conn.execute("SELECT COUNT(*) AS c FROM platform_admins").fetchone()["c"]
        if admin_count == 0:
            conn.execute("""
                INSERT INTO platform_admins (username, password_hash, name, is_super, created_at)
                VALUES (?, ?, ?, 1, ?)
            """, ("admin", generate_password_hash("admin@123"), "Super Admin", datetime.now(timezone.utc).isoformat()))

        # Seed sample wholesaler catalog
        wc_count = conn.execute("SELECT COUNT(*) AS c FROM wholesaler_catalog").fetchone()["c"]
        if wc_count == 0:
            sample_catalog = [
                (1, "SS & Co", "Dolo 650mg", "Paracetamol", "Tablet", 20.5, 28.0, 5.0, "10+1 Free on 10 strips", 10, 500, datetime.now(timezone.utc).isoformat()),
                (1, "SS & Co", "Augmentin 625", "Amoxicillin", "Tablet", 108.0, 142.0, 8.0, "None", 5, 200, datetime.now(timezone.utc).isoformat()),
                (1, "SS & Co", "Crocin 500mg", "Paracetamol", "Tablet", 18.0, 24.0, 5.0, "5% extra discount on 50+", 20, 1000, datetime.now(timezone.utc).isoformat()),
                (2, "Med Distributors", "Dolo 650mg", "Paracetamol", "Tablet", 19.0, 28.0, 7.0, "None", 5, 800, datetime.now(timezone.utc).isoformat()),
                (2, "Med Distributors", "Augmentin 625", "Amoxicillin", "Tablet", 105.0, 142.0, 10.0, "Buy 10 Get 1 Free", 10, 150, datetime.now(timezone.utc).isoformat()),
                (2, "Med Distributors", "Pan-D", "Pantoprazole", "Capsule", 65.0, 85.0, 8.0, "None", 10, 300, datetime.now(timezone.utc).isoformat()),
                (3, "Apollo Pharmacy WS", "Dolo 650mg", "Paracetamol", "Tablet", 21.0, 28.0, 3.0, "12+1 scheme on 100 strips", 50, 2000, datetime.now(timezone.utc).isoformat()),
                (3, "Apollo Pharmacy WS", "Pan-D", "Pantoprazole", "Capsule", 62.0, 85.0, 12.0, "10% on bulk order 500+", 20, 500, datetime.now(timezone.utc).isoformat()),
            ]
            conn.executemany("""
                INSERT INTO wholesaler_catalog (wholesaler_id, wholesaler_name, medicine_name, generic_name, category, price, mrp, discount_pct, free_offer, min_order_qty, stock_available, last_updated)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, sample_catalog)

        # ── Subscription plans ───────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS subscription_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                price_monthly REAL DEFAULT 0,
                price_yearly REAL DEFAULT 0,
                max_shops INTEGER DEFAULT 1,
                features TEXT DEFAULT '[]',
                is_active INTEGER DEFAULT 1,
                created_at TEXT
            )
        """)

        # ── Active subscriptions ─────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_type TEXT NOT NULL,
                account_id INTEGER NOT NULL,
                plan_id INTEGER,
                plan_name TEXT,
                status TEXT DEFAULT 'trial',
                start_date TEXT,
                end_date TEXT,
                amount_paid REAL DEFAULT 0,
                currency TEXT DEFAULT 'INR',
                razorpay_order_id TEXT,
                razorpay_payment_id TEXT,
                billing_cycle TEXT DEFAULT 'monthly',
                auto_renew INTEGER DEFAULT 1,
                created_at TEXT,
                updated_at TEXT
            )
        """)

        # ── Tally companies ──────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS tally_companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                address TEXT,
                city TEXT,
                state TEXT,
                pincode TEXT,
                gstin TEXT,
                pan TEXT,
                phone TEXT,
                email TEXT,
                bank_name TEXT,
                bank_account TEXT,
                bank_ifsc TEXT,
                opening_balance REAL DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                created_at TEXT
            )
        """)

        # ── Tally ledger entries ─────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS tally_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                entry_type TEXT NOT NULL,
                party_name TEXT,
                party_type TEXT DEFAULT 'supplier',
                amount REAL NOT NULL,
                gst_amount REAL DEFAULT 0,
                total_amount REAL NOT NULL,
                description TEXT,
                reference_no TEXT,
                invoice_no TEXT,
                date TEXT,
                due_date TEXT,
                payment_mode TEXT,
                bank_name TEXT,
                cheque_no TEXT,
                status TEXT DEFAULT 'pending',
                created_at TEXT,
                updated_at TEXT,
                created_by TEXT
            )
        """)

        # ── Bank statement uploads ───────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS bank_statements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                bank_name TEXT,
                account_no TEXT,
                statement_month TEXT,
                opening_balance REAL DEFAULT 0,
                closing_balance REAL DEFAULT 0,
                total_credits REAL DEFAULT 0,
                total_debits REAL DEFAULT 0,
                entries TEXT DEFAULT '[]',
                file_name TEXT,
                uploaded_at TEXT,
                uploaded_by TEXT
            )
        """)

        # ── Cheque register ──────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cheque_register (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                cheque_no TEXT,
                date TEXT,
                payee TEXT,
                bank_name TEXT,
                account_no TEXT,
                amount REAL NOT NULL,
                memo TEXT,
                status TEXT DEFAULT 'issued',
                entry_id INTEGER,
                created_at TEXT,
                printed_at TEXT
            )
        """)

        # ── Smart receiving orders ───────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS receiving_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shop_id INTEGER,
                shop_type TEXT DEFAULT 'retail',
                supplier_name TEXT,
                po_date TEXT,
                expected_date TEXT,
                items TEXT DEFAULT '[]',
                status TEXT DEFAULT 'pending',
                daily_code TEXT,
                received_by TEXT,
                received_at TEXT,
                notes TEXT,
                return_items TEXT DEFAULT '[]',
                created_at TEXT
            )
        """)

        # ── Delivery routes ──────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS delivery_routes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                route_name TEXT,
                delivery_boy TEXT,
                delivery_phone TEXT,
                delivery_date TEXT,
                shops TEXT DEFAULT '[]',
                status TEXT DEFAULT 'pending',
                started_at TEXT,
                completed_at TEXT,
                created_at TEXT,
                notes TEXT
            )
        """)

        # ── AI call logs ─────────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS call_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shop_id INTEGER,
                shop_name TEXT,
                phone TEXT,
                call_type TEXT DEFAULT 'medicine_reminder',
                status TEXT DEFAULT 'scheduled',
                scheduled_at TEXT,
                made_at TEXT,
                duration_secs INTEGER DEFAULT 0,
                call_response TEXT,
                medicine_name TEXT,
                days_of_stock REAL,
                follow_up_at TEXT,
                created_at TEXT
            )
        """)

        # Seed default subscription plans
        sp_count = conn.execute("SELECT COUNT(*) AS c FROM subscription_plans").fetchone()["c"]
        if sp_count == 0:
            now_ts = datetime.now(timezone.utc).isoformat()
            plans = [
                ("Starter",      499,  4990,  1, '["Basic Billing","Stock Management","20 Medicines","Email Support"]', 1, now_ts),
                ("Professional", 999,  9990,  5, '["Full Billing","AI Reorder","Wholesaler Compare","Excel Import","Tally Basic","100 Medicines","Phone Support"]', 1, now_ts),
                ("Enterprise",  2499, 24990, 99, '["All Features","Tally Full","Delivery Tracking","AI Auto-Caller","Smart Receiving","Unlimited Medicines","Dedicated Support","White-label"]', 1, now_ts),
            ]
            conn.executemany("""
                INSERT INTO subscription_plans (name, price_monthly, price_yearly, max_shops, features, is_active, created_at)
                VALUES (?,?,?,?,?,?,?)
            """, plans)

        # ── Staff members ────────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS staff (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT    NOT NULL,
                role        TEXT    DEFAULT 'Cashier',
                phone       TEXT    DEFAULT '',
                pin         TEXT    DEFAULT '',
                is_active   INTEGER DEFAULT 1,
                salary      REAL    DEFAULT 0,
                joined_date TEXT    DEFAULT '',
                created_at  TEXT
            )
        """)

        # Seed default owner staff
        st_count = conn.execute("SELECT COUNT(*) AS c FROM staff").fetchone()["c"]
        if st_count == 0:
            conn.execute("""
                INSERT INTO staff (name, role, phone, pin, is_active, created_at)
                VALUES ('Owner', 'Manager', '', '1234', 1, ?)
            """, (datetime.now(timezone.utc).isoformat(),))

        # ── Attendance logs ──────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS attendance_logs (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id     INTEGER NOT NULL,
                staff_name   TEXT,
                date         TEXT    NOT NULL,
                punch_in     TEXT,
                punch_out    TEXT,
                worked_hours REAL    DEFAULT 0,
                status       TEXT    DEFAULT 'present',
                notes        TEXT    DEFAULT '',
                created_at   TEXT
            )
        """)

        # ── WhatsApp message log ─────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS whatsapp_logs (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                bill_id     TEXT,
                phone       TEXT,
                message     TEXT,
                status      TEXT    DEFAULT 'pending',
                provider_id TEXT    DEFAULT '',
                sent_at     TEXT,
                error_msg   TEXT    DEFAULT ''
            )
        """)

        # ── Drug schedules (H / H1 / X / NDPS / OTC) ────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS drug_schedules (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                medicine_name    TEXT NOT NULL,
                generic_name     TEXT DEFAULT '',
                schedule_type    TEXT DEFAULT 'OTC',
                narcotic         INTEGER DEFAULT 0,
                psychotropic     INTEGER DEFAULT 0,
                requires_rx      INTEGER DEFAULT 0,
                max_qty_no_rx    INTEGER DEFAULT 0,
                inspector_log    INTEGER DEFAULT 0,
                notes            TEXT DEFAULT '',
                created_at       TEXT
            )
        """)

        # ── Drug combinations & 1mg-style alternatives ───────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS drug_combinations (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                brand_name       TEXT NOT NULL,
                generic_names    TEXT DEFAULT '[]',
                manufacturer     TEXT DEFAULT '',
                category         TEXT DEFAULT 'Tablet',
                schedule_type    TEXT DEFAULT 'OTC',
                alternatives     TEXT DEFAULT '[]',
                mrp              REAL  DEFAULT 0,
                pack_size        TEXT  DEFAULT '',
                uses             TEXT  DEFAULT '',
                contraindications TEXT DEFAULT '',
                is_otc           INTEGER DEFAULT 1,
                source           TEXT  DEFAULT 'local'
            )
        """)

        # ── Narcotic / Schedule-X daily register ─────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS narcotic_register (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                medicine_name    TEXT NOT NULL,
                generic_name     TEXT DEFAULT '',
                schedule_type    TEXT DEFAULT 'X',
                date             TEXT NOT NULL,
                bill_id          TEXT DEFAULT '',
                customer_name    TEXT DEFAULT '',
                customer_phone   TEXT DEFAULT '',
                doctor_name      TEXT DEFAULT '',
                doctor_reg_no    TEXT DEFAULT '',
                qty_dispensed    INTEGER DEFAULT 0,
                batch_no         TEXT DEFAULT '',
                opening_balance  INTEGER DEFAULT 0,
                closing_balance  INTEGER DEFAULT 0,
                purpose          TEXT DEFAULT '',
                created_at       TEXT,
                created_by       TEXT DEFAULT ''
            )
        """)

        # ── Patient medication reminders ──────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS patient_reminders (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_name    TEXT NOT NULL,
                phone            TEXT DEFAULT '',
                medicine_name    TEXT NOT NULL,
                last_bill_date   TEXT,
                days_supply      INTEGER DEFAULT 30,
                daily_dose       REAL    DEFAULT 1,
                qty_dispensed    INTEGER DEFAULT 30,
                expected_finish  TEXT,
                reminder_date    TEXT,
                status           TEXT DEFAULT 'pending',
                wa_sent          INTEGER DEFAULT 0,
                call_scheduled   INTEGER DEFAULT 0,
                wa_response      TEXT DEFAULT '',
                notes            TEXT DEFAULT '',
                created_at       TEXT
            )
        """)

        # ── Drug inspector visit log ──────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS inspector_visits (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                inspector_name   TEXT DEFAULT '',
                badge_no         TEXT DEFAULT '',
                visit_date       TEXT NOT NULL,
                visit_time       TEXT DEFAULT '',
                purpose          TEXT DEFAULT 'routine',
                items_checked    TEXT DEFAULT '[]',
                observations     TEXT DEFAULT '',
                compliance_status TEXT DEFAULT 'pass',
                next_visit_date  TEXT DEFAULT '',
                signature_obtained INTEGER DEFAULT 0,
                created_at       TEXT
            )
        """)

        # ── Seed comprehensive drug schedule data ─────────────────────────
        ds_count = conn.execute("SELECT COUNT(*) AS c FROM drug_schedules").fetchone()["c"]
        if ds_count == 0:
            now_ts = datetime.now(timezone.utc).isoformat()
            schedule_drugs = [
                # Schedule X (Narcotic / psychotropic — requires narcotic register)
                ("Morphine","Morphine Sulphate","X",1,0,1,0,1,"Opioid analgesic — strict register",now_ts),
                ("Pethidine","Pethidine HCl","X",1,0,1,0,1,"Opioid — narcotic register mandatory",now_ts),
                ("Codeine","Codeine Phosphate","X",1,0,1,0,1,"Narcotic — requires Schedule X licence",now_ts),
                ("Buprenorphine","Buprenorphine","X",1,0,1,0,1,"Opioid partial agonist",now_ts),
                ("Fentanyl","Fentanyl","X",1,0,1,0,1,"High-potency narcotic — ICU use",now_ts),
                ("Tramadol","Tramadol HCl","X",0,1,1,0,1,"Opioid analgesic — abuse potential",now_ts),
                ("Alprazolam","Alprazolam","X",0,1,1,0,1,"Benzodiazepine — Schedule X",now_ts),
                ("Clonazepam","Clonazepam","X",0,1,1,0,1,"Benzodiazepine — anticonvulsant",now_ts),
                ("Diazepam","Diazepam","X",0,1,1,0,1,"Benzodiazepine — anxiolytic",now_ts),
                ("Lorazepam","Lorazepam","X",0,1,1,0,1,"Benzodiazepine — sedative",now_ts),
                ("Zolpidem","Zolpidem","X",0,1,1,0,1,"Sleep disorder — Schedule X",now_ts),
                ("Nitrazepam","Nitrazepam","X",0,1,1,0,1,"Hypnotic benzodiazepine",now_ts),
                ("Phenobarbitone","Phenobarbital","X",0,1,1,0,1,"Barbiturate anticonvulsant",now_ts),
                ("Pregabalin","Pregabalin","X",0,1,1,0,1,"Gabapentinoid — Schedule X from 2023",now_ts),
                ("Gabapentin","Gabapentin","X",0,1,1,0,1,"Gabapentinoid — abuse potential",now_ts),
                # Schedule H1 (enhanced prescription monitoring)
                ("Ciprofloxacin","Ciprofloxacin","H1",0,0,1,0,0,"Fluoroquinolone antibiotic — H1",now_ts),
                ("Ofloxacin","Ofloxacin","H1",0,0,1,0,0,"Fluoroquinolone — H1",now_ts),
                ("Levofloxacin","Levofloxacin","H1",0,0,1,0,0,"Fluoroquinolone — H1",now_ts),
                ("Azithromycin","Azithromycin","H1",0,0,1,0,0,"Macrolide antibiotic — H1",now_ts),
                ("Amoxicillin","Amoxicillin","H1",0,0,1,0,0,"Penicillin — H1",now_ts),
                ("Augmentin","Amoxicillin+Clavulanate","H1",0,0,1,0,0,"Beta-lactam combination — H1",now_ts),
                ("Metronidazole","Metronidazole","H1",0,0,1,0,0,"Antiprotozoal — H1",now_ts),
                ("Fluconazole","Fluconazole","H1",0,0,1,0,0,"Antifungal — H1",now_ts),
                ("Ceftriaxone","Ceftriaxone","H1",0,0,1,0,0,"3rd gen cephalosporin — H1",now_ts),
                ("Cefixime","Cefixime","H1",0,0,1,0,0,"Oral cephalosporin — H1",now_ts),
                # Schedule H
                ("Atorvastatin","Atorvastatin","H",0,0,1,0,0,"Statin — requires prescription",now_ts),
                ("Metformin","Metformin HCl","H",0,0,1,0,0,"Antidiabetic — H",now_ts),
                ("Amlodipine","Amlodipine","H",0,0,1,0,0,"CCB antihypertensive — H",now_ts),
                ("Lisinopril","Lisinopril","H",0,0,1,0,0,"ACE inhibitor — H",now_ts),
                ("Omeprazole","Omeprazole","H",0,0,1,0,0,"PPI — H",now_ts),
                ("Pantoprazole","Pantoprazole","H",0,0,1,0,0,"PPI — H",now_ts),
                ("Atenolol","Atenolol","H",0,0,1,0,0,"Beta blocker — H",now_ts),
                ("Losartan","Losartan Potassium","H",0,0,1,0,0,"ARB — H",now_ts),
                ("Glimepiride","Glimepiride","H",0,0,1,0,0,"Sulfonylurea — H",now_ts),
                ("Levothyroxine","Levothyroxine","H",0,0,1,0,0,"Thyroid hormone — H",now_ts),
                ("Insulin","Insulin","H",0,0,1,0,0,"Insulin — H schedule",now_ts),
                ("Warfarin","Warfarin Sodium","H",0,0,1,0,0,"Anticoagulant — H",now_ts),
                ("Digoxin","Digoxin","H",0,0,1,0,0,"Cardiac glycoside — H narrow TI",now_ts),
                ("Phenytoin","Phenytoin","H",0,0,1,0,0,"Antiepileptic — H",now_ts),
                ("Carbamazepine","Carbamazepine","H",0,0,1,0,0,"Antiepileptic — H",now_ts),
                # OTC
                ("Paracetamol","Paracetamol","OTC",0,0,0,10,0,"Analgesic/antipyretic",now_ts),
                ("Aspirin","Aspirin","OTC",0,0,0,10,0,"Low dose — OTC",now_ts),
                ("Cetirizine","Cetirizine HCl","OTC",0,0,0,10,0,"Antihistamine — OTC",now_ts),
                ("ORS","Oral Rehydration Salts","OTC",0,0,0,0,0,"Electrolyte solution — OTC",now_ts),
                ("Antacid","Aluminium+Magnesium","OTC",0,0,0,0,0,"Antacid — OTC",now_ts),
            ]
            conn.executemany("""
                INSERT INTO drug_schedules
                (medicine_name,generic_name,schedule_type,narcotic,psychotropic,requires_rx,max_qty_no_rx,inspector_log,notes,created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, schedule_drugs)

        # ── Seed 1mg-style drug combinations ─────────────────────────────
        dc_count = conn.execute("SELECT COUNT(*) AS c FROM drug_combinations").fetchone()["c"]
        if dc_count == 0:
            now_ts = datetime.now(timezone.utc).isoformat()
            combos = [
              ("Dolo 650","[\"Paracetamol 650mg\"]","Micro Labs","Tablet","OTC","[\"Calpol 650\",\"Crocin 650\",\"Fepanil 650\"]",30.0,"1x15","Fever, Pain, Headache","Liver disease","yes"),
              ("Crocin 500","[\"Paracetamol 500mg\"]","GSK","Tablet","OTC","[\"Dolo 500\",\"Calpol 500\",\"P-500\"]",18.0,"1x15","Mild fever, Body pain","Hepatic impairment","yes"),
              ("Augmentin 625","[\"Amoxicillin 500mg\",\"Clavulanate 125mg\"]","GSK","Tablet","H1","[\"Moxclav 625\",\"Clavam 625\",\"Clavix 625\"]",142.0,"1x10","Bacterial infections","Penicillin allergy","no"),
              ("Azithral 500","[\"Azithromycin 500mg\"]","Alembic","Tablet","H1","[\"Zithromax 500\",\"Azee 500\",\"Azimax 500\"]",85.0,"1x5","RTI, STI, CAP","QT prolongation","no"),
              ("Pantop 40","[\"Pantoprazole 40mg\"]","Aristo","Tablet","H","[\"Pan 40\",\"Pantocid 40\",\"Rantac-D\"]",42.0,"1x15","Acid reflux, GERD, Ulcer","Hypersensitivity","no"),
              ("Pan-D","[\"Pantoprazole 40mg\",\"Domperidone 30mg\"]","Alkem","Capsule","H","[\"Nexpro-D\",\"Omesec-D\",\"Pantocid-D\"]",65.0,"1x15","GERD with nausea","Cardiac arrhythmia","no"),
              ("Metformin 500","[\"Metformin HCl 500mg\"]","USV","Tablet","H","[\"Glycomet 500\",\"Obimet 500\",\"Gluconorm 500\"]",22.0,"1x10","Type 2 Diabetes","Renal failure","no"),
              ("Glycomet GP1","[\"Glimepiride 1mg\",\"Metformin 500mg\"]","USV","Tablet","H","[\"Amaryl M 1\",\"Gemer 1\",\"Glynase MF 1\"]",48.0,"1x10","Type 2 Diabetes","Hypoglycaemia risk","no"),
              ("Amlodipine 5","[\"Amlodipine 5mg\"]","Sun Pharma","Tablet","H","[\"Amlip 5\",\"Stamlo 5\",\"Norvasc 5\"]",35.0,"1x15","Hypertension, Angina","Cardiogenic shock","no"),
              ("Telmisartan 40","[\"Telmisartan 40mg\"]","Glenmark","Tablet","H","[\"Telma 40\",\"Telmikind 40\",\"Inderal 40\"]",52.0,"1x15","Hypertension","Pregnancy","no"),
              ("Atorvastatin 10","[\"Atorvastatin 10mg\"]","Ranbaxy","Tablet","H","[\"Atorva 10\",\"Lipitor 10\",\"Storvas 10\"]",38.0,"1x15","Hypercholesterolaemia","Liver disease","no"),
              ("Rosuvastatin 10","[\"Rosuvastatin 10mg\"]","AstraZeneca","Tablet","H","[\"Crestor 10\",\"Rozucor 10\",\"Stator 10\"]",95.0,"1x10","Hyperlipidaemia","Myopathy","no"),
              ("Levothyroxine 50","[\"Levothyroxine 50mcg\"]","GSK","Tablet","H","[\"Thyronorm 50\",\"Eltroxin 50\",\"Thyrofit 50\"]",35.0,"1x120","Hypothyroidism","Hyperthyroidism","no"),
              ("Ecosprin 75","[\"Aspirin 75mg\"]","USV","Tablet","OTC","[\"Aspocid 75\",\"Disprin 75\",\"Cardioprin 75\"]",15.0,"1x14","Antiplatelet, CAD prevention","GI ulcer","yes"),
              ("Omeprazole 20","[\"Omeprazole 20mg\"]","Cipla","Capsule","H","[\"Omez 20\",\"Prilosec 20\",\"Ocid 20\"]",28.0,"1x15","Peptic ulcer, GERD","Hypomagnesaemia","no"),
              ("Cetirizine 10","[\"Cetirizine HCl 10mg\"]","Cipla","Tablet","OTC","[\"Zyrtec 10\",\"Okacet 10\",\"Alerid 10\"]",22.0,"1x10","Allergic rhinitis, Urticaria","Renal impairment","yes"),
              ("Montelukast 10","[\"Montelukast 10mg\"]","MSD","Tablet","H","[\"Montair 10\",\"Seroflo\",\"Singulair 10\"]",68.0,"1x15","Asthma, Allergic rhinitis","","no"),
              ("Salbutamol Inhaler","[\"Salbutamol 100mcg\"]","GSK","Inhaler","H","[\"Asthalin\",\"Ventolin\",\"Salmapace\"]",140.0,"200 doses","Asthma, COPD bronchospasm","Tachycardia","no"),
              ("Budesonide Inhaler","[\"Budesonide 200mcg\"]","AstraZeneca","Inhaler","H","[\"Pulmicort\",\"Budecort\",\"Foracort\"]",285.0,"200 doses","Asthma maintenance","Active TB","no"),
              ("Metronidazole 400","[\"Metronidazole 400mg\"]","Pfizer","Tablet","H1","[\"Flagyl 400\",\"Metrogyl 400\",\"Metro 400\"]",18.0,"1x15","Amoebiasis, H.pylori","First trimester pregnancy","no"),
              ("Fluconazole 150","[\"Fluconazole 150mg\"]","Pfizer","Capsule","H1","[\"Forcan 150\",\"Zocon 150\",\"Flucos 150\"]",42.0,"1","Vaginal candidiasis","Liver disease","no"),
              ("Ibuprofen 400","[\"Ibuprofen 400mg\"]","Abbott","Tablet","OTC","[\"Brufen 400\",\"Combiflam\",\"Advil 400\"]",25.0,"1x15","Pain, Fever, Inflammation","Peptic ulcer, Renal failure","yes"),
              ("Diclofenac 50","[\"Diclofenac 50mg\"]","Novartis","Tablet","H","[\"Voveran 50\",\"Diclomol 50\",\"Reactin 50\"]",18.0,"1x10","Musculoskeletal pain","CV disease, GI ulcer","no"),
              ("Rabeprazole 20","[\"Rabeprazole 20mg\"]","Eisai","Tablet","H","[\"Razo 20\",\"Rablet 20\",\"Pariet 20\"]",55.0,"1x15","GERD, H.pylori eradication","Hypersensitivity","no"),
              ("Vitamin D3 60k","[\"Cholecalciferol 60000IU\"]","Sun Pharma","Capsule","OTC","[\"Calcirol 60k\",\"D3 Must\",\"Tayo 60k\"]",55.0,"1","Vitamin D deficiency","Hypercalcaemia","yes"),
              ("B-Complex","[\"Vitamin B1\",\"B2\",\"B6\",\"B12\",\"Niacinamide\"]","Various","Tablet","OTC","[\"Neurobion\",\"Beplex\",\"Cobadex CZS\"]",28.0,"1x30","Vitamin B deficiency, Neuropathy","","yes"),
              ("Calcium + D3","[\"Calcium Carbonate 500mg\",\"Vitamin D3 250IU\"]","Procter","Tablet","OTC","[\"Shelcal 500\",\"Calcimax\",\"Ostocalcium\"]",85.0,"1x30","Osteoporosis, Calcium deficiency","Hypercalcaemia","yes"),
              ("Iron + Folic","[\"Ferrous Sulphate\",\"Folic Acid 5mg\"]","Various","Tablet","H","[\"Autrin\",\"Orofer\",\"Irofer\"]",25.0,"1x30","Anaemia in pregnancy","Haemochromatosis","no"),
              ("Ondansetron 4","[\"Ondansetron 4mg\"]","GSK","Tablet","H","[\"Zofran 4\",\"Emeset 4\",\"Vomikind 4\"]",35.0,"1x10","Nausea, Vomiting, Chemotherapy","QT prolongation","no"),
              ("Domperidone 10","[\"Domperidone 10mg\"]","Sanofi","Tablet","H","[\"Motilium 10\",\"Domstal 10\",\"Vomistop 10\"]",18.0,"1x15","Nausea, Gastroparesis","Cardiac risk","no"),
            ]
            conn.executemany("""
                INSERT INTO drug_combinations
                (brand_name,generic_names,manufacturer,category,schedule_type,alternatives,mrp,pack_size,uses,contraindications,is_otc)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, combos)

        # ── App settings (key-value store) ───────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key   TEXT PRIMARY KEY,
                value TEXT DEFAULT ''
            )
        """)

        # ── Performance indexes ──────────────────────────────────────────
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bills_ts ON bills(ts)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bills_date ON bills(date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_medicines_n ON medicines(n)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_customers_name ON customers(name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_regs_status ON platform_registrations(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_narcotic_date ON narcotic_register(date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_reminders_status ON patient_reminders(status)")

    migrate_db()


def migrate_db() -> None:
    with get_conn() as conn:
        # Purchases migration
        p_cols = table_columns(conn, "purchases")
        if "batch" not in p_cols:
            conn.execute("ALTER TABLE purchases ADD COLUMN batch TEXT DEFAULT ''")
        if "expiry" not in p_cols:
            conn.execute("ALTER TABLE purchases ADD COLUMN expiry TEXT DEFAULT ''")
        if "photo" not in p_cols:
            conn.execute("ALTER TABLE purchases ADD COLUMN photo TEXT DEFAULT ''")

        # Bills migration
        b_cols = table_columns(conn, "bills")
        if "rx" not in b_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN rx TEXT DEFAULT ''")
        if "prescription" not in b_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN prescription TEXT DEFAULT ''")
        if "bill_type" not in b_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN bill_type TEXT DEFAULT 'retail'")
        if "customer_type" not in b_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN customer_type TEXT DEFAULT 'customer'")
        if "whatsapp_sent" not in b_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN whatsapp_sent INTEGER DEFAULT 0")
        if "staff_name" not in b_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN staff_name TEXT DEFAULT ''")

        # Customers migration
        c_cols = table_columns(conn, "customers")
        if "address" not in c_cols:
            conn.execute("ALTER TABLE customers ADD COLUMN address TEXT DEFAULT ''")
        if "email" not in c_cols:
            conn.execute("ALTER TABLE customers ADD COLUMN email TEXT DEFAULT ''")
        if "face_vector" not in c_cols:
            conn.execute("ALTER TABLE customers ADD COLUMN face_vector TEXT DEFAULT ''")

        # Doctors migration
        d_cols = table_columns(conn, "doctors")
        if "email" not in d_cols:
            conn.execute("ALTER TABLE doctors ADD COLUMN email TEXT DEFAULT ''")

        # Suppliers migration
        s_cols = table_columns(conn, "suppliers")
        if "gst" not in s_cols:
            conn.execute("ALTER TABLE suppliers ADD COLUMN gst TEXT DEFAULT ''")
        if "last_order" not in s_cols:
            conn.execute("ALTER TABLE suppliers ADD COLUMN last_order TEXT DEFAULT '-'")
        if "status" not in s_cols:
            conn.execute("ALTER TABLE suppliers ADD COLUMN status TEXT DEFAULT 'Active'")

        # Medicines extra columns
        m_cols = table_columns(conn, "medicines")
        med_new_cols = [
            ("batch", "TEXT DEFAULT ''"),
            ("expiry", "TEXT DEFAULT ''"),
            ("p_rate", "REAL DEFAULT 0"),
            ("p_packing", "TEXT DEFAULT ''"),
            ("s_packing", "TEXT DEFAULT ''"),
            ("p_gst", "REAL DEFAULT 0"),
            ("s_gst", "REAL DEFAULT 0"),
            ("disc", "REAL DEFAULT 0"),
            ("offer", "TEXT DEFAULT ''"),
            ("reorder", "INTEGER DEFAULT 0"),
            ("max_qty", "INTEGER DEFAULT 0"),
        ]
        for col, col_type in med_new_cols:
            if col not in m_cols:
                conn.execute(f"ALTER TABLE medicines ADD COLUMN {col} {col_type}")

        # ── Expanded drug combinations (adds 60+ more if DB has <35) ────
        dc_count = conn.execute("SELECT COUNT(*) AS c FROM drug_combinations").fetchone()["c"]
        if dc_count < 35:
            now_ts = datetime.now(timezone.utc).isoformat()
            extra_combos = [
              # PAIN / NSAID
              ("Ketorolac 10","[\"Ketorolac Tromethamine 10mg\"]","Sun Pharma","Tablet","H","[\"Ketanov 10\",\"Toradol 10\",\"Ketlur 10\"]",38.0,"1x10","Short-term acute pain","Renal impairment, GI bleed","no"),
              ("Tramadol 50","[\"Tramadol HCl 50mg\"]","Grunenthal","Capsule","X","[\"Ultracet 50\",\"Tramazac 50\",\"Contramal 50\"]",28.0,"1x10","Moderate-severe pain","Opioid addiction, Seizures","no"),
              ("Tramadol+Para","[\"Tramadol 37.5mg\",\"Paracetamol 325mg\"]","Janssen","Tablet","X","[\"Ultracet\",\"Tramazac P\",\"Combifin P\"]",45.0,"1x10","Moderate pain","Same as Tramadol","no"),
              ("Aceclofenac 100","[\"Aceclofenac 100mg\"]","Intas","Tablet","H","[\"Hifenac 100\",\"Zerodol 100\",\"Acenac 100\"]",32.0,"1x10","Arthritis, Musculoskeletal pain","CV disease, GI ulcer","no"),
              ("Aceclofenac+PCM","[\"Aceclofenac 100mg\",\"Paracetamol 325mg\"]","Intas","Tablet","H","[\"Hifenac P\",\"Zerodol P\",\"Combigesic\"]",38.0,"1x10","Pain and fever combined","CV risk","no"),
              ("Etoricoxib 60","[\"Etoricoxib 60mg\"]","MSD","Tablet","H","[\"Nucoxia 60\",\"Etova 60\",\"Torocoxia 60\"]",68.0,"1x7","Arthritis, Acute gout","CV disease","no"),
              ("Meloxicam 15","[\"Meloxicam 15mg\"]","Boehringer","Tablet","H","[\"Muvera 15\",\"Mobizox 15\",\"Melonex 15\"]",28.0,"1x10","Osteoarthritis, RA","Renal failure, Asthma","no"),
              # ANTIBIOTICS
              ("Amoxicillin 500","[\"Amoxicillin 500mg\"]","GSK","Capsule","H1","[\"Mox 500\",\"Novamox 500\",\"Biomox 500\"]",42.0,"1x10","Bacterial infections","Penicillin allergy","no"),
              ("Doxycycline 100","[\"Doxycycline 100mg\"]","Pfizer","Capsule","H1","[\"Vibramycin 100\",\"Doxt 100\",\"Biodoxi 100\"]",22.0,"1x10","Infections, Malaria prophylaxis","Pregnancy, Children <8yr","no"),
              ("Ciprofloxacin 500","[\"Ciprofloxacin 500mg\"]","Bayer","Tablet","H1","[\"Cifran 500\",\"Ciplox 500\",\"Ciprolet 500\"]",48.0,"1x10","UTI, RTI, GI infections","Tendon rupture risk","no"),
              ("Levofloxacin 500","[\"Levofloxacin 500mg\"]","Sanofi","Tablet","H1","[\"Levoflox 500\",\"Tavanic 500\",\"Levox 500\"]",85.0,"1x5","Pneumonia, UTI, Skin infections","QT prolongation","no"),
              ("Cefixime 200","[\"Cefixime 200mg\"]","Alkem","Tablet","H1","[\"Taxim-O 200\",\"Zifi 200\",\"Magnex 200\"]",68.0,"1x10","UTI, RTI, STI","Cephalosporin allergy","no"),
              ("Cephalexin 500","[\"Cephalexin 500mg\"]","Ranbaxy","Capsule","H1","[\"Sporidex 500\",\"Cefadur 500\",\"Ceporex 500\"]",55.0,"1x10","Skin, bone, UTI infections","Cephalosporin allergy","no"),
              ("Clindamycin 300","[\"Clindamycin 300mg\"]","Pfizer","Capsule","H1","[\"Dalacin C 300\",\"Clintas 300\",\"Clindac 300\"]",92.0,"1x8","Anaerobic infections, Dental","Clostridium colitis","no"),
              ("Clarithromycin 500","[\"Clarithromycin 500mg\"]","Abbott","Tablet","H1","[\"Klacid 500\",\"Claribid 500\",\"Myclaris 500\"]",125.0,"1x10","H.pylori, RTI, Skin infections","QT prolongation, Liver disease","no"),
              # ANTIDIABETICS
              ("Voglibose 0.2","[\"Voglibose 0.2mg\"]","Takeda","Tablet","H","[\"Volix 0.2\",\"Vgb 0.2\",\"Glynase Vog\"]",32.0,"1x10","Type 2 diabetes post-meal sugar","GI disorders","no"),
              ("Sitagliptin 100","[\"Sitagliptin 100mg\"]","MSD","Tablet","H","[\"Januvia 100\",\"Istavel 100\",\"Zita 100\"]",148.0,"1x10","Type 2 diabetes","Pancreatitis","no"),
              ("Gliclazide 80","[\"Gliclazide 80mg\"]","Servier","Tablet","H","[\"Diamicron 80\",\"Glizide 80\",\"Orzid 80\"]",28.0,"1x10","Type 2 diabetes","Hypoglycaemia risk","no"),
              ("Dapagliflozin 10","[\"Dapagliflozin 10mg\"]","AstraZeneca","Tablet","H","[\"Forxiga 10\",\"Dapaglu 10\",\"Oxra 10\"]",185.0,"1x10","Type 2 DM, Heart failure","UTI, DKA risk","no"),
              ("Insulin Glargine","[\"Insulin Glargine 100U/mL\"]","Sanofi","Injection","H","[\"Lantus\",\"Toujeo\",\"Basalog\"]",680.0,"3mL pen","Diabetes (basal insulin)","Hypoglycaemia","no"),
              # ANTIHYPERTENSIVES
              ("Enalapril 5","[\"Enalapril Maleate 5mg\"]","MSD","Tablet","H","[\"Enam 5\",\"Envas 5\",\"Renitec 5\"]",22.0,"1x15","Hypertension, Heart failure","Pregnancy, Angioedema","no"),
              ("Ramipril 5","[\"Ramipril 5mg\"]","Sanofi","Capsule","H","[\"Cardace 5\",\"Rami 5\",\"Tritace 5\"]",38.0,"1x15","Hypertension, Post-MI","Pregnancy","no"),
              ("Olmesartan 20","[\"Olmesartan Medoxomil 20mg\"]","Daiichi","Tablet","H","[\"Olmetrack 20\",\"Olsar 20\",\"Benicar 20\"]",62.0,"1x15","Hypertension","Pregnancy, Biliary obstruction","no"),
              ("Bisoprolol 5","[\"Bisoprolol Fumarate 5mg\"]","Merck","Tablet","H","[\"Corbis 5\",\"Concor 5\",\"Bisocor 5\"]",45.0,"1x15","Hypertension, Heart failure, Angina","Asthma, Heart block","no"),
              ("Carvedilol 6.25","[\"Carvedilol 6.25mg\"]","Roche","Tablet","H","[\"Carca 6.25\",\"C-Carvedil 6.25\",\"Dilatrend 6.25\"]",38.0,"1x15","Hypertension, Heart failure","Asthma","no"),
              ("Spironolactone 25","[\"Spironolactone 25mg\"]","Pfizer","Tablet","H","[\"Aldactone 25\",\"Spiroderm 25\",\"Lasilactone\"]",28.0,"1x15","Heart failure, Hyperaldosteronism","Hyperkalaemia","no"),
              ("Furosemide 40","[\"Furosemide 40mg\"]","Sanofi","Tablet","H","[\"Lasix 40\",\"Frusamide 40\",\"Diurin 40\"]",8.0,"1x15","Oedema, Heart failure, Hypertension","Anuria, Dehydration","no"),
              ("Chlorthalidone 12.5","[\"Chlorthalidone 12.5mg\"]","MSD","Tablet","H","[\"Hygroton 12.5\",\"Thalitone\",\"Aquadon\"]",22.0,"1x15","Hypertension, Oedema","Anuria, Sulfonamide allergy","no"),
              # CARDIAC
              ("Clopidogrel 75","[\"Clopidogrel 75mg\"]","Sanofi","Tablet","H","[\"Plavix 75\",\"Clopilet 75\",\"Deplatt 75\"]",28.0,"1x15","Antiplatelet, ACS, Stroke prevention","Active bleeding, Peptic ulcer","no"),
              ("Digoxin 0.25","[\"Digoxin 0.25mg\"]","GSK","Tablet","H","[\"Lanoxin 0.25\",\"Digicor\",\"Lanicor\"]",12.0,"1x30","Atrial fibrillation, Heart failure","AV block, Hypokalaemia","no"),
              ("Isosorbide Mono","[\"Isosorbide Mononitrate 20mg\"]","Schwarz","Tablet","H","[\"Imdur 20\",\"Ismo 20\",\"Monosorb 20\"]",28.0,"1x15","Angina prophylaxis","Hypotension, Sildenafil use","no"),
              # CNS
              ("Alprazolam 0.25","[\"Alprazolam 0.25mg\"]","Pfizer","Tablet","H1","[\"Alprax 0.25\",\"Restyl 0.25\",\"Tranax 0.25\"]",8.0,"1x10","Anxiety, Panic disorder","Pregnancy, Substance abuse","no"),
              ("Clonazepam 0.5","[\"Clonazepam 0.5mg\"]","Roche","Tablet","H1","[\"Rivotril 0.5\",\"Lonazep 0.5\",\"Clonotril 0.5\"]",12.0,"1x15","Anxiety, Epilepsy, Panic disorder","Liver disease, Pregnancy","no"),
              ("Escitalopram 10","[\"Escitalopram 10mg\"]","Lundbeck","Tablet","H","[\"Nexito 10\",\"Rexipra 10\",\"Cipralex 10\"]",68.0,"1x15","Depression, Anxiety, OCD","MAOI use, QT prolongation","no"),
              ("Sertraline 50","[\"Sertraline 50mg\"]","Pfizer","Tablet","H","[\"Zoloft 50\",\"Serta 50\",\"Daxid 50\"]",45.0,"1x15","Depression, OCD, PTSD, Panic","MAOI, Pimozide use","no"),
              ("Amitriptyline 10","[\"Amitriptyline HCl 10mg\"]","Merck","Tablet","H","[\"Tryptomer 10\",\"Elavil 10\",\"Amitril 10\"]",12.0,"1x30","Depression, Neuropathic pain, Migraine","CV disease, Glaucoma","no"),
              ("Pregabalin 75","[\"Pregabalin 75mg\"]","Pfizer","Capsule","H1","[\"Lyrica 75\",\"Pregabid 75\",\"Nervigesic 75\"]",95.0,"1x15","Neuropathic pain, Fibromyalgia","Substance abuse risk","no"),
              ("Gabapentin 300","[\"Gabapentin 300mg\"]","Pfizer","Capsule","H","[\"Gabapin 300\",\"Neurontin 300\",\"Gabantin 300\"]",48.0,"1x10","Neuropathic pain, Epilepsy","Renal impairment","no"),
              ("Quetiapine 50","[\"Quetiapine 50mg\"]","AstraZeneca","Tablet","H","[\"Seroquel 50\",\"Qutan 50\",\"Qutipin 50\"]",85.0,"1x10","Schizophrenia, Bipolar, Depression","Diabetes, Elderly","no"),
              ("Olanzapine 5","[\"Olanzapine 5mg\"]","Eli Lilly","Tablet","H","[\"Oleanz 5\",\"Ozace 5\",\"Olanpine 5\"]",62.0,"1x10","Schizophrenia, Bipolar","Metabolic syndrome","no"),
              # THYROID / HORMONES
              ("Propylthiouracil 50","[\"Propylthiouracil 50mg\"]","Nycomed","Tablet","H","[\"PTU 50\",\"Propycil 50\",\"Thiouracil 50\"]",28.0,"1x30","Hyperthyroidism","Agranulocytosis","no"),
              ("Progesterone 200","[\"Progesterone 200mg\"]","IPCA","Capsule","H","[\"Susten 200\",\"Prometrium 200\",\"Naturogest 200\"]",95.0,"1x10","Luteal support, Threatened abortion","Liver disease","no"),
              ("Clomiphene 50","[\"Clomiphene Citrate 50mg\"]","MSD","Tablet","H","[\"Clofert 50\",\"Siphene 50\",\"Fertomid 50\"]",38.0,"1x10","Ovulation induction, PCOS","Ovarian cysts","no"),
              # GASTRO
              ("Esomeprazole 40","[\"Esomeprazole 40mg\"]","AstraZeneca","Capsule","H","[\"Nexium 40\",\"Esoz 40\",\"Nexpro 40\"]",88.0,"1x15","GERD, Peptic ulcer, Zollinger-Ellison","Atrophic gastritis","no"),
              ("Itopride 50","[\"Itopride HCl 50mg\"]","Abbott","Tablet","H","[\"Ganaton 50\",\"Torid 50\",\"Elipride 50\"]",48.0,"1x15","Functional dyspepsia, Gastroparesis","","no"),
              ("Mesalamine 400","[\"Mesalamine 400mg\"]","Ferring","Tablet","H","[\"Asacol 400\",\"Mesacol 400\",\"Pentasa 400\"]",145.0,"1x30","Ulcerative colitis, Crohn's disease","Salicylate allergy, Renal failure","no"),
              ("Lactulose Syrup","[\"Lactulose 10g/15mL\"]","Duphalac","Syrup","OTC","[\"Duphalac\",\"Lactohep\",\"Regulax\"]",180.0,"200mL","Constipation, Hepatic encephalopathy","Galactosaemia","yes"),
              ("Loperamide 2","[\"Loperamide HCl 2mg\"]","Janssen","Capsule","OTC","[\"Imodium 2\",\"Lopamide 2\",\"Eldoper 2\"]",22.0,"1x6","Acute diarrhoea","Bacterial colitis, Children <2yr","yes"),
              # RESPIRATORY
              ("Montelukast+Levocet","[\"Montelukast 10mg\",\"Levocetirizine 5mg\"]","MSD","Tablet","H","[\"Montair LC\",\"Aerocort Plus\",\"Levosiz M\"]",78.0,"1x15","Allergic rhinitis with asthma","Hepatic impairment","no"),
              ("Tiotropium Inhaler","[\"Tiotropium 18mcg\"]","Boehringer","Inhaler","H","[\"Spiriva\",\"Tiova\",\"Braltus\"]",450.0,"30 doses","COPD maintenance","Narrow-angle glaucoma","no"),
              ("N-Acetyl Cysteine","[\"N-Acetylcysteine 600mg\"]","Zambon","Tablet","H","[\"NAC 600\",\"Acetin 600\",\"Fluimucil 600\"]",65.0,"1x10","Mucolytic, Paracetamol overdose","Asthma","no"),
              ("Levosalbutamol","[\"Levosalbutamol 50mcg\"]","GSK","Inhaler","H","[\"Levolin\",\"Albulair\",\"Leva-Sal\"]",120.0,"200 doses","Asthma, Bronchospasm","Tachyarrhythmia","no"),
              # VITAMINS / SUPPLEMENTS
              ("Omega 3","[\"EPA 180mg\",\"DHA 120mg\"]","USV","Capsule","OTC","[\"Maxepa\",\"Seacod\",\"Biomega\"]",95.0,"1x30","Hyperlipidaemia, CV protection","Bleeding disorders","yes"),
              ("Zinc Sulphate","[\"Zinc Sulphate 20mg\"]","Dr Reddy's","Tablet","OTC","[\"Zincovit\",\"Zicovit\",\"Zinc-D\"]",28.0,"1x15","Zinc deficiency, Diarrhoea in children","Copper deficiency","yes"),
              ("Folic Acid 5","[\"Folic Acid 5mg\"]","Various","Tablet","OTC","[\"Folvite 5\",\"Folicin 5\",\"Folsafe 5\"]",12.0,"1x30","Megaloblastic anaemia, Pregnancy","Undiagnosed anaemia","yes"),
              ("Methylcobalamin 500","[\"Methylcobalamin 500mcg\"]","Ranbaxy","Tablet","OTC","[\"Methycobal 500\",\"Mecobalamin 500\",\"Cobadex 500\"]",38.0,"1x30","B12 deficiency, Neuropathy","Polycythaemia","yes"),
              ("Pyridoxine 40","[\"Pyridoxine HCl 40mg\"]","Various","Tablet","OTC","[\"Benadon 40\",\"P6 forte\",\"Vitabex\"]",18.0,"1x30","Morning sickness, B6 deficiency","","yes"),
              # SKIN / DERMA
              ("Clobetasol Cream","[\"Clobetasol Propionate 0.05%\"]","GSK","Cream","H","[\"Dermovate\",\"Tenovate\",\"Lobate\"]",68.0,"30g","Psoriasis, Eczema, Dermatitis","Infected skin, Face/groin","no"),
              ("Betamethasone Cream","[\"Betamethasone 0.1%\"]","GSK","Cream","H","[\"Betnovate\",\"Betnol\",\"Celestoderm\"]",48.0,"20g","Eczema, Dermatitis, Psoriasis","Fungal infections","no"),
              ("Clotrimazole Cream","[\"Clotrimazole 1%\"]","Bayer","Cream","OTC","[\"Canesten 1%\",\"Candid\",\"Clotrin\"]",35.0,"20g","Fungal skin infections, Ringworm","Hypersensitivity","yes"),
              ("Ketoconazole Shampoo","[\"Ketoconazole 2%\"]","Janssen","Shampoo","H","[\"Nizoral\",\"Ketomed\",\"Fungicide\"]",165.0,"75mL","Dandruff, Seborrhoeic dermatitis","Hepatic disease","no"),
              ("Isotretinoin 10","[\"Isotretinoin 10mg\"]","Roche","Capsule","H","[\"Accutane 10\",\"Isotroin 10\",\"Tretiva 10\"]",185.0,"1x10","Severe acne, Acne vulgaris","Pregnancy — CONTRAINDICATED","no"),
              ("Permethrin 5%","[\"Permethrin 5%\"]","GSK","Cream","H","[\"Lyclear 5%\",\"Scabisan 5%\",\"Permite 5%\"]",95.0,"60g","Scabies","Hypersensitivity","no"),
              # EYE / ENT
              ("Chloramphenicol Eye","[\"Chloramphenicol 0.5%\"]","Pfizer","Eye Drops","H","[\"Chloromycetin Eye\",\"Danocrine Eye\",\"Aquamycetin\"]",38.0,"10mL","Bacterial conjunctivitis","Aplastic anaemia risk","no"),
              ("Ciprofloxacin Eye","[\"Ciprofloxacin 0.3%\"]","Alcon","Eye Drops","H","[\"Ciplox Eye\",\"Ciloxan\",\"Ciprodex\"]",48.0,"10mL","Bacterial eye infections","Fungal keratitis","no"),
              ("Ofloxacin Ear","[\"Ofloxacin 0.3%\"]","Alcon","Ear Drops","H","[\"Panotile\",\"Floxin Otic\",\"Tarivid Ear\"]",55.0,"5mL","Otitis externa, Otitis media","Perforated eardrum (caution)","no"),
              ("Mometasone Nasal","[\"Mometasone Furoate 50mcg\"]","MSD","Nasal Spray","H","[\"Nasonex\",\"Momeflo\",\"Mometide\"]",285.0,"140 doses","Allergic rhinitis, Nasal polyps","Active nasal infection","no"),
              # GYNAEC / UROLOGY
              ("Tranexamic 500","[\"Tranexamic Acid 500mg\"]","Pfizer","Tablet","H","[\"Trapic 500\",\"Cyclokapron 500\",\"Traxyl 500\"]",28.0,"1x10","Menorrhagia, Fibrinolysis","Thromboembolic risk","no"),
              ("Tamsulosin 0.4","[\"Tamsulosin HCl 0.4mg\"]","Boehringer","Capsule","H","[\"Urimax 0.4\",\"Tamlet 0.4\",\"Flomax 0.4\"]",68.0,"1x15","BPH, Urinary retention","Orthostatic hypotension","no"),
              ("Sildenafil 50","[\"Sildenafil Citrate 50mg\"]","Pfizer","Tablet","H","[\"Viagra 50\",\"Penegra 50\",\"Manforce 50\"]",125.0,"1x4","Erectile dysfunction, PAH","Nitrate use, Hypotension","no"),
              # ANTI-PARASITIC / MALARIA
              ("Albendazole 400","[\"Albendazole 400mg\"]","GSK","Tablet","OTC","[\"Zentel 400\",\"Albenza 400\",\"Noworm 400\"]",18.0,"1","Intestinal worms, Giardia","Pregnancy first trimester","yes"),
              ("Ivermectin 6","[\"Ivermectin 6mg\"]","MSD","Tablet","H","[\"Mectizan 6\",\"Iverjohn 6\",\"Stromectol 6\"]",22.0,"1","Onchocerciasis, Scabies","CNS disorders","no"),
              ("Hydroxychloroquine 200","[\"Hydroxychloroquine 200mg\"]","Sanofi","Tablet","H","[\"Plaquenil 200\",\"HCQS 200\",\"Lariago 200\"]",45.0,"1x10","Malaria, Lupus, RA","Retinal disease, QT prolongation","no"),
              ("Artemether+Lumef","[\"Artemether 20mg\",\"Lumefantrine 120mg\"]","Novartis","Tablet","H","[\"Coartem\",\"Lumartem\",\"Synriam\"]",245.0,"24 tabs","Falciparum malaria","QT prolongation","no"),
              # VACCINES / BIOLOGICS
              ("Hepatitis B Vaccine","[\"HBsAg Recombinant\"]","Serum Institute","Injection","H","[\"Engerix-B\",\"Recombivax\",\"Shanvac-B\"]",125.0,"1 dose","Hepatitis B prevention","Yeast allergy","no"),
            ]
            conn.executemany("""
                INSERT INTO drug_combinations
                (brand_name,generic_names,manufacturer,category,schedule_type,alternatives,mrp,pack_size,uses,contraindications,is_otc)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, extra_combos)

        # ── Expanded narcotic schedules (adds more if count < 50) ─────────
        ds_count = conn.execute("SELECT COUNT(*) AS c FROM drug_schedules").fetchone()["c"]
        if ds_count < 50:
            now_ts = datetime.now(timezone.utc).isoformat()
            extra_schedules = [
              # Additional Schedule X (Narcotic/Psychotropic)
              ("Tramadol","Tramadol HCl","X",1,0,1,0,1,"Opioid analgesic — Schedule X, narcotic register mandatory",now_ts),
              ("Alprazolam","Alprazolam","X",0,1,1,0,1,"Benzodiazepine — Schedule X psychotropic",now_ts),
              ("Clonazepam","Clonazepam","X",0,1,1,0,1,"Benzodiazepine anticonvulsant — Schedule X",now_ts),
              ("Nitrazepam","Nitrazepam","X",0,1,1,0,1,"Benzodiazepine hypnotic — Schedule X",now_ts),
              ("Diazepam","Diazepam","X",0,1,1,0,1,"Benzodiazepine — Schedule X",now_ts),
              ("Lorazepam","Lorazepam","X",0,1,1,0,1,"Benzodiazepine — Schedule X",now_ts),
              ("Zolpidem","Zolpidem","X",0,1,1,0,1,"Non-BZD hypnotic — Schedule X",now_ts),
              ("Fentanyl","Fentanyl","X",1,0,1,0,1,"Opioid — Schedule X; patches need narcotic register",now_ts),
              ("Hydrocodone","Hydrocodone","X",1,0,1,0,1,"Opioid analgesic — Schedule X",now_ts),
              ("Oxycodone","Oxycodone","X",1,0,1,0,1,"Opioid analgesic — Schedule X",now_ts),
              ("Buprenorphine","Buprenorphine","X",1,0,1,0,1,"Opioid agonist/antagonist — Schedule X, narcotic register",now_ts),
              ("Pentazocine","Pentazocine","X",1,0,1,0,1,"Opioid agonist/antagonist — Schedule X",now_ts),
              ("Methylphenidate","Methylphenidate","X",0,1,1,0,1,"CNS stimulant — Schedule X psychotropic",now_ts),
              ("Amphetamine","Amphetamine","X",0,1,1,0,1,"CNS stimulant — Schedule X",now_ts),
              # Additional Schedule H1 (Antibiotics needing Rx + register)
              ("Metronidazole","Metronidazole","H1",0,0,1,0,0,"Nitroimidazole antibiotic — Schedule H1",now_ts),
              ("Tinidazole","Tinidazole","H1",0,0,1,0,0,"Nitroimidazole — Schedule H1",now_ts),
              ("Amikacin","Amikacin","H1",0,0,1,0,0,"Aminoglycoside — Schedule H1",now_ts),
              ("Gentamicin","Gentamicin","H1",0,0,1,0,0,"Aminoglycoside — Schedule H1",now_ts),
              ("Vancomycin","Vancomycin","H1",0,0,1,0,0,"Glycopeptide — Schedule H1; hospital use",now_ts),
              ("Linezolid","Linezolid","H1",0,0,1,0,0,"Oxazolidinone — Schedule H1; MRSA",now_ts),
              ("Meropenem","Meropenem","H1",0,0,1,0,0,"Carbapenem — Schedule H1; hospital use",now_ts),
              ("Teicoplanin","Teicoplanin","H1",0,0,1,0,0,"Glycopeptide — Schedule H1",now_ts),
              ("Piperacillin+Tazobactam","Piperacillin Tazobactam","H1",0,0,1,0,0,"Beta-lactam/BLI — Schedule H1",now_ts),
              ("Colistin","Colistin Sulphate","H1",0,0,1,0,0,"Polymyxin — Schedule H1; last resort antibiotic",now_ts),
              ("Pregabalin","Pregabalin","H1",0,0,1,0,0,"Gabapentinoid — Schedule H1; abuse potential",now_ts),
              ("Tapentadol","Tapentadol","H1",1,0,1,0,1,"Opioid — Schedule H1 in some states, X in others",now_ts),
            ]
            conn.executemany("""
                INSERT INTO drug_schedules
                (medicine_name,generic_name,schedule_type,narcotic,psychotropic,requires_rx,max_qty_no_rx,inspector_log,notes,created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, extra_schedules)


def normalize_bill_row(row: sqlite3.Row) -> dict[str, Any]:
    data = dict(row)
    items = safe_json_loads(data.get("items"), [])
    if not isinstance(items, list):
        items = []
    prescription = data.get("prescription", "") or data.get("rx", "") or ""
    return {
        "id": data.get("id"),
        "ts": data.get("ts"),
        "date": data.get("date"),
        "cust": data.get("cust"),
        "phone": data.get("phone"),
        "pay": data.get("pay"),
        "sub": data.get("sub"),
        "disc": data.get("disc"),
        "tax": data.get("tax"),
        "total": data.get("total"),
        "items": items,
        "doctor": data.get("doctor", "Self"),
        "prescription": prescription,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TEMPLATE ROUTES (HTML Pages)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/")
def root():
    """Public welcome/landing page"""
    return render_template("welcome.html")

@app.route("/internal")
def internal_login():
    """Login page for internal POS"""
    return render_template("login.html")


@app.route("/dashboard")
def dashboard():
    """Main internal POS/billing dashboard"""
    current_user = {
        "id":   session.get("staff_id", 0),
        "name": session.get("staff_name", "Owner"),
        "role": session.get("staff_role", "Manager"),
    }
    return render_template("dashboard.html", current_user=current_user)


@app.route("/wanted")
def wanted():
    """Render wanted (manual) page"""
    return render_template("wanted.html")


@app.route("/mfrchange2")
def mfr_change_detail_page():
    """Render manufacturer/location change page"""
    return render_template("mfrchange2.html")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# API ROUTES (JSON Data)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/login")
def login_redirect():
    """Friendly /login URL → redirect to the staff POS login screen."""
    return redirect("/internal")


@app.route("/portal-login")
def portal_login_page():
    """Retail portal login page (sel / demo123)."""
    return render_template("portal_login.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    """Staff login — checks PIN against staff table; fallback: owner/1234."""
    data     = request.get_json(silent=True) or {}
    username = (data.get("username","") or "").strip().lower()
    password = (data.get("password","") or "").strip()

    # Check staff PIN
    with get_conn() as conn:
        staff = [dict(r) for r in conn.execute(
            "SELECT * FROM staff WHERE is_active=1").fetchall()]

    # Match by name (case-insensitive) + PIN
    matched = None
    for s in staff:
        if s["name"].lower() == username and s.get("pin","") == password:
            matched = s
            break

    # Fallback: any staff member if username matches 'owner' with pin '1234'
    if not matched and username == "owner" and password == "1234":
        matched = next((s for s in staff if s["role"] in ("Manager","Owner")), staff[0] if staff else None)

    if matched:
        session["staff_id"]   = matched["id"]
        session["staff_name"] = matched["name"]
        session["staff_role"] = matched["role"]
        return jsonify({"status": "ok", "name": matched["name"], "role": matched["role"]})

    return jsonify({"status": "error", "message": "Invalid username or PIN"}), 401


@app.route("/api/health", methods=["GET"])
def api_health():
    """Health check endpoint"""
    with get_conn() as conn:
        med_count = conn.execute("SELECT COUNT(*) AS c FROM medicines").fetchone()["c"]
        bill_count = conn.execute("SELECT COUNT(*) AS c FROM bills").fetchone()["c"]
        now = datetime.now(timezone.utc).isoformat() + "Z"
    return jsonify(
        {
            "status": "ok",
            "database_path": DB_PATH,
            "medicines": med_count,
            "bills": bill_count,
            "time_utc": now,
        }
    )


@app.route("/api/kpis", methods=["GET"])
def get_kpis():
    """Lightweight KPI summary — no need to fetch all bills for dashboard cards."""
    date_str = request.args.get("date", "")
    try:
        d = datetime.strptime(date_str, "%d/%m/%Y") if date_str else datetime.now()
    except ValueError:
        d = datetime.now()

    ds = f"{d.day:02d}/{d.month:02d}/{d.year}"
    yd = d - timedelta(days=1)
    yds = f"{yd.day:02d}/{yd.month:02d}/{yd.year}"
    cutoff_90 = (datetime.now() + timedelta(days=90)).strftime("%Y-%m-%d")
    today_iso = datetime.now().strftime("%Y-%m-%d")

    with get_conn() as conn:
        today = conn.execute(
            "SELECT SUM(total) AS t, COUNT(*) AS c FROM bills WHERE date LIKE ?",
            (f"{ds}%",),
        ).fetchone()
        yesterday = conn.execute(
            "SELECT SUM(total) AS t, COUNT(*) AS c FROM bills WHERE date LIKE ?",
            (f"{yds}%",),
        ).fetchone()
        total_bills = conn.execute("SELECT COUNT(*) AS c FROM bills").fetchone()["c"]
        low_stock = conn.execute(
            "SELECT COUNT(*) AS c FROM medicines WHERE s < 15"
        ).fetchone()["c"]
        expiry_alerts = conn.execute(
            "SELECT COUNT(*) AS c FROM medicines WHERE expiry != '' AND expiry <= ? AND expiry >= ?",
            (cutoff_90, today_iso),
        ).fetchone()["c"]

    return jsonify({
        "today_bills": today["c"] or 0,
        "today_revenue": float(today["t"] or 0),
        "yesterday_bills": yesterday["c"] or 0,
        "yesterday_revenue": float(yesterday["t"] or 0),
        "total_bills": total_bills,
        "low_stock_count": low_stock,
        "expiry_alert_count": expiry_alerts,
    })


@app.route("/api/backup")
def backup_db():
    """Download database backup"""
    return send_file(DB_PATH, as_attachment=True)


# --- BILLS ---
@app.route("/api/bills", methods=["GET"])
def get_bills():
    """Get bills with optional pagination. Without ?limit returns all (backward compat)."""
    limit = request.args.get("limit", type=int)
    offset = request.args.get("offset", 0, type=int)
    with get_conn() as conn:
        if limit is not None:
            rows = conn.execute(
                "SELECT * FROM bills ORDER BY ts DESC LIMIT ? OFFSET ?",
                (limit, offset),
            ).fetchall()
            total = conn.execute("SELECT COUNT(*) AS c FROM bills").fetchone()["c"]
            return jsonify({
                "bills": [normalize_bill_row(r) for r in rows],
                "total": total,
                "limit": limit,
                "offset": offset,
            })
        rows = conn.execute("SELECT * FROM bills ORDER BY ts DESC").fetchall()
    return jsonify([normalize_bill_row(row) for row in rows])


@app.route("/api/bills", methods=["POST"])
def save_bill():
    """Create a new bill"""
    data = request.get_json(silent=True) or {}
    required = ["id", "ts", "date", "cust", "phone", "pay", "sub", "disc", "tax", "total", "items"]
    missing = required_fields(data, required)
    if missing:
        return json_error("Missing required bill fields", 400, missing)
    if not isinstance(data["items"], list) or len(data["items"]) == 0:
        return json_error("Bill must include at least one item", 400)

    try:
        with get_conn() as conn:
            bill_cols = table_columns(conn, "bills")
            insert_cols = [
                "id", "ts", "date", "cust", "phone", "pay",
                "sub", "disc", "tax", "total", "items", "doctor",
            ]
            insert_values: list[Any] = [
                data["id"],
                data["ts"],
                data["date"],
                data["cust"],
                data["phone"],
                data["pay"],
                data["sub"],
                data["disc"],
                data["tax"],
                data["total"],
                json.dumps(data["items"]),
                data.get("doctor", "Self"),
            ]
            if "rx" in bill_cols:
                insert_cols.append("rx")
                insert_values.append(data.get("rx", ""))
            if "prescription" in bill_cols:
                insert_cols.append("prescription")
                insert_values.append(data.get("prescription", ""))
            if "bill_type" in bill_cols:
                insert_cols.append("bill_type")
                insert_values.append(data.get("bill_type", "retail"))
            if "customer_type" in bill_cols:
                insert_cols.append("customer_type")
                insert_values.append(data.get("customer_type", "customer"))
            if "staff_name" in bill_cols:
                insert_cols.append("staff_name")
                insert_values.append(data.get("staff_name", ""))

            placeholders = ",".join(["?"] * len(insert_cols))
            conn.execute(
                f"INSERT INTO bills ({','.join(insert_cols)}) VALUES ({placeholders})",
                tuple(insert_values),
            )

            # Update customer records
            customer_name = str(data["cust"]).strip()
            customer_phone = str(data["phone"]).strip()
            face_vector = data.get("face_vector", "")
            total_value = float(data["total"])
            customer = conn.execute(
                "SELECT * FROM customers WHERE LOWER(name)=LOWER(?)",
                (customer_name,),
            ).fetchone()
            if customer:
                update_parts = ["visits = visits + 1", "total = total + ?"]
                update_values: list[Any] = [total_value]
                if face_vector and not customer["face_vector"]:
                    update_parts.append("face_vector = ?")
                    update_values.append(face_vector)
                update_values.append(customer["id"])
                conn.execute(
                    f"UPDATE customers SET {', '.join(update_parts)} WHERE id = ?",
                    tuple(update_values),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO customers (name, phone, visits, total, address, email, face_vector)
                    VALUES (?, ?, 1, ?, '', '', ?)
                    """,
                    (customer_name, customer_phone, total_value, face_vector),
                )

            # Update doctor records
            doctor_name = str(data.get("doctor", "Self")).strip()
            if doctor_name and doctor_name.lower() != "self":
                doctor = conn.execute(
                    "SELECT id FROM doctors WHERE LOWER(name)=LOWER(?)",
                    (doctor_name,),
                ).fetchone()
                if not doctor:
                    conn.execute(
                        """
                        INSERT INTO doctors (name, specialty, hospital, phone, email)
                        VALUES (?, '', '', '', '')
                        """,
                        (doctor_name,),
                    )

            # Update medicine stock (skip non-inventory/manual lines gracefully)
            for item in data["items"]:
                med_id = str(item.get("id", "")).strip()
                qty = int(item.get("qty", 0) or 0)
                if not med_id or qty <= 0:
                    continue
                med = conn.execute("SELECT s FROM medicines WHERE id = ?", (med_id,)).fetchone()
                if not med:
                    continue
                current_stock = int(med["s"] or 0)
                next_stock = max(0, current_stock - qty)
                conn.execute("UPDATE medicines SET s = ? WHERE id = ?", (next_stock, med_id))

        # ── Optional auto-WhatsApp on bill save ──────────────────────────
        wa_result = None
        if data.get("send_whatsapp") and customer_phone:
            bill_snapshot = normalize_bill_row(conn.execute(
                "SELECT * FROM bills WHERE id=?", (data["id"],)
            ).fetchone())
            msg       = build_bill_receipt_message(bill_snapshot)
            wa_result = send_whatsapp(customer_phone, msg)
            now_ts    = datetime.now(timezone.utc).isoformat()
            conn.execute("""
                INSERT INTO whatsapp_logs (bill_id, phone, message, status, provider_id, sent_at)
                VALUES (?,?,?,?,?,?)
            """, (data["id"], customer_phone, msg, wa_result["status"],
                  wa_result.get("sid", ""), now_ts))
            if wa_result["status"] in ("sent", "mocked"):
                conn.execute("UPDATE bills SET whatsapp_sent=1 WHERE id=?", (data["id"],))

        # ── Stage 14: Log per-bill commission ────────────────────────────
        try:
            sname = str(data.get("staff_name", "")).strip()
            btotal = float(data.get("total", 0) or 0)
            if sname and btotal > 0:
                with get_conn() as cconn:
                    cols14 = table_columns(cconn, "staff_commission_log") if cconn.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name='staff_commission_log'"
                    ).fetchone() else set()
                    if cols14:
                        sf = cconn.execute(
                            "SELECT id FROM staff WHERE LOWER(name)=LOWER(?) AND is_active=1 LIMIT 1",
                            (sname,)
                        ).fetchone()
                        s_id = sf["id"] if sf else 0
                        tgt = cconn.execute(
                            "SELECT commission_pct FROM staff_targets WHERE staff_id=? AND month=? LIMIT 1",
                            (s_id, data.get("date", "")[:7])
                        ).fetchone() if s_id else None
                        comm_pct = float(tgt["commission_pct"] if tgt else 0)
                        comm_amt = round(btotal * comm_pct / 100, 2)
                        cconn.execute("""
                            INSERT INTO staff_commission_log
                                (bill_id, staff_id, staff_name, bill_total, commission_pct,
                                 commission_earned, bill_date, created_at)
                            VALUES (?,?,?,?,?,?,?,?)
                        """, (data["id"], s_id, sname, btotal, comm_pct, comm_amt,
                              data.get("date", ""), datetime.now(timezone.utc).isoformat()))
        except Exception:
            pass  # commission log is non-critical

        # ── Stage 15: track bill usage ────────────────────────────────────
        try:
            shop_id = session.get("portal_user")
            if shop_id:
                track_usage(int(shop_id), "bills", 1)
        except Exception:
            pass

        return jsonify({"status": "success", "whatsapp": wa_result})
    except sqlite3.IntegrityError:
        return json_error("Bill ID already exists", 409, {"id": data.get("id")})
    except (ValueError, TypeError) as err:
        return json_error("Invalid bill data", 400, str(err))
    except Exception as err:
        return json_error("Failed to save bill", 500, str(err))


# --- MEDICINES (INVENTORY) ---
@app.route("/api/medicines", methods=["GET"])
def get_meds():
    """Get medicines. Supports ?q= search and ?limit/?offset pagination."""
    q = request.args.get("q", "").strip()
    limit = request.args.get("limit", type=int)
    offset = request.args.get("offset", 0, type=int)
    keys = [
        "id", "n", "g", "c", "p", "s", "batch", "expiry",
        "p_rate", "p_packing", "s_packing", "p_gst", "s_gst",
        "disc", "offer", "reorder", "max_qty",
    ]
    with get_conn() as conn:
        if q and limit is not None:
            rows = conn.execute(
                "SELECT * FROM medicines WHERE n LIKE ? OR g LIKE ? LIMIT ? OFFSET ?",
                (f"%{q}%", f"%{q}%", limit, offset),
            ).fetchall()
        elif q:
            rows = conn.execute(
                "SELECT * FROM medicines WHERE n LIKE ? OR g LIKE ?",
                (f"%{q}%", f"%{q}%"),
            ).fetchall()
        elif limit is not None:
            rows = conn.execute(
                "SELECT * FROM medicines LIMIT ? OFFSET ?",
                (limit, offset),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM medicines").fetchall()
    return jsonify([dict(zip(keys, tuple(row))) for row in rows])


@app.route("/api/medicines/alerts", methods=["GET"])
def medicine_alerts():
    """Get low stock and expiry alerts"""
    low_stock_threshold = int(request.args.get("low_stock", 15))
    expiry_days = int(request.args.get("expiry_days", 90))
    now = datetime.now().date()

    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT id, n, s, expiry, reorder
            FROM medicines
            ORDER BY n COLLATE NOCASE ASC
            """
        ).fetchall()

    low_stock: list[dict[str, Any]] = []
    expiring_soon: list[dict[str, Any]] = []
    for row in rows:
        med = dict(row)
        stock = int(med.get("s") or 0)
        reorder_level = int(med.get("reorder") or 0)
        threshold = reorder_level if reorder_level > 0 else low_stock_threshold
        if stock <= threshold:
            med["threshold"] = threshold
            low_stock.append(med)

        expiry_raw = (med.get("expiry") or "").strip()
        if expiry_raw:
            try:
                exp_date = datetime.strptime(expiry_raw, "%Y-%m-%d").date()
                days_left = (exp_date - now).days
                if days_left <= expiry_days:
                    expiring_soon.append(
                        {
                            "id": med.get("id"),
                            "n": med.get("n"),
                            "s": stock,
                            "expiry": expiry_raw,
                            "days_left": days_left,
                        }
                    )
            except ValueError:
                pass

    return jsonify(
        {
            "low_stock": low_stock,
            "expiring_soon": sorted(expiring_soon, key=lambda x: x["days_left"]),
            "config": {"low_stock": low_stock_threshold, "expiry_days": expiry_days},
        }
    )


@app.route("/api/medicines", methods=["POST"])
def update_med():
    """Create or update medicine"""
    data = request.get_json(silent=True) or {}
    missing = required_fields(data, ["id", "n", "p", "s"])
    if missing:
        return json_error("Missing required medicine fields", 400, missing)

    try:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO medicines
                (id, n, g, c, p, s, batch, expiry, p_rate, p_packing, s_packing, p_gst, s_gst, disc, offer, reorder, max_qty)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    data["id"],
                    data["n"],
                    data.get("g", "Generic"),
                    data.get("c", "Tablet"),
                    float(data["p"]),
                    int(data["s"]),
                    data.get("batch", ""),
                    data.get("expiry", ""),
                    float(data.get("p_rate", 0) or 0),
                    data.get("p_packing", ""),
                    data.get("s_packing", ""),
                    float(data.get("p_gst", 0) or 0),
                    float(data.get("s_gst", 0) or 0),
                    float(data.get("disc", 0) or 0),
                    data.get("offer", ""),
                    int(data.get("reorder", 0) or 0),
                    int(data.get("max_qty", 0) or 0),
                ),
            )
        return jsonify({"status": "success"})
    except (ValueError, TypeError) as err:
        return json_error("Invalid medicine payload", 400, str(err))
    except Exception as err:
        return json_error("Failed to save medicine", 500, str(err))


@app.route("/api/medicines/<id>", methods=["DELETE"])
def delete_med(id):
    """Delete medicine by ID"""
    with get_conn() as conn:
        conn.execute("DELETE FROM medicines WHERE id = ?", (id,))
    return jsonify({"status": "success"})


# --- PURCHASES ---
@app.route("/api/purchases", methods=["GET"])
def get_purchases():
    """Get all purchases"""
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM purchases").fetchall()
    return jsonify(
        [
            {
                "id": row["id"],
                "supplier": row["supplier"],
                "items": row["items"],
                "amount": row["amount"],
                "date": row["date"],
                "status": row["status"],
                "batch": row["batch"],
                "expiry": row["expiry"],
                "photo": row["photo"],
            }
            for row in rows
        ]
    )


@app.route("/api/purchases", methods=["POST"])
def add_purchase():
    """Create a new purchase"""
    data = request.get_json(silent=True) or {}
    missing = required_fields(data, ["id", "supplier", "items", "amount", "date", "status"])
    if missing:
        return json_error("Missing required purchase fields", 400, missing)

    try:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO purchases
                (id, supplier, items, amount, date, status, batch, expiry, photo)
                VALUES (?,?,?,?,?,?,?,?,?)
                """,
                (
                    data["id"],
                    data["supplier"],
                    data["items"],
                    float(data["amount"]),
                    data["date"],
                    data["status"],
                    data.get("batch", ""),
                    data.get("expiry", ""),
                    data.get("photo", ""),
                ),
            )

            supplier_name = str(data.get("supplier", "")).strip()
            if supplier_name:
                existing = conn.execute(
                    "SELECT id FROM suppliers WHERE LOWER(name)=LOWER(?)",
                    (supplier_name,),
                ).fetchone()
                if existing:
                    conn.execute(
                        "UPDATE suppliers SET last_order = ? WHERE id = ?",
                        (data["date"], existing["id"]),
                    )
                else:
                    conn.execute(
                        """
                        INSERT INTO suppliers (name, phone, gst, last_order, status)
                        VALUES (?, '', '', ?, 'Active')
                        """,
                        (supplier_name, data["date"]),
                    )

        return jsonify({"status": "success"})
    except (ValueError, TypeError) as err:
        return json_error("Invalid purchase payload", 400, str(err))
    except Exception as err:
        return json_error("Failed to save purchase", 500, str(err))


# --- MASTERS (CUSTOMERS / DOCTORS / SUPPLIERS) ---
@app.route("/api/suppliers", methods=["GET"])
def get_suppliers():
    """Get all suppliers"""
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM suppliers").fetchall()
    return jsonify(
        [
            {
                "id": row["id"],
                "name": row["name"],
                "phone": row["phone"],
                "gst": row["gst"],
                "last_order": row["last_order"],
                "status": row["status"],
            }
            for row in rows
        ]
    )


@app.route("/api/suppliers", methods=["POST"])
def add_supplier():
    """Create or update supplier"""
    data = request.get_json(silent=True) or {}
    missing = required_fields(data, ["name", "phone"])
    if missing:
        return json_error("Missing required supplier fields", 400, missing)
    try:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO suppliers (id, name, phone, gst, last_order, status)
                VALUES (?,?,?,?,?,?)
                """,
                (
                    data.get("id"),
                    data["name"],
                    data["phone"],
                    data.get("gst", ""),
                    data.get("last_order", "-"),
                    data.get("status", "Active"),
                ),
            )
        return jsonify({"status": "success"})
    except Exception as err:
        return json_error("Failed to save supplier", 500, str(err))


@app.route("/api/customers", methods=["GET"])
def get_customers():
    """Get all customers"""
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM customers").fetchall()
    return jsonify(
        [
            {
                "id": row["id"],
                "name": row["name"],
                "phone": row["phone"],
                "visits": row["visits"],
                "total_spend": row["total"],
                "address": row["address"],
                "email": row["email"],
                "face_vector": row["face_vector"],
            }
            for row in rows
        ]
    )


@app.route("/api/customers", methods=["POST"])
def add_customer():
    """Create or update customer"""
    data = request.get_json(silent=True) or {}
    missing = required_fields(data, ["name", "phone"])
    if missing:
        return json_error("Missing required customer fields", 400, missing)
    try:
        with get_conn() as conn:
            # Normalise family fields
            fh_id = data.get("family_head_id")
            try:
                fh_id = int(fh_id) if fh_id not in (None, "", "null") else None
            except (TypeError, ValueError):
                fh_id = None
            fam_rel = (data.get("family_relation") or "").strip()
            is_chronic = 1 if (data.get("is_chronic") in (1, "1", True, "true", "yes")) else 0

            conn.execute(
                """
                INSERT OR REPLACE INTO customers
                (id, name, phone, visits, total, address, email, face_vector,
                 family_head_id, family_relation, is_chronic)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    data.get("id"),
                    data["name"],
                    data["phone"],
                    int(data.get("visits", 1) or 1),
                    float(data.get("total", 0) or 0),
                    data.get("address", ""),
                    data.get("email", ""),
                    data.get("face_vector", ""),
                    fh_id,
                    fam_rel,
                    is_chronic,
                ),
            )
        return jsonify({"status": "success"})
    except (ValueError, TypeError) as err:
        return json_error("Invalid customer payload", 400, str(err))
    except Exception as err:
        return json_error("Failed to save customer", 500, str(err))


@app.route("/api/doctors", methods=["GET"])
def get_doctors():
    """Get all doctors"""
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM doctors").fetchall()
    return jsonify(
        [
            {
                "id": row["id"],
                "name": row["name"],
                "specialty": row["specialty"],
                "hospital": row["hospital"],
                "phone": row["phone"],
                "email": row["email"],
            }
            for row in rows
        ]
    )


@app.route("/api/doctors", methods=["POST"])
def add_doctor():
    """Create or update doctor"""
    data = request.get_json(silent=True) or {}
    missing = required_fields(data, ["name", "specialty", "hospital", "phone"])
    if missing:
        return json_error("Missing required doctor fields", 400, missing)
    try:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO doctors (id, name, specialty, hospital, phone, email)
                VALUES (?,?,?,?,?,?)
                """,
                (
                    data.get("id"),
                    data["name"],
                    data["specialty"],
                    data["hospital"],
                    data["phone"],
                    data.get("email", ""),
                ),
            )
        return jsonify({"status": "success"})
    except Exception as err:
        return json_error("Failed to save doctor", 500, str(err))


@app.route("/api/suppliers/<id>", methods=["DELETE"])
def delete_supplier(id):
    """Delete supplier by ID"""
    with get_conn() as conn:
        conn.execute("DELETE FROM suppliers WHERE id = ?", (id,))
    return jsonify({"status": "success"})


@app.route("/api/customers/<id>", methods=["DELETE"])
def delete_customer(id):
    """Delete customer by ID"""
    with get_conn() as conn:
        conn.execute("DELETE FROM customers WHERE id = ?", (id,))
    return jsonify({"status": "success"})


@app.route("/api/doctors/<id>", methods=["DELETE"])
def delete_doctor(id):
    """Delete doctor by ID"""
    with get_conn() as conn:
        conn.execute("DELETE FROM doctors WHERE id = ?", (id,))
    return jsonify({"status": "success"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PLATFORM ROUTES — Welcome, Registration, Admin
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/welcome")
def welcome():
    return render_template("welcome.html")

@app.route("/register")
def register_choice():
    return render_template("register.html")

@app.route("/register/wholesale")
def register_wholesale():
    return render_template("register_wholesale.html")

@app.route("/register/retail")
def register_retail():
    return render_template("register_retail.html")

@app.route("/admin")
def admin_panel():
    if not session.get("is_platform_admin"):
        return redirect(url_for("welcome"))
    return render_template("admin_dashboard.html")

@app.route("/portal")
def portal():
    if not session.get("portal_user"):
        return redirect("/portal-login")
    return render_template("portal.html")

@app.route("/portal/billing")
def portal_billing():
    if not session.get("portal_user"):
        return redirect("/portal-login")
    return render_template("portal_billing.html")

@app.route("/portal/stock")
def portal_stock():
    if not session.get("portal_user"):
        return redirect("/portal-login")
    return render_template("portal_stock.html")

@app.route("/portal/ai-reorder")
def portal_ai_reorder():
    if not session.get("portal_user"):
        return redirect("/portal-login")
    # Plan gate: Professional+ only
    try:
        allowed, pi = check_feature("ai_reorder", int(session["portal_user"]))
        if not allowed:
            return redirect("/portal/subscription?upgrade=1&feature=ai_reorder")
    except Exception:
        pass
    return render_template("portal_ai_reorder.html")

@app.route("/portal/import")
def portal_import_page():
    if not session.get("portal_user") and not session.get("admin_user"):
        return redirect(url_for("welcome"))
    return render_template("portal_import.html")


# ── Admin auth ───────────────────────────────────────────────────────

@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    with get_conn() as conn:
        admin = conn.execute(
            "SELECT * FROM platform_admins WHERE username=?", (username,)
        ).fetchone()
    if admin and check_password_hash(admin["password_hash"], password):
        session["is_platform_admin"] = True
        session["admin_name"] = admin["name"]
        session["admin_id"] = admin["id"]
        return jsonify({"status": "success", "name": admin["name"]})
    return jsonify({"status": "error", "message": "Invalid credentials"}), 401


@app.route("/api/admin/logout", methods=["POST"])
def admin_logout():
    session.clear()
    return jsonify({"status": "success"})


# ── Portal auth ──────────────────────────────────────────────────────

@app.route("/api/portal/login", methods=["POST"])
def portal_login():
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    with get_conn() as conn:
        # Check retail shops
        shop = conn.execute(
            "SELECT * FROM retail_shops WHERE username=? AND is_active=1", (username,)
        ).fetchone()
        if shop and check_password_hash(shop["password_hash"], password):
            session["portal_user"] = username
            session["portal_type"] = "retail"
            session["portal_id"] = shop["id"]
            session["portal_name"] = shop["shop_name"]
            return jsonify({"status": "success", "type": "retail", "name": shop["shop_name"]})
        # Check wholesale accounts
        ws = conn.execute(
            "SELECT * FROM wholesale_accounts WHERE username=? AND is_active=1", (username,)
        ).fetchone()
        if ws and check_password_hash(ws["password_hash"], password):
            session["portal_user"] = username
            session["portal_type"] = "wholesale"
            session["portal_id"] = ws["id"]
            session["portal_name"] = ws["business_name"]
            return jsonify({"status": "success", "type": "wholesale", "name": ws["business_name"]})
    return jsonify({"status": "error", "message": "Invalid credentials or account not yet approved"}), 401


@app.route("/api/portal/logout", methods=["POST"])
def portal_logout():
    session.clear()
    return jsonify({"status": "success"})


@app.route("/api/portal/me", methods=["GET"])
def portal_me():
    if not session.get("portal_user"):
        return jsonify({"status": "error"}), 401
    return jsonify({
        "username": session.get("portal_user"),
        "type": session.get("portal_type"),
        "id": session.get("portal_id"),
        "name": session.get("portal_name"),
    })


# ── Registration submission ──────────────────────────────────────────

@app.route("/api/register", methods=["POST"])
def submit_registration():
    data = request.get_json(silent=True) or {}
    reg_type = data.get("reg_type", "retail")
    required = ["business_name", "owner_name", "owner_phone"]
    missing = required_fields(data, required)
    if missing:
        return json_error("Missing required fields", 400, missing)
    try:
        with get_conn() as conn:
            conn.execute("""
                INSERT INTO platform_registrations
                (reg_type, business_name, owner_name, owner_phone, owner_email, alt_phone,
                 address, city, state, pincode, gstin, drug_license, pan,
                 bank_name, bank_account, bank_ifsc, monthly_turnover, no_of_staff,
                 no_of_shops, software_used, hear_from, status, submitted_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                reg_type,
                data.get("business_name"),
                data.get("owner_name"),
                data.get("owner_phone"),
                data.get("owner_email", ""),
                data.get("alt_phone", ""),
                data.get("address", ""),
                data.get("city", ""),
                data.get("state", ""),
                data.get("pincode", ""),
                data.get("gstin", ""),
                data.get("drug_license", ""),
                data.get("pan", ""),
                data.get("bank_name", ""),
                data.get("bank_account", ""),
                data.get("bank_ifsc", ""),
                data.get("monthly_turnover", ""),
                data.get("no_of_staff", ""),
                data.get("no_of_shops", ""),
                data.get("software_used", ""),
                data.get("hear_from", ""),
                "pending",
                datetime.now(timezone.utc).isoformat(),
            ))
        return jsonify({"status": "success", "message": "Registration submitted! We will verify and contact you within 24 hours."})
    except Exception as e:
        return json_error("Failed to submit registration", 500, str(e))


# ── Admin: list/approve/reject registrations ─────────────────────────

@app.route("/api/admin/registrations", methods=["GET"])
def admin_list_registrations():
    if not session.get("is_platform_admin"):
        return jsonify({"status": "error"}), 403
    status_filter = request.args.get("status", "pending")
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM platform_registrations WHERE status=? ORDER BY submitted_at DESC",
            (status_filter,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/registrations/<int:reg_id>/approve", methods=["POST"])
def admin_approve_registration(reg_id):
    if not session.get("is_platform_admin"):
        return jsonify({"status": "error"}), 403
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    if not username or not password:
        return json_error("Username and password required for approval", 400)
    with get_conn() as conn:
        reg = conn.execute(
            "SELECT * FROM platform_registrations WHERE id=?", (reg_id,)
        ).fetchone()
        if not reg:
            return json_error("Registration not found", 404)
        reg = dict(reg)
        now = datetime.now(timezone.utc).isoformat()
        pw_hash = generate_password_hash(password)
        if reg["reg_type"] == "wholesale":
            conn.execute("""
                INSERT OR REPLACE INTO wholesale_accounts
                (business_name, owner_name, phone, address, city, gstin, drug_license, username, password_hash, is_active, approved_at, reg_id)
                VALUES (?,?,?,?,?,?,?,?,?,1,?,?)
            """, (reg["business_name"], reg["owner_name"], reg["owner_phone"],
                  reg.get("address",""), reg.get("city",""), reg.get("gstin",""),
                  reg.get("drug_license",""), username, pw_hash, now, reg_id))
        else:
            conn.execute("""
                INSERT OR REPLACE INTO retail_shops
                (shop_name, owner_name, phone, address, city, gstin, drug_license, username, password_hash, is_active, approved_at, reg_id)
                VALUES (?,?,?,?,?,?,?,?,?,1,?,?)
            """, (reg["business_name"], reg["owner_name"], reg["owner_phone"],
                  reg.get("address",""), reg.get("city",""), reg.get("gstin",""),
                  reg.get("drug_license",""), username, pw_hash, now, reg_id))
        conn.execute("""
            UPDATE platform_registrations SET status='approved', reviewed_at=?, reviewed_by=?, username=?, password_hash=?
            WHERE id=?
        """, (now, session.get("admin_name","admin"), username, pw_hash, reg_id))
    return jsonify({"status": "success", "message": "Registration approved and account created."})


@app.route("/api/admin/registrations/<int:reg_id>/reject", methods=["POST"])
def admin_reject_registration(reg_id):
    if not session.get("is_platform_admin"):
        return jsonify({"status": "error"}), 403
    data = request.get_json(silent=True) or {}
    notes = data.get("notes", "")
    now = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        conn.execute("""
            UPDATE platform_registrations SET status='rejected', reviewed_at=?, reviewed_by=?, review_notes=?
            WHERE id=?
        """, (now, session.get("admin_name","admin"), notes, reg_id))
    return jsonify({"status": "success"})


@app.route("/api/admin/stats", methods=["GET"])
def admin_stats():
    if not session.get("is_platform_admin"):
        return jsonify({"status": "error"}), 403
    with get_conn() as conn:
        pending = conn.execute("SELECT COUNT(*) AS c FROM platform_registrations WHERE status='pending'").fetchone()["c"]
        approved = conn.execute("SELECT COUNT(*) AS c FROM platform_registrations WHERE status='approved'").fetchone()["c"]
        rejected = conn.execute("SELECT COUNT(*) AS c FROM platform_registrations WHERE status='rejected'").fetchone()["c"]
        wholesale_count = conn.execute("SELECT COUNT(*) AS c FROM wholesale_accounts WHERE is_active=1").fetchone()["c"]
        retail_count = conn.execute("SELECT COUNT(*) AS c FROM retail_shops WHERE is_active=1").fetchone()["c"]
        today_bills = conn.execute(
            "SELECT COUNT(*) AS c FROM bills WHERE date LIKE ?",
            (datetime.now().strftime("%d/%m/%Y") + "%",)
        ).fetchone()["c"]
    return jsonify({
        "pending_registrations": pending,
        "approved_registrations": approved,
        "rejected_registrations": rejected,
        "wholesale_accounts": wholesale_count,
        "retail_shops": retail_count,
        "today_bills": today_bills,
    })


@app.route("/api/admin/accounts", methods=["GET"])
def admin_list_accounts():
    if not session.get("is_platform_admin"):
        return jsonify({"status": "error"}), 403
    account_type = request.args.get("type", "all")
    result = {"wholesale": [], "retail": []}
    with get_conn() as conn:
        if account_type in ("all", "wholesale"):
            rows = conn.execute("SELECT * FROM wholesale_accounts ORDER BY approved_at DESC").fetchall()
            result["wholesale"] = [dict(r) for r in rows]
        if account_type in ("all", "retail"):
            rows = conn.execute("SELECT * FROM retail_shops ORDER BY approved_at DESC").fetchall()
            result["retail"] = [dict(r) for r in rows]
    return jsonify(result)


@app.route("/api/admin/accounts/<account_type>/<int:account_id>/toggle", methods=["POST"])
def admin_toggle_account(account_type, account_id):
    if not session.get("is_platform_admin"):
        return jsonify({"status": "error"}), 403
    table = "wholesale_accounts" if account_type == "wholesale" else "retail_shops"
    with get_conn() as conn:
        row = conn.execute(f"SELECT is_active FROM {table} WHERE id=?", (account_id,)).fetchone()
        if not row:
            return json_error("Account not found", 404)
        new_status = 0 if row["is_active"] else 1
        conn.execute(f"UPDATE {table} SET is_active=? WHERE id=?", (new_status, account_id))
    return jsonify({"status": "success", "is_active": bool(new_status)})


# ── AI Reorder suggestions ───────────────────────────────────────────

@app.route("/api/ai/reorder", methods=["GET"])
def ai_reorder_suggestions():
    days = int(request.args.get("days", 30))
    cutoff_ts = int((datetime.now() - timedelta(days=days)).timestamp() * 1000)
    with get_conn() as conn:
        bills = conn.execute(
            "SELECT items FROM bills WHERE ts >= ?", (cutoff_ts,)
        ).fetchall()
        medicines = conn.execute("SELECT id, n, s, reorder, max_qty FROM medicines").fetchall()
    sales_map: dict[str, float] = {}
    for bill in bills:
        items = safe_json_loads(bill["items"], [])
        for itm in items:
            mid = str(itm.get("id", ""))
            qty = float(itm.get("qty", 0) or 0)
            sales_map[mid] = sales_map.get(mid, 0) + qty
    suggestions = []
    for med in medicines:
        mid = med["id"]
        name = med["n"]
        current_stock = int(med["s"] or 0)
        manual_reorder = int(med["reorder"] or 0)
        manual_max = int(med["max_qty"] or 0)
        total_sold = sales_map.get(mid, 0)
        avg_daily = round(total_sold / max(days, 1), 2)
        ai_min = max(int(avg_daily * 7), 5)
        ai_max = max(int(avg_daily * 30), ai_min * 3)
        reorder_qty = max(ai_max - current_stock, 0)
        days_of_stock = round(current_stock / avg_daily, 1) if avg_daily > 0 else 999
        urgency = "critical" if days_of_stock < 3 else "low" if days_of_stock < 7 else "ok" if days_of_stock < 30 else "excess"
        suggestions.append({
            "id": mid,
            "name": name,
            "current_stock": current_stock,
            "sold_last_30d": round(total_sold, 0),
            "avg_daily_sales": avg_daily,
            "ai_min": ai_min,
            "ai_max": ai_max,
            "manual_reorder": manual_reorder,
            "manual_max": manual_max,
            "reorder_qty": reorder_qty,
            "days_of_stock": days_of_stock,
            "urgency": urgency,
        })
    suggestions.sort(key=lambda x: ({"critical": 0, "low": 1, "ok": 2, "excess": 3}.get(x["urgency"], 4)))
    return jsonify(suggestions)


@app.route("/api/ai/wholesaler-compare", methods=["GET"])
def ai_wholesaler_compare():
    medicine_name = request.args.get("q", "").strip()
    with get_conn() as conn:
        if medicine_name:
            rows = conn.execute(
                "SELECT * FROM wholesaler_catalog WHERE medicine_name LIKE ? ORDER BY price ASC",
                (f"%{medicine_name}%",)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM wholesaler_catalog ORDER BY medicine_name, price ASC"
            ).fetchall()
    result = [dict(r) for r in rows]
    for item in result:
        mrp = float(item.get("mrp") or 0)
        price = float(item.get("price") or 0)
        disc = float(item.get("discount_pct") or 0)
        effective_price = round(price * (1 - disc / 100), 2)
        margin_pct = round(((mrp - effective_price) / mrp * 100) if mrp > 0 else 0, 1)
        item["effective_price"] = effective_price
        item["margin_pct"] = margin_pct
        item["score"] = round(margin_pct + (5 if item.get("free_offer") and item["free_offer"] != "None" else 0), 1)
    return jsonify(result)


# ── Stock import (Excel/CSV paste) ───────────────────────────────────

@app.route("/api/stock/import", methods=["POST"])
def import_stock():
    data = request.get_json(silent=True) or {}
    items = data.get("items", [])
    import_date = data.get("date", datetime.now().strftime("%Y-%m-%d"))
    shop_id = data.get("shop_id", 0)
    shop_type = data.get("shop_type", "retail")
    if not items:
        return json_error("No items provided", 400)
    updated = 0
    created = 0
    with get_conn() as conn:
        for itm in items:
            name = str(itm.get("name", "")).strip()
            stock = itm.get("stock", 0)
            price = itm.get("price", 0)
            if not name:
                continue
            existing = conn.execute("SELECT id FROM medicines WHERE n LIKE ?", (name,)).fetchone()
            if existing:
                conn.execute("UPDATE medicines SET s=? WHERE id=?", (int(stock or 0), existing["id"]))
                updated += 1
            else:
                import uuid as _uuid
                new_id = str(_uuid.uuid4())[:8]
                conn.execute("""
                    INSERT OR IGNORE INTO medicines (id, n, g, c, p, s, reorder, max_qty)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (new_id, name, "", "Tablet", float(price or 0), int(stock or 0), 0, 0))
                created += 1
        conn.execute("""
            INSERT INTO stock_imports (shop_id, shop_type, import_date, item_count, imported_at, raw_data)
            VALUES (?,?,?,?,?,?)
        """, (shop_id, shop_type, import_date, len(items), datetime.now(timezone.utc).isoformat(), json.dumps(items)))
    return jsonify({"status": "success", "updated": updated, "created": created, "total": len(items)})


@app.route("/api/stock/imports", methods=["GET"])
def list_stock_imports():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, shop_id, shop_type, import_date, item_count, imported_at FROM stock_imports ORDER BY imported_at DESC LIMIT 50"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


# ── Shop reorder rules ───────────────────────────────────────────────

@app.route("/api/shop/reorder-rules", methods=["GET"])
def get_reorder_rules():
    shop_id = request.args.get("shop_id", 0, type=int)
    with get_conn() as conn:
        if shop_id:
            rows = conn.execute(
                "SELECT * FROM shop_reorder_rules WHERE shop_id=? ORDER BY medicine_name", (shop_id,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM shop_reorder_rules ORDER BY medicine_name").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/shop/reorder-rules", methods=["POST"])
def save_reorder_rule():
    data = request.get_json(silent=True) or {}
    now = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        existing = conn.execute(
            "SELECT id FROM shop_reorder_rules WHERE shop_id=? AND medicine_id=?",
            (data.get("shop_id",0), data.get("medicine_id",""))
        ).fetchone()
        if existing:
            conn.execute("""
                UPDATE shop_reorder_rules SET min_qty=?, max_qty=?, reorder_qty=?, last_updated=? WHERE id=?
            """, (data.get("min_qty",0), data.get("max_qty",0), data.get("reorder_qty",0), now, existing["id"]))
        else:
            conn.execute("""
                INSERT INTO shop_reorder_rules (shop_id, medicine_id, medicine_name, min_qty, max_qty, reorder_qty, last_updated)
                VALUES (?,?,?,?,?,?,?)
            """, (data.get("shop_id",0), data.get("medicine_id",""), data.get("medicine_name",""),
                  data.get("min_qty",0), data.get("max_qty",0), data.get("reorder_qty",0), now))
    return jsonify({"status": "success"})


# ── Stock import: file upload + parse ────────────────────────────────

def _guess_category(name: str) -> str:
    n = name.upper()
    if any(x in n for x in [" TAB", "TAB ", "TAB.", "TABLET"]):
        return "Tablet"
    if any(x in n for x in [" CAP", "CAP ", "CAPS", "CAPSULE"]):
        return "Capsule"
    if any(x in n for x in ["SYP", "SYR", "SYRUP"]):
        return "Syrup"
    if "INJ" in n:
        return "Injection"
    if any(x in n for x in ["DROP", " EYE", " EAR"]):
        return "Drops"
    if any(x in n for x in ["CREAM", "CRM", "OINT", " GEL"]):
        return "Topical"
    if "SUSP" in n:
        return "Suspension"
    if any(x in n for x in ["POWDER", " PDR"]):
        return "Powder"
    if "LOTION" in n:
        return "Lotion"
    if any(x in n for x in ["SOAP", "SHAMPOO"]):
        return "Personal Care"
    return "Other"


def _parse_xls_bytes(file_bytes: bytes) -> list[dict]:
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)
    rows = []
    for i in range(ws.nrows):
        rows.append([ws.cell_value(i, j) for j in range(ws.ncols)])
    return rows


def _parse_xlsx_bytes(file_bytes: bytes) -> list[list]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _parse_csv_bytes(file_bytes: bytes) -> list[list]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


def _detect_header_row(rows: list[list]) -> int:
    """Find the row index that looks like a header (contains 'name' or 'item')."""
    for i, row in enumerate(rows[:5]):
        row_str = " ".join(str(c).lower() for c in row if c)
        if any(k in row_str for k in ["item name", "name", "medicine", "drug"]):
            return i
    return 0


def _map_ss_co_format(rows: list[list], header_row: int) -> list[dict]:
    """Map SS & Co style: Sno|Item Name|Qty|MRP|Selprice|S.Value|NetPur.Rate|Pur.Rate|Packing|P.Value|Netval"""
    items = []
    for row in rows[header_row + 1:]:
        if not row or not row[1]:
            continue
        name = str(row[1]).strip().strip("###").strip()
        if not name or len(name) < 2:
            continue
        try:
            qty     = max(0, int(float(row[2] or 0)))
            mrp     = float(row[3] or 0)
            selprice = float(row[4] or 0)
            net_pur  = float(row[6] or 0)
            packing  = int(float(row[8])) if len(row) > 8 and row[8] else 1
            packing  = max(1, packing)
            items.append({
                "name": name,
                "stock": qty,
                "price": round(selprice if selprice > 0 else mrp, 2),
                "mrp": round(mrp, 2),
                "p_rate": round(net_pur, 2),
                "packing": f"1x{packing}" if packing > 1 else "1",
                "category": _guess_category(name),
            })
        except (ValueError, TypeError, IndexError):
            continue
    return items


def _map_generic_format(rows: list[list], header_row: int, col_map: dict) -> list[dict]:
    """Map any CSV/Excel where caller specifies which column index = which field."""
    name_col  = col_map.get("name", 0)
    stock_col = col_map.get("stock")
    price_col = col_map.get("price")
    prate_col = col_map.get("p_rate")
    items = []
    for row in rows[header_row + 1:]:
        if not row or (name_col >= len(row)):
            continue
        name = str(row[name_col]).strip()
        if not name or len(name) < 2:
            continue
        def _val(idx):
            if idx is None or idx >= len(row):
                return 0
            try:
                return float(row[idx] or 0)
            except (ValueError, TypeError):
                return 0
        items.append({
            "name": name,
            "stock": max(0, int(_val(stock_col))),
            "price": round(_val(price_col), 2),
            "mrp": round(_val(price_col), 2),
            "p_rate": round(_val(prate_col), 2),
            "packing": "1",
            "category": _guess_category(name),
        })
    return items


@app.route("/api/stock/parse-file", methods=["POST"])
def parse_stock_file():
    """Accept XLS/XLSX/CSV upload, return parsed preview + column headers."""
    if "file" not in request.files:
        return json_error("No file uploaded", 400)
    f = request.files["file"]
    fname = f.filename.lower()
    raw = f.read()
    if not raw:
        return json_error("Empty file", 400)

    try:
        if fname.endswith(".xls"):
            rows = _parse_xls_bytes(raw)
        elif fname.endswith(".xlsx"):
            rows = _parse_xlsx_bytes(raw)
        elif fname.endswith(".csv"):
            rows = _parse_csv_bytes(raw)
        else:
            return json_error("Unsupported format. Use XLS, XLSX or CSV.", 400)
    except Exception as e:
        return json_error(f"Could not parse file: {e}", 400)

    header_row = _detect_header_row(rows)
    headers = [str(c) for c in (rows[header_row] if rows else [])]

    # Auto-detect SS & Co format
    header_str = " ".join(headers).lower()
    is_ssco = "selprice" in header_str or "netpur" in header_str or "pur.rate" in header_str
    if is_ssco:
        items = _map_ss_co_format(rows, header_row)
        format_detected = "SS & Co / Somasundaram & Co"
    else:
        items = []
        format_detected = "generic"

    preview = items[:20]
    return jsonify({
        "format": format_detected,
        "header_row": header_row,
        "headers": headers,
        "total_rows": len(rows) - header_row - 1,
        "preview": preview,
        "items": items,
        "is_ssco": is_ssco,
    })


@app.route("/api/stock/import-file", methods=["POST"])
def import_stock_file():
    """Import parsed items (JSON body) into medicines table with full field mapping."""
    data   = request.get_json(silent=True) or {}
    items  = data.get("items", [])
    source = data.get("source", "file_import")
    if not items:
        return json_error("No items provided", 400)

    created = updated = skipped = 0
    with get_conn() as conn:
        for itm in items:
            name = str(itm.get("name", "")).strip()
            if not name:
                skipped += 1
                continue
            stock   = max(0, int(itm.get("stock", 0) or 0))
            price   = float(itm.get("price", 0) or 0)
            p_rate  = float(itm.get("p_rate", 0) or 0)
            packing = str(itm.get("packing", "1") or "1")
            cat     = itm.get("category") or _guess_category(name)

            existing = conn.execute("SELECT id FROM medicines WHERE n = ?", (name,)).fetchone()
            if existing:
                conn.execute(
                    "UPDATE medicines SET s=?, p=?, p_rate=?, p_packing=?, s_packing=?, c=? WHERE id=?",
                    (stock, price, p_rate, packing, packing, cat, existing["id"])
                )
                updated += 1
            else:
                new_id = "imp_" + hashlib.md5(name.encode()).hexdigest()[:8]
                conn.execute(
                    """INSERT OR IGNORE INTO medicines
                       (id, n, g, c, p, s, p_rate, p_packing, s_packing, reorder, max_qty)
                       VALUES (?,?,?,?,?,?,?,?,?,0,0)""",
                    (new_id, name, "", cat, price, stock, p_rate, packing, packing)
                )
                created += 1

        try:
            conn.execute(
                """INSERT INTO stock_imports (shop_id, shop_type, import_date, item_count, imported_at, imported_by)
                   VALUES (0,'file_import',?,?,?,?)""",
                (datetime.now().strftime("%Y-%m-%d"), len(items),
                 datetime.now(timezone.utc).isoformat(), source)
            )
        except Exception:
            pass

    return jsonify({"status": "success", "created": created, "updated": updated,
                    "skipped": skipped, "total": len(items)})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PHASE 2 — PAGE ROUTES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/portal/subscription")
def portal_subscription():
    if not session.get("portal_user"):
        return redirect("/portal")
    return render_template("portal_subscription.html")

@app.route("/portal/tally")
def portal_tally():
    if not session.get("portal_user"):
        return redirect("/portal")
    try:
        allowed, _ = check_feature("tally", int(session["portal_user"]))
        if not allowed:
            return redirect("/portal/subscription?upgrade=1&feature=tally")
    except Exception:
        pass
    return render_template("portal_tally.html")

@app.route("/portal/receiving")
def portal_receiving():
    if not session.get("portal_user"):
        return redirect("/portal")
    try:
        allowed, _ = check_feature("receiving", int(session["portal_user"]))
        if not allowed:
            return redirect("/portal/subscription?upgrade=1&feature=receiving")
    except Exception:
        pass
    return render_template("portal_receiving.html")

@app.route("/portal/delivery")
def portal_delivery():
    if not session.get("portal_user"):
        return redirect("/portal")
    try:
        allowed, _ = check_feature("delivery", int(session["portal_user"]))
        if not allowed:
            return redirect("/portal/subscription?upgrade=1&feature=delivery")
    except Exception:
        pass
    return render_template("portal_delivery.html")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PHASE 2 — SUBSCRIPTION API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/subscriptions/plans", methods=["GET"])
def get_subscription_plans():
    with get_conn() as conn:
        plans = [dict(r) for r in conn.execute(
            "SELECT * FROM subscription_plans WHERE is_active=1 ORDER BY price_monthly").fetchall()]
    for p in plans:
        p["features"] = safe_json_loads(p.get("features"), [])
    return jsonify(plans)


@app.route("/api/subscriptions/my", methods=["GET"])
def my_subscription():
    user = session.get("portal_user")
    if not user:
        return json_error("Not authenticated", 401)
    acc_type = user.get("account_type", "retail")
    acc_id   = user.get("id", 0)
    with get_conn() as conn:
        sub = conn.execute("""
            SELECT s.*, sp.name as plan_display_name, sp.features as plan_features,
                   sp.max_shops, sp.price_monthly, sp.price_yearly
            FROM subscriptions s
            LEFT JOIN subscription_plans sp ON sp.id = s.plan_id
            WHERE s.account_type=? AND s.account_id=?
            ORDER BY s.id DESC LIMIT 1
        """, (acc_type, acc_id)).fetchone()
    if not sub:
        return jsonify({"status": "no_subscription", "plan": None})
    d = dict(sub)
    d["plan_features"] = safe_json_loads(d.get("plan_features"), [])
    # check expiry
    try:
        end = datetime.fromisoformat(d["end_date"])
        d["days_remaining"] = (end - datetime.now(timezone.utc)).days
        d["is_active"] = d["days_remaining"] > 0 and d["status"] in ("active", "trial")
    except Exception:
        d["days_remaining"] = 0
        d["is_active"] = False
    return jsonify(d)


@app.route("/api/subscriptions/create-order", methods=["POST"])
def create_subscription_order():
    """Create a Razorpay order (or simulate when no API key configured)."""
    user = session.get("portal_user")
    if not user:
        return json_error("Not authenticated", 401)
    data     = request.get_json(silent=True) or {}
    plan_id  = data.get("plan_id")
    billing  = data.get("billing_cycle", "monthly")
    if not plan_id:
        return json_error("plan_id required", 400)

    with get_conn() as conn:
        plan = conn.execute("SELECT * FROM subscription_plans WHERE id=?", (plan_id,)).fetchone()
    if not plan:
        return json_error("Plan not found", 404)

    plan = dict(plan)
    amount_inr = plan["price_yearly"] if billing == "yearly" else plan["price_monthly"]
    amount_paise = int(amount_inr * 100)

    rzp_key = os.environ.get("RAZORPAY_KEY_ID", "")
    rzp_secret = os.environ.get("RAZORPAY_KEY_SECRET", "")

    if rzp_key and rzp_secret:
        try:
            import razorpay
            client = razorpay.Client(auth=(rzp_key, rzp_secret))
            order = client.order.create({
                "amount": amount_paise,
                "currency": "INR",
                "receipt": f"sub_{user['id']}_{plan_id}",
                "notes": {"account_type": user.get("account_type"), "plan": plan["name"]}
            })
            return jsonify({
                "order_id": order["id"],
                "amount": amount_paise,
                "currency": "INR",
                "key": rzp_key,
                "plan": plan["name"],
                "billing": billing
            })
        except Exception as e:
            return json_error(f"Razorpay error: {e}", 500)
    else:
        # Simulation mode — no Razorpay keys configured
        fake_order_id = f"order_sim_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
        return jsonify({
            "order_id": fake_order_id,
            "amount": amount_paise,
            "currency": "INR",
            "key": "rzp_test_simulation",
            "plan": plan["name"],
            "billing": billing,
            "simulation": True,
            "note": "Configure RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET env vars for live payments"
        })


@app.route("/api/subscriptions/activate", methods=["POST"])
def activate_subscription():
    """Activate subscription after payment (or manual activation in simulation)."""
    user = session.get("portal_user")
    if not user:
        return json_error("Not authenticated", 401)
    data     = request.get_json(silent=True) or {}
    plan_id  = data.get("plan_id")
    billing  = data.get("billing_cycle", "monthly")
    rzp_order_id   = data.get("razorpay_order_id", "")
    rzp_payment_id = data.get("razorpay_payment_id", "")
    is_sim   = data.get("simulation", False)

    with get_conn() as conn:
        plan = conn.execute("SELECT * FROM subscription_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return json_error("Plan not found", 404)
        plan = dict(plan)

        now = datetime.now(timezone.utc)
        days = 365 if billing == "yearly" else 30
        end_date = (now + timedelta(days=days)).isoformat()
        amount = plan["price_yearly"] if billing == "yearly" else plan["price_monthly"]

        acc_type = user.get("account_type", "retail")
        acc_id   = user.get("id", 0)

        existing = conn.execute(
            "SELECT id FROM subscriptions WHERE account_type=? AND account_id=?",
            (acc_type, acc_id)
        ).fetchone()

        if existing:
            conn.execute("""
                UPDATE subscriptions SET plan_id=?, plan_name=?, status='active',
                start_date=?, end_date=?, amount_paid=?, billing_cycle=?,
                razorpay_order_id=?, razorpay_payment_id=?, updated_at=?
                WHERE id=?
            """, (plan_id, plan["name"], now.isoformat(), end_date, amount, billing,
                  rzp_order_id, rzp_payment_id, now.isoformat(), existing["id"]))
        else:
            conn.execute("""
                INSERT INTO subscriptions (account_type, account_id, plan_id, plan_name,
                status, start_date, end_date, amount_paid, billing_cycle,
                razorpay_order_id, razorpay_payment_id, created_at, updated_at)
                VALUES (?,?,?,?,'active',?,?,?,?,?,?,?,?)
            """, (acc_type, acc_id, plan_id, plan["name"], now.isoformat(), end_date,
                  amount, billing, rzp_order_id, rzp_payment_id, now.isoformat(), now.isoformat()))

    return jsonify({"status": "success", "plan": plan["name"], "end_date": end_date})


@app.route("/api/admin/subscriptions", methods=["GET"])
def admin_list_subscriptions():
    if not session.get("admin_user"):
        return json_error("Forbidden", 403)
    with get_conn() as conn:
        subs = [dict(r) for r in conn.execute("""
            SELECT s.*, sp.name as plan_display
            FROM subscriptions s
            LEFT JOIN subscription_plans sp ON sp.id = s.plan_id
            ORDER BY s.id DESC
        """).fetchall()]
    return jsonify(subs)


@app.route("/api/admin/subscriptions/<int:sub_id>/extend", methods=["POST"])
def admin_extend_subscription(sub_id):
    if not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    days = int(data.get("days", 30))
    with get_conn() as conn:
        sub = conn.execute("SELECT * FROM subscriptions WHERE id=?", (sub_id,)).fetchone()
        if not sub:
            return json_error("Not found", 404)
        try:
            current_end = datetime.fromisoformat(sub["end_date"])
        except Exception:
            current_end = datetime.now(timezone.utc)
        new_end = (current_end + timedelta(days=days)).isoformat()
        conn.execute("UPDATE subscriptions SET end_date=?, status='active', updated_at=? WHERE id=?",
                     (new_end, datetime.now(timezone.utc).isoformat(), sub_id))
    return jsonify({"status": "success", "new_end_date": new_end})


@app.route("/api/admin/subscriptions/<int:sub_id>/cancel", methods=["POST"])
def admin_cancel_subscription(sub_id):
    if not session.get("admin_user"):
        return json_error("Forbidden", 403)
    with get_conn() as conn:
        conn.execute("UPDATE subscriptions SET status='cancelled', updated_at=? WHERE id=?",
                     (datetime.now(timezone.utc).isoformat(), sub_id))
    return jsonify({"status": "success"})


@app.route("/api/admin/plans", methods=["GET"])
def admin_list_plans():
    if not session.get("admin_user"):
        return json_error("Forbidden", 403)
    with get_conn() as conn:
        plans = [dict(r) for r in conn.execute("SELECT * FROM subscription_plans ORDER BY price_monthly").fetchall()]
    for p in plans:
        p["features"] = safe_json_loads(p.get("features"), [])
    return jsonify(plans)


@app.route("/api/admin/plans", methods=["POST"])
def admin_create_plan():
    if not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return json_error("name required", 400)
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO subscription_plans (name, price_monthly, price_yearly, max_shops, features, is_active, created_at)
            VALUES (?,?,?,?,?,1,?)
        """, (name, float(data.get("price_monthly", 0)),
              float(data.get("price_yearly", 0)),
              int(data.get("max_shops", 1)),
              json.dumps(data.get("features", [])),
              datetime.now(timezone.utc).isoformat()))
    return jsonify({"status": "success"})


@app.route("/api/admin/plans/<int:plan_id>", methods=["PUT"])
def admin_update_plan(plan_id):
    if not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    with get_conn() as conn:
        conn.execute("""
            UPDATE subscription_plans SET name=?, price_monthly=?, price_yearly=?,
            max_shops=?, features=?, is_active=? WHERE id=?
        """, (data.get("name"), float(data.get("price_monthly", 0)),
              float(data.get("price_yearly", 0)), int(data.get("max_shops", 1)),
              json.dumps(data.get("features", [])), int(data.get("is_active", 1)), plan_id))
    return jsonify({"status": "success"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PHASE 2 — TALLY / ACCOUNTS API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/tally/companies", methods=["GET"])
def tally_list_companies():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    with get_conn() as conn:
        companies = [dict(r) for r in conn.execute(
            "SELECT * FROM tally_companies WHERE is_active=1 ORDER BY name").fetchall()]
    return jsonify(companies)


@app.route("/api/tally/companies", methods=["POST"])
def tally_create_company():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return json_error("Company name required", 400)
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO tally_companies (name, address, city, state, pincode, gstin, pan,
            phone, email, bank_name, bank_account, bank_ifsc, opening_balance, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (name, data.get("address",""), data.get("city",""), data.get("state",""),
              data.get("pincode",""), data.get("gstin",""), data.get("pan",""),
              data.get("phone",""), data.get("email",""), data.get("bank_name",""),
              data.get("bank_account",""), data.get("bank_ifsc",""),
              float(data.get("opening_balance", 0)),
              datetime.now(timezone.utc).isoformat()))
    return jsonify({"status": "success"})


@app.route("/api/tally/companies/<int:company_id>", methods=["PUT"])
def tally_update_company(company_id):
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    with get_conn() as conn:
        conn.execute("""
            UPDATE tally_companies SET name=?, address=?, city=?, state=?, pincode=?,
            gstin=?, pan=?, phone=?, email=?, bank_name=?, bank_account=?, bank_ifsc=?,
            opening_balance=? WHERE id=?
        """, (data.get("name"), data.get("address",""), data.get("city",""),
              data.get("state",""), data.get("pincode",""), data.get("gstin",""),
              data.get("pan",""), data.get("phone",""), data.get("email",""),
              data.get("bank_name",""), data.get("bank_account",""), data.get("bank_ifsc",""),
              float(data.get("opening_balance", 0)), company_id))
    return jsonify({"status": "success"})


@app.route("/api/tally/companies/<int:company_id>", methods=["DELETE"])
def tally_delete_company(company_id):
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    with get_conn() as conn:
        conn.execute("UPDATE tally_companies SET is_active=0 WHERE id=?", (company_id,))
    return jsonify({"status": "success"})


@app.route("/api/tally/entries", methods=["GET"])
def tally_list_entries():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    company_id = request.args.get("company_id")
    entry_type = request.args.get("entry_type")
    party_name = request.args.get("party_name", "")
    date_from  = request.args.get("date_from", "")
    date_to    = request.args.get("date_to", "")
    limit      = int(request.args.get("limit", 100))

    query = "SELECT * FROM tally_entries WHERE 1=1"
    params: list = []
    if company_id:
        query += " AND company_id=?"
        params.append(int(company_id))
    if entry_type:
        query += " AND entry_type=?"
        params.append(entry_type)
    if party_name:
        query += " AND party_name LIKE ?"
        params.append(f"%{party_name}%")
    if date_from:
        query += " AND date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND date <= ?"
        params.append(date_to)
    query += " ORDER BY date DESC, id DESC LIMIT ?"
    params.append(limit)

    with get_conn() as conn:
        entries = [dict(r) for r in conn.execute(query, params).fetchall()]
    return jsonify(entries)


@app.route("/api/tally/entries", methods=["POST"])
def tally_create_entry():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    # session stores username STRING — not a dict
    user_raw = session.get("portal_user") or session.get("admin_user") or "system"
    if isinstance(user_raw, dict):
        created_by = user_raw.get("name", "system")
    else:
        created_by = str(user_raw) or "system"

    missing = required_fields(data, ["entry_type", "amount", "date"])
    if missing:
        return json_error(f"Missing: {', '.join(missing)}", 400)

    try:
        amount      = float(data.get("amount", 0) or 0)
        gst_amount  = float(data.get("gst_amount", 0) or 0)
        total       = float(data.get("total_amount", amount + gst_amount) or amount + gst_amount)
    except (TypeError, ValueError):
        return json_error("Invalid number in amount/gst_amount/total_amount", 400)

    # Normalise company_id (can be null/empty)
    company_id = data.get("company_id")
    if company_id in ("", "null", None):
        company_id = None
    else:
        try: company_id = int(company_id)
        except (TypeError, ValueError): company_id = None

    now = datetime.now(timezone.utc).isoformat()

    try:
        with get_conn() as conn:
            cur = conn.execute("""
                INSERT INTO tally_entries (company_id, entry_type, party_name, party_type,
                amount, gst_amount, total_amount, description, reference_no, invoice_no,
                date, due_date, payment_mode, bank_name, cheque_no, status, created_at, updated_at, created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (company_id, data.get("entry_type"), data.get("party_name",""),
                  data.get("party_type","supplier"), amount, gst_amount, total,
                  data.get("description",""), data.get("reference_no",""), data.get("invoice_no",""),
                  data.get("date"), data.get("due_date",""), data.get("payment_mode","cash"),
                  data.get("bank_name",""), data.get("cheque_no",""),
                  data.get("status","pending"), now, now, created_by))
            new_id = cur.lastrowid
        _audit("tally_entry_save", "tally", new_id, new=f"{data.get('entry_type')} ₹{total} ({data.get('party_name')})")
        return jsonify({"status": "success", "id": new_id})
    except Exception as e:
        return json_error(f"Save failed: {e}", 500)


@app.route("/api/tally/entries/<int:entry_id>", methods=["PUT"])
def tally_update_entry(entry_id):
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    now  = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        entry = conn.execute("SELECT * FROM tally_entries WHERE id=?", (entry_id,)).fetchone()
        if not entry:
            return json_error("Entry not found", 404)
        amount     = float(data.get("amount", entry["amount"]))
        gst_amount = float(data.get("gst_amount", entry["gst_amount"]))
        total      = float(data.get("total_amount", amount + gst_amount))
        conn.execute("""
            UPDATE tally_entries SET entry_type=?, party_name=?, party_type=?,
            amount=?, gst_amount=?, total_amount=?, description=?, reference_no=?,
            invoice_no=?, date=?, due_date=?, payment_mode=?, bank_name=?, cheque_no=?,
            status=?, updated_at=? WHERE id=?
        """, (data.get("entry_type", entry["entry_type"]),
              data.get("party_name", entry["party_name"]),
              data.get("party_type", entry["party_type"]),
              amount, gst_amount, total,
              data.get("description", entry["description"]),
              data.get("reference_no", entry["reference_no"]),
              data.get("invoice_no", entry["invoice_no"]),
              data.get("date", entry["date"]),
              data.get("due_date", entry["due_date"]),
              data.get("payment_mode", entry["payment_mode"]),
              data.get("bank_name", entry["bank_name"]),
              data.get("cheque_no", entry["cheque_no"]),
              data.get("status", entry["status"]),
              now, entry_id))
    return jsonify({"status": "success"})


@app.route("/api/tally/entries/<int:entry_id>", methods=["DELETE"])
def tally_delete_entry(entry_id):
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    with get_conn() as conn:
        conn.execute("DELETE FROM tally_entries WHERE id=?", (entry_id,))
    return jsonify({"status": "success"})


@app.route("/api/tally/entries/recent", methods=["GET"])
def tally_recent_entries():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    with get_conn() as conn:
        entries = [dict(r) for r in conn.execute(
            "SELECT * FROM tally_entries ORDER BY updated_at DESC LIMIT 20").fetchall()]
    return jsonify(entries)


@app.route("/api/tally/summary", methods=["GET"])
def tally_summary():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    company_id = request.args.get("company_id")
    date_from  = request.args.get("date_from", datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d"))
    date_to    = request.args.get("date_to", datetime.now(timezone.utc).strftime("%Y-%m-%d"))

    q_base = "FROM tally_entries WHERE date BETWEEN ? AND ?"
    params_base = [date_from, date_to]
    if company_id:
        q_base += " AND company_id=?"
        params_base.append(int(company_id))

    with get_conn() as conn:
        # Purchase invoices (amount payable to suppliers)
        payable = conn.execute(
            f"SELECT COALESCE(SUM(total_amount),0) AS s {q_base} AND entry_type='purchase'",
            params_base).fetchone()["s"]
        # Payments made
        paid = conn.execute(
            f"SELECT COALESCE(SUM(total_amount),0) AS s {q_base} AND entry_type='payment'",
            params_base).fetchone()["s"]
        # Sales receipts from retail
        receipts = conn.execute(
            f"SELECT COALESCE(SUM(total_amount),0) AS s {q_base} AND entry_type='receipt'",
            params_base).fetchone()["s"]
        # Outstanding by supplier
        outstanding = [dict(r) for r in conn.execute(f"""
            SELECT party_name, SUM(CASE WHEN entry_type='purchase' THEN total_amount
                                        WHEN entry_type='payment' THEN -total_amount ELSE 0 END) AS outstanding
            {q_base} GROUP BY party_name ORDER BY outstanding DESC
        """, params_base).fetchall()]

        pending_cheques = conn.execute(
            "SELECT COUNT(*) AS c FROM cheque_register WHERE status='issued'").fetchone()["c"]

    return jsonify({
        "period": {"from": date_from, "to": date_to},
        "total_payable": round(payable, 2),
        "total_paid": round(paid, 2),
        "total_receipts": round(receipts, 2),
        "net_balance": round(payable - paid, 2),
        "outstanding_by_supplier": outstanding,
        "pending_cheques": pending_cheques
    })


@app.route("/api/tally/bank-statements", methods=["GET"])
def tally_list_statements():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    company_id = request.args.get("company_id")
    query = "SELECT id, company_id, bank_name, account_no, statement_month, opening_balance, closing_balance, total_credits, total_debits, file_name, uploaded_at FROM bank_statements"
    params: list = []
    if company_id:
        query += " WHERE company_id=?"
        params.append(int(company_id))
    query += " ORDER BY statement_month DESC LIMIT 50"
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    return jsonify(rows)


@app.route("/api/tally/bank-statements", methods=["POST"])
def tally_upload_statement():
    """Accept pasted CSV/TSV bank statement rows."""
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data     = request.get_json(silent=True) or {}
    entries  = data.get("entries", [])
    user     = session.get("portal_user") or session.get("admin_user") or {}

    total_cr = sum(float(e.get("credit", 0) or 0) for e in entries)
    total_dr = sum(float(e.get("debit", 0) or 0)  for e in entries)

    with get_conn() as conn:
        conn.execute("""
            INSERT INTO bank_statements (company_id, bank_name, account_no, statement_month,
            opening_balance, closing_balance, total_credits, total_debits, entries, file_name,
            uploaded_at, uploaded_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (data.get("company_id"), data.get("bank_name",""), data.get("account_no",""),
              data.get("statement_month", datetime.now(timezone.utc).strftime("%Y-%m")),
              float(data.get("opening_balance", 0)),
              float(data.get("closing_balance", 0)),
              total_cr, total_dr,
              json.dumps(entries), data.get("file_name","paste"),
              datetime.now(timezone.utc).isoformat(), user.get("name","system")))
    return jsonify({"status": "success", "entries": len(entries),
                    "total_credits": total_cr, "total_debits": total_dr})


@app.route("/api/tally/cheques", methods=["GET"])
def tally_list_cheques():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    company_id = request.args.get("company_id")
    status     = request.args.get("status")
    query = "SELECT * FROM cheque_register WHERE 1=1"
    params: list = []
    if company_id:
        query += " AND company_id=?"
        params.append(int(company_id))
    if status:
        query += " AND status=?"
        params.append(status)
    query += " ORDER BY date DESC LIMIT 200"
    with get_conn() as conn:
        cheques = [dict(r) for r in conn.execute(query, params).fetchall()]
    return jsonify(cheques)


@app.route("/api/tally/cheques", methods=["POST"])
def tally_issue_cheque():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    missing = required_fields(data, ["payee", "amount", "date"])
    if missing:
        return json_error(f"Missing: {', '.join(missing)}", 400)
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO cheque_register (company_id, cheque_no, date, payee, bank_name,
            account_no, amount, memo, status, entry_id, created_at)
            VALUES (?,?,?,?,?,?,?,?,'issued',?,?)
        """, (data.get("company_id"), data.get("cheque_no",""),
              data.get("date"), data.get("payee"),
              data.get("bank_name",""), data.get("account_no",""),
              float(data.get("amount", 0)), data.get("memo",""),
              data.get("entry_id"), datetime.now(timezone.utc).isoformat()))
    return jsonify({"status": "success"})


@app.route("/api/tally/cheques/<int:cheque_id>", methods=["PUT"])
def tally_update_cheque(cheque_id):
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data = request.get_json(silent=True) or {}
    new_status = data.get("status", "issued")
    printed_at = datetime.now(timezone.utc).isoformat() if new_status == "printed" else None
    with get_conn() as conn:
        if printed_at:
            conn.execute("UPDATE cheque_register SET status=?, printed_at=? WHERE id=?",
                         (new_status, printed_at, cheque_id))
        else:
            conn.execute("UPDATE cheque_register SET status=? WHERE id=?",
                         (new_status, cheque_id))
    return jsonify({"status": "success"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PHASE 2 — SMART RECEIVING API (portal-gated, kept for backward compat)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/portal/receiving/orders", methods=["GET"])
def receiving_list_orders():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    status = request.args.get("status")
    query = "SELECT * FROM receiving_orders WHERE 1=1"
    params: list = []
    if status:
        query += " AND status=?"
        params.append(status)
    query += " ORDER BY id DESC LIMIT 100"
    with get_conn() as conn:
        orders = [dict(r) for r in conn.execute(query, params).fetchall()]
    for o in orders:
        o["items"] = safe_json_loads(o.get("items"), [])
        o["return_items"] = safe_json_loads(o.get("return_items"), [])
    return jsonify(orders)


@app.route("/api/receiving/orders", methods=["POST"])
def receiving_create_order():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data  = request.get_json(silent=True) or {}
    user  = session.get("portal_user") or {}
    items = data.get("items", [])

    # Generate daily code (6-char alphanumeric, date-seeded)
    import hashlib as _hl
    daily_seed = datetime.now(timezone.utc).strftime("%Y%m%d") + str(user.get("id","0"))
    daily_code = _hl.md5(daily_seed.encode()).hexdigest()[:6].upper()

    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO receiving_orders (shop_id, shop_type, supplier_name, po_date,
            expected_date, items, status, daily_code, created_at)
            VALUES (?,?,?,?,?,?,'pending',?,?)
        """, (user.get("id", 0), user.get("account_type","retail"),
              data.get("supplier_name",""), data.get("po_date",""),
              data.get("expected_date",""), json.dumps(items), daily_code,
              datetime.now(timezone.utc).isoformat()))
        order_id = cur.lastrowid
    return jsonify({"status": "success", "order_id": order_id, "daily_code": daily_code})


@app.route("/api/receiving/orders/<int:order_id>/items", methods=["PUT"])
def receiving_update_items(order_id):
    """Staff ticks items off the checklist — update received quantities."""
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data  = request.get_json(silent=True) or {}
    items = data.get("items", [])
    with get_conn() as conn:
        conn.execute("UPDATE receiving_orders SET items=?, status='receiving' WHERE id=?",
                     (json.dumps(items), order_id))
    return jsonify({"status": "success"})


@app.route("/api/receiving/orders/<int:order_id>/complete", methods=["POST"])
def receiving_complete_order(order_id):
    """Complete receiving — auto-update medicine stock, flag returns."""
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data         = request.get_json(silent=True) or {}
    user         = session.get("portal_user") or {}
    return_items = data.get("return_items", [])
    now          = datetime.now(timezone.utc).isoformat()

    with get_conn() as conn:
        order = conn.execute("SELECT * FROM receiving_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return json_error("Order not found", 404)

        items = safe_json_loads(order["items"], [])
        updated_meds = 0

        for item in items:
            if not item.get("received", False):
                continue
            name     = str(item.get("name","")).strip()
            qty_recv = int(item.get("qty_received", item.get("qty_ordered", 0)))
            price    = float(item.get("price", 0))
            if not name:
                continue
            existing = conn.execute("SELECT id, s FROM medicines WHERE n=?", (name,)).fetchone()
            if existing:
                new_stock = existing["s"] + qty_recv
                conn.execute("UPDATE medicines SET s=?, p_rate=? WHERE id=?",
                             (new_stock, price, existing["id"]))
            else:
                new_id = "recv_" + _uuid_mod.uuid4().hex[:8]
                conn.execute("""
                    INSERT OR IGNORE INTO medicines (id, n, c, p, s, p_rate, reorder, max_qty)
                    VALUES (?,?,?,?,?,?,0,0)
                """, (new_id, name, "General", price, qty_recv, price))
            updated_meds += 1

        conn.execute("""
            UPDATE receiving_orders SET status='completed', received_by=?, received_at=?,
            return_items=?, notes=? WHERE id=?
        """, (user.get("name","staff"), now, json.dumps(return_items),
              data.get("notes",""), order_id))

    return jsonify({"status": "success", "medicines_updated": updated_meds,
                    "returns": len(return_items)})


@app.route("/api/receiving/verify-code", methods=["POST"])
def receiving_verify_code():
    """Verify daily access code for staff receiving."""
    data = request.get_json(silent=True) or {}
    code = str(data.get("code","")).strip().upper()
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    with get_conn() as conn:
        order = conn.execute(
            "SELECT * FROM receiving_orders WHERE daily_code=? AND po_date=?",
            (code, today)).fetchone()
        if not order:
            # also check by code alone today
            order = conn.execute(
                "SELECT * FROM receiving_orders WHERE daily_code=? AND date(created_at)=date('now')",
                (code,)).fetchone()
    if not order:
        return json_error("Invalid or expired code", 403)
    o = dict(order)
    o["items"] = safe_json_loads(o.get("items"), [])
    return jsonify({"status": "valid", "order": o})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PHASE 2 — DELIVERY TRACKING API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/delivery/routes", methods=["GET"])
def delivery_list_routes():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    date_filter = request.args.get("date", "")
    query = "SELECT * FROM delivery_routes WHERE 1=1"
    params: list = []
    if date_filter:
        query += " AND delivery_date=?"
        params.append(date_filter)
    query += " ORDER BY id DESC LIMIT 100"
    with get_conn() as conn:
        routes = [dict(r) for r in conn.execute(query, params).fetchall()]
    for r in routes:
        r["shops"] = safe_json_loads(r.get("shops"), [])
    return jsonify(routes)


@app.route("/api/delivery/routes", methods=["POST"])
def delivery_create_route():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data  = request.get_json(silent=True) or {}
    shops = data.get("shops", [])
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO delivery_routes (route_name, delivery_boy, delivery_phone,
            delivery_date, shops, status, created_at, notes)
            VALUES (?,?,?,?,?,'pending',?,?)
        """, (data.get("route_name","Route-1"), data.get("delivery_boy",""),
              data.get("delivery_phone",""),
              data.get("delivery_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
              json.dumps(shops), datetime.now(timezone.utc).isoformat(),
              data.get("notes","")))
        route_id = cur.lastrowid
    return jsonify({"status": "success", "route_id": route_id})


@app.route("/api/delivery/routes/<int:route_id>", methods=["PUT"])
def delivery_update_route(route_id):
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data   = request.get_json(silent=True) or {}
    status = data.get("status", "pending")
    shops  = data.get("shops")
    now    = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        if status == "started":
            conn.execute("UPDATE delivery_routes SET status=?, started_at=? WHERE id=?",
                         (status, now, route_id))
        elif status == "completed":
            conn.execute("UPDATE delivery_routes SET status=?, completed_at=? WHERE id=?",
                         (status, now, route_id))
        elif shops is not None:
            conn.execute("UPDATE delivery_routes SET shops=? WHERE id=?",
                         (json.dumps(shops), route_id))
        else:
            conn.execute("UPDATE delivery_routes SET status=? WHERE id=?",
                         (status, route_id))
    return jsonify({"status": "success"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PHASE 2 — AI CALLER API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/calls/scheduled", methods=["GET"])
def calls_list():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    status = request.args.get("status")
    query  = "SELECT * FROM call_logs WHERE 1=1"
    params: list = []
    if status:
        query += " AND status=?"
        params.append(status)
    query += " ORDER BY id DESC LIMIT 200"
    with get_conn() as conn:
        logs = [dict(r) for r in conn.execute(query, params).fetchall()]
    return jsonify(logs)


@app.route("/api/calls/schedule", methods=["POST"])
def calls_schedule():
    """Auto-generate call tasks for shops with critically low stock."""
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    data      = request.get_json(silent=True) or {}
    call_type = data.get("call_type", "medicine_reminder")
    now       = datetime.now(timezone.utc)
    scheduled = []

    with get_conn() as conn:
        if call_type == "medicine_reminder":
            # Find medicines with stock < reorder point for current shop
            user     = session.get("portal_user") or {}
            shop_id  = user.get("id", 0)
            shop_name = user.get("name","")
            phone    = user.get("phone", data.get("phone",""))

            meds = conn.execute("""
                SELECT n, s, reorder, p FROM medicines
                WHERE s > 0 AND reorder > 0 AND s <= reorder * 1.5
                ORDER BY (CAST(s AS REAL)/CAST(reorder AS REAL)) ASC LIMIT 10
            """).fetchall()

            for med in meds:
                avg_daily = float(med["reorder"]) / 7 if med["reorder"] > 0 else 0.1
                days_left = round(med["s"] / avg_daily, 1) if avg_daily > 0 else 99
                scheduled_at = (now + timedelta(hours=1)).isoformat()
                conn.execute("""
                    INSERT INTO call_logs (shop_id, shop_name, phone, call_type, status,
                    scheduled_at, medicine_name, days_of_stock, created_at)
                    VALUES (?,?,?,'medicine_reminder','scheduled',?,?,?,?)
                """, (shop_id, shop_name, phone, scheduled_at, med["n"], days_left,
                      now.isoformat()))
                scheduled.append({"medicine": med["n"], "days_left": days_left})

        elif call_type == "payment_reminder":
            # Schedule calls for overdue tally entries
            overdue = conn.execute("""
                SELECT party_name, SUM(total_amount) AS total_due
                FROM tally_entries
                WHERE entry_type='purchase' AND status='pending'
                AND due_date < date('now')
                GROUP BY party_name
                HAVING total_due > 0
                LIMIT 10
            """).fetchall()

            user  = session.get("portal_user") or {}
            phone = user.get("phone", data.get("phone",""))
            for row in overdue:
                conn.execute("""
                    INSERT INTO call_logs (shop_id, shop_name, phone, call_type, status,
                    scheduled_at, medicine_name, created_at)
                    VALUES (?,?,?,'payment_reminder','scheduled',?,?,?)
                """, (user.get("id",0), row["party_name"], phone,
                      (now + timedelta(hours=2)).isoformat(),
                      f"Payment due: ₹{row['total_due']:.0f}", now.isoformat()))
                scheduled.append({"party": row["party_name"], "due": row["total_due"]})

    return jsonify({"status": "success", "scheduled": len(scheduled), "items": scheduled})


@app.route("/api/calls/<int:call_id>/response", methods=["PUT"])
def calls_record_response(call_id):
    """Record the response from an AI call (press 1=order, 2=reject, 3=later)."""
    data     = request.get_json(silent=True) or {}
    response = data.get("response", "")  # "1", "2", "3", or "no_answer"
    now      = datetime.now(timezone.utc)

    follow_up = None
    new_status = "completed"
    if response == "3":   # Later
        follow_up = (now + timedelta(hours=24)).isoformat()
        new_status = "follow_up"
    elif response == "no_answer":
        follow_up = (now + timedelta(hours=4)).isoformat()
        new_status = "follow_up"

    with get_conn() as conn:
        call = conn.execute("SELECT * FROM call_logs WHERE id=?", (call_id,)).fetchone()
        if not call:
            return json_error("Call not found", 404)
        conn.execute("""
            UPDATE call_logs SET call_response=?, status=?, made_at=?, follow_up_at=?
            WHERE id=?
        """, (response, new_status, now.isoformat(), follow_up, call_id))

        action_taken = None
        if response == "1" and call["call_type"] == "medicine_reminder":
            # Auto-create reorder suggestion (would trigger actual order in full integration)
            action_taken = f"Reorder initiated for {call['medicine_name']}"

    return jsonify({"status": "success", "action": action_taken, "follow_up": follow_up})


@app.route("/api/calls/stats", methods=["GET"])
def calls_stats():
    if not session.get("portal_user") and not session.get("admin_user"):
        return json_error("Forbidden", 403)
    with get_conn() as conn:
        total     = conn.execute("SELECT COUNT(*) AS c FROM call_logs").fetchone()["c"]
        completed = conn.execute("SELECT COUNT(*) AS c FROM call_logs WHERE status='completed'").fetchone()["c"]
        ordered   = conn.execute("SELECT COUNT(*) AS c FROM call_logs WHERE call_response='1'").fetchone()["c"]
        rejected  = conn.execute("SELECT COUNT(*) AS c FROM call_logs WHERE call_response='2'").fetchone()["c"]
        no_answer = conn.execute("SELECT COUNT(*) AS c FROM call_logs WHERE call_response='no_answer'").fetchone()["c"]
    return jsonify({
        "total": total, "completed": completed,
        "ordered": ordered, "rejected": rejected, "no_answer": no_answer,
        "order_rate": round(ordered / max(completed,1) * 100, 1)
    })


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# WHATSAPP
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/bills/<bill_id>/whatsapp", methods=["POST"])
def bill_send_whatsapp(bill_id):
    """Send a WhatsApp bill receipt to the customer's number."""
    data  = request.get_json(silent=True) or {}
    phone = data.get("phone", "").strip()

    with get_conn() as conn:
        row = conn.execute("SELECT * FROM bills WHERE id=?", (bill_id,)).fetchone()
        if not row:
            return json_error("Bill not found", 404)
        bill = normalize_bill_row(row)

    if not phone:
        phone = bill.get("phone", "")
    if not phone:
        return json_error("No phone number on bill", 400)

    msg    = data.get("message") or build_bill_receipt_message(bill)
    result = send_whatsapp(phone, msg)
    now    = datetime.now(timezone.utc).isoformat()

    with get_conn() as conn:
        conn.execute("""
            INSERT INTO whatsapp_logs (bill_id, phone, message, status, provider_id, sent_at, error_msg)
            VALUES (?,?,?,?,?,?,?)
        """, (bill_id, phone, msg, result["status"],
              result.get("sid", ""), now, result.get("error", "")))
        if result["status"] in ("sent", "mocked"):
            conn.execute("UPDATE bills SET whatsapp_sent=1 WHERE id=?", (bill_id,))

    return jsonify({"status": result["status"], "sid": result.get("sid", ""),
                    "phone": phone, "note": result.get("note", "")})


@app.route("/api/whatsapp/logs", methods=["GET"])
def whatsapp_logs():
    """List all WhatsApp message logs."""
    limit  = request.args.get("limit", 50, type=int)
    offset = request.args.get("offset", 0,  type=int)
    with get_conn() as conn:
        rows  = conn.execute(
            "SELECT * FROM whatsapp_logs ORDER BY id DESC LIMIT ? OFFSET ?",
            (limit, offset)
        ).fetchall()
        total = conn.execute("SELECT COUNT(*) AS c FROM whatsapp_logs").fetchone()["c"]
    return jsonify({
        "logs":   [dict(r) for r in rows],
        "total":  total,
        "limit":  limit,
        "offset": offset,
    })


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# FACE RECOGNITION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/customers/face-match", methods=["POST"])
def customer_face_match_route():
    """
    Match a 128-dim face vector against all stored customer vectors.
    Body: { "vector": [...128 floats...], "threshold": 0.50 }
    Returns: matched customer info + confidence, or 404 if no match.
    """
    data      = request.get_json(silent=True) or {}
    vector    = data.get("vector") or data.get("descriptor") or []
    threshold = float(data.get("threshold", 0.50))

    if isinstance(vector, str):
        try: vector = json.loads(vector)
        except Exception: vector = []

    if not vector or not isinstance(vector, list):
        return json_error("Provide a 'vector' or 'descriptor' array of floats", 400)

    # Coerce numbers
    try:
        vector = [float(x) for x in vector]
    except (TypeError, ValueError):
        return json_error("vector must contain numbers only", 400)

    match = face_match_customer(vector, threshold)
    if not match:
        return jsonify({"matched": False}), 200

    return jsonify({"matched": True, **match})


@app.route("/api/customers/<int:cust_id>/face", methods=["PATCH"])
def update_customer_face(cust_id):
    """Store or replace the face vector for a customer.
    Accepts any reasonable vector length (32–1024) — older 128-d face-api.js and newer 512-d ArcFace both work."""
    data   = request.get_json(silent=True) or {}
    vector = data.get("vector") or data.get("descriptor") or data.get("face_vector") or []

    if isinstance(vector, str):
        try: vector = json.loads(vector)
        except Exception: vector = []

    if not isinstance(vector, list) or not (32 <= len(vector) <= 1024):
        return json_error(f"Provide a 'vector' array (32-1024 floats). Got: {len(vector) if isinstance(vector, list) else 'invalid'}", 400)

    # Coerce to floats
    try:
        vector = [float(x) for x in vector]
    except (TypeError, ValueError):
        return json_error("vector must contain numbers only", 400)

    with get_conn() as conn:
        row = conn.execute("SELECT id FROM customers WHERE id=?", (cust_id,)).fetchone()
        if not row:
            return json_error("Customer not found", 404)
        conn.execute(
            "UPDATE customers SET face_vector=? WHERE id=?",
            (json.dumps(vector), cust_id)
        )
    return jsonify({"status": "success", "vector_length": len(vector)})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STAFF MANAGEMENT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/staff", methods=["GET"])
def get_staff():
    """List all staff members."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, name, role, phone, pin, is_active, salary, joined_date, created_at "
            "FROM staff ORDER BY name COLLATE NOCASE"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/staff", methods=["POST"])
def save_staff():
    """Create or update a staff member. Pass 'id' to update."""
    data    = request.get_json(silent=True) or {}
    name    = str(data.get("name", "")).strip()
    if not name:
        return json_error("Name is required", 400)

    role        = data.get("role", "Cashier")
    phone       = data.get("phone", "")
    pin         = str(data.get("pin", "")).strip()
    salary      = float(data.get("salary", 0) or 0)
    joined_date = data.get("joined_date", "")
    is_active   = int(data.get("is_active", 1))
    now         = datetime.now(timezone.utc).isoformat()
    staff_id    = data.get("id")

    with get_conn() as conn:
        if staff_id:
            conn.execute("""
                UPDATE staff SET name=?, role=?, phone=?, pin=?, is_active=?, salary=?, joined_date=?
                WHERE id=?
            """, (name, role, phone, pin, is_active, salary, joined_date, staff_id))
            sid = staff_id
        else:
            cur = conn.execute("""
                INSERT INTO staff (name, role, phone, pin, is_active, salary, joined_date, created_at)
                VALUES (?,?,?,?,?,?,?,?)
            """, (name, role, phone, pin, is_active, salary, joined_date, now))
            sid = cur.lastrowid

    return jsonify({"status": "success", "id": sid})


@app.route("/api/staff/<int:staff_id>", methods=["DELETE"])
def delete_staff(staff_id):
    """Deactivate (soft-delete) a staff member."""
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM staff WHERE id=?", (staff_id,)).fetchone()
        if not row:
            return json_error("Staff not found", 404)
        conn.execute("UPDATE staff SET is_active=0 WHERE id=?", (staff_id,))
    return jsonify({"status": "success"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ATTENDANCE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/attendance/punch", methods=["POST"])
def attendance_punch():
    """
    Punch in or out. Automatically decides direction.
    Body: { "pin": "1234" }  OR  { "staff_id": 1 }
    If staff has no punch-in today  → punch IN.
    If staff already punched in     → punch OUT and calc hours.
    """
    data     = request.get_json(silent=True) or {}
    pin      = str(data.get("pin", "")).strip()
    staff_id = data.get("staff_id")
    today    = datetime.now().strftime("%Y-%m-%d")
    now_time = datetime.now().strftime("%H:%M:%S")

    with get_conn() as conn:
        if pin:
            staff = conn.execute(
                "SELECT * FROM staff WHERE pin=? AND is_active=1", (pin,)
            ).fetchone()
        elif staff_id:
            staff = conn.execute(
                "SELECT * FROM staff WHERE id=? AND is_active=1", (staff_id,)
            ).fetchone()
        else:
            return json_error("Provide 'pin' or 'staff_id'", 400)

        if not staff:
            return json_error("Staff not found or PIN incorrect", 404)

        log = conn.execute(
            "SELECT * FROM attendance_logs WHERE staff_id=? AND date=?",
            (staff["id"], today)
        ).fetchone()

        if not log:
            # First punch = IN
            conn.execute("""
                INSERT INTO attendance_logs
                (staff_id, staff_name, date, punch_in, status, created_at)
                VALUES (?,?,?,?,?,?)
            """, (staff["id"], staff["name"], today, now_time,
                  "late" if now_time > "09:30:00" else "present",
                  datetime.now(timezone.utc).isoformat()))
            direction = "in"
            worked    = None
        elif not log["punch_out"]:
            # Second punch = OUT
            punch_in_dt  = datetime.strptime(f"{today} {log['punch_in']}", "%Y-%m-%d %H:%M:%S")
            punch_out_dt = datetime.strptime(f"{today} {now_time}", "%Y-%m-%d %H:%M:%S")
            worked       = round((punch_out_dt - punch_in_dt).seconds / 3600, 2)
            conn.execute("""
                UPDATE attendance_logs
                SET punch_out=?, worked_hours=?
                WHERE staff_id=? AND date=?
            """, (now_time, worked, staff["id"], today))
            direction = "out"
        else:
            return jsonify({
                "status": "already_complete",
                "message": f"{staff['name']} already punched out at {log['punch_out']}",
                "staff": dict(staff),
            })

    return jsonify({
        "status":    "success",
        "direction": direction,
        "staff":     {"id": staff["id"], "name": staff["name"], "role": staff["role"]},
        "time":      now_time,
        "worked_hours": worked,
    })


@app.route("/api/attendance", methods=["GET"])
def get_attendance():
    """
    List attendance logs.
    ?date=YYYY-MM-DD   — filter by date (defaults to today)
    ?staff_id=1        — filter by staff
    ?month=YYYY-MM     — full month view
    """
    date_filter  = request.args.get("date", "")
    month_filter = request.args.get("month", "")
    staff_filter = request.args.get("staff_id", type=int)

    with get_conn() as conn:
        where, params = [], []

        if date_filter:
            where.append("a.date=?")
            params.append(date_filter)
        elif month_filter:
            where.append("a.date LIKE ?")
            params.append(f"{month_filter}%")
        else:
            where.append("a.date=?")
            params.append(datetime.now().strftime("%Y-%m-%d"))

        if staff_filter:
            where.append("a.staff_id=?")
            params.append(staff_filter)

        clause = f"WHERE {' AND '.join(where)}" if where else ""
        rows   = conn.execute(f"""
            SELECT a.*, s.role
            FROM attendance_logs a
            LEFT JOIN staff s ON a.staff_id = s.id
            {clause}
            ORDER BY a.date DESC, a.punch_in ASC
        """, params).fetchall()

    return jsonify([dict(r) for r in rows])


@app.route("/api/attendance/summary", methods=["GET"])
def attendance_summary():
    """Daily summary: total staff, present, absent, late count."""
    date = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    with get_conn() as conn:
        total_staff = conn.execute(
            "SELECT COUNT(*) AS c FROM staff WHERE is_active=1"
        ).fetchone()["c"]
        present  = conn.execute(
            "SELECT COUNT(*) AS c FROM attendance_logs WHERE date=?", (date,)
        ).fetchone()["c"]
        late     = conn.execute(
            "SELECT COUNT(*) AS c FROM attendance_logs WHERE date=? AND status='late'", (date,)
        ).fetchone()["c"]
        punched_out = conn.execute(
            "SELECT COUNT(*) AS c FROM attendance_logs WHERE date=? AND punch_out IS NOT NULL", (date,)
        ).fetchone()["c"]

    return jsonify({
        "date":        date,
        "total_staff": total_staff,
        "present":     present,
        "absent":      total_staff - present,
        "late":        late,
        "punched_out": punched_out,
        "still_in":    present - punched_out,
    })


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BILLS HISTORY + INVOICE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/bills/search", methods=["GET"])
def bills_search():
    """
    Advanced bill search.
    Params: q (customer/doctor text), date_from, date_to, pay, bill_type, limit, offset
    """
    q          = request.args.get("q", "").strip()
    date_from  = request.args.get("date_from", "")
    date_to    = request.args.get("date_to", "")
    pay        = request.args.get("pay", "")
    bill_type  = request.args.get("bill_type", "")
    limit      = request.args.get("limit",  50, type=int)
    offset     = request.args.get("offset",  0, type=int)

    where, params = [], []

    if q:
        where.append("(cust LIKE ? OR doctor LIKE ? OR id LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if date_from:
        # bills.date is stored as "DD/MM/YYYY HH:MM" so compare ts
        try:
            ts_from = int(datetime.strptime(date_from, "%Y-%m-%d").timestamp() * 1000)
            where.append("ts >= ?")
            params.append(ts_from)
        except ValueError:
            pass
    if date_to:
        try:
            ts_to = int((datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)).timestamp() * 1000)
            where.append("ts < ?")
            params.append(ts_to)
        except ValueError:
            pass
    if pay:
        where.append("pay=?")
        params.append(pay)
    if bill_type:
        where.append("bill_type=?")
        params.append(bill_type)

    clause = f"WHERE {' AND '.join(where)}" if where else ""

    with get_conn() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) AS c FROM bills {clause}", params
        ).fetchone()["c"]
        rows = conn.execute(
            f"SELECT * FROM bills {clause} ORDER BY ts DESC LIMIT ? OFFSET ?",
            params + [limit, offset]
        ).fetchall()

    return jsonify({
        "bills":  [normalize_bill_row(r) for r in rows],
        "total":  total,
        "limit":  limit,
        "offset": offset,
    })


@app.route("/api/bills/<bill_id>", methods=["GET"])
def get_single_bill(bill_id):
    """Fetch one bill by ID."""
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM bills WHERE id=?", (bill_id,)).fetchone()
    if not row:
        return json_error("Bill not found", 404)
    return jsonify(normalize_bill_row(row))


@app.route("/api/bills/<bill_id>", methods=["DELETE"])
def cancel_bill(bill_id):
    """Cancel (soft-delete) a bill and restore stock."""
    data   = request.get_json(silent=True) or {}
    reason = data.get("reason", "Cancelled")

    with get_conn() as conn:
        row = conn.execute("SELECT * FROM bills WHERE id=?", (bill_id,)).fetchone()
        if not row:
            return json_error("Bill not found", 404)

        bill  = normalize_bill_row(row)
        items = bill.get("items", [])

        # Restore stock
        for item in items:
            med_id = str(item.get("id", "")).strip()
            qty    = int(item.get("qty", 0) or 0)
            if med_id and qty > 0:
                conn.execute("UPDATE medicines SET s = s + ? WHERE id=?", (qty, med_id))

        # Mark bill cancelled via a prefix on customer name
        conn.execute(
            "UPDATE bills SET cust=?, doctor=? WHERE id=?",
            (f"[CANCELLED] {bill.get('cust', '')}", reason, bill_id)
        )

    return jsonify({"status": "success", "message": f"Bill {bill_id} cancelled"})


# ── PAGE ROUTES (HTML) ────────────────────────────────────────────────

@app.route("/bills/history")
def bills_history_page():
    """Bill history & search page."""
    return render_template("bills_history.html")


@app.route("/bills/<bill_id>/invoice")
def bill_invoice_page(bill_id):
    """Printable invoice for a single bill."""
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM bills WHERE id=?", (bill_id,)).fetchone()
    if not row:
        return "Bill not found", 404
    bill = normalize_bill_row(row)
    # Attach extra columns if present
    bill["bill_type"]     = dict(row).get("bill_type", "retail")
    bill["customer_type"] = dict(row).get("customer_type", "customer")
    bill["whatsapp_sent"] = dict(row).get("whatsapp_sent", 0)
    return render_template("bill_invoice.html", bill=bill)


@app.route("/attendance")
def attendance_page():
    """Staff attendance management page."""
    return render_template("attendance.html")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SPRINT 3 — SMART STOCK RECEIVING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/receiving/orders", methods=["GET"])
def receiving_list():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM receiving_orders ORDER BY id DESC LIMIT 100"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/receiving/orders", methods=["POST"])
def receiving_create():
    data     = request.get_json(silent=True) or {}
    supplier = data.get("supplier_name","").strip()
    if not supplier:
        return json_error("supplier_name required", 400)

    items   = data.get("items", [])
    code    = str(int(datetime.now().timestamp()))[-6:]   # 6-digit daily code
    now     = datetime.now(timezone.utc).isoformat()

    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO receiving_orders
            (supplier_name, po_date, expected_date, items, status, daily_code, notes, created_at)
            VALUES (?,?,?,?,?,?,?,?)
        """, (supplier, data.get("po_date", now[:10]),
              data.get("expected_date", now[:10]),
              json.dumps(items), "pending", code,
              data.get("notes",""), now))
        oid = cur.lastrowid
    return jsonify({"status":"success","id":oid,"daily_code":code})


@app.route("/api/receiving/orders/<int:oid>", methods=["GET"])
def receiving_get(oid):
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM receiving_orders WHERE id=?", (oid,)).fetchone()
    if not row:
        return json_error("Order not found", 404)
    d = dict(row)
    d["items"]        = safe_json_loads(d.get("items"), [])
    d["return_items"] = safe_json_loads(d.get("return_items"), [])
    return jsonify(d)


@app.route("/api/receiving/orders/<int:oid>/verify-item", methods=["PUT"])
def receiving_verify_item(oid):
    """Mark one item line as verified/unverified with actuals."""
    data  = request.get_json(silent=True) or {}
    idx   = data.get("index")
    if idx is None:
        return json_error("index required", 400)

    with get_conn() as conn:
        row = conn.execute("SELECT * FROM receiving_orders WHERE id=?", (oid,)).fetchone()
        if not row:
            return json_error("Order not found", 404)

        items = safe_json_loads(row["items"], [])
        if idx < 0 or idx >= len(items):
            return json_error("Invalid item index", 400)

        items[idx]["verified"]      = data.get("verified", True)
        items[idx]["actual_qty"]    = data.get("actual_qty",    items[idx].get("qty", 0))
        items[idx]["actual_batch"]  = data.get("actual_batch",  items[idx].get("batch",""))
        items[idx]["actual_expiry"] = data.get("actual_expiry", items[idx].get("expiry",""))
        items[idx]["actual_mrp"]    = data.get("actual_mrp",    items[idx].get("mrp", 0))

        conn.execute("UPDATE receiving_orders SET items=? WHERE id=?",
                     (json.dumps(items), oid))

    return jsonify({"status":"success","item": items[idx]})


@app.route("/api/receiving/orders/<int:oid>/complete", methods=["POST"])
def receiving_complete(oid):
    """
    Complete receiving: auto-push verified items to medicines stock,
    generate return note for extras, log who received.
    """
    data     = request.get_json(silent=True) or {}
    receiver = data.get("received_by", "Staff")
    now      = datetime.now(timezone.utc).isoformat()

    with get_conn() as conn:
        row = conn.execute("SELECT * FROM receiving_orders WHERE id=?", (oid,)).fetchone()
        if not row:
            return json_error("Order not found", 404)
        if row["status"] == "completed":
            return json_error("Order already completed", 409)

        items        = safe_json_loads(row["items"], [])
        return_items = []
        imported     = []

        for item in items:
            if not item.get("verified"):
                continue
            med_name   = str(item.get("name","")).strip()
            actual_qty = int(item.get("actual_qty", item.get("qty", 0)) or 0)
            exp_qty    = int(item.get("qty", 0) or 0)
            batch      = item.get("actual_batch", item.get("batch",""))
            expiry     = item.get("actual_expiry", item.get("expiry",""))
            mrp        = float(item.get("actual_mrp", item.get("mrp", 0)) or 0)

            if not med_name or actual_qty <= 0:
                continue

            # Find medicine by name
            med = conn.execute(
                "SELECT id, s FROM medicines WHERE LOWER(n)=LOWER(?)", (med_name,)
            ).fetchone()

            if med:
                conn.execute("UPDATE medicines SET s=s+?, batch=?, expiry=? WHERE id=?",
                             (actual_qty, batch, expiry, med["id"]))
            else:
                # Auto-create new medicine
                new_id = str(_uuid_mod.uuid4())[:8]
                conn.execute("""
                    INSERT INTO medicines (id,n,g,c,p,s,batch,expiry,p_rate)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (new_id, med_name, "", "Tablet", mrp, actual_qty, batch, expiry, mrp*0.8))

            imported.append({"name": med_name, "qty": actual_qty})

            # Extra items → return note
            if actual_qty > exp_qty and exp_qty > 0:
                extra = actual_qty - exp_qty
                return_items.append({"name": med_name, "extra_qty": extra, "reason": "Excess received"})

        conn.execute("""
            UPDATE receiving_orders
            SET status='completed', received_by=?, received_at=?, return_items=?
            WHERE id=?
        """, (receiver, now, json.dumps(return_items), oid))

    return jsonify({
        "status":       "success",
        "imported":     len(imported),
        "items":        imported,
        "return_note":  return_items,
    })


@app.route("/stock/receiving")
def receiving_page():
    return render_template("stock_receiving.html")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SPRINT 5 — AI REMINDERS (WhatsApp + Call)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/reminders/compute", methods=["POST"])
def reminders_compute():
    """
    Scan recent bills → find chronic patients → compute expected finish date
    → create reminder 2 days before.
    Looks at bills in last 45 days, medicines with qty >= 14 (multi-day supply).
    """
    now  = datetime.now()
    cutoff = (now - timedelta(days=45)).strftime("%d/%m/%Y")
    created = 0
    skipped = 0

    with get_conn() as conn:
        bills = conn.execute(
            "SELECT * FROM bills WHERE cust != '' AND cust NOT LIKE '[CANCELLED]%' "
            "ORDER BY ts DESC LIMIT 500"
        ).fetchall()

        for bill in bills:
            b = normalize_bill_row(bill)
            if not b.get("phone"):
                continue
            for item in b.get("items", []):
                qty  = int(item.get("qty", item.get("s", 0)) or 0)
                if qty < 14:
                    continue   # Less than 14 = probably not daily medicine

                med_name = item.get("n", item.get("name", ""))
                if not med_name:
                    continue

                # Assume daily dose = 1–2 tablets, days supply = qty
                daily_dose = 2 if qty >= 60 else 1
                days_supply = qty // daily_dose

                # Parse bill date
                try:
                    bill_dt = datetime.strptime(b["date"][:10], "%d/%m/%Y")
                except Exception:
                    continue

                finish_dt   = bill_dt + timedelta(days=days_supply)
                reminder_dt = finish_dt - timedelta(days=2)

                if reminder_dt.date() < now.date():
                    skipped += 1
                    continue

                # Skip if reminder already exists
                exists = conn.execute(
                    "SELECT id FROM patient_reminders "
                    "WHERE customer_name=? AND medicine_name=? AND expected_finish=?",
                    (b["cust"], med_name, finish_dt.strftime("%Y-%m-%d"))
                ).fetchone()
                if exists:
                    skipped += 1
                    continue

                conn.execute("""
                    INSERT INTO patient_reminders
                    (customer_name,phone,medicine_name,last_bill_date,days_supply,
                     daily_dose,qty_dispensed,expected_finish,reminder_date,status,created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, (b["cust"], b["phone"], med_name,
                      bill_dt.strftime("%Y-%m-%d"), days_supply,
                      daily_dose, qty,
                      finish_dt.strftime("%Y-%m-%d"),
                      reminder_dt.strftime("%Y-%m-%d"),
                      "pending", now.isoformat()))
                created += 1

    return jsonify({"status":"success","created":created,"skipped":skipped})


@app.route("/api/reminders", methods=["GET"])
def reminders_list():
    status = request.args.get("status","")
    days   = request.args.get("days", 7, type=int)
    limit  = request.args.get("limit", 100, type=int)
    today  = datetime.now().strftime("%Y-%m-%d")
    future = (datetime.now()+timedelta(days=days)).strftime("%Y-%m-%d")

    with get_conn() as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM patient_reminders WHERE status=? ORDER BY reminder_date ASC LIMIT ?",
                (status, limit)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM patient_reminders WHERE reminder_date BETWEEN ? AND ? "
                "ORDER BY reminder_date ASC LIMIT ?",
                (today, future, limit)
            ).fetchall()
        total = conn.execute("SELECT COUNT(*) AS c FROM patient_reminders WHERE status='pending'").fetchone()["c"]

    return jsonify({"reminders":[dict(r) for r in rows],"pending_total":total})


@app.route("/api/reminders/<int:rid>/send-whatsapp", methods=["POST"])
def reminder_send_wa(rid):
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM patient_reminders WHERE id=?", (rid,)).fetchone()
        if not row:
            return json_error("Reminder not found", 404)
        r    = dict(row)
        days = r.get("days_supply", 30)
        msg  = (
            f"🏥 *Selvam Medicals* — Medicine Reminder\n"
            f"Dear {r['customer_name']},\n\n"
            f"⚠️ Your *{r['medicine_name']}* supply will run out in ~2 days "
            f"(expected finish: {r['expected_finish']}).\n\n"
            f"📲 Reply YES to auto-place a refill order, or visit us.\n"
            f"📍 SS & Co — Selvam Medicals, Tamil Nadu.\n"
            f"Thank you! 💊"
        )
        result = send_whatsapp(r["phone"], msg)
        status = "sent" if result["status"] in ("sent","mocked") else "failed"
        conn.execute(
            "UPDATE patient_reminders SET wa_sent=1, status=?, wa_response=? WHERE id=?",
            (status, result.get("sid",""), rid)
        )
    return jsonify({"status":status,"phone":r["phone"],"sid":result.get("sid","")})


@app.route("/api/reminders/send-bulk", methods=["POST"])
def reminders_send_bulk():
    """Send WhatsApp to all pending reminders due today or overdue."""
    today  = datetime.now().strftime("%Y-%m-%d")
    sent   = 0
    failed = 0

    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM patient_reminders WHERE status='pending' AND reminder_date<=? AND phone!=''",
            (today,)
        ).fetchall()

        for row in rows:
            r   = dict(row)
            msg = (
                f"🏥 *Selvam Medicals* — Medicine Reminder\n"
                f"Dear {r['customer_name']}, your *{r['medicine_name']}* "
                f"runs out in ~2 days (by {r['expected_finish']}).\n"
                f"Please reorder. 💊  Reply YES to auto-order."
            )
            result = send_whatsapp(r["phone"], msg)
            ok     = result["status"] in ("sent","mocked")
            conn.execute(
                "UPDATE patient_reminders SET wa_sent=1,status=?,wa_response=? WHERE id=?",
                ("sent" if ok else "failed", result.get("sid",""), r["id"])
            )
            if ok: sent += 1
            else:  failed += 1

    return jsonify({"status":"success","sent":sent,"failed":failed})


@app.route("/api/reminders/<int:rid>/dismiss", methods=["POST"])
def reminder_dismiss(rid):
    with get_conn() as conn:
        conn.execute("UPDATE patient_reminders SET status='dismissed' WHERE id=?", (rid,))
    return jsonify({"status":"success"})


@app.route("/ai/reminders")
def reminders_page():
    return render_template("ai_reminders.html")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# NARCOTIC / DRUG SCHEDULE / INSPECTOR FEATURES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/narcotic-register", methods=["GET"])
@app.route("/api/drug-register", methods=["GET"])
def drug_register_list():
    date_from = request.args.get("from", datetime.now().strftime("%Y-%m-%d"))
    date_to   = request.args.get("to",   datetime.now().strftime("%Y-%m-%d"))
    schedule  = request.args.get("schedule","")
    with get_conn() as conn:
        if schedule:
            rows = conn.execute(
                "SELECT * FROM narcotic_register WHERE date BETWEEN ? AND ? AND schedule_type=? ORDER BY date DESC",
                (date_from, date_to, schedule)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM narcotic_register WHERE date BETWEEN ? AND ? ORDER BY date DESC",
                (date_from, date_to)
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/drug-register", methods=["POST"])
def drug_register_add():
    data = request.get_json(silent=True) or {}
    req  = ["medicine_name","qty_dispensed","date"]
    miss = required_fields(data, req)
    if miss:
        return json_error("Missing fields", 400, miss)
    now = datetime.now(timezone.utc).isoformat()

    with get_conn() as conn:
        # Get opening balance
        prev = conn.execute(
            "SELECT closing_balance FROM narcotic_register "
            "WHERE medicine_name=? ORDER BY id DESC LIMIT 1",
            (data["medicine_name"],)
        ).fetchone()
        opening = prev["closing_balance"] if prev else int(data.get("opening_balance", 0) or 0)
        closing = opening - int(data["qty_dispensed"])

        cur = conn.execute("""
            INSERT INTO narcotic_register
            (medicine_name,generic_name,schedule_type,date,bill_id,customer_name,
             customer_phone,doctor_name,doctor_reg_no,qty_dispensed,batch_no,
             opening_balance,closing_balance,purpose,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (data["medicine_name"], data.get("generic_name",""),
              data.get("schedule_type","X"), data["date"],
              data.get("bill_id",""), data.get("customer_name",""),
              data.get("customer_phone",""), data.get("doctor_name",""),
              data.get("doctor_reg_no",""), int(data["qty_dispensed"]),
              data.get("batch_no",""), opening, closing,
              data.get("purpose","dispensed"), now))
    return jsonify({"status":"success","id":cur.lastrowid,"closing_balance":closing})


@app.route("/api/drug-register/alerts", methods=["GET"])
def drug_register_alerts():
    """Scan bills for Schedule X/H1 medicines dispensed without register entry today."""
    today = datetime.now().strftime("%Y-%m-%d")
    with get_conn() as conn:
        schedule_names = [r["medicine_name"].lower() for r in
            conn.execute("SELECT medicine_name FROM drug_schedules WHERE schedule_type IN ('X','H1')").fetchall()]
        logged_today   = [r["medicine_name"].lower() for r in
            conn.execute("SELECT DISTINCT medicine_name FROM narcotic_register WHERE date=?", (today,)).fetchall()]

        # Find bills today with scheduled medicines not yet logged
        today_bills = conn.execute(
            "SELECT id, date, cust, items FROM bills WHERE date LIKE ?",
            (f"{datetime.now().strftime('%d/%m/%Y')}%",)
        ).fetchall()

    alerts = []
    for bill in today_bills:
        items = safe_json_loads(bill["items"], [])
        for item in items:
            name = str(item.get("n", item.get("name",""))).lower().strip()
            if any(name in sn or sn in name for sn in schedule_names):
                if not any(name in lg or lg in name for lg in logged_today):
                    alerts.append({
                        "bill_id":    bill["id"],
                        "date":       bill["date"],
                        "customer":   bill["cust"],
                        "medicine":   item.get("n", item.get("name","")),
                        "qty":        item.get("qty", 1),
                        "alert_type": "not_registered",
                    })
    return jsonify({"alerts":alerts,"count":len(alerts),"date":today})


@app.route("/api/drug-combinations", methods=["GET"])
def drug_combinations_list():
    """1mg-style alternatives — public API for drug register page."""
    q    = request.args.get("q","").strip()
    sched= request.args.get("schedule","")
    page = int(request.args.get("page",1))
    limit= int(request.args.get("limit",50))
    offset=(page-1)*limit
    with get_conn() as conn:
        params: list = []
        where = "WHERE 1=1"
        if q:
            where += " AND (brand_name LIKE ? OR generic_names LIKE ? OR uses LIKE ?)"; params += [f"%{q}%"]*3
        if sched:
            where += " AND schedule_type=?"; params.append(sched)
        rows = conn.execute(
            f"SELECT * FROM drug_combinations {where} ORDER BY brand_name LIMIT ? OFFSET ?",
            params + [limit, offset]
        ).fetchall()
        total = conn.execute(f"SELECT COUNT(*) AS c FROM drug_combinations {where}", params).fetchone()["c"]
    result = []
    for r in rows:
        d = dict(r)
        d["generic_names"]  = safe_json_loads(d.get("generic_names"), [])
        d["alternatives"]   = safe_json_loads(d.get("alternatives"), [])
        result.append(d)
    return jsonify({"data": result, "total": total, "page": page, "limit": limit})


@app.route("/api/drug-schedules", methods=["GET"])
def drug_schedules_list():
    q        = request.args.get("q","").strip()
    schedule = request.args.get("schedule","")
    with get_conn() as conn:
        if q and schedule:
            rows = conn.execute(
                "SELECT * FROM drug_schedules WHERE schedule_type=? AND (medicine_name LIKE ? OR generic_name LIKE ?)",
                (schedule, f"%{q}%", f"%{q}%")
            ).fetchall()
        elif schedule:
            rows = conn.execute(
                "SELECT * FROM drug_schedules WHERE schedule_type=? ORDER BY medicine_name", (schedule,)
            ).fetchall()
        elif q:
            rows = conn.execute(
                "SELECT * FROM drug_schedules WHERE medicine_name LIKE ? OR generic_name LIKE ? ORDER BY medicine_name",
                (f"%{q}%", f"%{q}%")
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM drug_schedules ORDER BY schedule_type, medicine_name").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/alternatives", methods=["GET"])
def drug_alternatives():
    """Search drug combinations and get 1mg-style alternatives."""
    q = request.args.get("q","").strip()
    if not q:
        with get_conn() as conn:
            rows = conn.execute("SELECT * FROM drug_combinations ORDER BY brand_name LIMIT 100").fetchall()
    else:
        with get_conn() as conn:
            rows = conn.execute("""
                SELECT * FROM drug_combinations
                WHERE brand_name LIKE ? OR generic_names LIKE ? OR uses LIKE ?
                ORDER BY brand_name LIMIT 50
            """, (f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["generic_names"]  = safe_json_loads(d.get("generic_names"),  [])
        d["alternatives"]   = safe_json_loads(d.get("alternatives"),   [])
        result.append(d)
    return jsonify(result)


@app.route("/api/inspector/visits", methods=["GET"])
def inspector_visits_list():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM inspector_visits ORDER BY visit_date DESC LIMIT 50").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/inspector/visits", methods=["POST"])
def inspector_visits_add():
    data = request.get_json(silent=True) or {}
    now  = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO inspector_visits
            (inspector_name,badge_no,visit_date,visit_time,purpose,items_checked,
             observations,compliance_status,next_visit_date,signature_obtained,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (data.get("inspector_name",""), data.get("badge_no",""),
              data.get("visit_date", now[:10]),
              data.get("visit_time", datetime.now().strftime("%H:%M")),
              data.get("purpose","routine"),
              json.dumps(data.get("items_checked",[])),
              data.get("observations",""),
              data.get("compliance_status","pass"),
              data.get("next_visit_date",""),
              int(data.get("signature_obtained",0)), now))
    return jsonify({"status":"success","id":cur.lastrowid})


@app.route("/drug-register")
def drug_register_page():
    return render_template("drug_register.html")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA INSIGHTS — CHARTS & ANALYTICS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/insights/revenue-trend", methods=["GET"])
def insights_revenue_trend():
    days = request.args.get("days", 30, type=int)
    result = []
    with get_conn() as conn:
        for i in range(days-1, -1, -1):
            d    = datetime.now() - timedelta(days=i)
            ds   = f"{d.day:02d}/{d.month:02d}/{d.year}"
            row  = conn.execute(
                "SELECT SUM(total) AS t, COUNT(*) AS c FROM bills WHERE date LIKE ?",
                (f"{ds}%",)
            ).fetchone()
            result.append({
                "date":    d.strftime("%Y-%m-%d"),
                "label":   d.strftime("%d %b"),
                "revenue": round(float(row["t"] or 0), 2),
                "bills":   int(row["c"] or 0),
            })
    return jsonify(result)


@app.route("/api/insights/top-medicines", methods=["GET"])
def insights_top_medicines():
    limit = request.args.get("limit", 15, type=int)
    with get_conn() as conn:
        bills = conn.execute("SELECT items FROM bills WHERE items IS NOT NULL").fetchall()

    counter: dict[str, dict] = {}
    for bill in bills:
        for item in safe_json_loads(bill["items"], []):
            name = str(item.get("n", item.get("name",""))).strip()
            if not name:
                continue
            qty    = int(item.get("qty", item.get("s", 1)) or 1)
            amount = float(item.get("p", 0) or 0) * qty
            if name not in counter:
                counter[name] = {"name": name, "qty": 0, "revenue": 0.0, "count": 0}
            counter[name]["qty"]     += qty
            counter[name]["revenue"] += amount
            counter[name]["count"]   += 1

    top = sorted(counter.values(), key=lambda x: x["revenue"], reverse=True)[:limit]
    return jsonify(top)


@app.route("/api/insights/payment-breakdown", methods=["GET"])
def insights_payment_breakdown():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT pay, COUNT(*) AS c, SUM(total) AS t FROM bills GROUP BY pay"
        ).fetchall()
    return jsonify([{"pay": r["pay"] or "cash", "count": r["c"], "revenue": round(float(r["t"] or 0),2)}
                    for r in rows])


@app.route("/api/insights/customer-analytics", methods=["GET"])
def insights_customer_analytics():
    with get_conn() as conn:
        total_cust  = conn.execute("SELECT COUNT(*) AS c FROM customers").fetchone()["c"]
        returning   = conn.execute("SELECT COUNT(*) AS c FROM customers WHERE visits > 1").fetchone()["c"]
        top_spend   = conn.execute(
            "SELECT name, phone, visits, total FROM customers ORDER BY total DESC LIMIT 10"
        ).fetchall()
        chronic     = conn.execute(
            "SELECT COUNT(DISTINCT customer_name) AS c FROM patient_reminders"
        ).fetchone()["c"]
        new_30d_ct  = conn.execute(
            "SELECT COUNT(DISTINCT cust) AS c FROM bills "
            "WHERE ts >= ? AND cust != ''",
            (int((datetime.now()-timedelta(days=30)).timestamp()*1000),)
        ).fetchone()["c"]

    return jsonify({
        "total_customers":   total_cust,
        "returning":         returning,
        "new_last_30_days":  new_30d_ct,
        "retention_rate":    round(returning/max(total_cust,1)*100,1),
        "chronic_patients":  chronic,
        "top_spenders":      [dict(r) for r in top_spend],
    })


@app.route("/api/insights/hourly-heatmap", methods=["GET"])
def insights_hourly():
    """Bills count by hour of day — last 30 days."""
    with get_conn() as conn:
        bills = conn.execute(
            "SELECT date FROM bills WHERE ts >= ?",
            (int((datetime.now()-timedelta(days=30)).timestamp()*1000),)
        ).fetchall()
    heatmap = [0]*24
    for b in bills:
        try:
            parts = str(b["date"]).split()
            if len(parts) >= 2:
                hour = int(parts[1].split(":")[0])
                heatmap[hour] += 1
        except Exception:
            pass
    return jsonify([{"hour": h, "label": f"{h:02d}:00", "count": heatmap[h]} for h in range(24)])


@app.route("/api/insights/stock-health", methods=["GET"])
def insights_stock_health():
    today  = datetime.now().date()
    cutoff = (today + timedelta(days=90)).strftime("%Y-%m-%d")
    with get_conn() as conn:
        total_meds  = conn.execute("SELECT COUNT(*) AS c FROM medicines").fetchone()["c"]
        zero_stock  = conn.execute("SELECT COUNT(*) AS c FROM medicines WHERE s<=0").fetchone()["c"]
        low_stock   = conn.execute("SELECT COUNT(*) AS c FROM medicines WHERE s>0 AND s<15").fetchone()["c"]
        expiring    = conn.execute(
            "SELECT COUNT(*) AS c FROM medicines WHERE expiry!='' AND expiry<=? AND expiry>=?",
            (cutoff, today.strftime("%Y-%m-%d"))
        ).fetchone()["c"]
        expired     = conn.execute(
            "SELECT COUNT(*) AS c FROM medicines WHERE expiry!='' AND expiry<?",
            (today.strftime("%Y-%m-%d"),)
        ).fetchone()["c"]
        total_val   = conn.execute(
            "SELECT SUM(p*s) AS v FROM medicines WHERE s>0"
        ).fetchone()["v"] or 0

    return jsonify({
        "total_medicines": total_meds,
        "zero_stock":      zero_stock,
        "low_stock":       low_stock,
        "expiring_soon":   expiring,
        "expired":         expired,
        "healthy":         total_meds - zero_stock - low_stock - expiring,
        "total_stock_value": round(float(total_val),2),
    })


@app.route("/insights")
def insights_page():
    return render_template("insights_dashboard.html")


@app.route("/tally")
def tally_page():
    return render_template("tally.html")


@app.route("/delivery")
def delivery_page():
    return render_template("delivery.html")


@app.route("/gst")
def gst_page():
    return render_template("gst.html")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GST RETURNS MODULE
# Pharma HSN: 3004 (most medicines 5% / 12%)
# Bill fields: sub=subtotal, disc=discount, tax=gst_amount, total=final
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _parse_bill_date(date_str: str):
    """Parse 'DD/MM/YYYY HH:MM' or 'YYYY-MM-DD' → (year, month, day) tuple."""
    if not date_str:
        return None, None, None
    try:
        if "/" in date_str:
            parts = date_str.split(" ")[0].split("/")
            return int(parts[2]), int(parts[1]), int(parts[0])
        else:
            parts = date_str.split("-")
            return int(parts[0]), int(parts[1]), int(parts[2])
    except Exception:
        return None, None, None


@app.route("/api/gst/summary", methods=["GET"])
def gst_summary():
    """Monthly GST summary — KPIs, B2C/B2B split, payment-wise."""
    month = request.args.get("month", "")          # "2026-05"
    year  = request.args.get("year",  "")          # "2026"

    with get_conn() as conn:
        bills = [dict(r) for r in conn.execute("SELECT * FROM bills ORDER BY ts DESC").fetchall()]

    # ── Filter by month / year ────────────────────────────────────────
    filtered = []
    for b in bills:
        yr, mo, _ = _parse_bill_date(b.get("date",""))
        if yr is None:
            continue
        if month and f"{yr}-{mo:02d}" != month:
            continue
        if year and not month and str(yr) != year:
            continue
        filtered.append(b)

    # ── Cancelled bills excluded ──────────────────────────────────────
    filtered = [b for b in filtered if not (b.get("cust","") or "").startswith("[CANCELLED]")]

    total_bills    = len(filtered)
    gross_sales    = sum(float(b.get("sub",0) or 0) for b in filtered)
    total_discount = sum(float(b.get("disc",0) or 0) for b in filtered)
    taxable_value  = gross_sales - total_discount
    total_gst      = sum(float(b.get("tax",0) or 0) for b in filtered)
    cgst           = round(total_gst / 2, 2)
    sgst           = round(total_gst / 2, 2)
    grand_total    = sum(float(b.get("total",0) or 0) for b in filtered)

    # ── B2C vs B2B split ─────────────────────────────────────────────
    b2c_bills = [b for b in filtered if (b.get("bill_type","retail") or "retail") == "retail"]
    b2b_bills = [b for b in filtered if (b.get("bill_type","retail") or "retail") == "wholesale"]
    b2c_taxable = sum(float(b.get("sub",0) or 0) - float(b.get("disc",0) or 0) for b in b2c_bills)
    b2b_taxable = sum(float(b.get("sub",0) or 0) - float(b.get("disc",0) or 0) for b in b2b_bills)
    b2c_gst     = sum(float(b.get("tax",0) or 0) for b in b2c_bills)
    b2b_gst     = sum(float(b.get("tax",0) or 0) for b in b2b_bills)

    # ── Payment-mode breakdown ────────────────────────────────────────
    pay_modes: dict = {}
    for b in filtered:
        pm = (b.get("pay","") or "cash").lower()
        pay_modes[pm] = pay_modes.get(pm, 0) + float(b.get("total",0) or 0)

    # ── Month-wise trend (last 6 months) ─────────────────────────────
    month_trend: dict = {}
    for b in bills:
        if (b.get("cust","") or "").startswith("[CANCELLED]"):
            continue
        yr, mo, _ = _parse_bill_date(b.get("date",""))
        if yr is None:
            continue
        key = f"{yr}-{mo:02d}"
        if key not in month_trend:
            month_trend[key] = {"month": key, "sales": 0, "gst": 0, "bills": 0}
        month_trend[key]["sales"] += float(b.get("sub",0) or 0)
        month_trend[key]["gst"]   += float(b.get("tax",0) or 0)
        month_trend[key]["bills"] += 1
    trend = sorted(month_trend.values(), key=lambda x: x["month"])[-6:]
    for t in trend:
        t["sales"] = round(t["sales"], 2)
        t["gst"]   = round(t["gst"], 2)

    return jsonify({
        "month":           month or "all",
        "total_bills":     total_bills,
        "gross_sales":     round(gross_sales, 2),
        "total_discount":  round(total_discount, 2),
        "taxable_value":   round(taxable_value, 2),
        "total_gst":       round(total_gst, 2),
        "cgst":            cgst,
        "sgst":            sgst,
        "grand_total":     round(grand_total, 2),
        "b2c": {"bills": len(b2c_bills), "taxable": round(b2c_taxable,2), "gst": round(b2c_gst,2)},
        "b2b": {"bills": len(b2b_bills), "taxable": round(b2b_taxable,2), "gst": round(b2b_gst,2)},
        "payment_modes":   {k: round(v,2) for k,v in pay_modes.items()},
        "monthly_trend":   trend,
    })


@app.route("/api/gst/gstr1", methods=["GET"])
def gst_gstr1():
    """GSTR-1 ready data — B2C large, B2C small, HSN summary."""
    month = request.args.get("month", "")

    with get_conn() as conn:
        bills = [dict(r) for r in conn.execute("SELECT * FROM bills ORDER BY ts DESC").fetchall()]

    filtered = []
    for b in bills:
        yr, mo, _ = _parse_bill_date(b.get("date",""))
        if yr is None:
            continue
        if month and f"{yr}-{mo:02d}" != month:
            continue
        if (b.get("cust","") or "").startswith("[CANCELLED]"):
            continue
        filtered.append(b)

    # ── B2CS (B2C Small — taxable value < 2.5L per rate) ─────────────
    b2cs_5pct  = {"rate": 5,  "taxable": 0.0, "cgst": 0.0, "sgst": 0.0}
    b2cs_12pct = {"rate": 12, "taxable": 0.0, "cgst": 0.0, "sgst": 0.0}
    b2cl_rows  = []   # B2CL — individual bills > ₹2.5L

    for b in filtered:
        if (b.get("bill_type","retail") or "retail") != "retail":
            continue
        taxable = float(b.get("sub",0) or 0) - float(b.get("disc",0) or 0)
        gst_amt = float(b.get("tax",0) or 0)
        # Determine effective rate
        eff_rate = round((gst_amt / taxable * 100) if taxable > 0 else 0)
        cgst = round(gst_amt / 2, 2)
        sgst = round(gst_amt / 2, 2)
        if float(b.get("total",0) or 0) >= 250000:
            b2cl_rows.append({
                "bill_no": b.get("id",""),
                "date":    b.get("date",""),
                "customer": b.get("cust",""),
                "phone":   b.get("phone",""),
                "taxable": round(taxable,2),
                "rate":    eff_rate,
                "cgst":    cgst,
                "sgst":    sgst,
                "total":   float(b.get("total",0) or 0),
            })
        elif eff_rate >= 10:
            b2cs_12pct["taxable"] += taxable
            b2cs_12pct["cgst"]    += cgst
            b2cs_12pct["sgst"]    += sgst
        else:
            b2cs_5pct["taxable"]  += taxable
            b2cs_5pct["cgst"]     += cgst
            b2cs_5pct["sgst"]     += sgst

    # ── HSN Summary ───────────────────────────────────────────────────
    # All pharma medicines → HSN 3004
    hsn_5pct  = {"hsn": "3004", "desc": "Medicaments (5% GST)", "uom": "NOS",
                  "qty": 0, "taxable": 0.0, "rate": 5, "cgst": 0.0, "sgst": 0.0}
    hsn_12pct = {"hsn": "3004", "desc": "Medicaments (12% GST)", "uom": "NOS",
                  "qty": 0, "taxable": 0.0, "rate": 12, "cgst": 0.0, "sgst": 0.0}
    hsn_0pct  = {"hsn": "3006", "desc": "Pharmaceutical goods (0% GST)", "uom": "NOS",
                  "qty": 0, "taxable": 0.0, "rate": 0, "cgst": 0.0, "sgst": 0.0}

    for b in filtered:
        items = safe_json_loads(b.get("items",""), [])
        gst_amt = float(b.get("tax",0) or 0)
        taxable = float(b.get("sub",0) or 0) - float(b.get("disc",0) or 0)
        eff_rate = round((gst_amt / taxable * 100) if taxable > 0 else 0)
        total_qty = sum(int(i.get("qty",1)) for i in items if isinstance(i,dict))
        cgst = round(gst_amt / 2, 2)
        sgst = round(gst_amt / 2, 2)
        if gst_amt == 0:
            hsn_0pct["qty"]     += total_qty
            hsn_0pct["taxable"] += taxable
        elif eff_rate >= 10:
            hsn_12pct["qty"]     += total_qty
            hsn_12pct["taxable"] += taxable
            hsn_12pct["cgst"]    += cgst
            hsn_12pct["sgst"]    += sgst
        else:
            hsn_5pct["qty"]      += total_qty
            hsn_5pct["taxable"]  += taxable
            hsn_5pct["cgst"]     += cgst
            hsn_5pct["sgst"]     += sgst

    for h in [hsn_5pct, hsn_12pct, hsn_0pct]:
        h["taxable"] = round(h["taxable"], 2)
        h["cgst"]    = round(h["cgst"], 2)
        h["sgst"]    = round(h["sgst"], 2)

    b2cs_5pct["taxable"]  = round(b2cs_5pct["taxable"], 2)
    b2cs_12pct["taxable"] = round(b2cs_12pct["taxable"], 2)

    return jsonify({
        "month":        month or "all",
        "b2cs":         [b2cs_5pct, b2cs_12pct],
        "b2cl":         b2cl_rows,
        "hsn_summary":  [hsn_5pct, hsn_12pct, hsn_0pct],
        "total_bills":  len(filtered),
    })


@app.route("/api/gst/bills", methods=["GET"])
def gst_bills_list():
    """Paginated bill-wise GST ledger for a given month."""
    month  = request.args.get("month","")
    page   = int(request.args.get("page",1))
    limit  = int(request.args.get("limit",50))

    with get_conn() as conn:
        all_bills = [dict(r) for r in conn.execute("SELECT * FROM bills ORDER BY ts DESC").fetchall()]

    filtered = []
    for b in all_bills:
        yr, mo, _ = _parse_bill_date(b.get("date",""))
        if yr is None:
            continue
        if month and f"{yr}-{mo:02d}" != month:
            continue
        if (b.get("cust","") or "").startswith("[CANCELLED]"):
            continue
        taxable = float(b.get("sub",0) or 0) - float(b.get("disc",0) or 0)
        gst_amt = float(b.get("tax",0) or 0)
        eff_rate = round((gst_amt / taxable * 100) if taxable > 0 else 0)
        filtered.append({
            "bill_no":    b.get("id",""),
            "date":       b.get("date",""),
            "customer":   b.get("cust",""),
            "phone":      b.get("phone",""),
            "bill_type":  b.get("bill_type","retail"),
            "payment":    b.get("pay","cash"),
            "gross":      float(b.get("sub",0) or 0),
            "discount":   float(b.get("disc",0) or 0),
            "taxable":    round(taxable,2),
            "gst_rate":   eff_rate,
            "gst_amount": round(gst_amt,2),
            "cgst":       round(gst_amt/2,2),
            "sgst":       round(gst_amt/2,2),
            "total":      float(b.get("total",0) or 0),
        })

    total = len(filtered)
    start = (page-1)*limit
    return jsonify({
        "data":  filtered[start:start+limit],
        "total": total,
        "page":  page,
        "limit": limit,
    })


@app.route("/api/gst/months", methods=["GET"])
def gst_available_months():
    """Returns list of months that have bills — for the month picker."""
    with get_conn() as conn:
        bills = conn.execute("SELECT date FROM bills").fetchall()
    months = set()
    for (d,) in bills:
        yr, mo, _ = _parse_bill_date(d or "")
        if yr:
            months.add(f"{yr}-{mo:02d}")
    return jsonify(sorted(months, reverse=True))


@app.route("/api/gst/settings", methods=["GET"])
def gst_get_settings():
    """Return saved GST settings (GSTIN, firm name, etc.)."""
    with get_conn() as conn:
        rows = conn.execute("SELECT key, value FROM app_settings WHERE key LIKE 'gst_%'").fetchall()
    return jsonify({r["key"]: r["value"] for r in rows})


@app.route("/api/gst/settings", methods=["POST"])
def gst_save_settings():
    """Save GST settings key-value pairs."""
    data = request.get_json(silent=True) or {}
    allowed = {"gst_gstin", "gst_firm_name", "gst_address", "gst_state_code", "gst_pan"}
    with get_conn() as conn:
        for key, value in data.items():
            if key in allowed:
                conn.execute(
                    "INSERT INTO app_settings (key, value) VALUES (?,?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (key, str(value))
                )
    return jsonify({"status": "saved"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# OPEN DELIVERY APIs (no portal gate — staff use delivery page)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/open/delivery/routes", methods=["GET"])
def open_delivery_list():
    date_filter = request.args.get("date", "")
    q = "SELECT * FROM delivery_routes WHERE 1=1"
    params: list = []
    if date_filter:
        q += " AND delivery_date=?"; params.append(date_filter)
    q += " ORDER BY id DESC LIMIT 200"
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    for r in rows:
        r["shops"] = safe_json_loads(r.get("shops"), [])
    return jsonify(rows)


@app.route("/api/open/delivery/routes", methods=["POST"])
def open_delivery_create():
    data  = request.get_json(silent=True) or {}
    shops = data.get("shops", [])
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO delivery_routes (route_name, delivery_boy, delivery_phone,
            delivery_date, shops, status, created_at, notes)
            VALUES (?,?,?,?,?,'pending',?,?)
        """, (data.get("route_name","Route-1"), data.get("delivery_boy",""),
              data.get("delivery_phone",""),
              data.get("delivery_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
              json.dumps(shops), datetime.now(timezone.utc).isoformat(),
              data.get("notes","")))
        route_id = cur.lastrowid
    return jsonify({"status":"success","route_id":route_id})


@app.route("/api/open/delivery/routes/<int:route_id>", methods=["PUT"])
def open_delivery_update(route_id):
    data   = request.get_json(silent=True) or {}
    status = data.get("status","pending")
    shops  = data.get("shops")
    notes  = data.get("notes")
    now    = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        if status == "started":
            conn.execute("UPDATE delivery_routes SET status=?,started_at=? WHERE id=?",
                         (status, now, route_id))
        elif status == "completed":
            conn.execute("UPDATE delivery_routes SET status=?,completed_at=? WHERE id=?",
                         (status, now, route_id))
        elif shops is not None:
            conn.execute("UPDATE delivery_routes SET shops=? WHERE id=?",
                         (json.dumps(shops), route_id))
        elif notes is not None:
            conn.execute("UPDATE delivery_routes SET notes=? WHERE id=?",
                         (notes, route_id))
        else:
            conn.execute("UPDATE delivery_routes SET status=? WHERE id=?",
                         (status, route_id))
    return jsonify({"status":"success"})


@app.route("/api/open/delivery/routes/<int:route_id>", methods=["DELETE"])
def open_delivery_delete(route_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM delivery_routes WHERE id=?", (route_id,))
    return jsonify({"status":"success"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# OPEN TALLY APIs — uses actual DB schema (tally_entries, cheque_register, bank_statements)
# Real columns: date, party_name, party_type, amount, total_amount, description, entry_type
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/open/tally/summary", methods=["GET"])
def open_tally_summary():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM tally_entries ORDER BY date DESC LIMIT 500").fetchall()
    entries = [dict(r) for r in rows]
    # Income = receipt/sale entries, Expense = payment/purchase entries
    income_types  = {"receipt","sales","sale"}
    expense_types = {"payment","purchase","expense"}
    total_income  = sum(float(e.get("total_amount") or e.get("amount") or 0)
                        for e in entries if (e.get("entry_type","") or "").lower() in income_types)
    total_expense = sum(float(e.get("total_amount") or e.get("amount") or 0)
                        for e in entries if (e.get("entry_type","") or "").lower() in expense_types)
    return jsonify({
        "total_income":  round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "net_balance":   round(total_income - total_expense, 2),
        "entry_count":   len(entries)
    })


@app.route("/api/open/tally/entries", methods=["GET"])
def open_tally_entries():
    date_from = request.args.get("from", "")
    date_to   = request.args.get("to", "")
    account   = request.args.get("account", "")
    tx_type   = request.args.get("type", "")
    q = "SELECT * FROM tally_entries WHERE 1=1"
    params: list = []
    if date_from:
        q += " AND date>=?"; params.append(date_from)
    if date_to:
        q += " AND date<=?"; params.append(date_to)
    if account:
        q += " AND party_name LIKE ?"; params.append(f"%{account}%")
    if tx_type:
        q += " AND entry_type=?"; params.append(tx_type)
    q += " ORDER BY date DESC, id DESC LIMIT 300"
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    # Add normalized fields for the UI
    for r in rows:
        r["entry_date"]     = r.get("date","")
        r["debit_account"]  = r.get("party_name","") if r.get("party_type","")=="vendor" else ""
        r["credit_account"] = r.get("party_name","") if r.get("party_type","")=="customer" else r.get("party_name","")
        r["debit"]          = float(r.get("amount",0)) if (r.get("entry_type","") or "").lower() in {"payment","purchase","expense"} else 0
        r["credit"]         = float(r.get("amount",0)) if (r.get("entry_type","") or "").lower() in {"receipt","sales","sale"} else 0
        r["narration"]      = r.get("description","")
    return jsonify(rows)


@app.route("/api/open/tally/entries", methods=["POST"])
def open_tally_add_entry():
    data = request.get_json(silent=True) or {}
    required = ["entry_date","amount","narration"]
    if not all(data.get(k) for k in required):
        return json_error("Missing required fields")
    entry_type = data.get("entry_type","journal")
    party_name = data.get("debit_account","") or data.get("credit_account","") or data.get("party_name","")
    party_type = "vendor" if entry_type in ("payment","purchase","expense") else "customer"
    amount     = float(data["amount"])
    gst        = round(amount * 0.05, 2) if entry_type in ("sales","purchase") else 0
    total      = round(amount + gst, 2)
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO tally_entries
            (company_id, entry_type, party_name, party_type, amount, gst_amount,
             total_amount, description, reference_no, invoice_no, date, payment_mode,
             bank_name, status, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (data.get("company_id",1),
              entry_type, party_name, party_type,
              amount, gst, total,
              data["narration"],
              data.get("reference_no",""),
              data.get("invoice_no",""),
              data["entry_date"],
              data.get("payment_mode","cash"),
              data.get("bank_name",""),
              "pending",
              datetime.now(timezone.utc).isoformat()))
        entry_id = cur.lastrowid
    return jsonify({"status":"success","entry_id":entry_id})


@app.route("/api/open/tally/entries/<int:entry_id>", methods=["DELETE"])
def open_tally_delete_entry(entry_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM tally_entries WHERE id=?", (entry_id,))
    return jsonify({"status":"success"})


@app.route("/api/open/tally/cheques", methods=["GET"])
def open_tally_cheques():
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM cheque_register ORDER BY date DESC LIMIT 200").fetchall()]
    # Normalize for UI
    for r in rows:
        r["cheque_date"] = r.get("date","")
        r["narration"]   = r.get("memo","")
        r["cheque_type"] = "issued"   # default; actual table has no cheque_type column
    return jsonify(rows)


@app.route("/api/open/tally/cheques", methods=["POST"])
def open_tally_add_cheque():
    data = request.get_json(silent=True) or {}
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO cheque_register
            (company_id, cheque_no, date, bank_name, account_no, payee,
             amount, memo, status, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (data.get("company_id",1),
              data.get("cheque_no",""),
              data.get("cheque_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
              data.get("bank_name",""),
              data.get("account_no",""),
              data.get("payee",""),
              float(data.get("amount",0)),
              data.get("narration",""),
              data.get("status","pending"),
              datetime.now(timezone.utc).isoformat()))
        cheque_id = cur.lastrowid
    return jsonify({"status":"success","cheque_id":cheque_id})


@app.route("/api/open/tally/cheques/<int:cheque_id>", methods=["PUT"])
def open_tally_update_cheque(cheque_id):
    data   = request.get_json(silent=True) or {}
    status = data.get("status","pending")
    with get_conn() as conn:
        conn.execute("UPDATE cheque_register SET status=? WHERE id=?", (status, cheque_id))
    return jsonify({"status":"success"})


@app.route("/api/open/tally/bank-statements", methods=["GET"])
def open_tally_bank_stmts():
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM bank_statements ORDER BY id DESC LIMIT 200").fetchall()]
    # Normalize for UI
    for r in rows:
        r["statement_date"] = r.get("statement_month","")
        r["description"]    = r.get("file_name","")
        r["debit"]          = float(r.get("total_debits",0))
        r["credit"]         = float(r.get("total_credits",0))
        r["balance"]        = float(r.get("closing_balance",0))
    return jsonify(rows)


@app.route("/api/open/tally/bank-statements", methods=["POST"])
def open_tally_add_bank_stmt():
    data = request.get_json(silent=True) or {}
    debit  = float(data.get("debit",0))
    credit = float(data.get("credit",0))
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO bank_statements
            (company_id, bank_name, account_no, statement_month,
             opening_balance, closing_balance, total_credits, total_debits,
             file_name, uploaded_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (data.get("company_id",1),
              data.get("bank_name",""),
              data.get("account_no",""),
              data.get("statement_date", datetime.now(timezone.utc).strftime("%Y-%m")),
              float(data.get("opening_balance",0)),
              float(data.get("balance",0)),
              credit, debit,
              data.get("description","Manual Entry"),
              datetime.now(timezone.utc).isoformat()))
    return jsonify({"status":"success"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# RAZORPAY PAYMENT GATEWAY (Sprint 1)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route("/api/payment/create-order", methods=["POST"])
def payment_create_order():
    """Create Razorpay order — falls back to mock if keys absent."""
    data       = request.get_json(silent=True) or {}
    amount_rs  = float(data.get("amount", 0))
    purpose    = data.get("purpose", "payment")
    customer   = data.get("customer_name", "")
    phone      = data.get("phone", "")
    bill_id    = data.get("bill_id", "")

    rz_key_id  = os.environ.get("RAZORPAY_KEY_ID", "")
    rz_key_sec = os.environ.get("RAZORPAY_KEY_SECRET", "")

    order_id = f"order_mock_{int(datetime.now().timestamp())}"
    mock_mode = not (rz_key_id and rz_key_sec)

    if not mock_mode:
        try:
            import base64 as _b64
            import urllib.request as _req, urllib.parse as _up
            payload_bytes = json.dumps({
                "amount":   int(amount_rs * 100),
                "currency": "INR",
                "receipt":  f"rcpt_{bill_id or int(datetime.now().timestamp())}",
                "notes":    {"customer": customer, "purpose": purpose}
            }).encode()
            creds = _b64.b64encode(f"{rz_key_id}:{rz_key_sec}".encode()).decode()
            rz_req = _req.Request(
                "https://api.razorpay.com/v1/orders",
                data=payload_bytes,
                headers={"Authorization": f"Basic {creds}",
                         "Content-Type": "application/json"})
            with _req.urlopen(rz_req, timeout=10) as resp:
                rz_data = json.loads(resp.read())
            order_id = rz_data.get("id", order_id)
        except Exception as e:
            mock_mode = True

    # Log payment intent
    with get_conn() as conn:
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS payment_orders (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id TEXT, bill_id TEXT, customer_name TEXT,
                    phone TEXT, amount REAL, purpose TEXT,
                    status TEXT DEFAULT 'created', mock INTEGER DEFAULT 0,
                    created_at TEXT, paid_at TEXT
                )""")
            conn.execute("""
                INSERT INTO payment_orders
                (order_id,bill_id,customer_name,phone,amount,purpose,mock,created_at)
                VALUES (?,?,?,?,?,?,?,?)
            """, (order_id, str(bill_id), customer, phone,
                  amount_rs, purpose, int(mock_mode),
                  datetime.now(timezone.utc).isoformat()))
        except Exception:
            pass

    return jsonify({
        "status":    "success",
        "order_id":  order_id,
        "amount":    amount_rs,
        "currency":  "INR",
        "key_id":    rz_key_id if not mock_mode else "rzp_test_mock",
        "mock":      mock_mode,
        "customer":  customer,
        "phone":     phone
    })


@app.route("/api/payment/verify", methods=["POST"])
def payment_verify():
    """Verify Razorpay payment signature and mark bill as paid."""
    data       = request.get_json(silent=True) or {}
    order_id   = data.get("razorpay_order_id","")
    payment_id = data.get("razorpay_payment_id","")
    signature  = data.get("razorpay_signature","")
    bill_id    = data.get("bill_id","")
    mock       = data.get("mock", False)

    verified = False
    if mock:
        verified = True
    else:
        rz_key_sec = os.environ.get("RAZORPAY_KEY_SECRET","")
        if rz_key_sec:
            import hmac as _hmac, hashlib as _hs
            expected = _hmac.new(
                rz_key_sec.encode(), f"{order_id}|{payment_id}".encode(),
                _hs.sha256).hexdigest()
            verified = (expected == signature)

    if verified:
        now = datetime.now(timezone.utc).isoformat()
        with get_conn() as conn:
            try:
                conn.execute("UPDATE payment_orders SET status='paid', paid_at=? WHERE order_id=?",
                             (now, order_id))
            except Exception:
                pass
            if bill_id:
                try:
                    conn.execute("UPDATE bills SET payment_mode='razorpay' WHERE id=?",
                                 (int(bill_id),))
                except Exception:
                    pass
        return jsonify({"status":"success","verified":True,"payment_id":payment_id})
    return json_error("Payment verification failed", 400)


@app.route("/api/payment/orders", methods=["GET"])
def payment_orders_list():
    with get_conn() as conn:
        try:
            rows = [dict(r) for r in conn.execute(
                "SELECT * FROM payment_orders ORDER BY id DESC LIMIT 200").fetchall()]
        except Exception:
            rows = []
    return jsonify(rows)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STAGE 6 — COMPLIANCE MODULE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def init_compliance_db():
    """Create Stage 6 compliance tables if they don't exist."""
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS expiry_returns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                return_no TEXT UNIQUE NOT NULL,
                return_date TEXT NOT NULL,
                supplier_id INTEGER,
                supplier_name TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'DRAFT',
                total_qty INTEGER NOT NULL DEFAULT 0,
                total_value REAL NOT NULL DEFAULT 0.0,
                credit_note_no TEXT,
                credit_note_date TEXT,
                credit_note_amount REAL,
                remarks TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS expiry_return_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                expiry_return_id INTEGER NOT NULL,
                item_name TEXT NOT NULL DEFAULT '',
                batch_no TEXT NOT NULL DEFAULT '',
                expiry_date TEXT NOT NULL DEFAULT '',
                qty_returned INTEGER NOT NULL DEFAULT 0,
                purchase_rate REAL NOT NULL DEFAULT 0.0,
                mrp REAL NOT NULL DEFAULT 0.0,
                return_value REAL NOT NULL DEFAULT 0.0,
                FOREIGN KEY (expiry_return_id) REFERENCES expiry_returns(id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS license_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                license_type TEXT NOT NULL,
                license_name TEXT NOT NULL,
                license_no TEXT NOT NULL,
                issuing_authority TEXT,
                issued_date TEXT,
                expiry_date TEXT NOT NULL,
                doc_url TEXT,
                notes TEXT,
                renewal_cost REAL,
                renewal_contact TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        conn.commit()


# ── Expiry Returns Page ────────────────────────────────────────
@app.route("/expiry-returns")
def expiry_returns_page():
    return render_template("expiry_returns.html")


@app.route("/api/expiry-returns/near-expiry-stock", methods=["GET"])
def near_expiry_for_return():
    days = int(request.args.get("days", 90))
    from datetime import date
    today = date.today()
    cutoff = (today + timedelta(days=days)).isoformat()
    today_str = today.isoformat()
    with get_conn() as conn:
        try:
            rows = conn.execute("""
                SELECT sb.id as stock_batch_id, sb.item_id, sb.batch_no,
                       sb.expiry_date, sb.current_qty, sb.purchase_rate, sb.mrp,
                       COALESCE(i.item_name, sb.item_id) as item_name
                FROM stock_batches sb
                LEFT JOIN items i ON i.item_id = sb.item_id
                WHERE sb.current_qty > 0
                  AND sb.expiry_date <= ? AND sb.expiry_date >= ?
                ORDER BY sb.expiry_date ASC
                LIMIT 200
            """, (cutoff, today_str)).fetchall()
        except Exception:
            rows = []
    result = []
    for r in rows:
        r = dict(r)
        try:
            from datetime import date as d
            exp = d.fromisoformat(r["expiry_date"])
            days_left = (exp - today).days
        except Exception:
            days_left = 0
        result.append({
            "stock_batch_id": r.get("stock_batch_id"),
            "item_id": r.get("item_id", ""),
            "item_name": r.get("item_name", ""),
            "batch_no": r.get("batch_no", ""),
            "expiry_date": r.get("expiry_date", "")[:7] if r.get("expiry_date") else "",
            "expiry_date_iso": r.get("expiry_date", ""),
            "days_left": days_left,
            "current_qty": r.get("current_qty", 0),
            "mrp": float(r.get("mrp", 0)),
            "purchase_rate": float(r.get("purchase_rate", 0)),
            "risk_value": round(float(r.get("purchase_rate", 0)) * int(r.get("current_qty", 0)), 2),
        })
    return jsonify(result)


@app.route("/api/expiry-returns/suppliers", methods=["GET"])
def expiry_return_suppliers():
    with get_conn() as conn:
        try:
            rows = conn.execute(
                "SELECT id as supplier_id, supplier_name, phone FROM suppliers WHERE is_active=1 ORDER BY supplier_name LIMIT 200"
            ).fetchall()
        except Exception:
            try:
                rows = conn.execute(
                    "SELECT DISTINCT supplier_name as supplier_name FROM purchase_invoices LIMIT 50"
                ).fetchall()
            except Exception:
                rows = []
    return jsonify([dict(r) for r in rows])


@app.route("/api/expiry-returns", methods=["GET"])
def list_expiry_returns():
    status = request.args.get("status", "").strip().upper()
    with get_conn() as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM expiry_returns WHERE status=? ORDER BY return_date DESC, id DESC",
                (status,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM expiry_returns ORDER BY return_date DESC, id DESC"
            ).fetchall()
    result = []
    for r in [dict(r) for r in rows]:
        result.append({
            "expiry_return_id": r["id"],
            "return_no": r["return_no"],
            "return_date": r["return_date"],
            "supplier_id": r.get("supplier_id"),
            "supplier_name": r.get("supplier_name", ""),
            "status": r["status"],
            "total_qty": r["total_qty"],
            "total_value": float(r["total_value"]),
            "credit_note_no": r.get("credit_note_no") or "",
            "credit_note_amount": float(r.get("credit_note_amount") or 0),
            "remarks": r.get("remarks") or "",
            "created_at": (r.get("created_at") or "")[:10],
        })
    return jsonify(result)


@app.route("/api/expiry-returns", methods=["POST"])
def create_expiry_return():
    data = request.get_json(silent=True) or {}
    items_data = data.get("items", [])
    supplier_id = data.get("supplier_id")
    supplier_name = data.get("supplier_name", "")
    if not items_data:
        return jsonify({"status": "error", "message": "No items provided"}), 400
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM expiry_returns").fetchone()[0] + 1
        from datetime import date
        return_no = f"EXR-{date.today().strftime('%Y%m')}-{count:04d}"
        return_date = data.get("return_date", date.today().isoformat())
        # Resolve supplier name if id given
        if supplier_id and not supplier_name:
            try:
                row = conn.execute("SELECT supplier_name FROM suppliers WHERE id=?", (supplier_id,)).fetchone()
                if row:
                    supplier_name = row[0]
            except Exception:
                pass
        total_qty = sum(int(it.get("qty_returned", 0)) for it in items_data)
        total_value = sum(
            float(it.get("qty_returned", 0)) * float(it.get("purchase_rate", 0))
            for it in items_data
        )
        conn.execute("""
            INSERT INTO expiry_returns (return_no, return_date, supplier_id, supplier_name,
                status, total_qty, total_value, remarks)
            VALUES (?,?,?,?,'DRAFT',?,?,?)
        """, (return_no, return_date, supplier_id, supplier_name,
              total_qty, round(total_value, 2), data.get("remarks", "")))
        er_id = conn.execute("SELECT id FROM expiry_returns WHERE return_no=?", (return_no,)).fetchone()[0]
        for it in items_data:
            qty = int(it.get("qty_returned", 0))
            rate = float(it.get("purchase_rate", 0))
            conn.execute("""
                INSERT INTO expiry_return_items
                    (expiry_return_id, item_name, batch_no, expiry_date, qty_returned, purchase_rate, mrp, return_value)
                VALUES (?,?,?,?,?,?,?,?)
            """, (er_id, it.get("item_name", ""), it.get("batch_no", ""),
                  it.get("expiry_date", ""), qty, rate,
                  float(it.get("mrp", 0)), round(qty * rate, 2)))
        conn.commit()
    return jsonify({"status": "success", "expiry_return_id": er_id, "return_no": return_no})


@app.route("/api/expiry-returns/<int:return_id>", methods=["GET"])
def get_expiry_return(return_id):
    with get_conn() as conn:
        r = conn.execute("SELECT * FROM expiry_returns WHERE id=?", (return_id,)).fetchone()
        if not r:
            return jsonify({"status": "error", "message": "Not found"}), 404
        r = dict(r)
        items = [dict(i) for i in conn.execute(
            "SELECT * FROM expiry_return_items WHERE expiry_return_id=?", (return_id,)
        ).fetchall()]
    line_items = [{"expiry_return_item_id": it["id"], "item_name": it["item_name"],
                   "batch_no": it["batch_no"], "expiry_date": it["expiry_date"][:7] if it["expiry_date"] else "",
                   "qty_returned": it["qty_returned"], "purchase_rate": float(it["purchase_rate"]),
                   "mrp": float(it["mrp"]), "return_value": float(it["return_value"])} for it in items]
    return jsonify({
        "expiry_return_id": r["id"], "return_no": r["return_no"],
        "return_date": r["return_date"], "supplier_name": r.get("supplier_name", ""),
        "status": r["status"], "total_qty": r["total_qty"],
        "total_value": float(r["total_value"]),
        "credit_note_no": r.get("credit_note_no") or "",
        "credit_note_date": r.get("credit_note_date") or "",
        "credit_note_amount": float(r.get("credit_note_amount") or 0),
        "remarks": r.get("remarks") or "", "items": line_items,
    })


@app.route("/api/expiry-returns/<int:return_id>/status", methods=["PUT"])
def update_expiry_return_status(return_id):
    data = request.get_json(silent=True) or {}
    new_status = str(data.get("status", "")).strip().upper()
    valid = {"DRAFT", "DISPATCHED", "CREDIT_NOTE_RECEIVED", "CLOSED"}
    if new_status not in valid:
        return jsonify({"status": "error", "message": "Invalid status"}), 400
    with get_conn() as conn:
        r = conn.execute("SELECT * FROM expiry_returns WHERE id=?", (return_id,)).fetchone()
        if not r:
            return jsonify({"status": "error", "message": "Not found"}), 404
        updates = {"status": new_status}
        if new_status == "CREDIT_NOTE_RECEIVED":
            updates["credit_note_no"] = data.get("credit_note_no", "")
            updates["credit_note_date"] = data.get("credit_note_date", "")
            updates["credit_note_amount"] = float(data.get("credit_note_amount") or 0)
        set_clause = ", ".join(f"{k}=?" for k in updates)
        conn.execute(f"UPDATE expiry_returns SET {set_clause} WHERE id=?",
                     (*updates.values(), return_id))
        conn.commit()
    return jsonify({"status": "success", "new_status": new_status})


# ── License Tracker Page ───────────────────────────────────────
@app.route("/licenses")
def licenses_page():
    return render_template("licenses.html")


@app.route("/api/licenses", methods=["GET"])
def list_licenses():
    from datetime import date
    today = date.today()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM license_documents WHERE is_active=1 ORDER BY expiry_date ASC"
        ).fetchall()
    result = []
    for r in [dict(r) for r in rows]:
        try:
            exp = date.fromisoformat(r["expiry_date"])
            days_left = (exp - today).days
            expiry_fmt = exp.strftime("%d/%m/%Y")
        except Exception:
            days_left = 0
            expiry_fmt = r["expiry_date"]
        if days_left < 0:
            level = "EXPIRED"
        elif days_left <= 30:
            level = "CRITICAL"
        elif days_left <= 60:
            level = "WARNING"
        elif days_left <= 90:
            level = "ATTENTION"
        else:
            level = "OK"
        result.append({
            "license_id": r["id"],
            "license_type": r.get("license_type", "OTHER"),
            "license_name": r.get("license_name", ""),
            "license_no": r.get("license_no", ""),
            "issuing_authority": r.get("issuing_authority") or "",
            "issued_date": r.get("issued_date") or "",
            "expiry_date": expiry_fmt,
            "expiry_date_iso": r.get("expiry_date", ""),
            "days_left": days_left,
            "alert_level": level,
            "doc_url": r.get("doc_url") or "",
            "notes": r.get("notes") or "",
            "renewal_cost": float(r.get("renewal_cost") or 0),
            "renewal_contact": r.get("renewal_contact") or "",
        })
    return jsonify(result)


@app.route("/api/licenses/alerts", methods=["GET"])
def license_alerts():
    from datetime import date
    today = date.today()
    cutoff = (today + timedelta(days=90)).isoformat()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM license_documents WHERE is_active=1 AND expiry_date<=? ORDER BY expiry_date ASC",
            (cutoff,)
        ).fetchall()
    result = []
    for r in [dict(r) for r in rows]:
        try:
            exp = date.fromisoformat(r["expiry_date"])
            days_left = (exp - today).days
        except Exception:
            days_left = 0
        result.append({
            "license_id": r["id"],
            "license_name": r.get("license_name", ""),
            "license_type": r.get("license_type", ""),
            "expiry_date": r.get("expiry_date", ""),
            "days_left": days_left,
            "alert_level": "EXPIRED" if days_left < 0 else ("CRITICAL" if days_left <= 30 else ("WARNING" if days_left <= 60 else "ATTENTION")),
        })
    return jsonify(result)


@app.route("/api/licenses", methods=["POST"])
def create_license():
    data = request.get_json(silent=True) or {}
    required = ["license_type", "license_name", "license_no", "expiry_date"]
    missing = [f for f in required if not str(data.get(f, "")).strip()]
    if missing:
        return jsonify({"status": "error", "message": "Missing fields", "details": missing}), 400
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO license_documents
                (license_type, license_name, license_no, issuing_authority,
                 issued_date, expiry_date, doc_url, notes, renewal_cost, renewal_contact)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            str(data["license_type"]).upper(),
            str(data["license_name"]).strip(),
            str(data["license_no"]).strip(),
            str(data.get("issuing_authority", "") or "").strip() or None,
            str(data.get("issued_date", "") or "").strip() or None,
            str(data["expiry_date"]).strip(),
            str(data.get("doc_url", "") or "").strip() or None,
            str(data.get("notes", "") or "").strip() or None,
            float(data["renewal_cost"]) if data.get("renewal_cost") else None,
            str(data.get("renewal_contact", "") or "").strip() or None,
        ))
        lid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()
    return jsonify({"status": "success", "license_id": lid})


@app.route("/api/licenses/<int:license_id>", methods=["PUT"])
def update_license(license_id):
    data = request.get_json(silent=True) or {}
    with get_conn() as conn:
        r = conn.execute("SELECT id FROM license_documents WHERE id=? AND is_active=1", (license_id,)).fetchone()
        if not r:
            return jsonify({"status": "error", "message": "Not found"}), 404
        fields = {}
        for k in ["license_type", "license_name", "license_no", "issuing_authority",
                  "issued_date", "expiry_date", "doc_url", "notes", "renewal_cost", "renewal_contact"]:
            if k in data:
                v = data[k]
                if k == "license_type" and v:
                    v = str(v).upper()
                fields[k] = v if v not in ("", None) else None
        if fields:
            fields["updated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            set_clause = ", ".join(f"{k}=?" for k in fields)
            conn.execute(f"UPDATE license_documents SET {set_clause} WHERE id=?",
                         (*fields.values(), license_id))
            conn.commit()
    return jsonify({"status": "success"})


@app.route("/api/licenses/<int:license_id>", methods=["DELETE"])
def delete_license(license_id):
    with get_conn() as conn:
        conn.execute("UPDATE license_documents SET is_active=0 WHERE id=?", (license_id,))
        conn.commit()
    return jsonify({"status": "success"})


# ═══════════════════════════════════════════════════════════════════════════
# STAGE 7 — AI PRESCRIPTION READER + DRUG INTERACTIONS + SMART REORDER AI
# ═══════════════════════════════════════════════════════════════════════════
import difflib
import re

# ── Drug interaction knowledge base ─────────────────────────────────────────
_INTERACTIONS = [
    ("warfarin",      "aspirin",        "SEVERE",   "Increased bleeding risk — both inhibit clotting. Monitor INR closely."),
    ("warfarin",      "ibuprofen",      "SEVERE",   "NSAIDs displace warfarin from protein binding, raising bleeding risk."),
    ("warfarin",      "naproxen",       "SEVERE",   "NSAIDs potentiate anticoagulant effect of warfarin — risk of haemorrhage."),
    ("metformin",     "alcohol",        "MODERATE", "Risk of lactic acidosis, especially with heavy or binge alcohol use."),
    ("amlodipine",    "simvastatin",    "MODERATE", "Amlodipine raises simvastatin AUC ~77% — increased myopathy risk."),
    ("atorvastatin",  "clarithromycin", "SEVERE",   "Macrolide inhibits CYP3A4 — statin levels surge, risk of rhabdomyolysis."),
    ("methotrexate",  "ibuprofen",      "SEVERE",   "NSAIDs reduce renal methotrexate clearance — severe toxicity risk."),
    ("digoxin",       "amiodarone",     "SEVERE",   "Amiodarone increases digoxin plasma levels — digoxin toxicity risk."),
    ("lisinopril",    "potassium",      "MODERATE", "ACE inhibitors raise serum K⁺ — hyperkalemia risk, especially in CKD."),
    ("ramipril",      "potassium",      "MODERATE", "ACE inhibitors raise serum K⁺ — hyperkalemia risk."),
    ("enalapril",     "potassium",      "MODERATE", "ACE inhibitors raise serum K⁺ — hyperkalemia risk."),
    ("clopidogrel",   "omeprazole",     "MODERATE", "Omeprazole inhibits CYP2C19, reducing clopidogrel activation by ~40%."),
    ("sildenafil",    "nitroglycerin",  "SEVERE",   "Severe hypotension — nitrates + PDE5 inhibitors are contraindicated."),
    ("sildenafil",    "isosorbide",     "SEVERE",   "Severe hypotension — nitrates + PDE5 inhibitors are contraindicated."),
    ("fluoxetine",    "tramadol",       "MODERATE", "Serotonin syndrome risk + lowered seizure threshold."),
    ("sertraline",    "tramadol",       "MODERATE", "Serotonin syndrome risk — monitor for agitation, fever, tremor."),
    ("ciprofloxacin", "antacid",        "MILD",     "Antacids form chelates with fluoroquinolones — take 2 h apart."),
    ("ciprofloxacin", "calcium",        "MILD",     "Calcium reduces ciprofloxacin absorption — take on empty stomach."),
    ("levothyroxine", "calcium",        "MILD",     "Calcium carbonate reduces T4 absorption — separate by ≥4 h."),
    ("levothyroxine", "iron",           "MILD",     "Iron salts reduce levothyroxine absorption — separate by ≥4 h."),
    ("tetracycline",  "calcium",        "MODERATE", "Divalent cations chelate tetracyclines — avoid dairy within 2 h."),
    ("tetracycline",  "antacid",        "MODERATE", "Antacids severely reduce tetracycline absorption."),
    ("azithromycin",  "amiodarone",     "SEVERE",   "Both prolong QT interval — risk of life-threatening arrhythmia."),
    ("metformin",     "contrast",       "SEVERE",   "Hold metformin 48 h before/after iodinated contrast — lactic acidosis risk."),
    ("aspirin",       "ibuprofen",      "MODERATE", "Ibuprofen may antagonise aspirin's antiplatelet effect — take aspirin first."),
    ("phenytoin",     "folic acid",     "MILD",     "Folate may lower phenytoin levels; phenytoin depletes folate stores."),
    ("amlodipine",    "grapefruit",     "MILD",     "Grapefruit juice inhibits CYP3A4 — may increase amlodipine levels."),
    ("atorvastatin",  "gemfibrozil",    "SEVERE",   "Fibrates + statins raise myopathy risk significantly — avoid combination."),
    ("glimepiride",   "fluconazole",    "SEVERE",   "Azoles inhibit CYP2C9 — sulphonylurea levels rise, hypoglycaemia risk."),
    ("metformin",     "furosemide",     "MILD",     "Loop diuretics may raise metformin plasma levels slightly."),
]

def _build_interaction_lookup():
    lookup = {}
    for d1, d2, sev, note in _INTERACTIONS:
        lookup.setdefault(d1, []).append((d2, sev, note))
        lookup.setdefault(d2, []).append((d1, sev, note))
    return lookup

_INT_LOOKUP = _build_interaction_lookup()


def _clean_token(raw: str) -> str:
    """Strip dosage/form suffixes from a medicine token."""
    t = raw.strip()
    # remove trailing dosage: 500mg, 10ml, 2.5%, etc.
    t = re.sub(r'\s*\d+(\.\d+)?\s*(mg|mcg|ml|g|iu|%|units?)\b', '', t, flags=re.I)
    # remove form words
    t = re.sub(r'\b(tab|tablet|cap|capsule|inj|injection|syrup|susp|suspension|drops?|cream|oint|gel|spray|sachet|powder|soln|solution)\b', '', t, flags=re.I)
    # remove leading numbers like "1.", "2)" etc.
    t = re.sub(r'^\s*\d+[.)]\s*', '', t)
    return t.strip()


def _fuzzy_match_medicines(conn, tokens):
    all_meds = conn.execute(
        "SELECT id, n, g, c, p, s, p_rate FROM medicines LIMIT 8000"
    ).fetchall()
    names = [m["n"] for m in all_meds]
    name_lower = [n.lower() for n in names]

    results = []
    for raw in tokens:
        if not raw or len(raw) < 3:
            continue
        cleaned = _clean_token(raw)
        if len(cleaned) < 2:
            cleaned = raw.strip()

        query_l = cleaned.lower()

        # 1. Exact / starts-with match (highest confidence)
        exact = next((m for m in all_meds if m["n"].lower() == query_l), None)
        if exact:
            results.append({"extracted": raw, "item_id": str(exact["id"]),
                             "name": exact["n"], "category": exact["c"],
                             "mrp": float(exact["p"] or 0),
                             "stock": int(exact["s"] or 0),
                             "purchase_rate": float(exact["p_rate"] or 0),
                             "score": 1.0})
            continue

        starts = [m for m in all_meds if m["n"].lower().startswith(query_l[:5])]
        if starts:
            m = starts[0]
            score = difflib.SequenceMatcher(None, query_l, m["n"].lower()).ratio()
            results.append({"extracted": raw, "item_id": str(m["id"]),
                             "name": m["n"], "category": m["c"],
                             "mrp": float(m["p"] or 0),
                             "stock": int(m["s"] or 0),
                             "purchase_rate": float(m["p_rate"] or 0),
                             "score": round(score, 2)})
            continue

        # 2. Fuzzy via difflib
        close = difflib.get_close_matches(cleaned, names, n=1, cutoff=0.5)
        if close:
            m = next(x for x in all_meds if x["n"] == close[0])
            score = difflib.SequenceMatcher(None, query_l, close[0].lower()).ratio()
            results.append({"extracted": raw, "item_id": str(m["id"]),
                             "name": m["n"], "category": m["c"],
                             "mrp": float(m["p"] or 0),
                             "stock": int(m["s"] or 0),
                             "purchase_rate": float(m["p_rate"] or 0),
                             "score": round(score, 2)})
            continue

        # 3. Substring search
        sub = [m for m in all_meds if query_l in m["n"].lower()]
        if sub:
            m = sub[0]
            results.append({"extracted": raw, "item_id": str(m["id"]),
                             "name": m["n"], "category": m["c"],
                             "mrp": float(m["p"] or 0),
                             "stock": int(m["s"] or 0),
                             "purchase_rate": float(m["p_rate"] or 0),
                             "score": 0.6})
            continue

        # No match
        results.append({"extracted": raw, "item_id": None, "name": None,
                        "score": 0.0})

    return results


def _check_interactions(matches):
    """Check all extracted medicine names (matched or not) for known drug interactions."""
    found_drugs = []
    seen_keywords: set[str] = set()
    for m in matches:
        # Use matched DB name if available, else the raw extracted text
        display_name = m.get("name") or m.get("extracted", "")
        if not display_name:
            continue
        name_l = display_name.lower()
        for keyword in _INT_LOOKUP:
            if keyword in name_l and keyword not in seen_keywords:
                seen_keywords.add(keyword)
                found_drugs.append((keyword, display_name))

    interactions = []
    seen = set()
    for i, (kw1, name1) in enumerate(found_drugs):
        for kw2, name2 in found_drugs[i+1:]:
            if kw1 == kw2:
                continue
            pair_key = tuple(sorted([kw1, kw2]))
            if pair_key in seen:
                continue
            if kw2 in _INT_LOOKUP.get(kw1, [d[0] for d in []]):
                seen.add(pair_key)
            for (partner, sev, note) in _INT_LOOKUP.get(kw1, []):
                if partner == kw2:
                    if pair_key not in seen:
                        seen.add(pair_key)
                        interactions.append({
                            "drug1": name1, "drug2": name2,
                            "severity": sev, "note": note
                        })
    return interactions


# ── Page routes ──────────────────────────────────────────────────────────────
@app.route("/prescription")
def page_prescription():
    return render_template("prescription.html")


@app.route("/smart-reorder")
def page_smart_reorder():
    return render_template("smart_reorder.html")


# ── Prescription extract API ─────────────────────────────────────────────────
@app.route("/api/prescription/extract", methods=["POST"])
def api_prescription_extract():
    text = request.form.get("text", "").strip()
    img_file = request.files.get("image")

    ocr_note = None
    if img_file and img_file.filename:
        # Try pytesseract; gracefully fall back
        try:
            from PIL import Image
            import pytesseract
            img = Image.open(img_file.stream)
            ocr_text = pytesseract.image_to_string(img)
            if ocr_text.strip():
                text = ocr_text
                ocr_note = "OCR extracted text from image"
        except Exception:
            ocr_note = "OCR unavailable — using typed text (install pytesseract + Tesseract for image OCR)"

    if not text:
        return jsonify({"status": "error",
                        "message": "No text provided. Type medicine names or upload a prescription image."})

    # Parse tokens: split by newlines, commas, semicolons, bullets
    raw_tokens = re.split(r'[\n\r,;•·\-]+', text)
    tokens = []
    for t in raw_tokens:
        t = t.strip().strip('*').strip()
        if len(t) >= 3:
            tokens.append(t)

    if not tokens:
        return jsonify({"status": "error", "message": "Could not parse any medicine names from the input."})

    with get_conn() as conn:
        matches = _fuzzy_match_medicines(conn, tokens[:30])  # cap at 30 items

    interactions = _check_interactions(matches)

    return jsonify({
        "status": "ok",
        "ocr_note": ocr_note,
        "matches": matches,
        "interactions": interactions,
    })


# ── Smart Reorder Analysis API ───────────────────────────────────────────────
@app.route("/api/smart-reorder/analysis")
def api_smart_reorder_analysis():
    from datetime import datetime, timedelta

    days_window = int(request.args.get("days", 90))
    cutoff_date = (datetime.utcnow() - timedelta(days=days_window)).strftime("%Y-%m-%d")

    with get_conn() as conn:
        # All medicines with stock data
        medicines = conn.execute(
            "SELECT id, n, g, c, p, s, p_rate, reorder, max_qty FROM medicines"
        ).fetchall()

        # Recent bills for sales velocity
        recent_bills = conn.execute(
            "SELECT items, date FROM bills WHERE date >= ? AND items IS NOT NULL",
            (cutoff_date,)
        ).fetchall()

    # Tally qty sold per medicine id over the window
    sales_map: dict[str, float] = {}
    for bill in recent_bills:
        try:
            bill_items = json.loads(bill["items"]) if isinstance(bill["items"], str) else bill["items"]
            if not isinstance(bill_items, list):
                continue
            for bi in bill_items:
                mid = str(bi.get("id", ""))
                qty = float(bi.get("qty", 0))
                sales_map[mid] = sales_map.get(mid, 0) + qty
        except (json.JSONDecodeError, TypeError, AttributeError):
            continue

    items_out = []
    summary = {"critical": 0, "warning": 0, "attention": 0, "ok": 0,
               "total": 0, "reorder_value": 0.0}

    for m in medicines:
        mid = str(m["id"])
        current_stock = int(m["s"] or 0)
        mrp = float(m["p"] or 0)
        purchase_rate = float(m["p_rate"] or 0)
        min_reorder = int(m["reorder"] or 10)
        max_qty = int(m["max_qty"] or 0)

        # Velocity: units sold / day
        total_sold = sales_map.get(mid, 0)
        velocity = round(total_sold / days_window, 3)

        # Days to stockout
        if velocity > 0:
            days_to_stockout = int(current_stock / velocity)
        else:
            days_to_stockout = 9999  # no recent sales

        # Stock pct (relative to max_qty or 100)
        stock_cap = max_qty if max_qty > 0 else max(current_stock * 2, 50)
        stock_pct = min(100, int((current_stock / stock_cap) * 100)) if stock_cap else 0

        # Default min stock threshold (portal uses 15)
        default_min = max(min_reorder, 15)

        # Priority: velocity-based when we have sales data; stock-level fallback otherwise
        if velocity > 0:
            if days_to_stockout <= 7:
                priority = "critical"
            elif days_to_stockout <= 14:
                priority = "warning"
            elif days_to_stockout <= 30:
                priority = "attention"
            else:
                priority = "ok"
            confidence_base = 75
        else:
            # No sales history — use stock vs default_min
            if current_stock == 0:
                priority = "critical"
            elif current_stock < 5:
                priority = "warning"
            elif current_stock < default_min:
                priority = "attention"
            else:
                priority = "ok"
            confidence_base = 40

        # Reorder qty: velocity-based when possible, else fill to default_min
        if velocity > 0:
            reorder_qty = max(int(velocity * 30) - current_stock, min_reorder)
        elif current_stock < default_min:
            reorder_qty = default_min - current_stock
        else:
            reorder_qty = 0

        reorder_value = round(reorder_qty * purchase_rate, 2)

        # Confidence: higher when more sales data exists
        data_points = min(int(total_sold), 100)
        confidence = min(95, confidence_base + int(data_points * 0.55))

        # Skip items with adequate stock and no recent sales (noise reduction)
        if current_stock == 0 and total_sold == 0:
            continue

        summary[priority] += 1
        summary["total"] += 1
        if reorder_qty > 0:
            summary["reorder_value"] += reorder_value

        items_out.append({
            "item_id": mid,
            "name": m["n"],
            "generic": m["g"] or "",
            "category": m["c"] or "",
            "supplier": "",  # suppliers table not linked in SQLite demo
            "current_stock": current_stock,
            "min_stock": min_reorder,
            "stock_pct": stock_pct,
            "daily_velocity": velocity,
            "total_sold_90d": int(total_sold),
            "days_to_stockout": days_to_stockout,
            "priority": priority,
            "reorder_qty": reorder_qty,
            "purchase_rate": purchase_rate,
            "mrp": mrp,
            "reorder_value": reorder_value,
            "confidence": confidence,
        })

    # Sort by priority urgency then days
    priority_order = {"critical": 0, "warning": 1, "attention": 2, "ok": 3}
    items_out.sort(key=lambda x: (priority_order[x["priority"]], x["days_to_stockout"]))

    summary["reorder_value"] = round(summary["reorder_value"], 2)

    return jsonify({"status": "ok", "summary": summary, "items": items_out})


# ═══════════════════════════════════════════════════════════════════════════
# STAGE 8 — PATIENT HEALTH RECORDS + DOCTOR CRM + LOYALTY POINTS
# ═══════════════════════════════════════════════════════════════════════════

def init_stage8_db():
    with get_conn() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS patient_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            dob TEXT,
            blood_group TEXT,
            address TEXT DEFAULT '',
            email TEXT DEFAULT '',
            emergency_name TEXT DEFAULT '',
            emergency_phone TEXT DEFAULT '',
            allergies TEXT DEFAULT '[]',
            conditions TEXT DEFAULT '[]',
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS doctor_extra (
            id INTEGER PRIMARY KEY,
            reg_no TEXT DEFAULT '',
            address TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS loyalty_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            points_balance INTEGER DEFAULT 0,
            tier TEXT DEFAULT 'bronze',
            total_earned INTEGER DEFAULT 0,
            total_redeemed INTEGER DEFAULT 0,
            visits INTEGER DEFAULT 0,
            last_visit TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS loyalty_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT NOT NULL,
            bill_id TEXT,
            type TEXT NOT NULL,
            points INTEGER NOT NULL,
            balance_after INTEGER NOT NULL,
            note TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        conn.commit()


def _loyalty_tier(total_earned: int) -> str:
    if total_earned >= 15000: return 'platinum'
    if total_earned >= 5000:  return 'gold'
    if total_earned >= 1000:  return 'silver'
    return 'bronze'


def _tier_multiplier(tier: str) -> float:
    return {'bronze':1.0,'silver':1.25,'gold':1.5,'platinum':2.0}.get(tier,1.0)


# ── Patient pages & APIs ─────────────────────────────────────────────────────
@app.route("/patients")
def page_patients():
    return render_template("patients.html")

@app.route("/doctor-crm")
def page_doctor_crm():
    return render_template("doctor_crm.html")

@app.route("/loyalty")
def page_loyalty():
    return render_template("loyalty.html")


@app.route("/api/patients", methods=["GET"])
def api_patients_list():
    with get_conn() as conn:
        # Merge patient_records with customers table
        records = conn.execute("SELECT * FROM patient_records ORDER BY name").fetchall()
        patients = []
        now = datetime.utcnow().strftime("%Y-%m-%d")
        cutoff_30 = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
        allergy_count = chronic_count = active_30d = 0
        total_visits = 0
        for r in records:
            cust = conn.execute("SELECT visits,total FROM customers WHERE phone=?", (r["phone"],)).fetchone()
            last_bill = conn.execute(
                "SELECT MAX(date) as ld FROM bills WHERE phone=?", (r["phone"],)
            ).fetchone()
            allergies = json.loads(r["allergies"] or "[]")
            conditions = json.loads(r["conditions"] or "[]")
            last_visit = last_bill["ld"] if last_bill and last_bill["ld"] else None
            if allergies: allergy_count += 1
            if conditions: chronic_count += 1
            if last_visit and last_visit >= cutoff_30: active_30d += 1
            v = cust["visits"] if cust else 0
            total_visits += v
            patients.append({
                "phone": r["phone"], "name": r["name"], "dob": r["dob"],
                "blood_group": r["blood_group"], "allergies": allergies,
                "conditions": conditions, "visits": v,
                "total": float(cust["total"]) if cust else 0,
                "last_visit": last_visit,
                "address": r["address"], "email": r["email"],
            })
    return jsonify({
        "patients": patients,
        "summary": {
            "total": len(patients), "active_30d": active_30d,
            "total_visits": total_visits, "allergy_count": allergy_count,
            "chronic_count": chronic_count,
        }
    })


@app.route("/api/patients", methods=["POST"])
def api_patients_create():
    data = request.json or {}
    phone = data.get("phone","").strip()
    name = data.get("name","").strip()
    if not phone or not name:
        return jsonify({"status":"error","message":"name and phone required"}), 400
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO patient_records (phone,name,dob,blood_group,address,email,
                emergency_name,emergency_phone,allergies,conditions,notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(phone) DO UPDATE SET
                name=excluded.name, dob=excluded.dob, blood_group=excluded.blood_group,
                address=excluded.address, email=excluded.email,
                emergency_name=excluded.emergency_name, emergency_phone=excluded.emergency_phone,
                allergies=excluded.allergies, conditions=excluded.conditions,
                notes=excluded.notes, updated_at=datetime('now')
        """, (phone, name, data.get("dob"), data.get("blood_group"),
              data.get("address",""), data.get("email",""),
              data.get("emergency_name",""), data.get("emergency_phone",""),
              json.dumps(data.get("allergies",[])), json.dumps(data.get("conditions",[])),
              data.get("notes","")))
        # Also insert into customers table if not already there
        exists = conn.execute("SELECT id FROM customers WHERE phone=?", (phone,)).fetchone()
        if not exists:
            conn.execute("INSERT INTO customers (name,phone,visits,total,address,email) VALUES (?,?,0,0.0,?,?)",
                         (name, phone, data.get("address",""), data.get("email","")))
        conn.commit()
    return jsonify({"status":"ok"})


@app.route("/api/patients/<phone>", methods=["GET"])
def api_patient_detail(phone):
    with get_conn() as conn:
        r = conn.execute("SELECT * FROM patient_records WHERE phone=?", (phone,)).fetchone()
        if not r:
            return jsonify({"status":"error","message":"Not found"}), 404
        cust = conn.execute("SELECT * FROM customers WHERE phone=?", (phone,)).fetchone()
        bills = conn.execute(
            "SELECT id,ts,date,doctor,items,total FROM bills WHERE phone=? ORDER BY date DESC LIMIT 20",
            (phone,)
        ).fetchall()
        return jsonify({
            "patient": {
                "phone": r["phone"], "name": r["name"], "dob": r["dob"],
                "blood_group": r["blood_group"],
                "allergies": json.loads(r["allergies"] or "[]"),
                "conditions": json.loads(r["conditions"] or "[]"),
                "notes": r["notes"], "address": r["address"], "email": r["email"],
                "emergency_name": r["emergency_name"], "emergency_phone": r["emergency_phone"],
                "total": float(cust["total"]) if cust else 0,
                "visits": cust["visits"] if cust else 0,
                "last_visit": bills[0]["date"] if bills else None,
            },
            "bills": [dict(b) for b in bills]
        })


@app.route("/api/patients/<phone>", methods=["PUT"])
def api_patient_update(phone):
    data = request.json or {}
    with get_conn() as conn:
        conn.execute("""
            UPDATE patient_records SET name=?,dob=?,blood_group=?,address=?,email=?,
                emergency_name=?,emergency_phone=?,allergies=?,conditions=?,notes=?,
                updated_at=datetime('now') WHERE phone=?
        """, (data.get("name"), data.get("dob"), data.get("blood_group"),
              data.get("address",""), data.get("email",""),
              data.get("emergency_name",""), data.get("emergency_phone",""),
              json.dumps(data.get("allergies",[])), json.dumps(data.get("conditions",[])),
              data.get("notes",""), phone))
        conn.commit()
    return jsonify({"status":"ok"})


@app.route("/api/patients/<phone>", methods=["DELETE"])
def api_patient_delete(phone):
    """Hard-delete a patient record by phone."""
    with get_conn() as conn:
        cur = conn.execute("DELETE FROM patient_records WHERE phone=?", (phone,))
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({"status": "error", "message": "Patient not found"}), 404
    return jsonify({"status": "ok", "deleted": phone})


# ── Doctor CRM APIs ──────────────────────────────────────────────────────────
@app.route("/api/doctors-crm", methods=["GET"])
def api_doctors_crm_list():
    cutoff_30 = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    with get_conn() as conn:
        docs = conn.execute("SELECT * FROM doctors").fetchall()
        result, total_rx, total_rev, unique_pats, active_30d = [], 0, 0.0, set(), 0
        for d in docs:
            name = d["name"]
            bills = conn.execute(
                "SELECT id,cust,phone,date,total,items FROM bills WHERE doctor=?", (name,)
            ).fetchall()
            rx = len(bills)
            rev = sum(float(b["total"] or 0) for b in bills)
            pats = {b["phone"] for b in bills if b["phone"]}
            last_rx = max((b["date"] or "" for b in bills), default=None) or None
            is_active = any((b["date"] or "") >= cutoff_30 for b in bills)
            if is_active: active_30d += 1
            total_rx += rx; total_rev += rev; unique_pats.update(pats)
            extra = conn.execute("SELECT * FROM doctor_extra WHERE id=?", (d["id"],)).fetchone()
            result.append({
                "id": d["id"], "name": name, "specialty": d["specialty"],
                "hospital": d["hospital"], "phone": d["phone"], "email": d["email"],
                "reg_no": extra["reg_no"] if extra else "",
                "address": extra["address"] if extra else "",
                "notes": extra["notes"] if extra else "",
                "rx_count": rx, "revenue": round(rev,2),
                "patient_count": len(pats), "last_rx": last_rx,
            })
    result.sort(key=lambda x: x["rx_count"], reverse=True)
    return jsonify({
        "doctors": result,
        "summary": {
            "total": len(result), "active_30d": active_30d,
            "total_rx": total_rx, "revenue": round(total_rev,2),
            "unique_patients": len(unique_pats),
        }
    })


@app.route("/api/doctors-crm", methods=["POST"])
def api_doctors_crm_create():
    data = request.json or {}
    name = data.get("name","").strip()
    if not name: return jsonify({"status":"error","message":"name required"}), 400
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO doctors (name,specialty,hospital,phone,email) VALUES (?,?,?,?,?)",
            (name, data.get("specialty",""), data.get("hospital",""),
             data.get("phone",""), data.get("email",""))
        )
        doc_id = cur.lastrowid
        conn.execute(
            "INSERT OR IGNORE INTO doctor_extra (id,reg_no,address,notes) VALUES (?,?,?,?)",
            (doc_id, data.get("reg_no",""), data.get("address",""), data.get("notes",""))
        )
        conn.commit()
    return jsonify({"status":"ok"})


@app.route("/api/doctors-crm/<int:doc_id>", methods=["GET"])
def api_doctor_crm_detail(doc_id):
    with get_conn() as conn:
        d = conn.execute("SELECT * FROM doctors WHERE id=?", (doc_id,)).fetchone()
        if not d: return jsonify({"status":"error","message":"Not found"}), 404
        extra = conn.execute("SELECT * FROM doctor_extra WHERE id=?", (doc_id,)).fetchone()
        bills = conn.execute(
            "SELECT id,cust,phone,date,total,items FROM bills WHERE doctor=? ORDER BY date DESC LIMIT 50",
            (d["name"],)
        ).fetchall()
        # Count top medicines
        med_count: dict[str,int] = {}
        for b in bills:
            try:
                items = json.loads(b["items"]) if isinstance(b["items"],str) else b["items"]
                for it in (items or []):
                    n = it.get("name","")
                    if n: med_count[n] = med_count.get(n,0) + int(it.get("qty",1))
            except: pass
        top_meds = sorted([{"name":k,"count":v} for k,v in med_count.items()],
                          key=lambda x: x["count"], reverse=True)[:12]
        rev = sum(float(b["total"] or 0) for b in bills)
        pats = {b["phone"] for b in bills if b["phone"]}
        last_rx = max((b["date"] or "" for b in bills), default=None) or None
        return jsonify({
            "doctor": {
                "id": d["id"], "name": d["name"], "specialty": d["specialty"],
                "hospital": d["hospital"], "phone": d["phone"], "email": d["email"],
                "reg_no": extra["reg_no"] if extra else "",
                "address": extra["address"] if extra else "",
                "notes": extra["notes"] if extra else "",
                "rx_count": len(bills), "revenue": round(rev,2),
                "patient_count": len(pats), "last_rx": last_rx,
            },
            "bills": [dict(b) for b in bills[:20]],
            "top_medicines": top_meds,
        })


@app.route("/api/doctors-crm/<int:doc_id>", methods=["PUT"])
def api_doctor_crm_update(doc_id):
    data = request.json or {}
    with get_conn() as conn:
        conn.execute("""UPDATE doctors SET name=?,specialty=?,hospital=?,phone=?,email=? WHERE id=?""",
                     (data.get("name"), data.get("specialty",""), data.get("hospital",""),
                      data.get("phone",""), data.get("email",""), doc_id))
        conn.execute("""INSERT INTO doctor_extra (id,reg_no,address,notes) VALUES (?,?,?,?)
            ON CONFLICT(id) DO UPDATE SET reg_no=excluded.reg_no,address=excluded.address,
            notes=excluded.notes,updated_at=datetime('now')""",
                     (doc_id, data.get("reg_no",""), data.get("address",""), data.get("notes","")))
        conn.commit()
    return jsonify({"status":"ok"})


# ── Loyalty APIs ─────────────────────────────────────────────────────────────
@app.route("/api/loyalty", methods=["GET"])
def api_loyalty_list():
    with get_conn() as conn:
        members = conn.execute(
            "SELECT * FROM loyalty_accounts ORDER BY points_balance DESC"
        ).fetchall()
        result = [dict(m) for m in members]
        tier_counts = {"bronze":0,"silver":0,"gold":0,"platinum":0}
        total_earned = total_redeemed = total_balance = 0
        for m in result:
            tier_counts[m["tier"]] = tier_counts.get(m["tier"],0) + 1
            total_earned += m.get("total_earned",0)
            total_redeemed += m.get("total_redeemed",0)
            total_balance += m.get("points_balance",0)
        n = len(result)
    return jsonify({
        "members": result,
        "summary": {
            "total": n, "active_points": total_balance,
            "total_earned": total_earned, "total_redeemed": total_redeemed,
            "avg_points": round(total_balance / n, 1) if n else 0,
            "tier_counts": tier_counts,
        }
    })


@app.route("/api/loyalty/<phone>", methods=["GET"])
def api_loyalty_member(phone):
    with get_conn() as conn:
        m = conn.execute("SELECT * FROM loyalty_accounts WHERE phone=?", (phone,)).fetchone()
        if not m: return jsonify({"status":"error","message":"Not found"}), 404
        txns = conn.execute(
            "SELECT * FROM loyalty_transactions WHERE phone=? ORDER BY created_at DESC LIMIT 30",
            (phone,)
        ).fetchall()
    return jsonify({"member": dict(m), "transactions": [dict(t) for t in txns]})


@app.route("/api/loyalty/<phone>/redeem", methods=["POST"])
def api_loyalty_redeem(phone):
    data = request.json or {}
    pts = int(data.get("points",0))
    with get_conn() as conn:
        m = conn.execute("SELECT * FROM loyalty_accounts WHERE phone=?", (phone,)).fetchone()
        if not m: return jsonify({"status":"error","message":"Not found"}), 404
        if pts > m["points_balance"]: return jsonify({"status":"error","message":"Insufficient points"}), 400
        new_bal = m["points_balance"] - pts
        new_redeemed = m["total_redeemed"] + pts
        conn.execute("UPDATE loyalty_accounts SET points_balance=?,total_redeemed=?,updated_at=datetime('now') WHERE phone=?",
                     (new_bal, new_redeemed, phone))
        conn.execute("INSERT INTO loyalty_transactions (phone,type,points,balance_after,note) VALUES (?,?,?,?,?)",
                     (phone,'REDEEM', pts, new_bal, f'Redeemed {pts} pts = ₹{pts//100}'))
        conn.commit()
    return jsonify({"status":"ok","new_balance":new_bal})


@app.route("/api/loyalty/<phone>/bonus", methods=["POST"])
def api_loyalty_bonus(phone):
    data = request.json or {}
    pts = int(data.get("points",0))
    with get_conn() as conn:
        m = conn.execute("SELECT * FROM loyalty_accounts WHERE phone=?", (phone,)).fetchone()
        if not m: return jsonify({"status":"error","message":"Not found"}), 404
        new_bal = m["points_balance"] + pts
        new_earned = m["total_earned"] + pts
        new_tier = _loyalty_tier(new_earned)
        conn.execute("UPDATE loyalty_accounts SET points_balance=?,total_earned=?,tier=?,updated_at=datetime('now') WHERE phone=?",
                     (new_bal, new_earned, new_tier, phone))
        conn.execute("INSERT INTO loyalty_transactions (phone,type,points,balance_after,note) VALUES (?,?,?,?,?)",
                     (phone,'BONUS', pts, new_bal, 'Bonus points awarded by staff'))
        conn.commit()
    return jsonify({"status":"ok","new_balance":new_bal})


@app.route("/api/loyalty/sync", methods=["POST"])
def api_loyalty_sync():
    """Sync loyalty accounts from existing bills data."""
    with get_conn() as conn:
        bills = conn.execute(
            "SELECT cust,phone,date,total FROM bills WHERE phone IS NOT NULL AND phone != '' ORDER BY date"
        ).fetchall()
        # Group by phone
        phone_data: dict[str, dict] = {}
        for b in bills:
            ph = b["phone"].strip()
            if not ph: continue
            if ph not in phone_data:
                phone_data[ph] = {"name": b["cust"] or ph, "visits":0, "total":0.0, "last_visit":None}
            phone_data[ph]["visits"] += 1
            phone_data[ph]["total"] += float(b["total"] or 0)
            phone_data[ph]["last_visit"] = b["date"]

        for ph, info in phone_data.items():
            # Base points: ₹10 = 1 point, multiplied by tier
            existing = conn.execute("SELECT * FROM loyalty_accounts WHERE phone=?", (ph,)).fetchone()
            if existing:
                continue  # Don't overwrite manually managed accounts
            base_pts = int(info["total"] / 10)
            tier = _loyalty_tier(base_pts)
            mult = _tier_multiplier(tier)
            earned = int(base_pts * mult)
            conn.execute("""
                INSERT INTO loyalty_accounts (phone,name,points_balance,tier,total_earned,total_redeemed,visits,last_visit)
                VALUES (?,?,?,?,?,0,?,?)
            """, (ph, info["name"], earned, tier, earned, info["visits"], info["last_visit"]))
            # Add summary transaction
            conn.execute("INSERT INTO loyalty_transactions (phone,type,points,balance_after,note) VALUES (?,?,?,?,?)",
                         (ph,'EARN',earned,earned,f'Auto-synced from {info["visits"]} bills (₹{info["total"]:.0f} total spend)'))
        conn.commit()
    return jsonify({"status":"ok"})


# ═══════════════════════════════════════════════════════════════════
#  STAGE 9 — MULTI-BRANCH SUPPORT + WHATSAPP PRODUCT CATALOG
# ═══════════════════════════════════════════════════════════════════

def init_stage9_db():
    """Create Stage 9 tables: branches, stock_transfers, catalog_items."""
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS branches (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                name         TEXT NOT NULL,
                type         TEXT DEFAULT 'retail',
                address      TEXT,
                phone        TEXT,
                manager      TEXT,
                gstin        TEXT,
                drug_license TEXT,
                is_active    INTEGER DEFAULT 1,
                created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS stock_transfers (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                from_branch_id  INTEGER,
                to_branch_id    INTEGER,
                from_branch     TEXT,
                to_branch       TEXT,
                medicine_id     INTEGER,
                medicine_name   TEXT NOT NULL,
                qty             INTEGER DEFAULT 0,
                note            TEXT,
                transferred_by  TEXT,
                status          TEXT DEFAULT 'completed',
                created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS catalog_items (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                medicine_id   INTEGER UNIQUE,
                medicine_name TEXT NOT NULL,
                generic_name  TEXT,
                category      TEXT,
                mrp           REAL DEFAULT 0,
                offer_price   REAL,
                description   TEXT,
                is_active     INTEGER DEFAULT 1,
                in_stock      INTEGER DEFAULT 1,
                image_url     TEXT,
                updated_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

        # Seed demo branches if empty
        existing = conn.execute("SELECT COUNT(*) FROM branches").fetchone()[0]
        if existing == 0:
            demo = [
                ("Selvam Medicals – Main Branch",   "main",      "12, Bazaar Street, Tirunelveli Junction, Tamil Nadu 627001",   "04622-234567", "Selvam Ramaswamy", "29AAACS1234A1Z5", "TN-DL-001-2023"),
                ("Selvam Medicals – Palayamkottai",  "retail",    "45, Shankar Nagar, Palayamkottai, Tirunelveli 627002",          "04622-345678", "Arun Kumar",       "29AAACS1234A1Z6", "TN-DL-002-2023"),
                ("Selvam Medicals – Wholesale Hub",  "wholesale", "8, Industrial Estate, Tirunelveli 627003",                     "04622-456789", "Vijay Shankar",    "29AAACS1234A1Z7", "TN-DL-003-2023"),
            ]
            for d in demo:
                conn.execute(
                    "INSERT INTO branches (name,type,address,phone,manager,gstin,drug_license) VALUES (?,?,?,?,?,?,?)",
                    d
                )
            conn.commit()


# ── Page routes ──────────────────────────────────────────────────

@app.route("/branches")
def page_branches():
    return render_template("branches.html")


@app.route("/catalog-manage")
def page_catalog_manage():
    return render_template("catalog_manage.html")


@app.route("/catalog")
def page_catalog():
    return render_template("catalog.html")


# ── Branches API ─────────────────────────────────────────────────

@app.route("/api/branches", methods=["GET"])
def api_branches_list():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM branches ORDER BY type='main' DESC, name"
        ).fetchall()

        result = []
        for r in rows:
            b = dict(r)
            # Attach bill stats from bills table
            stats = conn.execute(
                "SELECT COUNT(*) as bill_count, COALESCE(SUM(total),0) as revenue FROM bills"
            ).fetchone()
            # In demo, main branch owns all bills; other branches get proportional mock data
            if b["type"] == "main":
                b["bill_count"] = stats["bill_count"]
                b["revenue"]    = round(stats["revenue"], 2)
            elif b["type"] == "wholesale":
                b["bill_count"] = max(0, stats["bill_count"] // 4)
                b["revenue"]    = round(stats["revenue"] * 0.18, 2)
            else:
                b["bill_count"] = max(0, stats["bill_count"] // 6)
                b["revenue"]    = round(stats["revenue"] * 0.12, 2)
            result.append(b)

        # KPI totals
        total_revenue = sum(b["revenue"] for b in result)
        total_bills   = sum(b["bill_count"] for b in result)
        top_branch    = max(result, key=lambda x: x["revenue"])["name"] if result else "—"
        transfers     = conn.execute("SELECT COUNT(*) FROM stock_transfers").fetchone()[0]

        return jsonify({
            "branches":      result,
            "total_revenue": round(total_revenue, 2),
            "total_bills":   total_bills,
            "top_branch":    top_branch,
            "transfers":     transfers
        })


@app.route("/api/branches", methods=["POST"])
def api_branches_create():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Branch name required"}), 400
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO branches (name,type,address,phone,manager,gstin,drug_license,is_active) VALUES (?,?,?,?,?,?,?,?)",
            (name, data.get("type","retail"), data.get("address",""),
             data.get("phone",""), data.get("manager",""),
             data.get("gstin",""), data.get("drug_license",""),
             1 if data.get("is_active", True) else 0)
        )
        conn.commit()
        return jsonify({"id": cur.lastrowid, "status": "ok"})


@app.route("/api/branches/<int:bid>", methods=["PUT"])
def api_branches_update(bid):
    data = request.json or {}
    with get_conn() as conn:
        conn.execute("""
            UPDATE branches SET
              name=COALESCE(?,name), type=COALESCE(?,type),
              address=COALESCE(?,address), phone=COALESCE(?,phone),
              manager=COALESCE(?,manager), gstin=COALESCE(?,gstin),
              drug_license=COALESCE(?,drug_license),
              is_active=COALESCE(?,is_active)
            WHERE id=?
        """, (data.get("name"), data.get("type"), data.get("address"),
              data.get("phone"), data.get("manager"), data.get("gstin"),
              data.get("drug_license"), data.get("is_active"), bid))
        conn.commit()
        return jsonify({"status": "ok"})


# ── Stock Transfers API ──────────────────────────────────────────

@app.route("/api/stock-transfers", methods=["GET"])
def api_stock_transfers_list():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM stock_transfers ORDER BY created_at DESC LIMIT 100"
        ).fetchall()
        return jsonify({"transfers": [dict(r) for r in rows]})


@app.route("/api/stock-transfers", methods=["POST"])
def api_stock_transfers_create():
    data = request.json or {}
    med_name = (data.get("medicine_name") or "").strip()
    qty      = int(data.get("qty") or 0)
    if not med_name or qty <= 0:
        return jsonify({"error": "Medicine name and quantity required"}), 400

    # Look up branch names
    with get_conn() as conn:
        fb = conn.execute("SELECT name FROM branches WHERE id=?", (data.get("from_branch_id"),)).fetchone()
        tb = conn.execute("SELECT name FROM branches WHERE id=?", (data.get("to_branch_id"),)).fetchone()
        cur = conn.execute(
            """INSERT INTO stock_transfers
               (from_branch_id,to_branch_id,from_branch,to_branch,medicine_name,qty,note,transferred_by,status)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (data.get("from_branch_id"), data.get("to_branch_id"),
             fb["name"] if fb else "Unknown", tb["name"] if tb else "Unknown",
             med_name, qty,
             data.get("note",""), data.get("transferred_by","Staff"),
             "completed")
        )
        conn.commit()
        return jsonify({"id": cur.lastrowid, "status": "ok"})


# ── Medicine Search API (for transfer autocomplete) ──────────────

@app.route("/api/medicines/search")
def api_medicines_search():
    q     = (request.args.get("q") or "").strip()
    limit = min(int(request.args.get("limit", 20)), 50)
    if not q or len(q) < 2:
        return jsonify({"medicines": []})
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, n as name, g as generic, c as category, p as mrp, s as stock "
            "FROM medicines WHERE n LIKE ? LIMIT ?",
            (f"%{q}%", limit)
        ).fetchall()
        return jsonify({"medicines": [dict(r) for r in rows]})


# ── Catalog API ──────────────────────────────────────────────────

@app.route("/api/catalog", methods=["GET"])
def api_catalog_list():
    public = request.args.get("public") == "1"
    with get_conn() as conn:
        if public:
            rows = conn.execute(
                "SELECT * FROM catalog_items WHERE is_active=1 ORDER BY medicine_name"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM catalog_items ORDER BY medicine_name"
            ).fetchall()
        return jsonify({"items": [dict(r) for r in rows]})


@app.route("/api/catalog", methods=["POST"])
def api_catalog_create():
    data = request.json or {}
    name = (data.get("medicine_name") or "").strip()
    if not name:
        return jsonify({"error": "Medicine name required"}), 400
    with get_conn() as conn:
        try:
            cur = conn.execute(
                """INSERT INTO catalog_items
                   (medicine_id,medicine_name,generic_name,category,mrp,offer_price,
                    description,is_active,in_stock,image_url)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (data.get("medicine_id"), name,
                 data.get("generic_name",""), data.get("category",""),
                 float(data.get("mrp") or 0), data.get("offer_price"),
                 data.get("description",""),
                 int(data.get("is_active",1)), int(data.get("in_stock",1)),
                 data.get("image_url",""))
            )
            conn.commit()
            return jsonify({"id": cur.lastrowid, "status": "ok"})
        except Exception as e:
            return jsonify({"error": str(e)}), 400


@app.route("/api/catalog/<int:cid>", methods=["PUT"])
def api_catalog_update(cid):
    data = request.json or {}
    with get_conn() as conn:
        conn.execute("""
            UPDATE catalog_items SET
              medicine_name = COALESCE(?,medicine_name),
              generic_name  = COALESCE(?,generic_name),
              category      = COALESCE(?,category),
              mrp           = COALESCE(?,mrp),
              offer_price   = ?,
              description   = COALESCE(?,description),
              is_active     = COALESCE(?,is_active),
              in_stock      = COALESCE(?,in_stock),
              image_url     = COALESCE(?,image_url),
              updated_at    = CURRENT_TIMESTAMP
            WHERE id=?
        """, (data.get("medicine_name"), data.get("generic_name"),
              data.get("category"), data.get("mrp"),
              data.get("offer_price"),           # allow explicit NULL
              data.get("description"),
              data.get("is_active"), data.get("in_stock"),
              data.get("image_url"), cid))
        conn.commit()
        return jsonify({"status": "ok"})


@app.route("/api/catalog/<int:cid>", methods=["DELETE"])
def api_catalog_delete(cid):
    with get_conn() as conn:
        conn.execute("DELETE FROM catalog_items WHERE id=?", (cid,))
        conn.commit()
        return jsonify({"status": "ok"})


@app.route("/api/catalog/sync", methods=["POST"])
def api_catalog_sync():
    """Pull all in-stock medicines from inventory into catalog_items."""
    with get_conn() as conn:
        meds = conn.execute(
            "SELECT id, n as name, g as generic, c as category, p as mrp, s as stock "
            "FROM medicines WHERE s > 0 ORDER BY n"
        ).fetchall()

        added   = 0
        skipped = 0
        for m in meds:
            existing = conn.execute(
                "SELECT id FROM catalog_items WHERE medicine_id=?", (m["id"],)
            ).fetchone()
            if existing:
                skipped += 1
                continue
            conn.execute(
                """INSERT INTO catalog_items
                   (medicine_id,medicine_name,generic_name,category,mrp,in_stock,is_active)
                   VALUES (?,?,?,?,?,1,1)""",
                (m["id"], m["name"], m["generic"] or "", m["category"] or "", float(m["mrp"] or 0))
            )
            added += 1

        conn.commit()
        return jsonify({"status": "ok", "added": added, "skipped": skipped})


# ═══════════════════════════════════════════════════════════════════
#  STAGE 10 — WHOLESALE DISTRIBUTION PORTAL
# ═══════════════════════════════════════════════════════════════════

def init_wholesale_db():
    """Create wholesale tables: ws_shops, ws_orders, ws_order_items, ws_invoices,
       ws_payments, ws_order_timeline."""
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_shops (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                name          TEXT NOT NULL,
                owner         TEXT,
                phone         TEXT,
                gst_no        TEXT,
                drug_license  TEXT,
                address       TEXT,
                city          TEXT,
                credit_limit  REAL DEFAULT 50000,
                credit_days   INTEGER DEFAULT 30,
                discount      REAL DEFAULT 0,
                is_active     INTEGER DEFAULT 1,
                created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_orders (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no         TEXT UNIQUE NOT NULL,
                shop_id          INTEGER,
                shop_name        TEXT,
                status           TEXT DEFAULT 'pending',
                delivery_date    TEXT,
                delivery_address TEXT,
                notes            TEXT,
                subtotal         REAL DEFAULT 0,
                gst_amount       REAL DEFAULT 0,
                total            REAL DEFAULT 0,
                item_count       INTEGER DEFAULT 0,
                created_by       TEXT,
                created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_order_items (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id      INTEGER NOT NULL,
                medicine_id   INTEGER,
                medicine_name TEXT NOT NULL,
                generic_name  TEXT,
                qty           INTEGER DEFAULT 1,
                mrp           REAL DEFAULT 0,
                ws_rate       REAL DEFAULT 0,
                total         REAL DEFAULT 0
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_invoices (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_no    TEXT UNIQUE NOT NULL,
                order_id      INTEGER,
                order_no      TEXT,
                shop_id       INTEGER,
                shop_name     TEXT,
                subtotal      REAL DEFAULT 0,
                gst_amount    REAL DEFAULT 0,
                total         REAL DEFAULT 0,
                paid_amount   REAL DEFAULT 0,
                invoice_date  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_payments (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                shop_id     INTEGER,
                shop_name   TEXT,
                invoice_id  INTEGER,
                amount      REAL DEFAULT 0,
                mode        TEXT DEFAULT 'Cash',
                reference   TEXT,
                notes       TEXT,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_order_timeline (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id   INTEGER NOT NULL,
                note       TEXT,
                icon       TEXT DEFAULT '•',
                color      TEXT DEFAULT 'blue',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

        # Seed demo retail shops if empty
        if conn.execute("SELECT COUNT(*) FROM ws_shops").fetchone()[0] == 0:
            shops = [
                # Wholesale clients (3)
                ("Sri Ram Medical Hall",      "Ramesh Kumar",    "9876543001", "33AACCS1001A1Z1", "TN-DL-WS-001", "12, Main Road, Nagercoil, TN 629001",          "Nagercoil",   75000,  30, 5.0),
                ("Apollo Pharmacy – Madurai", "Suresh A.",       "9876543002", "33AACCS1002A1Z2", "TN-DL-WS-002", "45, Anna Nagar, Madurai, TN 625020",            "Madurai",     100000, 45, 8.0),
                ("City Medicals Pvt Ltd",     "Bala Subramani",  "9876543003", "33AACCS1003A1Z3", "TN-DL-WS-003", "7, Industrial Area, Coimbatore, TN 641003",     "Coimbatore",  150000, 60, 10.0),
                # Retail clients (10)
                ("Murugan Medical Store",     "Murugan P.",      "9876543010", "33AACCS1010A1Z4", "TN-DL-RT-001", "5, Temple Street, Tirunelveli, TN 627001",      "Tirunelveli", 25000,  21, 0.0),
                ("Vinayaga Medicals",         "Vinayaga R.",     "9876543011", "33AACCS1011A1Z5", "TN-DL-RT-002", "18, Gandhi Nagar, Tirunelveli, TN 627002",      "Tirunelveli", 20000,  21, 2.0),
                ("Sri Lakshmi Pharmacy",      "Lakshmi D.",      "9876543012", "33AACCS1012A1Z6", "TN-DL-RT-003", "33, Bazaar Road, Tuticorin, TN 628001",         "Tuticorin",   30000,  30, 3.0),
                ("Annamalai Medical Hall",    "Annamalai S.",    "9876543013", "33AACCS1013A1Z7", "TN-DL-RT-004", "2, KK Nagar, Tirunelveli, TN 627005",           "Tirunelveli", 20000,  21, 0.0),
                ("PSG Medicals",              "Prasanna G.",     "9876543014", "33AACCS1014A1Z8", "TN-DL-RT-005", "88, Palayamkottai Rd, Tirunelveli, TN 627002",  "Tirunelveli", 35000,  30, 4.0),
                ("Karuna Pharmacy",           "Karuna V.",       "9876543015", "33AACCS1015A1Z9", "TN-DL-RT-006", "14, Hospital Rd, Nellai Nagar, TN 627003",      "Tirunelveli", 15000,  15, 0.0),
                ("Sakthi Medicals",           "Sakthi R.",       "9876543016", "33AACCS1016A1ZA", "TN-DL-RT-007", "67, North St, Tenkasi, TN 627811",              "Tenkasi",     20000,  21, 1.5),
                ("Priya Drug Store",          "Priya S.",        "9876543017", "33AACCS1017A1ZB", "TN-DL-RT-008", "3, Anna St, Ambasamudram, TN 627401",           "Ambasamudram",15000,  21, 0.0),
                ("Thiruvalluvar Medicals",    "Thiruvalluvar K.","9876543018", "33AACCS1018A1ZC", "TN-DL-RT-009", "9, Kamaraj Nagar, Valliyoor, TN 627117",        "Valliyoor",   20000,  30, 2.0),
                ("HealthPlus Pharmacy",       "Karthik M.",      "9876543019", "33AACCS1019A1ZD", "TN-DL-RT-010", "22, NH 7, Sankarankovil, TN 627756",            "Sankarankovil",25000, 30, 3.0),
            ]
            for s in shops:
                conn.execute(
                    "INSERT INTO ws_shops (name,owner,phone,gst_no,drug_license,address,city,credit_limit,credit_days,discount) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    s
                )

        # Seed demo orders if empty
        if conn.execute("SELECT COUNT(*) FROM ws_orders").fetchone()[0] == 0:
            import random, datetime
            shop_rows = conn.execute("SELECT id,name FROM ws_shops LIMIT 6").fetchall()
            statuses = ['pending','confirmed','picking','dispatched','delivered','invoiced']
            meds = conn.execute("SELECT id,n,g,p FROM medicines WHERE p>0 LIMIT 30").fetchall()
            if meds:
                for i in range(18):
                    shop = shop_rows[i % len(shop_rows)]
                    dt = datetime.datetime.now() - datetime.timedelta(days=random.randint(0,20))
                    ono = f"WS{dt.strftime('%Y%m%d')}{1000+i}"
                    status = statuses[min(i//3, len(statuses)-1)]
                    sel_meds = random.sample(meds, random.randint(3,8))
                    items = [(m, random.randint(12,60)) for m in sel_meds]
                    sub = sum((m["p"] or 0)*0.85*q for m,q in items)
                    gst = sub*0.12
                    total = sub+gst
                    cur = conn.execute(
                        "INSERT INTO ws_orders (order_no,shop_id,shop_name,status,delivery_date,delivery_address,subtotal,gst_amount,total,item_count,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (ono, shop["id"], shop["name"], status,
                         (dt+datetime.timedelta(days=2)).strftime("%Y-%m-%d"),
                         f"Shop address, {shop['name']}", round(sub,2), round(gst,2), round(total,2), len(items), dt.strftime("%Y-%m-%d %H:%M:%S"))
                    )
                    oid = cur.lastrowid
                    for m, qty in items:
                        ws_rate = round((m["p"] or 0)*0.85, 2)
                        conn.execute(
                            "INSERT INTO ws_order_items (order_id,medicine_id,medicine_name,generic_name,qty,mrp,ws_rate,total) VALUES (?,?,?,?,?,?,?,?)",
                            (oid, m["id"], m["n"], m["g"] or "", qty, m["p"] or 0, ws_rate, round(ws_rate*qty, 2))
                        )
                    # timeline entry
                    conn.execute("INSERT INTO ws_order_timeline (order_id,note,icon,color) VALUES (?,?,?,?)",
                                 (oid, f"Order placed for {shop['name']}", "📦", "blue"))
                    if status in ['confirmed','picking','dispatched','delivered','invoiced']:
                        conn.execute("INSERT INTO ws_order_timeline (order_id,note,icon,color) VALUES (?,?,?,?)",
                                     (oid, "Order confirmed by wholesale team", "✓", "green"))
                    if status in ['invoiced','delivered']:
                        conn.execute("INSERT INTO ws_order_timeline (order_id,note,icon,color) VALUES (?,?,?,?)",
                                     (oid, "Goods dispatched", "🚚", "orange"))
                    if status == 'invoiced':
                        inv_no = f"INV{dt.strftime('%Y%m%d')}{1000+i}"
                        conn.execute(
                            "INSERT INTO ws_invoices (invoice_no,order_id,order_no,shop_id,shop_name,subtotal,gst_amount,total,paid_amount) VALUES (?,?,?,?,?,?,?,?,?)",
                            (inv_no, oid, ono, shop["id"], shop["name"], round(sub,2), round(gst,2), round(total,2), round(total,2) if i%3!=0 else 0)
                        )
        conn.commit()


# ── Page routes ──────────────────────────────────────────────────

@app.route("/wholesale-login")
def page_wholesale_login():
    return render_template("wholesale_login.html")


@app.route("/wholesale")
def page_wholesale():
    return render_template("wholesale_portal.html")


# ── Wholesale Shops API ──────────────────────────────────────────

@app.route("/api/ws/shops", methods=["GET"])
def api_ws_shops_list():
    with get_conn() as conn:
        shops = conn.execute("SELECT * FROM ws_shops WHERE is_active=1 ORDER BY name").fetchall()
        result = []
        for s in shops:
            sd = dict(s)
            stats = conn.execute(
                "SELECT COUNT(*) as cnt, COALESCE(SUM(total),0) as val FROM ws_orders WHERE shop_id=?",
                (s["id"],)
            ).fetchone()
            sd["total_orders"] = stats["cnt"]
            sd["total_value"]  = round(stats["val"], 2)
            # outstanding = total invoiced - total paid
            inv = conn.execute(
                "SELECT COALESCE(SUM(total),0) as ti, COALESCE(SUM(paid_amount),0) as tp FROM ws_invoices WHERE shop_id=?",
                (s["id"],)
            ).fetchone()
            sd["outstanding"] = round(inv["ti"] - inv["tp"], 2)
            result.append(sd)
        return jsonify({"shops": result})


@app.route("/api/ws/shops", methods=["POST"])
def api_ws_shops_create():
    d = request.json or {}
    if not d.get("name"):
        return jsonify({"error": "Name required"}), 400
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO ws_shops (name,owner,phone,gst_no,drug_license,address,city,credit_limit,credit_days,discount) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (d["name"], d.get("owner",""), d.get("phone",""), d.get("gst_no",""),
             d.get("drug_license",""), d.get("address",""), d.get("city",""),
             float(d.get("credit_limit",50000)), int(d.get("credit_days",30)),
             float(d.get("discount",0)))
        )
        conn.commit()
        return jsonify({"id": cur.lastrowid, "status": "ok"})


@app.route("/api/ws/shops/<int:sid>", methods=["PUT"])
def api_ws_shops_update(sid):
    d = request.json or {}
    with get_conn() as conn:
        conn.execute("""
            UPDATE ws_shops SET
              name=COALESCE(?,name), owner=COALESCE(?,owner), phone=COALESCE(?,phone),
              gst_no=COALESCE(?,gst_no), drug_license=COALESCE(?,drug_license),
              address=COALESCE(?,address), city=COALESCE(?,city),
              credit_limit=COALESCE(?,credit_limit), credit_days=COALESCE(?,credit_days),
              discount=COALESCE(?,discount)
            WHERE id=?
        """, (d.get("name"), d.get("owner"), d.get("phone"), d.get("gst_no"),
              d.get("drug_license"), d.get("address"), d.get("city"),
              d.get("credit_limit"), d.get("credit_days"), d.get("discount"), sid))
        conn.commit()
        return jsonify({"status": "ok"})


# ── Wholesale Orders API ─────────────────────────────────────────

@app.route("/api/ws/orders", methods=["GET"])
def api_ws_orders_list():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM ws_orders ORDER BY created_at DESC LIMIT 200"
        ).fetchall()
        return jsonify({"orders": [dict(r) for r in rows]})


@app.route("/api/ws/orders", methods=["POST"])
def api_ws_orders_create():
    d = request.json or {}
    if not d.get("shop_id"):
        return jsonify({"error": "shop_id required"}), 400
    items = d.get("items", [])
    if not items:
        return jsonify({"error": "No items"}), 400

    import datetime as dt
    now = dt.datetime.now()
    order_no = f"WS{now.strftime('%Y%m%d%H%M%S')}"

    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO ws_orders
               (order_no,shop_id,shop_name,status,delivery_date,delivery_address,
                notes,subtotal,gst_amount,total,item_count,created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (order_no, d["shop_id"], d.get("shop_name",""),
             "pending", d.get("delivery_date",""), d.get("delivery_address",""),
             d.get("notes",""), round(d.get("subtotal",0),2),
             round(d.get("gst_amount",0),2), round(d.get("total",0),2),
             len(items), now.strftime("%Y-%m-%d %H:%M:%S"))
        )
        oid = cur.lastrowid
        for it in items:
            conn.execute(
                "INSERT INTO ws_order_items (order_id,medicine_id,medicine_name,generic_name,qty,mrp,ws_rate,total) VALUES (?,?,?,?,?,?,?,?)",
                (oid, it.get("med_id"), it["medicine_name"], it.get("generic_name",""),
                 it.get("qty",1), it.get("mrp",0), it.get("ws_rate",0),
                 round((it.get("ws_rate",0))*(it.get("qty",1)),2))
            )
        conn.execute(
            "INSERT INTO ws_order_timeline (order_id,note,icon,color) VALUES (?,?,?,?)",
            (oid, f"Order {order_no} created", "📦", "blue")
        )
        conn.commit()
        return jsonify({"order_no": order_no, "id": oid, "status": "ok"})


@app.route("/api/ws/orders/<int:oid>", methods=["GET"])
def api_ws_orders_detail(oid):
    with get_conn() as conn:
        order = conn.execute("SELECT * FROM ws_orders WHERE id=?", (oid,)).fetchone()
        if not order:
            return jsonify({"error": "Not found"}), 404
        items    = conn.execute("SELECT * FROM ws_order_items WHERE order_id=?", (oid,)).fetchall()
        timeline = conn.execute("SELECT * FROM ws_order_timeline WHERE order_id=? ORDER BY created_at", (oid,)).fetchall()
        return jsonify({
            "order":    dict(order),
            "items":    [dict(i) for i in items],
            "timeline": [dict(t) for t in timeline]
        })


@app.route("/api/ws/orders/<int:oid>/status", methods=["PUT"])
def api_ws_orders_status(oid):
    d = request.json or {}
    new_status = d.get("status","")
    valid = ["pending","confirmed","picking","dispatched","delivered","invoiced"]
    if new_status not in valid:
        return jsonify({"error": "Invalid status"}), 400
    labels = {
        "confirmed":"Order confirmed by wholesale team",
        "picking":"Picking started — warehouse team notified",
        "dispatched":"Goods dispatched via delivery",
        "delivered":"Delivery confirmed by shop",
        "invoiced":"Invoice generated"
    }
    icons  = {"confirmed":"✓","picking":"📋","dispatched":"🚚","delivered":"✅","invoiced":"🧾"}
    colors = {"confirmed":"green","picking":"orange","dispatched":"purple","delivered":"green","invoiced":"blue"}
    with get_conn() as conn:
        conn.execute(
            "UPDATE ws_orders SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (new_status, oid)
        )
        conn.execute(
            "INSERT INTO ws_order_timeline (order_id,note,icon,color) VALUES (?,?,?,?)",
            (oid, labels.get(new_status, f"Status → {new_status}"),
             icons.get(new_status,"•"), colors.get(new_status,"blue"))
        )
        conn.commit()
        return jsonify({"status": "ok"})


@app.route("/api/ws/orders/<int:oid>/invoice", methods=["POST"])
def api_ws_orders_invoice(oid):
    import datetime as dt
    with get_conn() as conn:
        order = conn.execute("SELECT * FROM ws_orders WHERE id=?", (oid,)).fetchone()
        if not order:
            return jsonify({"error": "Not found"}), 404
        inv_no = f"INV{dt.datetime.now().strftime('%Y%m%d%H%M%S')}"
        conn.execute(
            "INSERT INTO ws_invoices (invoice_no,order_id,order_no,shop_id,shop_name,subtotal,gst_amount,total,paid_amount) VALUES (?,?,?,?,?,?,?,?,?)",
            (inv_no, oid, order["order_no"], order["shop_id"], order["shop_name"],
             order["subtotal"], order["gst_amount"], order["total"], 0)
        )
        conn.execute(
            "UPDATE ws_orders SET status='invoiced', updated_at=CURRENT_TIMESTAMP WHERE id=?", (oid,)
        )
        conn.execute(
            "INSERT INTO ws_order_timeline (order_id,note,icon,color) VALUES (?,?,?,?)",
            (oid, f"Invoice {inv_no} generated", "🧾", "blue")
        )
        conn.commit()
        return jsonify({"invoice_no": inv_no, "status": "ok"})


# ── Invoices API ─────────────────────────────────────────────────

@app.route("/api/ws/invoices")
def api_ws_invoices_list():
    month = request.args.get("month", "")
    with get_conn() as conn:
        if month:
            rows = conn.execute(
                "SELECT * FROM ws_invoices WHERE strftime('%Y-%m', invoice_date)=? ORDER BY invoice_date DESC",
                (month,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM ws_invoices ORDER BY invoice_date DESC LIMIT 100"
            ).fetchall()
        return jsonify({"invoices": [dict(r) for r in rows]})


# ── Payments API ─────────────────────────────────────────────────

@app.route("/api/ws/payments", methods=["GET"])
def api_ws_payments_list():
    """List wholesale payments — optional shop_id filter."""
    shop_id = request.args.get("shop_id")
    with get_conn() as conn:
        if shop_id:
            rows = conn.execute(
                "SELECT * FROM ws_payments WHERE shop_id=? ORDER BY created_at DESC LIMIT 200",
                (shop_id,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM ws_payments ORDER BY created_at DESC LIMIT 200"
            ).fetchall()
        return jsonify({"payments": [dict(r) for r in rows]})


@app.route("/api/ws/payments", methods=["POST"])
def api_ws_payments_record():
    d = request.json or {}
    shop_id = d.get("shop_id")
    amount  = float(d.get("amount", 0))
    if not shop_id or amount <= 0:
        return jsonify({"error": "shop_id and amount required"}), 400
    with get_conn() as conn:
        shop = conn.execute("SELECT name FROM ws_shops WHERE id=?", (shop_id,)).fetchone()
        conn.execute(
            "INSERT INTO ws_payments (shop_id,shop_name,amount,mode,reference,notes) VALUES (?,?,?,?,?,?)",
            (shop_id, shop["name"] if shop else "", amount,
             d.get("mode","Cash"), d.get("reference",""), d.get("notes",""))
        )
        # Update oldest unpaid invoice
        unpaid = conn.execute(
            "SELECT id, total, paid_amount FROM ws_invoices WHERE shop_id=? AND paid_amount<total ORDER BY invoice_date",
            (shop_id,)
        ).fetchall()
        remaining = amount
        for inv in unpaid:
            if remaining <= 0:
                break
            owed = inv["total"] - inv["paid_amount"]
            pay  = min(owed, remaining)
            conn.execute(
                "UPDATE ws_invoices SET paid_amount=paid_amount+? WHERE id=?",
                (pay, inv["id"])
            )
            remaining -= pay
        conn.commit()
        return jsonify({"status": "ok"})


# ── Ledger API ───────────────────────────────────────────────────

@app.route("/api/ws/ledger/<int:shop_id>")
def api_ws_ledger(shop_id):
    with get_conn() as conn:
        entries = []
        # Invoices as debits
        invoices = conn.execute(
            "SELECT invoice_date as date, invoice_no, total, paid_amount FROM ws_invoices WHERE shop_id=? ORDER BY invoice_date",
            (shop_id,)
        ).fetchall()
        for inv in invoices:
            entries.append({
                "date": inv["date"], "type": "invoice",
                "description": f"Invoice {inv['invoice_no']}",
                "debit": inv["total"], "credit": 0
            })
            if inv["paid_amount"] > 0:
                entries.append({
                    "date": inv["date"], "type": "payment",
                    "description": f"Payment against {inv['invoice_no']}",
                    "debit": 0, "credit": inv["paid_amount"]
                })
        # Extra payments
        payments = conn.execute(
            "SELECT created_at as date, amount, mode, reference FROM ws_payments WHERE shop_id=? ORDER BY created_at",
            (shop_id,)
        ).fetchall()
        for p in payments:
            entries.append({
                "date": p["date"], "type": "payment",
                "description": f"Payment — {p['mode']} {p['reference'] or ''}".strip(),
                "debit": 0, "credit": p["amount"]
            })
        entries.sort(key=lambda x: x["date"] or "")

        # Summary
        total_inv = sum(e["debit"] for e in entries if e["type"]=="invoice")
        total_paid = sum(e["credit"] for e in entries)
        return jsonify({
            "entries": entries,
            "summary": {
                "total_invoiced": round(total_inv, 2),
                "total_paid":     round(total_paid, 2),
                "outstanding":    round(total_inv - total_paid, 2)
            }
        })


# ── Receivables API ──────────────────────────────────────────────

@app.route("/api/ws/receivables")
def api_ws_receivables():
    import datetime as dt
    today = dt.date.today()
    with get_conn() as conn:
        shops = conn.execute("SELECT * FROM ws_shops WHERE is_active=1").fetchall()
        result = []
        for s in shops:
            inv = conn.execute(
                "SELECT COALESCE(SUM(total),0) ti, COALESCE(SUM(paid_amount),0) tp, MIN(invoice_date) oldest FROM ws_invoices WHERE shop_id=?",
                (s["id"],)
            ).fetchone()
            outstanding = round(inv["ti"] - inv["tp"], 2)
            oldest_date = inv["oldest"]
            overdue_days = 0
            if oldest_date and outstanding > 0:
                try:
                    od = dt.datetime.strptime(oldest_date[:10], "%Y-%m-%d").date()
                    overdue_days = max(0, (today - od).days - s["credit_days"])
                except Exception:
                    pass
            result.append({
                "id": s["id"], "name": s["name"], "city": s["city"],
                "total_invoiced": round(inv["ti"], 2),
                "total_paid":     round(inv["tp"], 2),
                "outstanding":    outstanding,
                "oldest_invoice_date": oldest_date,
                "overdue_days":   overdue_days
            })
        result.sort(key=lambda x: x["outstanding"], reverse=True)
        return jsonify({"shops": result})


# ── Wholesale Stock API ──────────────────────────────────────────

@app.route("/api/ws/stock")
def api_ws_stock():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, n as name, g as generic, c as category, p as mrp, s as stock FROM medicines ORDER BY n LIMIT 500"
        ).fetchall()
        return jsonify({"items": [dict(r) for r in rows]})


# ─── Wholesale stock with limit param ───────────────────────────────
# Override the earlier /api/ws/stock to support limit param
# (already defined above — no duplicate needed)


# ═══════════════════════════════════════════════════════════════════
#  STAGE 11 — SHOP PARTNER PORTAL + WHOLESALE B2B ENHANCEMENTS
# ═══════════════════════════════════════════════════════════════════

def init_stage11_db():
    """Create Stage 11 tables: ws_schemes, ws_returns, ws_delivery_assignments."""
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_schemes (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                title        TEXT NOT NULL,
                description  TEXT,
                scheme_type  TEXT DEFAULT 'buy_x_get_y',
                icon         TEXT DEFAULT '🎁',
                applies_to   TEXT DEFAULT 'all',
                scheme_value TEXT,
                min_order    TEXT,
                valid_from   TEXT,
                valid_till   TEXT,
                eligible     TEXT DEFAULT 'all',
                is_active    INTEGER DEFAULT 1,
                created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_returns (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                return_no     TEXT UNIQUE NOT NULL,
                shop_id       INTEGER,
                shop_name     TEXT,
                order_id      INTEGER,
                order_no      TEXT,
                reason        TEXT,
                items_json    TEXT DEFAULT '[]',
                total_value   REAL DEFAULT 0,
                status        TEXT DEFAULT 'requested',
                credit_note   TEXT,
                created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ws_delivery (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id       INTEGER UNIQUE,
                order_no       TEXT,
                shop_name      TEXT,
                delivery_agent TEXT,
                vehicle_no     TEXT,
                expected_date  TEXT,
                actual_date    TEXT,
                status         TEXT DEFAULT 'scheduled',
                notes          TEXT,
                created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Add shop_code + password columns to ws_shops if not exist
        cols = {r["name"] for r in conn.execute("PRAGMA table_info(ws_shops)").fetchall()}
        if "shop_code" not in cols:
            conn.execute("ALTER TABLE ws_shops ADD COLUMN shop_code TEXT")
        if "shop_password" not in cols:
            conn.execute("ALTER TABLE ws_shops ADD COLUMN shop_password TEXT")
        # Add source column to ws_orders
        ocols = {r["name"] for r in conn.execute("PRAGMA table_info(ws_orders)").fetchall()}
        if "source" not in ocols:
            conn.execute("ALTER TABLE ws_orders ADD COLUMN source TEXT DEFAULT 'wholesale'")
        conn.commit()

        # Seed demo schemes if empty
        if conn.execute("SELECT COUNT(*) FROM ws_schemes").fetchone()[0] == 0:
            demo_schemes = [
                ("Buy 10 Get 1 Free — Antibiotics", "Order 10 strips of any antibiotic and get 1 strip FREE. Valid on all antibiotic strips.", "buy_x_get_y", "🦠", "Antibiotic", "10+1", "10 strips", None, None, "all"),
                ("15% Off on Vitamins & Supplements", "Flat 15% discount on all vitamin and supplement orders above ₹2000.", "percent_off", "🍊", "Vitamin", "15%", "₹2000", None, None, "all"),
                ("Free Sample Pack — New Launches", "Get a free sample pack of 3 new medicines with every order above ₹5000.", "free_sample", "🎁", "all", "3 samples", "₹5000", None, None, "all"),
                ("Diabetes Combo — 5% Extra Off", "Additional 5% off on all antidiabetic medicines for Gold-tier shops.", "percent_off", "🩸", "Diabetes", "5%", "₹1000", None, None, "wholesale"),
            ]
            import datetime as _dt
            till = (_dt.date.today() + _dt.timedelta(days=30)).isoformat()
            frm  = _dt.date.today().isoformat()
            for s in demo_schemes:
                conn.execute(
                    "INSERT INTO ws_schemes (title,description,scheme_type,icon,applies_to,scheme_value,min_order,valid_from,valid_till,eligible) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (s[0],s[1],s[2],s[3],s[4],s[5],s[6],frm,till,s[9])
                )

        # Seed shop codes/passwords if blank
        shops = conn.execute("SELECT id, name FROM ws_shops WHERE shop_code IS NULL OR shop_code=''").fetchall()
        for i, s in enumerate(shops, 1):
            code = f"SHOP{i:03d}"
            pwd  = f"shop@{i:03d}"
            conn.execute("UPDATE ws_shops SET shop_code=?, shop_password=? WHERE id=?", (code, pwd, s["id"]))
        conn.commit()


# ── Page routes ──────────────────────────────────────────────────

@app.route("/shop-login")
def page_shop_login():
    return render_template("shop_login.html")


@app.route("/shop-portal")
def page_shop_portal():
    return render_template("shop_portal.html")


@app.route("/ws-schemes")
def page_ws_schemes():
    return render_template("ws_schemes.html")


# ── Shop Auth API ─────────────────────────────────────────────────

@app.route("/api/ws/shop-auth", methods=["POST"])
def api_ws_shop_auth():
    d = request.json or {}
    code = (d.get("code") or "").strip().upper()
    pwd  = (d.get("password") or "").strip()
    if not code or not pwd:
        return jsonify({"error": "Missing credentials"}), 400
    with get_conn() as conn:
        shop = conn.execute(
            "SELECT * FROM ws_shops WHERE UPPER(shop_code)=? AND shop_password=? AND is_active=1",
            (code, pwd)
        ).fetchone()
        if not shop:
            return jsonify({"error": "Invalid shop code or password"}), 401
        # Return safe shop info (no password)
        sd = dict(shop)
        sd.pop("shop_password", None)
        return jsonify({"shop": sd})


# ── Schemes API ───────────────────────────────────────────────────

@app.route("/api/ws/schemes", methods=["GET"])
def api_ws_schemes_list():
    show_all = request.args.get("all") == "1"
    with get_conn() as conn:
        if show_all:
            rows = conn.execute("SELECT * FROM ws_schemes ORDER BY is_active DESC, created_at DESC").fetchall()
        else:
            rows = conn.execute("SELECT * FROM ws_schemes WHERE is_active=1 ORDER BY created_at DESC").fetchall()
        return jsonify({"schemes": [dict(r) for r in rows]})


@app.route("/api/ws/schemes", methods=["POST"])
def api_ws_schemes_create():
    d = request.json or {}
    if not d.get("title"):
        return jsonify({"error": "Title required"}), 400
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO ws_schemes (title,description,scheme_type,icon,applies_to,scheme_value,min_order,valid_from,valid_till,eligible,is_active) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (d["title"], d.get("description",""), d.get("scheme_type","buy_x_get_y"),
             d.get("icon","🎁"), d.get("applies_to","all"), d.get("scheme_value",""),
             d.get("min_order",""), d.get("valid_from",""), d.get("valid_till",""),
             d.get("eligible","all"), int(d.get("is_active",1)))
        )
        conn.commit()
        return jsonify({"id": cur.lastrowid, "status": "ok"})


@app.route("/api/ws/schemes/<int:sid>", methods=["PUT"])
def api_ws_schemes_update(sid):
    d = request.json or {}
    with get_conn() as conn:
        conn.execute("""
            UPDATE ws_schemes SET
              title=COALESCE(?,title), description=COALESCE(?,description),
              scheme_type=COALESCE(?,scheme_type), icon=COALESCE(?,icon),
              applies_to=COALESCE(?,applies_to), scheme_value=COALESCE(?,scheme_value),
              min_order=COALESCE(?,min_order), valid_from=COALESCE(?,valid_from),
              valid_till=COALESCE(?,valid_till), eligible=COALESCE(?,eligible),
              is_active=COALESCE(?,is_active)
            WHERE id=?
        """, (d.get("title"), d.get("description"), d.get("scheme_type"), d.get("icon"),
              d.get("applies_to"), d.get("scheme_value"), d.get("min_order"),
              d.get("valid_from"), d.get("valid_till"), d.get("eligible"),
              d.get("is_active"), sid))
        conn.commit()
        return jsonify({"status": "ok"})


@app.route("/api/ws/schemes/<int:sid>", methods=["DELETE"])
def api_ws_schemes_delete(sid):
    """Soft-delete (deactivate) a wholesale scheme."""
    with get_conn() as conn:
        cur = conn.execute("UPDATE ws_schemes SET is_active=0 WHERE id=?", (sid,))
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({"status": "error", "message": "Scheme not found"}), 404
    return jsonify({"status": "ok", "deactivated": sid})


# ── Wholesale Orders — filter by shop_id ─────────────────────────
# Patch the existing orders list to support shop_id filter
@app.route("/api/ws/orders/shop/<int:shop_id>", methods=["GET"])
def api_ws_orders_by_shop(shop_id):
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM ws_orders WHERE shop_id=? ORDER BY created_at DESC LIMIT 100",
            (shop_id,)
        ).fetchall()
        return jsonify({"orders": [dict(r) for r in rows]})


# ── Returns API ───────────────────────────────────────────────────

@app.route("/api/ws/returns", methods=["GET"])
def api_ws_returns_list():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM ws_returns ORDER BY created_at DESC LIMIT 100").fetchall()
        return jsonify({"returns": [dict(r) for r in rows]})


@app.route("/api/ws/returns", methods=["POST"])
def api_ws_returns_create():
    import json as _json, datetime as _dt
    d = request.json or {}
    if not d.get("shop_id"):
        return jsonify({"error": "shop_id required"}), 400
    rno = f"RTN{_dt.datetime.now().strftime('%Y%m%d%H%M%S')}"
    with get_conn() as conn:
        shop = conn.execute("SELECT name FROM ws_shops WHERE id=?", (d["shop_id"],)).fetchone()
        cur = conn.execute(
            "INSERT INTO ws_returns (return_no,shop_id,shop_name,order_id,order_no,reason,items_json,total_value,status) VALUES (?,?,?,?,?,?,?,?,?)",
            (rno, d["shop_id"], shop["name"] if shop else "", d.get("order_id"),
             d.get("order_no",""), d.get("reason",""), _json.dumps(d.get("items",[])),
             float(d.get("total_value",0)), "requested")
        )
        conn.commit()
        return jsonify({"return_no": rno, "id": cur.lastrowid, "status": "ok"})


@app.route("/api/ws/returns/<int:rid>/approve", methods=["POST"])
def api_ws_returns_approve(rid):
    import datetime as _dt
    cn = f"CN{_dt.datetime.now().strftime('%Y%m%d%H%M%S')}"
    with get_conn() as conn:
        conn.execute(
            "UPDATE ws_returns SET status='approved', credit_note=? WHERE id=?",
            (cn, rid)
        )
        conn.commit()
        return jsonify({"credit_note": cn, "status": "ok"})


# ── Delivery Assignment API ───────────────────────────────────────

@app.route("/api/ws/delivery", methods=["GET"])
def api_ws_delivery_list():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT d.*, o.shop_name, o.delivery_address, o.total FROM ws_delivery d "
            "LEFT JOIN ws_orders o ON d.order_id=o.id "
            "ORDER BY d.created_at DESC LIMIT 50"
        ).fetchall()
        return jsonify({"deliveries": [dict(r) for r in rows]})


@app.route("/api/ws/delivery", methods=["POST"])
def api_ws_delivery_assign():
    d = request.json or {}
    if not d.get("order_id"):
        return jsonify({"error": "order_id required"}), 400
    with get_conn() as conn:
        order = conn.execute("SELECT order_no, shop_name FROM ws_orders WHERE id=?", (d["order_id"],)).fetchone()
        # Upsert
        existing = conn.execute("SELECT id FROM ws_delivery WHERE order_id=?", (d["order_id"],)).fetchone()
        if existing:
            conn.execute(
                "UPDATE ws_delivery SET delivery_agent=?,vehicle_no=?,expected_date=?,status=?,notes=? WHERE order_id=?",
                (d.get("delivery_agent",""), d.get("vehicle_no",""), d.get("expected_date",""),
                 d.get("status","scheduled"), d.get("notes",""), d["order_id"])
            )
        else:
            conn.execute(
                "INSERT INTO ws_delivery (order_id,order_no,shop_name,delivery_agent,vehicle_no,expected_date,status,notes) VALUES (?,?,?,?,?,?,?,?)",
                (d["order_id"], order["order_no"] if order else "", order["shop_name"] if order else "",
                 d.get("delivery_agent",""), d.get("vehicle_no",""), d.get("expected_date",""),
                 d.get("status","scheduled"), d.get("notes",""))
            )
        conn.commit()
        return jsonify({"status": "ok"})


@app.route("/api/ws/delivery/<int:did>/update", methods=["PUT"])
def api_ws_delivery_update(did):
    d = request.json or {}
    with get_conn() as conn:
        conn.execute(
            "UPDATE ws_delivery SET status=COALESCE(?,status), actual_date=COALESCE(?,actual_date), notes=COALESCE(?,notes) WHERE id=?",
            (d.get("status"), d.get("actual_date"), d.get("notes"), did)
        )
        conn.commit()
        return jsonify({"status": "ok"})


# ── WhatsApp Notification helpers ─────────────────────────────────

@app.route("/api/ws/notify-whatsapp", methods=["POST"])
def api_ws_notify_whatsapp():
    """Generate WhatsApp link to notify shop about order status."""
    d = request.json or {}
    order_no = d.get("order_no","")
    shop_phone = d.get("phone","")
    msg_type = d.get("type","confirmation")  # confirmation|dispatch|invoice|reminder
    templates = {
        "confirmation": f"Hello! Your order *{order_no}* has been confirmed at Selvam Medicals. We'll process and dispatch it shortly. 📦",
        "dispatch":     f"Your order *{order_no}* has been dispatched! 🚚 Expected delivery: {d.get('delivery_date','tomorrow')}. Track your order at our portal.",
        "invoice":      f"Invoice generated for order *{order_no}*. 🧾 Amount: ₹{d.get('amount',0)}. Please arrange payment within {d.get('credit_days',30)} days.",
        "reminder":     f"Friendly reminder 🔔 — Your account with Selvam Medicals has an outstanding balance of ₹{d.get('amount',0)}. Please clear at your earliest convenience.",
    }
    msg = templates.get(msg_type, templates["confirmation"])
    wa_url = f"https://wa.me/{shop_phone}?text={msg}" if shop_phone else None
    return jsonify({"wa_url": wa_url, "message": msg, "status": "ok"})


# ── Outstanding Reminder API ──────────────────────────────────────

@app.route("/api/ws/send-reminders", methods=["POST"])
def api_ws_send_reminders():
    """Generate WhatsApp reminder links for all shops with outstanding > 0."""
    with get_conn() as conn:
        shops = conn.execute(
            "SELECT s.id, s.name, s.phone, "
            "COALESCE(SUM(i.total),0)-COALESCE(SUM(i.paid_amount),0) as outstanding "
            "FROM ws_shops s LEFT JOIN ws_invoices i ON i.shop_id=s.id "
            "GROUP BY s.id HAVING outstanding > 0"
        ).fetchall()
        reminders = []
        for sh in shops:
            msg = f"Dear {sh['name']}, your outstanding balance with Selvam Medicals is ₹{sh['outstanding']:.0f}. Kindly clear at the earliest. Thank you!"
            wa_url = f"https://wa.me/{sh['phone']}?text={msg}" if sh["phone"] else None
            reminders.append({"shop": sh["name"], "phone": sh["phone"], "outstanding": sh["outstanding"], "wa_url": wa_url})
        return jsonify({"reminders": reminders, "count": len(reminders)})


# ── Price Tier API ────────────────────────────────────────────────

@app.route("/api/ws/price-tier/<int:shop_id>")
def api_ws_price_tier(shop_id):
    """Return personalised WS price for a shop based on their discount %."""
    with get_conn() as conn:
        shop = conn.execute("SELECT discount, credit_limit FROM ws_shops WHERE id=?", (shop_id,)).fetchone()
        disc = shop["discount"] if shop else 0
        meds = conn.execute(
            "SELECT id, n as name, g as generic, c as category, p as mrp, s as stock FROM medicines WHERE s>0 ORDER BY n LIMIT 200"
        ).fetchall()
        items = []
        for m in meds:
            base_ws = (m["mrp"] or 0) * 0.85
            shop_ws = round(base_ws * (1 - disc / 100), 2)
            items.append({"id": m["id"], "name": m["name"], "generic": m["generic"], "category": m["category"],
                          "mrp": m["mrp"], "ws_rate": shop_ws, "discount_applied": disc, "stock": m["stock"]})
        return jsonify({"items": items, "discount": disc})


# ── Patch ws/orders list to support shop_id filter param ─────────
# (Frontend shop_portal calls /api/ws/orders?shop_id=X)
_orig_ws_orders_list = api_ws_orders_list.__wrapped__ if hasattr(api_ws_orders_list,'__wrapped__') else None

# Extend existing orders route: if shop_id param present, filter
# We do this by rebuilding the handler:
app.view_functions.pop('api_ws_orders_list', None)

@app.route("/api/ws/orders", methods=["GET"])
def api_ws_orders_list():          # noqa: F811
    shop_id = request.args.get("shop_id")
    with get_conn() as conn:
        if shop_id:
            rows = conn.execute(
                "SELECT * FROM ws_orders WHERE shop_id=? ORDER BY created_at DESC LIMIT 200",
                (int(shop_id),)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM ws_orders ORDER BY created_at DESC LIMIT 200"
            ).fetchall()
        return jsonify({"orders": [dict(r) for r in rows]})


# ── Patch ws/stock to support limit param ────────────────────────
app.view_functions.pop('api_ws_stock', None)

@app.route("/api/ws/stock")
def api_ws_stock():                # noqa: F811
    limit = min(int(request.args.get("limit", 500)), 5000)
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, n as name, g as generic, c as category, p as mrp, s as stock FROM medicines ORDER BY n LIMIT ?",
            (limit,)
        ).fetchall()
        return jsonify({"items": [dict(r) for r in rows]})



# ════════════════════════════════════════════════════════════════════════════════
#  STAGE 12 — Pricing Intelligence + Symptom Advisor
# ════════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════════

# ── Symptom → Medicine Knowledge Base (Indian pharmacy brand-name keywords) ──
# keywords: name fragments matched via SQL LIKE UPPER(n)
_SYMPTOM_MEDICINE_MAP = {
    "fever": {
        "label": "Fever",
        "keywords": ["DOLO","CROCIN","COMBIFLAM","PARACETAMOL","CALPOL","BRUFEN","NIMULID"],
        "confidence": 92,
        "safety": {
            "pregnant":      {"block": ["BRUFEN","COMBIFLAM","NIMULID"], "warn": []},
            "child_under5":  {"block": ["NIMULID"], "warn": ["BRUFEN"]},
            "renal":         {"block": ["BRUFEN","NIMULID"], "warn": []},
            "allergic_nsaid":{"block": ["BRUFEN","COMBIFLAM","NIMULID","VOVERAN","DYNAPAR"], "warn": []},
        }
    },
    "cough_dry": {
        "label": "Dry Cough",
        "keywords": ["ALEX","COFSILS","TOSSEX","KOFEX","PHENSEDYL","BENADRYL","HONITUS","SEPTILIN"],
        "confidence": 85,
        "safety": {
            "pregnant":     {"block": ["PHENSEDYL"], "warn": ["BENADRYL"]},
            "child_under5": {"block": ["PHENSEDYL"], "warn": []},
        }
    },
    "cough_wet": {
        "label": "Productive Cough",
        "keywords": ["AMBROLITE","MUCOLITE","ALEX","BROMHEXINE","AMBROX","ASCORIL","GRILINCTUS"],
        "confidence": 87,
        "safety": {}
    },
    "cold": {
        "label": "Cold / Runny Nose",
        "keywords": ["ALEX","HH LEVO","TELEKAST","MONTAIR","CETRIZ","ALLEGRA","SINAREST","NASIVION"],
        "confidence": 83,
        "safety": {
            "pregnant":     {"block": [], "warn": ["SINAREST"]},
            "hypertensive": {"block": ["SINAREST"], "warn": []},
            "bp_high":      {"block": ["SINAREST"], "warn": []},
        }
    },
    "headache": {
        "label": "Headache",
        "keywords": ["DOLO","CROCIN","COMBIFLAM","BRUFEN","MEFTAL","SARIDON","ANACIN","SUMO"],
        "confidence": 89,
        "safety": {
            "pregnant":      {"block": ["BRUFEN","COMBIFLAM","MEFTAL"], "warn": []},
            "renal":         {"block": ["BRUFEN","MEFTAL"], "warn": []},
            "allergic_nsaid":{"block": ["BRUFEN","COMBIFLAM","MEFTAL","VOVERAN"], "warn": []},
        }
    },
    "migraine": {
        "label": "Migraine",
        "keywords": ["SUMINAT","SUMO","RIZACT","ERGOT","DOLO","BRUFEN","CAFERGOT"],
        "confidence": 78,
        "safety": {
            "pregnant": {"block": ["SUMINAT","CAFERGOT","ERGOT"], "warn": []},
            "bp_high":  {"block": [], "warn": ["CAFERGOT","ERGOT"]},
        }
    },
    "body_pain": {
        "label": "Body / Muscle Pain",
        "keywords": ["VOVERAN","DYNAPAR","COMBIFLAM","BRUFEN","MEFTAL","DOLO","ACECLOF","THIOCOLCHICOSID"],
        "confidence": 84,
        "safety": {
            "pregnant":      {"block": ["VOVERAN","DYNAPAR","BRUFEN","MEFTAL"], "warn": []},
            "renal":         {"block": ["VOVERAN","DYNAPAR","BRUFEN","MEFTAL"], "warn": []},
            "allergic_nsaid":{"block": ["VOVERAN","DYNAPAR","BRUFEN","MEFTAL","ACECLOF"], "warn": []},
        }
    },
    "sore_throat": {
        "label": "Sore Throat",
        "keywords": ["BETADINE GARGLE","STREPSILS","COFSILS","ALEX LOZENGES","AZITHRAL","AUGMENTIN","DOLO"],
        "confidence": 81,
        "safety": {
            "pregnant": {"block": [], "warn": ["AZITHRAL"]},
            "renal":    {"block": [], "warn": ["AUGMENTIN"]},
        }
    },
    "stomach_pain": {
        "label": "Stomach Pain / Cramps",
        "keywords": ["MEFTAL SPAS","BUSCOPAN","DROTIKA","PANTOP","OMEZ","NEXPRO"],
        "confidence": 77,
        "safety": {
            "pregnant": {"block": ["MEFTAL","BUSCOPAN"], "warn": []},
        }
    },
    "acidity": {
        "label": "Acidity / Heartburn",
        "keywords": ["PANTOP","OMEZ","NEXPRO","RABLET","RANTAC","ACILOC","GAVISCON","GELUSIL","DIGENE"],
        "confidence": 93,
        "safety": {}
    },
    "nausea": {
        "label": "Nausea / Vomiting",
        "keywords": ["DOMSTAL","VOMIKIND","EMESET","PERINORM","ONDANSET","AVOMINE","GRAVOL"],
        "confidence": 89,
        "safety": {
            "pregnant":     {"block": ["PERINORM"], "warn": ["DOMSTAL"]},
            "child_under5": {"block": ["AVOMINE"], "warn": ["PERINORM"]},
        }
    },
    "diarrhea": {
        "label": "Diarrhea",
        "keywords": ["ELECTRAL","NORFLOX","METROGYL","ECONORM","LOPERAMIDE","RACIPER","ENTEROGERMINA"],
        "confidence": 86,
        "safety": {
            "child_under5": {"block": ["LOPERAMIDE"], "warn": ["NORFLOX"]},
        }
    },
    "constipation": {
        "label": "Constipation",
        "keywords": ["DULCOFLEX","CASTOR OIL","CREMAFFIN","DUPHALAC","ISABGOL","GLYCERIN"],
        "confidence": 88,
        "safety": {
            "pregnant": {"block": ["DULCOFLEX"], "warn": []},
        }
    },
    "skin_rash": {
        "label": "Skin Rash / Itching",
        "keywords": ["CALAMINE","CANDID B","HH LEVO","CETRIZ","ALLEGRA","BETNESOL","DERMAC"],
        "confidence": 74,
        "safety": {
            "child_under5": {"block": [], "warn": ["BETNESOL"]},
        }
    },
    "eye_irritation": {
        "label": "Eye Irritation",
        "keywords": ["OPTIVE","TEARS NATURALE","TOBREX","SYSTANE","REFRESH","GENTEAL"],
        "confidence": 84,
        "safety": {}
    },
    "ear_pain": {
        "label": "Ear Pain",
        "keywords": ["WAXSOL","EAREX","OTOREX","CANDIBIOTIC","SOFRADEX"],
        "confidence": 80,
        "safety": {}
    },
    "toothache": {
        "label": "Toothache",
        "keywords": ["DOLO","BRUFEN","COMBIFLAM","MEFTAL","VOVERAN","BETADINE GARGLE","CLOVE"],
        "confidence": 82,
        "safety": {
            "renal":         {"block": ["BRUFEN","MEFTAL","VOVERAN"], "warn": []},
            "allergic_nsaid":{"block": ["BRUFEN","MEFTAL","VOVERAN","COMBIFLAM"], "warn": []},
        }
    },
    "uti": {
        "label": "UTI / Burning Urination",
        "keywords": ["NORFLOX","CIPRODAC","BACTRIM","NITROFURANTOIN","FOSFOMYCIN"],
        "confidence": 81,
        "safety": {
            "pregnant": {"block": ["NORFLOX","CIPRODAC","BACTRIM"], "warn": []},
            "renal":    {"block": ["NITROFURANTOIN"], "warn": ["CIPRODAC"]},
        }
    },
    "allergy": {
        "label": "Allergy",
        "keywords": ["MONTAIR","TELEKAST","HH LEVO","CETRIZ","ALLEGRA","BETNESOL","PREDNISOLONE"],
        "confidence": 87,
        "safety": {
            "diabetic": {"block": [], "warn": ["BETNESOL","PREDNISOLONE"]},
            "diabetes": {"block": [], "warn": ["BETNESOL","PREDNISOLONE"]},
        }
    },
    "back_pain": {
        "label": "Back Pain",
        "keywords": ["VOVERAN","DYNAPAR","COMBIFLAM","BRUFEN","ACECLOF","THIOCOLCHICOSID","MEFTAL"],
        "confidence": 83,
        "safety": {
            "renal":         {"block": ["VOVERAN","DYNAPAR","BRUFEN","MEFTAL"], "warn": []},
            "allergic_nsaid":{"block": ["VOVERAN","DYNAPAR","BRUFEN","MEFTAL","ACECLOF"], "warn": []},
        }
    },
    "joint_pain": {
        "label": "Joint Pain",
        "keywords": ["VOVERAN","DYNAPAR","BRUFEN","MEFTAL","NUCOXIA","HIFENAC","ACECLOF","GLUCOSAMINE"],
        "confidence": 80,
        "safety": {
            "renal":         {"block": ["VOVERAN","DYNAPAR","BRUFEN","MEFTAL","NUCOXIA"], "warn": []},
            "allergic_nsaid":{"block": ["VOVERAN","DYNAPAR","BRUFEN","NUCOXIA","HIFENAC"], "warn": []},
        }
    },
    "insomnia": {
        "label": "Insomnia / Poor Sleep",
        "keywords": ["RESTYL","TRIKA","MELATONIN","BENADRYL","ALPRAX"],
        "confidence": 74,
        "safety": {
            "pregnant":     {"block": ["RESTYL","TRIKA","ALPRAX"], "warn": []},
            "child_under5": {"block": ["RESTYL","TRIKA","ALPRAX"], "warn": []},
        }
    },
    "anxiety": {
        "label": "Anxiety / Stress",
        "keywords": ["RESTYL","TRIKA","ALPRAX","NEXITO","STALOPAM","SERTRALINE"],
        "confidence": 70,
        "safety": {
            "pregnant":     {"block": ["RESTYL","TRIKA","ALPRAX"], "warn": []},
            "child_under5": {"block": ["RESTYL","TRIKA","ALPRAX"], "warn": []},
        }
    },
    "wound": {
        "label": "Wound / Cut",
        "keywords": ["BETADINE","CIPLADINE","SOFRAMYCIN","POVIDONE","BACTROBAN","SAVLON"],
        "confidence": 92,
        "safety": {}
    },
    "fungal": {
        "label": "Fungal Infection",
        "keywords": ["CANDID","TERBINA","HH ZOLE","CLOTRIM","NIZORAL","FLUCONAZOLE","ITRACONAZOLE","ONABET"],
        "confidence": 87,
        "safety": {
            "pregnant": {"block": ["FLUCONAZOLE","ITRACONAZOLE"], "warn": ["CANDID"]},
            "liver":    {"block": ["NIZORAL","ITRACONAZOLE"], "warn": ["FLUCONAZOLE"]},
        }
    },
    "vitamin_def": {
        "label": "Vitamin / Mineral Deficiency",
        "keywords": ["BECOSULES","BECADEX","NEUROBION","REVITAL","LIMCEE",
                     "CALCIROL","SHELCAL","ZINCOVIT","CALDIKIND"],
        "confidence": 89,
        "safety": {}
    },
    "menstrual": {
        "label": "Menstrual Cramps",
        "keywords": ["MEFTAL SPAS","MEFTAL","DROTIKA","DOLO","BRUFEN","REGESTRONE","PRIMOLUT"],
        "confidence": 85,
        "safety": {
            "renal":         {"block": ["MEFTAL","BRUFEN"], "warn": []},
            "allergic_nsaid":{"block": ["MEFTAL","BRUFEN"], "warn": []},
        }
    },
    "bp_high": {
        "label": "High Blood Pressure",
        "keywords": ["AMLODAC","TELMISIM","LOSARTAR","ATENOLOL","METOPROLOL","TELMA","STAMLO","LOSAR"],
        "confidence": 71,
        "safety": {
            "pregnant": {"block": ["TELMISIM","LOSARTAR","TELMA","LOSAR"], "warn": ["AMLODAC"]},
        }
    },
    "diabetes": {
        "label": "Diabetes",
        "keywords": ["GLYCOMET","GALVUS","JANUVIA","TRAJENTA","GLUCOBAY","INSULIN","HUMINSULIN","GLIMEPIRIDE"],
        "confidence": 71,
        "safety": {
            "renal": {"block": ["GLYCOMET"], "warn": ["GLIMEPIRIDE"]},
        }
    },
    "insect_bite": {
        "label": "Insect Bite",
        "keywords": ["CALAMINE","HH LEVO","CETRIZ","ALLEGRA","BETNESOL CREAM","SOFRAMYCIN","BETADINE"],
        "confidence": 86,
        "safety": {}
    },
}

# Condition display labels
_CONDITION_WARNINGS = {
    "pregnant":      "Avoid in pregnancy",
    "diabetic":      "Monitor blood sugar",
    "diabetes":      "Monitor blood sugar",
    "hypertensive":  "Check BP impact",
    "bp_high":       "Check BP impact",
    "child_under5":  "Paediatric dose needed",
    "renal":         "Renal impairment",
    "liver":         "Hepatic impairment",
    "allergic_pcm":  "Paracetamol allergy",
    "allergic_nsaid":"NSAID allergy",
    "breastfeeding": "Check compatibility",
}


def _apply_safety(med_name, conditions, selected_symptoms):
    warnings, cautions = [], []
    name_up = med_name.upper()
    for sym_id, info in selected_symptoms.items():
        for cond in conditions:
            rules = info["safety"].get(cond, {})
            for frag in rules.get("block", []):
                if frag.upper() in name_up:
                    label = _CONDITION_WARNINGS.get(cond, cond)
                    msg = f"⚠ {label} — avoid {frag}"
                    if msg not in warnings:
                        warnings.append(msg)
            for frag in rules.get("warn", []):
                if frag.upper() in name_up:
                    label = _CONDITION_WARNINGS.get(cond, cond)
                    msg = f"⚡ {label} with {frag}"
                    if msg not in cautions:
                        cautions.append(msg)
    return warnings, cautions


def _score_medicine_sql(med_name, selected_symptoms, conditions,
                        age_group, gender, severity, duration):
    name_up = med_name.upper()
    matched_syms, total_conf, max_conf = [], 0, 0
    for sym_id, info in selected_symptoms.items():
        base = info["confidence"]
        max_conf += base
        hits = [kw for kw in info["keywords"] if kw.upper() in name_up]
        if hits:
            total_conf += base + min(10, len(hits) * 5)
            matched_syms.append(sym_id)
    if max_conf == 0 or total_conf == 0:
        return 0, []
    conf = min(97, int((total_conf / max_conf) * 100))
    if severity == "severe":
        conf = min(97, conf + 5)
    if duration in ("7+", "4-7"):
        conf = min(97, conf + 2)
    if age_group in ("infant", "child"):
        if any(f in name_up for f in ["SYP","SUS","DROPS","DT ","JUNIOR","PAED"]):
            conf = min(97, conf + 8)
        if "INJ" in name_up:
            conf = max(10, conf - 15)
    elif age_group == "senior" and "INJ" in name_up:
        conf = max(10, conf - 5)
    if gender == "female" and any(s in ["menstrual","uti","vitamin_def"] for s in matched_syms):
        conf = min(97, conf + 3)
    return conf, matched_syms


# ── Page routes ────────────────────────────────────────────────────────────

@app.route("/pricing-engine")
def page_pricing_engine():
    return render_template("pricing_engine.html")


@app.route("/symptom-advisor")
def page_symptom_advisor():
    return render_template("symptom_advisor.html")


# ── Symptom API ────────────────────────────────────────────────────────────

@app.route("/api/symptom/suggest", methods=["POST"])
def api_symptom_suggest():
    data         = request.get_json(force=True) or {}
    symptoms_sel = data.get("symptoms", [])
    conditions   = data.get("conditions", [])
    age_group    = data.get("age_group", "adult")
    gender       = data.get("gender", "other")
    severity     = data.get("severity", "mild")
    duration     = data.get("duration", "1-3")

    if not symptoms_sel:
        return jsonify({"medicines": [], "alert": "Please select at least one symptom."})

    selected = {s: _SYMPTOM_MEDICINE_MAP[s]
                for s in symptoms_sel if s in _SYMPTOM_MEDICINE_MAP}
    if not selected:
        return jsonify({"medicines": [], "alert": "Unknown symptom IDs."})

    # Collect unique keywords across selected symptoms
    all_kws, seen_kw = [], set()
    for info in selected.values():
        for kw in info["keywords"]:
            if kw.upper() not in seen_kw:
                seen_kw.add(kw.upper())
                all_kws.append(kw)

    if not all_kws:
        return jsonify({"medicines": [], "alert": "No keywords for selected symptoms."})

    # Pre-filter via SQL LIKE (case-insensitive)
    like_clauses = " OR ".join(["UPPER(n) LIKE ?" for _ in all_kws])
    params = [f"%{kw.upper()}%" for kw in all_kws]

    with get_conn() as conn:
        sql = (
            f"SELECT id, n as name, g as generic, c as category, p as mrp, s as stock "
            f"FROM medicines WHERE ({like_clauses}) "
            f"ORDER BY n LIMIT 500"
        )
        rows = conn.execute(sql, params).fetchall()

    results = []
    seen_names = set()
    for row in rows:
        med = dict(row)
        mname = med["name"]
        if mname in seen_names:
            continue
        seen_names.add(mname)

        conf, matched_syms = _score_medicine_sql(
            mname, selected, conditions, age_group, gender, severity, duration
        )
        if conf < 35:
            continue

        warnings, cautions = _apply_safety(mname, conditions, selected)
        if warnings:
            conf = min(conf, 40)

        sym_labels = [selected[s]["label"] for s in matched_syms if s in selected]
        results.append({
            "name": mname,
            "generic": med.get("generic") or "",
            "category": med.get("category") or "",
            "mrp": round(float(med.get("mrp") or 0), 2),
            "stock": int(med.get("stock") or 0),
            "confidence": conf,
            "matched_symptoms": matched_syms,
            "reason": f"Matches: {', '.join(sym_labels)}" if sym_labels else "Keyword match",
            "warnings": warnings,
            "cautions": cautions,
        })

    results.sort(key=lambda x: (-x["confidence"], x["name"]))

    return jsonify({
        "medicines": results[:80],
        "total_found": len(results),
        "symptoms_queried": list(selected.keys()),
    })


# ── Purpose-based medicine categories ─────────────────────────────────────
_PURPOSE_KEYWORDS = {
    "bp": {
        "label": "BP / Hypertension", "emoji": "💓", "color": "#ef4444",
        "generics": ["amlodipine","atenolol","losartan","telmisartan","enalapril","ramipril",
                     "lisinopril","metoprolol","bisoprolol","nifedipine","verapamil",
                     "hydrochlorothiazide","furosemide","spironolactone","clonidine"],
        "categories": ["Antihypertensives","Calcium Channel Blockers","Beta Blockers",
                       "ACE Inhibitors","ARBs","Diuretics"],
    },
    "diabetes": {
        "label": "Diabetes / Sugar", "emoji": "🩸", "color": "#06b6d4",
        "generics": ["metformin","glimepiride","glibenclamide","gliclazide","sitagliptin",
                     "voglibose","dapagliflozin","empagliflozin","insulin","pioglitazone",
                     "repaglinide","acarbose","liraglutide"],
        "categories": ["Antidiabetics","Oral Hypoglycaemics","Insulin","DPP-4 Inhibitors",
                       "SGLT2 Inhibitors","GLP-1 Agonists"],
    },
    "immediate": {
        "label": "Immediate Relief", "emoji": "⚡", "color": "#f97316",
        "generics": ["paracetamol","ibuprofen","diclofenac","dicyclomine","drotaverine",
                     "ondansetron","domperidone","ranitidine","aluminum hydroxide","loperamide",
                     "cetirizine","chlorpheniramine","salbutamol","ors","antacid"],
        "categories": ["Analgesics","Antacids","Antiemetics","Antispasmodics",
                       "Antihistamines","ORS","Antipyretics"],
    },
    "life_saving": {
        "label": "Life-Saving Drugs", "emoji": "🚨", "color": "#a855f7",
        "generics": ["adrenaline","epinephrine","atropine","digoxin","nitroglycerin",
                     "hydrocortisone","dexamethasone","dopamine","norepinephrine",
                     "magnesium sulphate","furosemide","mannitol","activated charcoal",
                     "naloxone","sodium bicarbonate","glucose","glucagon","aspirin"],
        "categories": ["Emergency","Cardiac","Corticosteroids","Resuscitation",
                       "Vasopressors","Antidotes","Anticoagulants"],
    },
    "doctor_rx": {
        "label": "Doctor / Rx Only", "emoji": "📋", "color": "#3b82f6",
        "generics": ["amoxicillin","azithromycin","ciprofloxacin","levofloxacin",
                     "metronidazole","fluconazole","tramadol","morphine","codeine",
                     "alprazolam","clonazepam","warfarin","phenytoin","carbamazepine",
                     "prednisolone","dexamethasone","levothyroxine"],
        "categories": ["Antibiotics","Antifungals","Opioids","Benzodiazepines",
                       "Anticoagulants","Antiepileptics","Steroids","Schedule H",
                       "Schedule H1","Schedule X"],
    },
    "otc": {
        "label": "OTC / No Rx", "emoji": "🌿", "color": "#22c55e",
        "generics": ["paracetamol","aspirin","cetirizine","loratadine","antacid",
                     "ors","vitamins","calcium","iron","zinc","povidone iodine",
                     "clotrimazole","permethrin","calamine","antiseptic"],
        "categories": ["Vitamins & Supplements","OTC Analgesics","Antiseptics",
                       "Antacids OTC","ORS","Antipyretics","First Aid","Antihistamines OTC"],
    },
}


@app.route("/api/symptom/medicines-by-purpose", methods=["GET"])
def api_medicines_by_purpose():
    """Return all stock medicines tagged by purpose category."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, n as name, g as generic, c as category, p as mrp, s as stock "
            "FROM medicines ORDER BY n LIMIT 3000"
        ).fetchall()

    all_meds = [dict(r) for r in rows]
    result = {}

    for pkey, meta in _PURPOSE_KEYWORDS.items():
        matched = []
        for m in all_meds:
            name_l    = (m["name"] or "").lower()
            generic_l = (m["generic"] or "").lower()
            cat_l     = (m["category"] or "").lower()
            hit = any(g in generic_l or g in name_l for g in meta["generics"])
            hit = hit or any(c.lower() in cat_l for c in meta["categories"])
            if hit:
                matched.append({
                    "id":       m["id"],
                    "name":     m["name"],
                    "generic":  m["generic"] or "",
                    "category": m["category"] or "",
                    "mrp":      float(m["mrp"] or 0),
                    "stock":    int(m["stock"] or 0),
                    "in_stock": int(m["stock"] or 0) > 0,
                })
        result[pkey] = {
            "label":    meta["label"],
            "emoji":    meta["emoji"],
            "color":    meta["color"],
            "count":    len(matched),
            "in_stock": sum(1 for x in matched if x["in_stock"]),
            "medicines": sorted(matched, key=lambda x: (-x["stock"], x["name"]))[:100],
        }

    all_list = [{
        "id":       m["id"],
        "name":     m["name"],
        "generic":  m["generic"] or "",
        "category": m["category"] or "",
        "mrp":      float(m["mrp"] or 0),
        "stock":    int(m["stock"] or 0),
        "in_stock": int(m["stock"] or 0) > 0,
    } for m in all_meds]
    result["all"] = {
        "label":    "All Medicines",
        "emoji":    "💊",
        "color":    "#64748b",
        "count":    len(all_list),
        "in_stock": sum(1 for x in all_list if x["in_stock"]),
        "medicines": sorted(all_list, key=lambda x: (-x["stock"], x["name"]))[:200],
    }
    return jsonify(result)


@app.route("/api/symptom/protocols", methods=["GET"])
def api_get_protocols():
    """Get saved custom symptom-medicine protocols."""
    with get_conn() as conn:
        try:
            rows = conn.execute(
                "SELECT * FROM symptom_protocols ORDER BY id DESC LIMIT 200"
            ).fetchall()
            return jsonify([dict(r) for r in rows])
        except Exception:
            return jsonify([])


@app.route("/api/symptom/protocols", methods=["POST"])
def api_save_protocol():
    """Save a custom symptom-medicine protocol."""
    data = request.get_json(silent=True) or {}
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS symptom_protocols (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT NOT NULL,
                purpose     TEXT DEFAULT 'otc',
                symptoms    TEXT DEFAULT '[]',
                medicines   TEXT DEFAULT '[]',
                age_group   TEXT DEFAULT 'all',
                notes       TEXT DEFAULT '',
                created_at  TEXT,
                created_by  TEXT DEFAULT ''
            )
        """)
        now = datetime.now(timezone.utc).isoformat()
        conn.execute("""
            INSERT INTO symptom_protocols
            (name,purpose,symptoms,medicines,age_group,notes,created_at,created_by)
            VALUES (?,?,?,?,?,?,?,?)
        """, (
            data.get("name","Unnamed"),
            data.get("purpose","otc"),
            json.dumps(data.get("symptoms",[])),
            json.dumps(data.get("medicines",[])),
            data.get("age_group","all"),
            data.get("notes",""),
            now,
            data.get("created_by","staff"),
        ))
        new_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return jsonify({"status":"saved","id":new_id})


@app.route("/api/symptom/protocols/<int:pid>", methods=["DELETE"])
def api_delete_protocol(pid):
    with get_conn() as conn:
        try:
            conn.execute("DELETE FROM symptom_protocols WHERE id=?", (pid,))
        except Exception:
            pass
    return jsonify({"status":"deleted"})


# ── Pricing Intelligence APIs ──────────────────────────────────────────────

def _get_abc_data(conn):
    """Compute ABC analysis from medicines table."""
    rows = conn.execute(
        "SELECT id, n as name, g as generic, c as category, p as mrp, s as stock "
        "FROM medicines WHERE s > 0 ORDER BY (p * s) DESC"
    ).fetchall()
    total_value = sum((r["mrp"] or 0) * (r["stock"] or 0) for r in rows)
    if total_value == 0:
        return [], {"A": 0, "B": 0, "C": 0}

    cum = 0
    abc_list = []
    a_count = b_count = c_count = 0
    for i, r in enumerate(rows):
        sv = (r["mrp"] or 0) * (r["stock"] or 0)
        cum += sv
        cum_pct = (cum / total_value) * 100
        if cum_pct <= 70:
            cls = "A"
            a_count += 1
        elif cum_pct <= 90:
            cls = "B"
            b_count += 1
        else:
            cls = "C"
            c_count += 1
        abc_list.append({
            "rank": i + 1,
            "name": r["name"],
            "generic": r["generic"] or "",
            "category": r["category"] or "",
            "stock": r["stock"],
            "mrp": round(r["mrp"] or 0, 2),
            "stock_value": round(sv, 2),
            "cum_pct": round(cum_pct, 2),
            "abc_class": cls,
        })

    return abc_list, {"A": a_count, "B": b_count, "C": c_count}


def _get_velocity_data(conn):
    """Compute velocity from bill_items + bills (last 90 days)."""
    import datetime as _dt
    cutoff = (_dt.date.today() - _dt.timedelta(days=90)).isoformat()

    try:
        sale_rows = conn.execute("""
            SELECT bi.name, SUM(bi.qty) as total_sold
            FROM bill_items bi
            JOIN bills b ON b.id = bi.bill_id
            WHERE b.created_at >= ?
            GROUP BY bi.name
        """, (cutoff,)).fetchall()
        sales_map = {r["name"]: r["total_sold"] for r in sale_rows}
    except Exception:
        sales_map = {}

    meds = conn.execute(
        "SELECT n as name, c as category, s as stock FROM medicines WHERE s >= 0 LIMIT 3000"
    ).fetchall()

    velocity_list = []
    fast = medium = slow = dead = 0

    for m in meds:
        total = sales_map.get(m["name"], 0)
        daily_vel = round(total / 90, 3)
        days_cover = round(m["stock"] / daily_vel, 1) if daily_vel > 0 else 9999

        if daily_vel > 5:
            band = "Fast"
            fast += 1
        elif daily_vel >= 1:
            band = "Medium"
            medium += 1
        elif daily_vel > 0:
            band = "Slow"
            slow += 1
        else:
            band = "Dead"
            dead += 1

        velocity_list.append({
            "name": m["name"],
            "category": m["category"] or "",
            "stock": m["stock"],
            "velocity": daily_vel,
            "days_cover": days_cover if days_cover < 9999 else None,
            "band": band,
        })

    velocity_list.sort(key=lambda x: -x["velocity"])
    return velocity_list, {"fast": fast, "medium": medium, "slow": slow, "dead": dead}


def _get_recommendations(abc_list, velocity_list, conn):
    """Generate price recommendations by combining ABC + velocity."""
    vel_map = {v["name"]: v for v in velocity_list}
    recs = []

    for item in abc_list[:500]:  # only top 500 by value
        v = vel_map.get(item["name"])
        if not v:
            continue

        band = v["band"]
        cls = item["abc_class"]
        mrp = item["mrp"]
        if mrp <= 0:
            continue

        if cls == "A" and band == "Fast":
            # High value, fast moving → nudge margin up slightly
            rec_type = "margin"
            rec_label = "Increase Margin"
            reason = "A-class fast-mover — can carry 2–3% margin increase"
            suggested = round(mrp * 1.03, 2)
            direction = "up"
            change_pct = 3.0
        elif cls == "C" and band == "Dead":
            # Low value, dead stock → clearance
            rec_type = "clear"
            rec_label = "Clearance Offer"
            reason = "C-class dead stock — offer 15% discount to liquidate"
            suggested = round(mrp * 0.85, 2)
            direction = "down"
            change_pct = -15.0
        elif band == "Slow" and item["stock"] > 50:
            # Slow moving with high stock → discount push
            rec_type = "discount"
            rec_label = "Discount Push"
            reason = "Slow-mover with excess stock — 8% off to stimulate demand"
            suggested = round(mrp * 0.92, 2)
            direction = "down"
            change_pct = -8.0
        elif cls == "B" and band == "Medium":
            # Combo opportunity
            rec_type = "combo"
            rec_label = "Bundle Offer"
            reason = "B-class medium-mover — bundle with complementary product"
            suggested = round(mrp * 0.95, 2)
            direction = "down"
            change_pct = -5.0
        elif band == "Fast" and v["days_cover"] and v["days_cover"] < 14:
            # Fast but low stock → reorder, no price change but flag
            rec_type = "reorder"
            rec_label = "Reorder Alert"
            reason = f"Only {v['days_cover']} days cover left — reorder urgently"
            suggested = mrp
            direction = "neutral"
            change_pct = 0
        else:
            continue

        recs.append({
            "name": item["name"],
            "category": item["category"],
            "type": rec_type,
            "type_label": rec_label,
            "reason": reason,
            "current_price": mrp,
            "suggested_price": suggested,
            "direction": direction,
            "change_pct": change_pct,
            "abc_class": cls,
            "velocity_band": band,
            "stock": item["stock"],
        })

    recs.sort(key=lambda x: abs(x["change_pct"]), reverse=True)
    return recs[:100]


def _get_category_breakdown(conn):
    rows = conn.execute("""
        SELECT c as category,
               COUNT(*) as sku_count,
               SUM(s) as total_stock,
               ROUND(SUM(p * s), 2) as stock_value,
               ROUND(AVG(p), 2) as avg_mrp
        FROM medicines
        WHERE s > 0
        GROUP BY c
        ORDER BY stock_value DESC
        LIMIT 20
    """).fetchall()
    return [dict(r) for r in rows]


def _get_forecast(velocity_list, conn):
    """Simple forecast: 30-day projection with ±20% seasonal variation."""
    import datetime as _dt
    import math as _math

    month = _dt.date.today().month
    # Rough seasonal indices (pharmacy demand peaks in monsoon/winter)
    seasonal = [0.85, 0.80, 0.90, 0.95, 1.00, 1.05, 1.20, 1.25, 1.15, 1.10, 1.05, 1.00]
    idx = seasonal[month - 1]

    top_fast = [v for v in velocity_list if v["band"] == "Fast"][:20]
    forecast_cards = []
    for v in top_fast:
        base_30 = v["velocity"] * 30
        forecast_30 = round(base_30 * idx, 1)
        gap = v["stock"] - forecast_30
        status = "OK" if gap > 10 else ("Low Stock" if gap > 0 else "Stockout Risk")
        forecast_cards.append({
            "name": v["name"],
            "category": v["category"],
            "current_30d": round(base_30, 1),
            "forecast_30d": forecast_30,
            "current_stock": v["stock"],
            "gap": round(gap, 1),
            "status": status,
        })

    # Category-level forecast
    cat_vel = {}
    for v in velocity_list:
        cat = v["category"] or "Other"
        cat_vel[cat] = cat_vel.get(cat, 0) + v["velocity"]

    cat_forecast = [
        {"category": k, "current_30d": round(v * 30, 0), "forecast_30d": round(v * 30 * idx, 0)}
        for k, v in sorted(cat_vel.items(), key=lambda x: -x[1])[:12]
    ]

    monthly_index = [{"month": i + 1, "index": seasonal[i]} for i in range(12)]

    # Alerts
    high_demand = [f["name"] for f in forecast_cards if f["status"] == "Stockout Risk"]
    overstock = [v["name"] for v in velocity_list
                 if v["band"] == "Dead" and v["stock"] > 100][:10]

    return {
        "categories": cat_forecast,
        "cards": forecast_cards,
        "monthly_index": monthly_index,
        "high_demand_alerts": high_demand[:10],
        "overstock_risk": overstock,
    }


@app.route("/api/pricing/insights")
def api_pricing_insights():
    with get_conn() as conn:
        abc_list, abc_summary = _get_abc_data(conn)
        velocity_list, vel_summary = _get_velocity_data(conn)
        recommendations = _get_recommendations(abc_list, velocity_list, conn)
        cat_breakdown = _get_category_breakdown(conn)
        forecast = _get_forecast(velocity_list, conn)

        total_skus = conn.execute("SELECT COUNT(*) FROM medicines WHERE s > 0").fetchone()[0]
        inventory_value = sum(i["stock_value"] for i in abc_list)
        dead_stock_value = sum(
            i["stock_value"] for i in abc_list
            if velocity_list[min(i["rank"] - 1, len(velocity_list) - 1)].get("band") == "Dead"
        ) if velocity_list else 0

        # Revenue from ws_invoices
        try:
            ws_rev = conn.execute(
                "SELECT COALESCE(SUM(total_amount),0) FROM ws_invoices"
            ).fetchone()[0]
        except Exception:
            ws_rev = 0

        avg_margin = 18.5  # typical pharma retail margin %

    return jsonify({
        "stats": {
            "total_skus": total_skus,
            "inventory_value": round(inventory_value, 2),
            "ws_revenue": round(float(ws_rev), 2),
            "dead_stock_value": round(dead_stock_value, 2),
            "avg_margin": avg_margin,
        },
        "abc": abc_list[:200],
        "velocity": velocity_list[:200],
        "recommendations": recommendations,
        "category_breakdown": cat_breakdown,
        "abc_summary": abc_summary,
        "velocity_summary": vel_summary,
        "forecast": forecast,
    })


@app.route("/api/pricing/flow")
def api_pricing_flow():
    with get_conn() as conn:
        # Order stats
        try:
            stats_row = conn.execute("""
                SELECT
                    COUNT(*) as total_orders,
                    SUM(CASE WHEN status='dispatched' THEN 1 ELSE 0 END) as dispatched,
                    SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) as delivered,
                    COALESCE(SUM(total_amount),0) as revenue
                FROM ws_orders
            """).fetchone()
            stats = dict(stats_row)
        except Exception:
            stats = {"total_orders": 0, "dispatched": 0, "delivered": 0, "revenue": 0}

        # Outstanding (invoiced but not fully paid)
        try:
            outstanding_row = conn.execute("""
                SELECT COALESCE(SUM(i.total_amount - COALESCE(p.paid,0)),0) as outstanding
                FROM ws_invoices i
                LEFT JOIN (
                    SELECT invoice_id, SUM(amount) as paid FROM ws_payments GROUP BY invoice_id
                ) p ON p.invoice_id = i.id
            """).fetchone()
            stats["outstanding"] = round(float(outstanding_row["outstanding"]), 2)
        except Exception:
            stats["outstanding"] = 0

        stats["revenue"] = round(float(stats.get("revenue") or 0), 2)

        # Recent orders with shop name
        try:
            recent_rows = conn.execute("""
                SELECT o.id, o.shop_id, s.name as shop_name,
                       o.status, o.total_amount, o.created_at,
                       o.source
                FROM ws_orders o
                LEFT JOIN ws_shops s ON s.id = o.shop_id
                ORDER BY o.created_at DESC LIMIT 20
            """).fetchall()
            recent_orders = [dict(r) for r in recent_rows]
        except Exception:
            recent_orders = []

        # Top demanded medicines (from ws_order_items)
        try:
            top_med_rows = conn.execute("""
                SELECT oi.name, SUM(oi.qty) as total_qty,
                       SUM(oi.qty * oi.price) as total_value,
                       COUNT(DISTINCT oi.order_id) as order_count
                FROM ws_order_items oi
                GROUP BY oi.name
                ORDER BY total_qty DESC LIMIT 15
            """).fetchall()
            top_medicines = [dict(r) for r in top_med_rows]
        except Exception:
            top_medicines = []

        # Shop performance
        try:
            shop_rows = conn.execute("""
                SELECT s.id, s.name, s.city, s.discount,
                       COUNT(o.id) as order_count,
                       COALESCE(SUM(o.total_amount),0) as total_revenue,
                       MAX(o.created_at) as last_order
                FROM ws_shops s
                LEFT JOIN ws_orders o ON o.shop_id = s.id
                GROUP BY s.id
                ORDER BY total_revenue DESC LIMIT 20
            """).fetchall()
            shop_performance = [dict(r) for r in shop_rows]
        except Exception:
            shop_performance = []

        # Intake trend (orders per day last 30 days)
        try:
            trend_rows = conn.execute("""
                SELECT DATE(created_at) as day, COUNT(*) as orders,
                       COALESCE(SUM(total_amount),0) as revenue
                FROM ws_orders
                WHERE created_at >= DATE('now','-30 days')
                GROUP BY day ORDER BY day
            """).fetchall()
            intake_trend = [dict(r) for r in trend_rows]
        except Exception:
            intake_trend = []

    return jsonify({
        "stats": stats,
        "recent_orders": recent_orders,
        "top_medicines": top_medicines,
        "shop_performance": shop_performance,
        "intake_trend": intake_trend,
    })


# ── Stage 12 DB init (no new tables needed — uses existing data) ──────────
def init_stage12_db():
    """Stage 12 uses existing medicines, bills, ws_orders tables. No new tables."""
    pass  # All analysis is computed on-the-fly from existing tables


# ════════════════════════════════════════════════════════════════════
# PHASE 3 — PROFIT ENGINE, STAFF TARGETS, CUSTOMER TIERS, SETTINGS
# ════════════════════════════════════════════════════════════════════

def init_phase3_db() -> None:
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS staff_targets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER NOT NULL,
                staff_name TEXT,
                month TEXT NOT NULL,
                sales_target REAL DEFAULT 0,
                profit_target REAL DEFAULT 0,
                base_salary REAL DEFAULT 0,
                commission_pct REAL DEFAULT 0,
                bonus_at_pct REAL DEFAULT 80,
                bonus_amount REAL DEFAULT 0,
                set_by TEXT DEFAULT 'owner',
                notes TEXT,
                created_at TEXT,
                UNIQUE(staff_id, month)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS staff_salary (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER NOT NULL,
                staff_name TEXT,
                month TEXT NOT NULL,
                base_salary REAL DEFAULT 0,
                sales_achieved REAL DEFAULT 0,
                profit_generated REAL DEFAULT 0,
                target_achievement_pct REAL DEFAULT 0,
                commission_earned REAL DEFAULT 0,
                bonus_earned REAL DEFAULT 0,
                deductions REAL DEFAULT 0,
                advance_taken REAL DEFAULT 0,
                net_salary REAL DEFAULT 0,
                status TEXT DEFAULT 'pending',
                paid_date TEXT,
                notes TEXT,
                created_at TEXT,
                UNIQUE(staff_id, month)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS customer_tiers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER UNIQUE,
                customer_name TEXT,
                tier TEXT DEFAULT 'silver',
                avg_margin_pct REAL DEFAULT 0,
                total_revenue REAL DEFAULT 0,
                total_profit REAL DEFAULT 0,
                visit_count INTEGER DEFAULT 0,
                max_stock_reserve INTEGER DEFAULT 0,
                notes TEXT,
                updated_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS medicine_profit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                medicine_id TEXT UNIQUE,
                medicine_name TEXT,
                purchase_price REAL DEFAULT 0,
                selling_price REAL DEFAULT 0,
                mrp REAL DEFAULT 0,
                margin_pct REAL DEFAULT 0,
                margin_per_unit REAL DEFAULT 0,
                units_sold_30d INTEGER DEFAULT 0,
                profit_30d REAL DEFAULT 0,
                priority TEXT DEFAULT 'normal',
                updated_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS wholesaler_performance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                wholesaler_name TEXT NOT NULL UNIQUE,
                total_orders INTEGER DEFAULT 0,
                on_time_deliveries INTEGER DEFAULT 0,
                late_deliveries INTEGER DEFAULT 0,
                avg_delivery_days REAL DEFAULT 0,
                total_value REAL DEFAULT 0,
                return_count INTEGER DEFAULT 0,
                quality_score REAL DEFAULT 10,
                preferred_meds TEXT DEFAULT '[]',
                last_order_date TEXT,
                updated_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS password_change_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_type TEXT, user_id INTEGER, user_name TEXT,
                changed_at TEXT, changed_by TEXT
            )
        """)
        # Migrations
        b_cols = table_columns(conn, "bills")
        if "staff_id" not in b_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN staff_id INTEGER DEFAULT 0")
        if "profit" not in b_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN profit REAL DEFAULT 0")
        s_cols = table_columns(conn, "staff")
        if "salary" not in s_cols:
            conn.execute("ALTER TABLE staff ADD COLUMN salary REAL DEFAULT 0")
        if "advance" not in s_cols:
            conn.execute("ALTER TABLE staff ADD COLUMN advance REAL DEFAULT 0")
        if "joined_date" not in s_cols:
            conn.execute("ALTER TABLE staff ADD COLUMN joined_date TEXT DEFAULT ''")


# ── Profit engine helpers ─────────────────────────────────────────────

def _calc_bill_profit(items: list, conn) -> float:
    total = 0.0
    for item in items:
        qty  = float(item.get("qty",1) or 1)
        sell = float(item.get("price",0) or 0)
        disc = float(item.get("disc",0) or 0)
        eff  = sell * (1 - disc/100)
        name = str(item.get("name","")).strip()
        med  = conn.execute("SELECT p_rate,p FROM medicines WHERE n=? LIMIT 1",(name,)).fetchone()
        buy  = float(med["p_rate"] or med["p"]*0.7) if med else eff*0.7
        total += (eff - buy) * qty
    return round(total, 2)


def _refresh_medicine_profit_cache(conn, medicine_id=None):
    meds = conn.execute("SELECT * FROM medicines WHERE id=?" ,(medicine_id,)).fetchall() \
           if medicine_id else conn.execute("SELECT * FROM medicines").fetchall()
    now  = datetime.now(timezone.utc).isoformat()
    cutoff_ts = int((datetime.now(timezone.utc) - timedelta(days=30)).timestamp())
    for m in meds:
        sell  = float(m["p"] or 0)
        buy   = float(m["p_rate"] or sell*0.7)
        margin = round((sell-buy)/sell*100,1) if sell > 0 else 0
        rows  = conn.execute("SELECT items FROM bills WHERE ts>=?",(cutoff_ts,)).fetchall()
        units_30 = 0; profit_30 = 0.0
        for r in rows:
            for it in safe_json_loads(r["items"],[]):
                if str(it.get("name","")).strip() == m["n"]:
                    q = float(it.get("qty",1) or 1)
                    s = float(it.get("price",0) or 0)
                    units_30 += int(q); profit_30 += (s - buy)*q
        priority = "high" if margin>=25 else "low" if margin<10 else "normal"
        conn.execute("""
            INSERT INTO medicine_profit
            (medicine_id,medicine_name,purchase_price,selling_price,mrp,margin_pct,
             margin_per_unit,units_sold_30d,profit_30d,priority,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(medicine_id) DO UPDATE SET
            purchase_price=excluded.purchase_price,selling_price=excluded.selling_price,
            margin_pct=excluded.margin_pct,margin_per_unit=excluded.margin_per_unit,
            units_sold_30d=excluded.units_sold_30d,profit_30d=excluded.profit_30d,
            priority=excluded.priority,updated_at=excluded.updated_at
        """, (m["id"],m["n"],buy,sell,sell,margin,round(sell-buy,2),units_30,round(profit_30,2),priority,now))


# ── Phase 3 page routes ───────────────────────────────────────────────

@app.route("/portal/profit")
def portal_profit():
    if not session.get("portal_user") and not session.get("staff_id"):
        return redirect("/portal")
    try:
        sid = session.get("portal_user")
        if sid:
            allowed, _ = check_feature("profit_analytics", int(sid))
            if not allowed:
                return redirect("/portal/subscription?upgrade=1&feature=profit_analytics")
    except Exception:
        pass
    return render_template("portal_profit.html")

@app.route("/portal/staff-targets")
def portal_staff_targets():
    if not session.get("portal_user") and not session.get("staff_id"):
        return redirect("/portal")
    try:
        sid = session.get("portal_user")
        if sid:
            allowed, _ = check_feature("staff_targets", int(sid))
            if not allowed:
                return redirect("/portal/subscription?upgrade=1&feature=staff_targets")
    except Exception:
        pass
    return render_template("portal_staff_targets.html")

@app.route("/portal/settings")
def portal_settings_page():
    if not session.get("portal_user") and not session.get("staff_id"):
        return redirect("/portal")
    return render_template("portal_settings.html")

@app.route("/settings")
def internal_settings():
    if not session.get("staff_id"):
        return redirect("/internal")
    return render_template("portal_settings.html")


# ── Password / PIN change API ─────────────────────────────────────────

@app.route("/api/settings/change-password", methods=["POST"])
def change_password():
    data      = request.get_json(silent=True) or {}
    user_type = data.get("user_type","staff")
    old_pass  = str(data.get("old_password","")).strip()
    new_pass  = str(data.get("new_password","")).strip()
    if not old_pass or not new_pass:
        return json_error("Old and new password required",400)
    if len(new_pass) < 4:
        return json_error("Minimum 4 characters",400)
    now = datetime.now(timezone.utc).isoformat()

    if user_type == "admin":
        admin = session.get("admin_user")
        if not admin: return json_error("Not authenticated",401)
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM platform_admins WHERE username=?",
                               (admin.get("username",""),)).fetchone()
            if not row or not check_password_hash(row["password_hash"],old_pass):
                return json_error("Current password incorrect",401)
            conn.execute("UPDATE platform_admins SET password_hash=? WHERE id=?",
                         (generate_password_hash(new_pass),row["id"]))
            conn.execute("INSERT INTO password_change_log VALUES (NULL,?,?,?,?,?)",
                         ("admin",row["id"],admin.get("name",""),now,admin.get("username","")))
        return jsonify({"status":"success","message":"Admin password updated"})

    elif user_type == "portal":
        user = session.get("portal_user")
        if not user: return json_error("Not authenticated",401)
        table = "retail_shops" if user.get("account_type","retail")=="retail" else "wholesale_accounts"
        with get_conn() as conn:
            row = conn.execute(f"SELECT * FROM {table} WHERE id=?",(user.get("id"),)).fetchone()
            if not row or not check_password_hash(row["password_hash"],old_pass):
                return json_error("Current password incorrect",401)
            conn.execute(f"UPDATE {table} SET password_hash=? WHERE id=?",
                         (generate_password_hash(new_pass),row["id"]))
            conn.execute("INSERT INTO password_change_log VALUES (NULL,?,?,?,?,?)",
                         (table,row["id"],user.get("name",""),now,user.get("username","")))
        return jsonify({"status":"success","message":"Password updated"})

    elif user_type == "staff":
        sid = session.get("staff_id")
        if not sid: return json_error("Not authenticated",401)
        with get_conn() as conn:
            s = conn.execute("SELECT * FROM staff WHERE id=?",(sid,)).fetchone()
            if not s or s["pin"] != old_pass:
                return json_error("Current PIN incorrect",401)
            conn.execute("UPDATE staff SET pin=? WHERE id=?",(new_pass,sid))
            conn.execute("INSERT INTO password_change_log VALUES (NULL,?,?,?,?,?)",
                         ("staff",sid,s["name"],now,s["name"]))
        return jsonify({"status":"success","message":"PIN updated"})

    return json_error("Unknown user_type",400)


@app.route("/api/settings/admin-reset-pin", methods=["POST"])
def admin_reset_staff_pin():
    if not session.get("admin_user") and not session.get("staff_id"):
        return json_error("Forbidden",403)
    data    = request.get_json(silent=True) or {}
    sid     = data.get("staff_id")
    new_pin = str(data.get("new_pin","")).strip()
    if not sid or not new_pin: return json_error("staff_id and new_pin required",400)
    with get_conn() as conn:
        s = conn.execute("SELECT name FROM staff WHERE id=?",(sid,)).fetchone()
        if not s: return json_error("Staff not found",404)
        conn.execute("UPDATE staff SET pin=? WHERE id=?",(new_pin,sid))
        conn.execute("INSERT INTO password_change_log VALUES (NULL,?,?,?,?,?)",
                     ("staff_reset",sid,s["name"],datetime.now(timezone.utc).isoformat(),"admin"))
    return jsonify({"status":"success","message":f"PIN reset for {s['name']}"})


@app.route("/api/settings/profile", methods=["GET"])
def get_my_profile():
    if session.get("admin_user"):   return jsonify({"user_type":"admin",   **session["admin_user"]})
    if session.get("portal_user"):  return jsonify({"user_type":"portal",  **session["portal_user"]})
    if session.get("staff_id"):
        return jsonify({"user_type":"staff","id":session["staff_id"],
                        "name":session.get("staff_name",""),"role":session.get("staff_role","")})
    return json_error("Not authenticated",401)


@app.route("/api/settings/company", methods=["GET"])
def get_company_settings():
    with get_conn() as conn:
        rows = conn.execute("SELECT key, value FROM app_settings").fetchall()
    return jsonify({r["key"]:r["value"] for r in rows})


@app.route("/api/settings/company", methods=["POST"])
def save_company_settings():
    if not session.get("admin_user") and not session.get("staff_id"):
        return json_error("Forbidden",403)
    data = request.get_json(silent=True) or {}
    now  = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        for k,v in data.items():
            if isinstance(k,str) and k:
                conn.execute("""INSERT INTO app_settings(key,value,updated_at) VALUES(?,?,?)
                    ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=excluded.updated_at""",
                    (k,str(v),now))
    return jsonify({"status":"success"})


# ── Profit analytics API ──────────────────────────────────────────────

@app.route("/api/profit/summary", methods=["GET"])
def profit_summary():
    import calendar as _cal
    month    = request.args.get("month", datetime.now(timezone.utc).strftime("%Y-%m"))
    d_from   = request.args.get("from", month+"-01")
    y,m_     = int(month.split("-")[0]), int(month.split("-")[1])
    last_day = _cal.monthrange(y,m_)[1]
    d_to     = request.args.get("to", f"{month}-{last_day:02d}")

    # Convert to dd/mm/yyyy for bills table
    def iso_to_dmy(s):
        try: p=s.split("-"); return f"{p[2]}/{p[1]}/{p[0]}"
        except: return s

    df_dmy = iso_to_dmy(d_from); dt_dmy = iso_to_dmy(d_to)
    cutoff_ts = int(datetime(y,m_,1).timestamp())
    end_ts    = int(datetime(y,m_,last_day,23,59,59).timestamp())

    with get_conn() as conn:
        bills = [dict(r) for r in conn.execute("""
            SELECT id,total,items,staff_name,date,pay FROM bills
            WHERE ts BETWEEN ? AND ?
        """,(cutoff_ts,end_ts)).fetchall()]

        total_rev=0.0; total_cost=0.0; total_profit=0.0
        by_staff: dict = {}; by_day: dict = {}; by_pay: dict = {}

        for bill in bills:
            items = safe_json_loads(bill.get("items"),[])
            bp=0.0; bc=0.0
            for it in items:
                qty=float(it.get("qty",1) or 1); sell=float(it.get("price",0) or 0)
                disc=float(it.get("disc",0) or 0); eff=sell*(1-disc/100)
                nm=str(it.get("name","")).strip()
                med=conn.execute("SELECT p_rate,p FROM medicines WHERE n=? LIMIT 1",(nm,)).fetchone()
                buy=float(med["p_rate"] or med["p"]*0.7) if med else eff*0.7
                bp+=(eff-buy)*qty; bc+=buy*qty

            rev=float(bill.get("total",0) or 0)
            total_rev+=rev; total_cost+=bc; total_profit+=bp

            sn=bill.get("staff_name") or "Unknown"
            if sn not in by_staff: by_staff[sn]={"name":sn,"bills":0,"revenue":0,"profit":0}
            by_staff[sn]["bills"]+=1; by_staff[sn]["revenue"]+=rev; by_staff[sn]["profit"]+=bp

            day=bill.get("date","")[:10]
            if day not in by_day: by_day[day]={"date":day,"revenue":0,"profit":0}
            by_day[day]["revenue"]+=rev; by_day[day]["profit"]+=bp

            pay=bill.get("pay","cash") or "cash"
            if pay not in by_pay: by_pay[pay]={"mode":pay,"revenue":0,"count":0}
            by_pay[pay]["revenue"]+=rev; by_pay[pay]["count"]+=1

    margin = round(total_profit/total_rev*100,1) if total_rev>0 else 0
    for s in by_staff.values():
        s["revenue"]=round(s["revenue"],2); s["profit"]=round(s["profit"],2)
        s["margin_pct"]=round(s["profit"]/s["revenue"]*100,1) if s["revenue"]>0 else 0
    days_sorted=[{**v,"revenue":round(v["revenue"],2),"profit":round(v["profit"],2)}
                 for v in sorted(by_day.values(),key=lambda x:x["date"])]

    return jsonify({
        "period":        {"from":d_from,"to":d_to},
        "total_revenue": round(total_rev,2),
        "total_cost":    round(total_cost,2),
        "total_profit":  round(total_profit,2),
        "margin_pct":    margin,
        "bill_count":    len(bills),
        "avg_per_bill":  round(total_profit/max(len(bills),1),2),
        "by_staff":      sorted(by_staff.values(),key=lambda x:-x["profit"]),
        "by_day":        days_sorted,
        "by_payment":    list(by_pay.values())
    })


@app.route("/api/profit/medicines", methods=["GET"])
def profit_by_medicine():
    sort_by = request.args.get("sort","profit")
    with get_conn() as conn:
        _refresh_medicine_profit_cache(conn)
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM medicine_profit ORDER BY profit_30d DESC LIMIT 100").fetchall()]
    if sort_by=="margin": rows.sort(key=lambda x:-x.get("margin_pct",0))
    elif sort_by=="volume": rows.sort(key=lambda x:-x.get("units_sold_30d",0))
    return jsonify(rows)


@app.route("/api/profit/customers", methods=["GET"])
def profit_by_customer():
    limit = int(request.args.get("limit",50))
    with get_conn() as conn:
        custs = [dict(r) for r in conn.execute(
            "SELECT id,name,phone,visits,total FROM customers ORDER BY total DESC LIMIT ?",(limit,)).fetchall()]
        result=[]
        for c in custs:
            bills=[dict(r) for r in conn.execute(
                "SELECT items,total FROM bills WHERE cust=? OR phone=?",
                (c["name"],c["phone"] or "")).fetchall()]
            crev=0.0; cpro=0.0
            for b in bills:
                crev+=float(b.get("total",0) or 0)
                for it in safe_json_loads(b.get("items"),[]):
                    qty=float(it.get("qty",1) or 1); sell=float(it.get("price",0) or 0)
                    disc=float(it.get("disc",0) or 0); eff=sell*(1-disc/100)
                    nm=str(it.get("name","")).strip()
                    med=conn.execute("SELECT p_rate,p FROM medicines WHERE n=? LIMIT 1",(nm,)).fetchone()
                    buy=float(med["p_rate"] or med["p"]*0.7) if med else eff*0.7
                    cpro+=(eff-buy)*qty
            margin=round(cpro/crev*100,1) if crev>0 else 0
            t_row=conn.execute("SELECT tier FROM customer_tiers WHERE customer_id=?",(c["id"],)).fetchone()
            tier=t_row["tier"] if t_row else ("gold" if margin>=25 else "bronze" if margin<10 else "silver")
            result.append({"id":c["id"],"name":c["name"],"phone":c["phone"],"visits":c["visits"],
                "revenue":round(crev,2),"profit":round(cpro,2),"margin_pct":margin,
                "tier":tier,"bill_count":len(bills),
                "priority":"high" if margin>=25 else "low" if margin<10 else "normal"})
        result.sort(key=lambda x:-x["profit"])
        now=datetime.now(timezone.utc).isoformat()
        for r in result:
            conn.execute("""INSERT INTO customer_tiers
                (customer_id,customer_name,tier,avg_margin_pct,total_revenue,total_profit,visit_count,updated_at)
                VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(customer_id) DO UPDATE SET
                tier=excluded.tier,avg_margin_pct=excluded.avg_margin_pct,
                total_revenue=excluded.total_revenue,total_profit=excluded.total_profit,
                visit_count=excluded.visit_count,updated_at=excluded.updated_at""",
                (r["id"],r["name"],r["tier"],r["margin_pct"],r["revenue"],r["profit"],r["visits"],now))
    return jsonify(result)


@app.route("/api/profit/customers/<int:cid>/tier", methods=["PUT"])
def update_customer_tier(cid):
    data=request.get_json(silent=True) or {}
    with get_conn() as conn:
        c=conn.execute("SELECT name FROM customers WHERE id=?",(cid,)).fetchone()
        if not c: return json_error("Not found",404)
        conn.execute("""INSERT INTO customer_tiers
            (customer_id,customer_name,tier,max_stock_reserve,notes,updated_at)
            VALUES(?,?,?,?,?,?) ON CONFLICT(customer_id) DO UPDATE SET
            tier=excluded.tier,max_stock_reserve=excluded.max_stock_reserve,
            notes=excluded.notes,updated_at=excluded.updated_at""",
            (cid,c["name"],data.get("tier","silver"),int(data.get("max_stock_reserve",0)),
             data.get("notes",""),datetime.now(timezone.utc).isoformat()))
    return jsonify({"status":"success"})


@app.route("/api/profit/stock-alert", methods=["POST"])
def stock_priority_alert():
    data=request.get_json(silent=True) or {}
    cust_id=data.get("customer_id"); items=data.get("items",[]); warnings=[]
    with get_conn() as conn:
        cust_tier="silver"
        if cust_id:
            t=conn.execute("SELECT tier FROM customer_tiers WHERE customer_id=?",(cust_id,)).fetchone()
            cust_tier=t["tier"] if t else "silver"
        for item in items:
            name=str(item.get("name","")).strip(); qty=int(item.get("qty",1) or 1)
            med=conn.execute("SELECT s,reorder,n FROM medicines WHERE n=? LIMIT 1",(name,)).fetchone()
            if not med: continue
            stock_after=med["s"]-qty; reorder=med["reorder"] or 0
            if stock_after<=reorder and cust_tier in ("silver","bronze"):
                gold=conn.execute("""SELECT ct.customer_name FROM customer_tiers ct
                    WHERE ct.tier='gold' AND ct.customer_id IN
                    (SELECT DISTINCT c.id FROM customers c JOIN bills b ON b.cust=c.name
                    WHERE b.items LIKE ?) LIMIT 3""",(f"%{name}%",)).fetchall()
                if gold:
                    warnings.append({"medicine":name,"stock_after":stock_after,
                        "reorder_point":reorder,"current_stock":med["s"],"qty_requested":qty,
                        "gold_customers":[r["customer_name"] for r in gold],
                        "warning":f"⚠ Stock hits reorder. Gold customers also buy {name}!",
                        "severity":"high" if stock_after<=0 else "medium"})
    return jsonify({"warnings":warnings,"customer_tier":cust_tier})


@app.route("/api/profit/wholesaler-performance", methods=["GET"])
def wholesaler_performance():
    with get_conn() as conn:
        rows=[dict(r) for r in conn.execute(
            "SELECT * FROM wholesaler_performance ORDER BY total_value DESC").fetchall()]
    for r in rows: r["preferred_meds"]=safe_json_loads(r.get("preferred_meds"),[])
    return jsonify(rows)


@app.route("/api/profit/wholesaler-performance", methods=["POST"])
def update_wholesaler_perf():
    data=request.get_json(silent=True) or {}; name=data.get("wholesaler_name","").strip()
    if not name: return json_error("wholesaler_name required",400)
    now=datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        conn.execute("""INSERT INTO wholesaler_performance
            (wholesaler_name,total_orders,on_time_deliveries,late_deliveries,avg_delivery_days,
             total_value,return_count,quality_score,preferred_meds,last_order_date,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(wholesaler_name) DO UPDATE SET
            total_orders=total_orders+excluded.total_orders,
            on_time_deliveries=on_time_deliveries+excluded.on_time_deliveries,
            late_deliveries=late_deliveries+excluded.late_deliveries,
            total_value=total_value+excluded.total_value,
            quality_score=excluded.quality_score,last_order_date=excluded.last_order_date,
            updated_at=excluded.updated_at""",
            (name,int(data.get("total_orders",1)),int(data.get("on_time",1)),int(data.get("late",0)),
             float(data.get("avg_days",1)),float(data.get("value",0)),int(data.get("returns",0)),
             float(data.get("quality",10)),json.dumps(data.get("preferred_meds",[])),
             data.get("last_order_date",now[:10]),now))
    return jsonify({"status":"success"})


# ── Staff list & update ───────────────────────────────────────────────

@app.route("/api/staff/list", methods=["GET"])
def list_staff_p3():
    with get_conn() as conn:
        rows=[dict(r) for r in conn.execute(
            "SELECT id,name,role,phone,pin,salary,advance,joined_date,is_active FROM staff ORDER BY name"
        ).fetchall()]
    return jsonify(rows)

@app.route("/api/staff/<int:sid>/update", methods=["PUT"])
def update_staff_p3(sid):
    if not session.get("admin_user") and not session.get("staff_id"):
        return json_error("Forbidden",403)
    data=request.get_json(silent=True) or {}
    with get_conn() as conn:
        conn.execute("""UPDATE staff SET name=?,role=?,phone=?,salary=?,advance=?,
            joined_date=?,is_active=? WHERE id=?""",
            (data.get("name"),data.get("role"),data.get("phone",""),
             float(data.get("salary",0)),float(data.get("advance",0)),
             data.get("joined_date",""),int(data.get("is_active",1)),sid))
    return jsonify({"status":"success"})


# ── Staff targets & salary API ────────────────────────────────────────

@app.route("/api/staff/targets", methods=["GET"])
def get_staff_targets():
    month=request.args.get("month",datetime.now(timezone.utc).strftime("%Y-%m"))
    with get_conn() as conn:
        targets=[dict(r) for r in conn.execute(
            "SELECT * FROM staff_targets WHERE month=? ORDER BY staff_name",(month,)).fetchall()]
        staff=[dict(r) for r in conn.execute(
            "SELECT id,name,role,salary FROM staff WHERE is_active=1").fetchall()]
    return jsonify({"targets":targets,"staff":staff,"month":month})


@app.route("/api/staff/targets", methods=["POST"])
def set_staff_target():
    if not session.get("admin_user") and not session.get("staff_id"):
        return json_error("Forbidden",403)
    data=request.get_json(silent=True) or {}
    actor=session.get("admin_user",{}).get("name","") or session.get("staff_name","owner")
    with get_conn() as conn:
        conn.execute("""INSERT INTO staff_targets
            (staff_id,staff_name,month,sales_target,profit_target,base_salary,
             commission_pct,bonus_at_pct,bonus_amount,set_by,notes,created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(staff_id,month) DO UPDATE SET
            sales_target=excluded.sales_target,profit_target=excluded.profit_target,
            base_salary=excluded.base_salary,commission_pct=excluded.commission_pct,
            bonus_at_pct=excluded.bonus_at_pct,bonus_amount=excluded.bonus_amount,
            set_by=excluded.set_by,notes=excluded.notes""",
            (data.get("staff_id"),data.get("staff_name",""),
             data.get("month",datetime.now(timezone.utc).strftime("%Y-%m")),
             float(data.get("sales_target",0)),float(data.get("profit_target",0)),
             float(data.get("base_salary",0)),float(data.get("commission_pct",0)),
             float(data.get("bonus_at_pct",80)),float(data.get("bonus_amount",0)),
             actor,data.get("notes",""),datetime.now(timezone.utc).isoformat()))
    return jsonify({"status":"success"})


@app.route("/api/staff/salary-calc", methods=["GET"])
def calc_staff_salary():
    import calendar as _cal
    month=request.args.get("month",datetime.now(timezone.utc).strftime("%Y-%m"))
    y,m_=int(month.split("-")[0]),int(month.split("-")[1])
    last=_cal.monthrange(y,m_)[1]
    cutoff_ts=int(datetime(y,m_,1).timestamp())
    end_ts=int(datetime(y,m_,last,23,59,59).timestamp())
    with get_conn() as conn:
        targets={r["staff_id"]:dict(r) for r in conn.execute(
            "SELECT * FROM staff_targets WHERE month=?",(month,)).fetchall()}
        staff=[dict(r) for r in conn.execute(
            "SELECT id,name,role,salary,advance FROM staff WHERE is_active=1").fetchall()]
        results=[]
        for s in staff:
            bills=[dict(r) for r in conn.execute(
                "SELECT total,items FROM bills WHERE staff_name=? AND ts BETWEEN ? AND ?",
                (s["name"],cutoff_ts,end_ts)).fetchall()]
            sales=sum(float(b.get("total",0) or 0) for b in bills)
            profit=0.0
            for b in bills:
                for it in safe_json_loads(b.get("items"),[]):
                    qty=float(it.get("qty",1) or 1); sell=float(it.get("price",0) or 0)
                    disc=float(it.get("disc",0) or 0); eff=sell*(1-disc/100)
                    nm=str(it.get("name","")).strip()
                    med=conn.execute("SELECT p_rate,p FROM medicines WHERE n=? LIMIT 1",(nm,)).fetchone()
                    buy=float(med["p_rate"] or med["p"]*0.7) if med else eff*0.7
                    profit+=(eff-buy)*qty
            tgt=targets.get(s["id"],{})
            base=float(tgt.get("base_salary",s.get("salary",0)) or 0)
            s_tgt=float(tgt.get("sales_target",0) or 0)
            ach_pct=round(sales/s_tgt*100,1) if s_tgt>0 else 0
            comm_pct=float(tgt.get("commission_pct",0) or 0)
            commission=round(profit*comm_pct/100,2)
            bonus=float(tgt.get("bonus_amount",0) or 0) if ach_pct>=float(tgt.get("bonus_at_pct",80) or 80) else 0
            advance=float(s.get("advance",0) or 0)
            net=round(base+commission+bonus-advance,2)
            results.append({"staff_id":s["id"],"staff_name":s["name"],"role":s["role"],
                "month":month,"base_salary":base,"sales_target":s_tgt,
                "sales_achieved":round(sales,2),"profit_generated":round(profit,2),
                "target_pct":ach_pct,"commission_pct":comm_pct,"commission_earned":commission,
                "bonus_threshold":float(tgt.get("bonus_at_pct",80) or 80),
                "bonus_earned":bonus,"advance_deducted":advance,"net_salary":max(0,net),
                "bills_count":len(bills),"has_target":s["id"] in targets})
        results.sort(key=lambda x:-x["sales_achieved"])
    return jsonify({"month":month,"staff":results})


@app.route("/api/staff/salary-save", methods=["POST"])
def save_staff_salary():
    if not session.get("admin_user") and not session.get("staff_id"):
        return json_error("Forbidden",403)
    data=request.get_json(silent=True) or {}; records=data.get("records",[])
    now=datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        for r in records:
            conn.execute("""INSERT INTO staff_salary
                (staff_id,staff_name,month,base_salary,sales_achieved,profit_generated,
                 target_achievement_pct,commission_earned,bonus_earned,advance_taken,
                 net_salary,status,notes,created_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(staff_id,month) DO UPDATE SET
                sales_achieved=excluded.sales_achieved,profit_generated=excluded.profit_generated,
                target_achievement_pct=excluded.target_achievement_pct,
                commission_earned=excluded.commission_earned,bonus_earned=excluded.bonus_earned,
                net_salary=excluded.net_salary,status=excluded.status,notes=excluded.notes""",
                (r.get("staff_id"),r.get("staff_name"),r.get("month"),
                 r.get("base_salary",0),r.get("sales_achieved",0),r.get("profit_generated",0),
                 r.get("target_pct",0),r.get("commission_earned",0),r.get("bonus_earned",0),
                 r.get("advance_deducted",0),r.get("net_salary",0),
                 data.get("status","pending"),r.get("notes",""),now))
    return jsonify({"status":"success","saved":len(records)})


@app.route("/api/staff/salary-history", methods=["GET"])
def salary_history():
    sid=request.args.get("staff_id"); q="SELECT * FROM staff_salary WHERE 1=1"; p: list=[]
    if sid: q+=" AND staff_id=?"; p.append(int(sid))
    q+=" ORDER BY month DESC LIMIT 24"
    with get_conn() as conn: rows=[dict(r) for r in conn.execute(q,p).fetchall()]
    return jsonify(rows)


@app.route("/api/staff/demand-allocation", methods=["GET"])
def demand_allocation():
    with get_conn() as conn:
        meds=[dict(r) for r in conn.execute(
            "SELECT * FROM medicine_profit WHERE units_sold_30d>0 ORDER BY profit_30d DESC LIMIT 30"
        ).fetchall()]
        gold=[dict(r) for r in conn.execute("""
            SELECT ct.*,c.phone FROM customer_tiers ct JOIN customers c ON c.id=ct.customer_id
            WHERE ct.tier='gold' ORDER BY ct.total_profit DESC LIMIT 20""").fetchall()]
    return jsonify({"top_profit_medicines":meds,"gold_customers":gold,
        "insight":f"{len(gold)} gold customers, {len(meds)} high-demand medicines tracked"})


# ════════════════════════════════════════════════════════════════════
#  STAGE 14 — Staff Commission Tracker + Birthday WhatsApp
# ════════════════════════════════════════════════════════════════════

def init_stage14_db():
    with get_conn() as conn:
        # Add birthday column to staff if missing
        sc = table_columns(conn, "staff")
        if "birthday" not in sc:
            conn.execute("ALTER TABLE staff ADD COLUMN birthday TEXT DEFAULT ''")
        # Per-bill commission log
        conn.execute("""
            CREATE TABLE IF NOT EXISTS staff_commission_log (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                bill_id          TEXT    NOT NULL,
                staff_id         INTEGER DEFAULT 0,
                staff_name       TEXT    DEFAULT '',
                bill_total       REAL    DEFAULT 0,
                commission_pct   REAL    DEFAULT 0,
                commission_earned REAL   DEFAULT 0,
                bill_date        TEXT    DEFAULT '',
                created_at       TEXT    DEFAULT ''
            )
        """)
        # Birthday WhatsApp log
        conn.execute("""
            CREATE TABLE IF NOT EXISTS birthday_wa_log (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id   INTEGER,
                staff_name TEXT,
                phone      TEXT,
                message    TEXT,
                status     TEXT,
                sent_at    TEXT
            )
        """)


@app.route("/staff-commission")
def staff_commission_page():
    return render_template("staff_commission.html")


# ── Commission Summary (month aggregate) ────────────────────────────
@app.route("/api/stage14/commission-summary", methods=["GET"])
def stage14_commission_summary():
    import calendar as _cal
    month = request.args.get("month", datetime.now(timezone.utc).strftime("%Y-%m"))
    try:
        y, m_ = int(month[:4]), int(month[5:7])
    except Exception:
        return json_error("Invalid month format")
    last = _cal.monthrange(y, m_)[1]
    cutoff_ts = int(datetime(y, m_, 1).timestamp())
    end_ts    = int(datetime(y, m_, last, 23, 59, 59).timestamp())

    with get_conn() as conn:
        targets = {r["staff_id"]: dict(r) for r in conn.execute(
            "SELECT * FROM staff_targets WHERE month=?", (month,)).fetchall()}
        staff   = [dict(r) for r in conn.execute(
            "SELECT id,name,role,salary,birthday FROM staff WHERE is_active=1").fetchall()]

        results = []
        for s in staff:
            bills = [dict(r) for r in conn.execute(
                "SELECT total FROM bills WHERE staff_name=? AND ts BETWEEN ? AND ?",
                (s["name"], cutoff_ts, end_ts)).fetchall()]
            sales = sum(float(b.get("total", 0) or 0) for b in bills)
            tgt   = targets.get(s["id"], {})
            s_tgt = float(tgt.get("sales_target", 0) or 0)
            ach_pct = round(sales / s_tgt * 100, 1) if s_tgt > 0 else 0
            comm_pct = float(tgt.get("commission_pct", 0) or 0)
            commission = round(sales * comm_pct / 100, 2)
            bonus = float(tgt.get("bonus_amount", 0) or 0) if ach_pct >= float(tgt.get("bonus_at_pct", 80) or 80) else 0
            # check commission log for saved value
            log_sum = conn.execute(
                "SELECT COALESCE(SUM(commission_earned),0) AS s FROM staff_commission_log "
                "WHERE staff_id=? AND bill_date LIKE ?", (s["id"], f"{month}%")
            ).fetchone()["s"]
            results.append({
                "staff_id": s["id"], "staff_name": s["name"], "role": s["role"],
                "birthday": s.get("birthday", ""),
                "month": month, "bills_count": len(bills),
                "sales_achieved": round(sales, 2), "sales_target": s_tgt,
                "target_pct": ach_pct, "commission_pct": comm_pct,
                "commission_earned": round(log_sum or commission, 2),
                "bonus_earned": bonus,
                "has_target": s["id"] in targets,
            })

        results.sort(key=lambda x: -x["sales_achieved"])
        total_commission = round(sum(r["commission_earned"] for r in results), 2)
        total_bills      = sum(r["bills_count"] for r in results)
        top_earner       = results[0]["staff_name"] if results else "—"

    return jsonify({
        "month": month, "staff": results,
        "summary": {
            "total_commission": total_commission,
            "total_bills": total_bills,
            "top_earner": top_earner,
            "staff_count": len(results),
        }
    })


# ── Per-bill commission log ──────────────────────────────────────────
@app.route("/api/stage14/commission-log", methods=["GET"])
def stage14_commission_log():
    sid   = request.args.get("staff_id")
    month = request.args.get("month", "")
    q     = "SELECT * FROM staff_commission_log WHERE 1=1"
    p: list = []
    if sid:
        q += " AND staff_id=?"; p.append(int(sid))
    if month:
        q += " AND bill_date LIKE ?"; p.append(f"{month}%")
    q += " ORDER BY created_at DESC LIMIT 200"
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(q, p).fetchall()]
    return jsonify(rows)


# ── Today's birthdays ────────────────────────────────────────────────
@app.route("/api/stage14/birthday-check", methods=["GET"])
def stage14_birthday_check():
    today = datetime.now().strftime("%m-%d")   # MM-DD
    with get_conn() as conn:
        staff = [dict(r) for r in conn.execute(
            "SELECT id,name,role,phone,birthday FROM staff WHERE is_active=1 AND birthday!=''").fetchall()]
    birthdays = []
    for s in staff:
        bd = str(s.get("birthday", "")).strip()
        # support YYYY-MM-DD or MM-DD
        mmdd = bd[5:] if len(bd) == 10 else bd
        if mmdd == today:
            birthdays.append(s)
    return jsonify({"today": today, "birthdays": birthdays, "count": len(birthdays)})


# ── Send birthday WhatsApp ───────────────────────────────────────────
@app.route("/api/stage14/birthday-wa-send", methods=["POST"])
def stage14_birthday_wa_send():
    data = request.get_json(silent=True) or {}
    staff_id = data.get("staff_id")
    custom_msg = data.get("message", "").strip()

    with get_conn() as conn:
        s = conn.execute(
            "SELECT id,name,phone,role,birthday FROM staff WHERE id=?", (staff_id,)
        ).fetchone()
        if not s:
            return json_error("Staff not found", 404)
        s = dict(s)

    phone = s.get("phone", "").strip()
    if not phone:
        return json_error("Staff has no phone number", 400)

    msg = custom_msg or (
        f"🎂 Happy Birthday, {s['name']}! 🎉\n\n"
        f"Wishing you a wonderful birthday and a great year ahead!\n\n"
        f"— Team Selvam Medicals 💊"
    )

    result = send_whatsapp(phone, msg)

    with get_conn() as conn:
        conn.execute("""
            INSERT INTO birthday_wa_log (staff_id, staff_name, phone, message, status, sent_at)
            VALUES (?,?,?,?,?,?)
        """, (s["id"], s["name"], phone, msg, result["status"],
              datetime.now(timezone.utc).isoformat()))

    return jsonify({"status": result["status"], "staff": s["name"], "wa": result})


# ── Update staff birthday ────────────────────────────────────────────
@app.route("/api/staff/<int:sid>/birthday", methods=["PUT"])
def update_staff_birthday(sid):
    data = request.get_json(silent=True) or {}
    birthday = str(data.get("birthday", "")).strip()
    with get_conn() as conn:
        conn.execute("UPDATE staff SET birthday=? WHERE id=?", (birthday, sid))
    return jsonify({"status": "ok", "staff_id": sid, "birthday": birthday})



# ════════════════════════════════════════════════════════════════════════════════
#  STAGE 13 — Voice Billing + PWA
# ════════════════════════════════════════════════════════════════════════════════

import difflib as _difflib

# Number words → int
_NUM_WORDS = {
    "zero":0,"one":1,"two":2,"three":3,"four":4,"five":5,
    "six":6,"seven":7,"eight":8,"nine":9,"ten":10,
    "eleven":11,"twelve":12,"fifteen":15,"twenty":20,
    "thirty":30,"fifty":50,"hundred":100,
    "a":1,"an":1,"half":1,
}

def _extract_qty(text: str):
    """Pull out the first number (digit or word) from text; return (qty, cleaned_text)."""
    import re
    # Digit first
    m = re.search(r'\b(\d+)\b', text)
    if m:
        return int(m.group(1)), (text[:m.start()] + text[m.end():]).strip()
    # Word number
    for word, num in sorted(_NUM_WORDS.items(), key=lambda x: -len(x[0])):
        pattern = r'\b' + re.escape(word) + r'\b'
        if re.search(pattern, text, re.I):
            cleaned = re.sub(pattern, '', text, flags=re.I).strip()
            return num, cleaned
    return 1, text


def _voice_fuzzy_match(query: str, conn, top_n: int = 5):
    """Fuzzy-match query string against medicine names; return top hits."""
    if not query:
        return []
    rows = conn.execute(
        "SELECT n as name, g as generic, c as category, p as mrp, s as stock "
        "FROM medicines ORDER BY n LIMIT 5000"
    ).fetchall()
    meds = [dict(r) for r in rows]
    names_lower = [m["name"].lower() for m in meds]

    # 1. exact substring match (highest priority)
    exact = [m for m in meds if query.lower() in m["name"].lower()]
    if exact:
        for m in exact[:top_n]:
            m["confidence"] = 97
        return exact[:top_n]

    # 2. difflib close matches
    close = _difflib.get_close_matches(query.lower(), names_lower, n=top_n, cutoff=0.45)
    results = []
    for match in close:
        for m in meds:
            if m["name"].lower() == match:
                score = _difflib.SequenceMatcher(None, query.lower(), m["name"].lower()).ratio()
                m = dict(m)
                m["confidence"] = round(score * 100)
                results.append(m)
                break
    return results


# ── Page routes ───────────────────────────────────────────────────────────────

@app.route("/voice-billing")
def page_voice_billing():
    return render_template("voice_billing.html")


@app.route("/offline")
def page_offline():
    return render_template("offline.html")


# ── PWA manifest ───────────────────────────────────────────────────────────────

@app.route("/manifest.json")
def pwa_manifest():
    manifest = {
        "name": "MediVision AI",
        "short_name": "MediVision",
        "description": "Complete Pharmacy Management Platform for Indian Retail & Wholesale",
        "start_url": "/welcome",
        "display": "standalone",
        "orientation": "portrait-primary",
        "background_color": "#0b0e16",
        "theme_color": "#f97316",
        "lang": "en-IN",
        "categories": ["medical", "business", "productivity"],
        "icons": [
            {
                "src": "data:image/svg+xml;base64,PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHZpZXdCb3g9JzAgMCAxMDAgMTAwJz48cmVjdCB3aWR0aD0nMTAwJyBoZWlnaHQ9JzEwMCcgcng9JzIwJyBmaWxsPScjMGIwZTE2Jy8+PHJlY3QgeD0nNDInIHk9JzIwJyB3aWR0aD0nMTYnIGhlaWdodD0nNjAnIHJ4PSc4JyBmaWxsPScjZjk3MzE2Jy8+PHJlY3QgeD0nMjAnIHk9JzQyJyB3aWR0aD0nNjAnIGhlaWdodD0nMTYnIHJ4PSc4JyBmaWxsPScjZjk3MzE2Jy8+PC9zdmc+",
                "sizes": "any",
                "type": "image/svg+xml",
                "purpose": "any maskable"
            }
        ],
        "shortcuts": [
            {
                "name": "New Bill",
                "short_name": "Bill",
                "url": "/portal/billing",
                "description": "Open billing screen"
            },
            {
                "name": "Voice Billing",
                "short_name": "Voice",
                "url": "/voice-billing",
                "description": "Voice-activated billing"
            },
            {
                "name": "Stock",
                "short_name": "Stock",
                "url": "/portal/stock",
                "description": "Check stock levels"
            },
            {
                "name": "Symptom Advisor",
                "short_name": "Symptoms",
                "url": "/symptom-advisor",
                "description": "Medicine by symptoms"
            },
        ],
        "screenshots": [],
        "related_applications": [],
        "prefer_related_applications": False,
    }
    resp = jsonify(manifest)
    resp.headers["Cache-Control"] = "no-cache"
    return resp


# ── Service Worker (served at root for full-scope control) ──────────────────

@app.route("/sw.js")
def service_worker():
    """Serve the service worker from root scope."""
    import os as _os
    sw_path = _os.path.join(STATIC_DIR, "sw.js")
    if _os.path.exists(sw_path):
        with open(sw_path, "r", encoding="utf-8") as f:
            sw_content = f.read()
        resp = app.response_class(sw_content, mimetype="application/javascript")
        resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        resp.headers["Service-Worker-Allowed"] = "/"
        return resp
    return ("// sw.js not found", 404, {"Content-Type": "application/javascript"})


# ── Voice Parse API ────────────────────────────────────────────────────────────

@app.route("/api/voice/parse", methods=["POST"])
def api_voice_parse():
    data = request.get_json(force=True) or {}
    text = (data.get("text") or "").strip()

    if not text:
        return jsonify({"matches": [], "raw": text})

    import re

    # Strip filler words
    filler = r'\b(add|give|me|please|the|some|strips|tablets|tablet|capsules|capsule|pieces|piece|units|unit|syrup|bottle|box|pack|sachet|injection|inj|ml|mg|gm)\b'
    cleaned = re.sub(filler, ' ', text.lower(), flags=re.I)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    qty, med_query = _extract_qty(cleaned)
    med_query = med_query.strip()

    if not med_query:
        return jsonify({"matches": [], "raw": text, "parsed_qty": qty})

    with get_conn() as conn:
        matches = _voice_fuzzy_match(med_query, conn, top_n=3)

    results = []
    for m in matches:
        results.append({
            "name":       m["name"],
            "generic":    m.get("generic") or "",
            "category":   m.get("category") or "",
            "mrp":        round(float(m.get("mrp") or 0), 2),
            "stock":      int(m.get("stock") or 0),
            "qty":        qty,
            "confidence": m.get("confidence", 75),
            "amount":     round(float(m.get("mrp") or 0) * qty, 2),
        })

    return jsonify({
        "matches":    results,
        "raw":        text,
        "parsed_qty": qty,
        "query":      med_query,
    })


# ── Stage 13 DB init (no new tables needed) ────────────────────────────────────
def init_stage13_db():
    """Stage 13 — Voice Billing + PWA. No new tables required."""
    pass


# ════════════════════════════════════════════════════════════════════
#  STAGE 15: Multi-tenant SaaS Billing + Plan Gating
# ════════════════════════════════════════════════════════════════════

PLAN_FEATURES: dict = {
    "starter": {
        "basic_billing": True, "stock_management": True, "reports": True,
        "ai_reorder": False, "tally": False, "delivery": False,
        "receiving": False, "wholesale": False, "multi_branch": False,
        "profit_analytics": False, "staff_targets": False, "crm": False,
        "white_label": False, "api_access": False,
        "bills_per_month": 300, "max_staff": 2, "max_medicines": 500,
    },
    "professional": {
        "basic_billing": True, "stock_management": True, "reports": True,
        "ai_reorder": True, "tally": True, "delivery": False,
        "receiving": True, "wholesale": False, "multi_branch": False,
        "profit_analytics": True, "staff_targets": True, "crm": True,
        "white_label": False, "api_access": False,
        "bills_per_month": 2000, "max_staff": 10, "max_medicines": -1,
    },
    "enterprise": {
        "basic_billing": True, "stock_management": True, "reports": True,
        "ai_reorder": True, "tally": True, "delivery": True,
        "receiving": True, "wholesale": True, "multi_branch": True,
        "profit_analytics": True, "staff_targets": True, "crm": True,
        "white_label": True, "api_access": True,
        "bills_per_month": -1, "max_staff": -1, "max_medicines": -1,
    },
}

PLAN_ORDER = {"starter": 0, "professional": 1, "enterprise": 2}

FEATURE_LABELS = {
    "basic_billing": "Basic Billing", "stock_management": "Stock Management",
    "reports": "Reports & Exports", "ai_reorder": "AI Auto-Reorder",
    "tally": "Tally / Accounts", "delivery": "Delivery Tracking",
    "receiving": "Smart Receiving", "wholesale": "Wholesale Portal",
    "multi_branch": "Multi-Branch", "profit_analytics": "Profit Analytics",
    "staff_targets": "Staff & Salary", "crm": "CRM / Patients",
    "white_label": "White-label", "api_access": "API Access",
}


def _normalize_plan(name: str) -> str:
    n = (name or "starter").lower()
    if "enterprise" in n:  return "enterprise"
    if "professional" in n or "pro" in n: return "professional"
    return "starter"


def get_shop_plan(shop_id: int, account_type: str = "retail") -> dict:
    with get_conn() as conn:
        sub = conn.execute("""
            SELECT s.*, sp.name AS plan_name_full, sp.price_monthly
            FROM subscriptions s
            LEFT JOIN subscription_plans sp ON sp.id = s.plan_id
            WHERE s.account_id = ? AND s.account_type = ?
            ORDER BY s.id DESC LIMIT 1
        """, (shop_id, account_type)).fetchone()
    if not sub:
        return {"plan": "starter", "status": "trial", "end_date": None,
                "features": PLAN_FEATURES["starter"], "sub": None}
    d = dict(sub)
    now_iso = datetime.now(timezone.utc).isoformat()
    end_date = d.get("end_date") or ""
    raw_status = d.get("status", "trial")
    if raw_status == "active" and end_date and end_date < now_iso:
        try:
            exp_dt = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
            if datetime.now(timezone.utc) <= exp_dt + timedelta(days=7):
                raw_status = "grace"
            else:
                raw_status = "expired"
        except Exception:
            raw_status = "expired"
    plan_key = _normalize_plan(d.get("plan_name") or d.get("plan_name_full") or "starter")
    if raw_status in ("trial", "active", "grace"):
        features = dict(PLAN_FEATURES.get(plan_key, PLAN_FEATURES["starter"]))
    else:
        features = {**PLAN_FEATURES["starter"], "bills_per_month": 50, "max_medicines": 100}
    d["status"] = raw_status
    return {"plan": plan_key, "status": raw_status, "end_date": end_date,
            "features": features, "sub": d}


def check_feature(feature_key: str, shop_id: int = None, account_type: str = "retail") -> tuple:
    if shop_id is None:
        shop_id = session.get("portal_user") or session.get("wholesale_user") or 0
    pi = get_shop_plan(int(shop_id), account_type)
    return bool(pi["features"].get(feature_key, False)), pi


def require_feature(feature_key: str, redirect_upgrade: bool = False):
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            shop_id = session.get("portal_user")
            if not shop_id:
                return redirect("/portal?login=1")
            allowed, pi = check_feature(feature_key, int(shop_id))
            if not allowed:
                if redirect_upgrade:
                    return redirect(f"/portal/subscription?upgrade=1&feature={feature_key}")
                return jsonify({
                    "error": "upgrade_required",
                    "feature": feature_key,
                    "feature_label": FEATURE_LABELS.get(feature_key, feature_key),
                    "current_plan": pi["plan"],
                    "message": (f"'{FEATURE_LABELS.get(feature_key, feature_key)}' requires a "
                                f"higher plan. You are on {pi['plan'].title()}.")
                }), 403
            return f(*args, **kwargs)
        return wrapped
    return decorator


def track_usage(shop_id: int, metric: str, amount: int = 1, account_type: str = "retail") -> None:
    col_map = {"bills": "bills_count", "ai_calls": "ai_calls",
               "api_calls": "api_calls", "staff_logins": "staff_logins"}
    col = col_map.get(metric, "api_calls")
    month = datetime.now(timezone.utc).strftime("%Y-%m")
    try:
        with get_conn() as conn:
            conn.execute(f"""
                INSERT INTO shop_usage (shop_id, account_type, month, {col})
                VALUES (?, ?, ?, ?)
                ON CONFLICT(shop_id, month) DO UPDATE SET {col} = {col} + ?
            """, (shop_id, account_type, month, amount, amount))
    except Exception:
        pass


def init_stage15_db() -> None:
    with get_conn() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS shop_usage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER NOT NULL, account_type TEXT DEFAULT 'retail',
            month TEXT NOT NULL, bills_count INTEGER DEFAULT 0,
            ai_calls INTEGER DEFAULT 0, api_calls INTEGER DEFAULT 0,
            staff_logins INTEGER DEFAULT 0,
            UNIQUE(shop_id, month))""")
        conn.execute("""CREATE TABLE IF NOT EXISTS saas_invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER NOT NULL, account_type TEXT DEFAULT 'retail',
            subscription_id INTEGER, plan_name TEXT, amount REAL NOT NULL,
            currency TEXT DEFAULT 'INR', billing_cycle TEXT DEFAULT 'monthly',
            period_start TEXT, period_end TEXT, status TEXT DEFAULT 'pending',
            razorpay_payment_id TEXT, invoice_number TEXT UNIQUE,
            paid_at TEXT, created_at TEXT)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS saas_trials (
            shop_id INTEGER NOT NULL, account_type TEXT DEFAULT 'retail',
            trial_start TEXT, trial_end TEXT, converted INTEGER DEFAULT 0,
            converted_plan TEXT, converted_at TEXT,
            PRIMARY KEY(shop_id, account_type))""")
        conn.execute("""CREATE TABLE IF NOT EXISTS plan_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER NOT NULL, account_type TEXT DEFAULT 'retail',
            event_type TEXT, from_plan TEXT, to_plan TEXT,
            amount REAL DEFAULT 0, razorpay_payment_id TEXT,
            notes TEXT, created_at TEXT)""")
        sub_cols = table_columns(conn, "subscriptions")
        if "grace_ends" not in sub_cols:
            conn.execute("ALTER TABLE subscriptions ADD COLUMN grace_ends TEXT DEFAULT ''")
        if "notes" not in sub_cols:
            conn.execute("ALTER TABLE subscriptions ADD COLUMN notes TEXT DEFAULT ''")


@app.route("/admin/saas")
def admin_saas_page():
    if not session.get("admin_user"):
        return redirect("/admin?login=1")
    return render_template("admin_saas.html")


@app.route("/api/saas/dashboard")
def saas_dashboard():
    if not session.get("admin_user"):
        return jsonify({"error": "unauthorized"}), 403
    month = datetime.now(timezone.utc).strftime("%Y-%m")
    with get_conn() as conn:
        active = conn.execute(
            "SELECT COUNT(*) AS c FROM subscriptions WHERE status IN ('active','trial','grace')"
        ).fetchone()["c"]
        mrr_row = conn.execute("""
            SELECT COALESCE(SUM(sp.price_monthly),0) AS mrr FROM subscriptions s
            JOIN subscription_plans sp ON sp.id=s.plan_id WHERE s.status='active'
        """).fetchone()
        mrr = float(mrr_row["mrr"] or 0)
        plan_dist = [dict(r) for r in conn.execute("""
            SELECT COALESCE(plan_name,'Unknown') AS plan_name, COUNT(*) AS cnt
            FROM subscriptions WHERE status IN ('active','trial','grace')
            GROUP BY plan_name ORDER BY cnt DESC
        """).fetchall()]
        rev_months = [dict(r) for r in conn.execute("""
            SELECT strftime('%Y-%m', paid_at) AS month, SUM(amount) AS revenue, COUNT(*) AS invoices
            FROM saas_invoices WHERE status='paid' GROUP BY month ORDER BY month DESC LIMIT 6
        """).fetchall()]
        trials    = conn.execute("SELECT COUNT(*) AS c FROM saas_trials WHERE converted=0").fetchone()["c"]
        converted = conn.execute("SELECT COUNT(*) AS c FROM saas_trials WHERE converted=1").fetchone()["c"]
        churn     = conn.execute(
            "SELECT COUNT(*) AS c FROM subscriptions WHERE status IN ('cancelled','expired')"
        ).fetchone()["c"]
        tenants = [dict(r) for r in conn.execute("""
            SELECT s.id, s.account_id, s.account_type, s.plan_name, s.status,
                   s.start_date, s.end_date, s.amount_paid, s.created_at,
                   COALESCE(sp.price_monthly,0) AS price_monthly,
                   COALESCE(rs.shop_name, wa.business_name, 'Shop #'||s.account_id) AS shop_name,
                   COALESCE(u.bills_count,0) AS bills_this_month
            FROM subscriptions s
            LEFT JOIN subscription_plans sp ON sp.id=s.plan_id
            LEFT JOIN retail_shops rs     ON rs.id=s.account_id AND s.account_type='retail'
            LEFT JOIN wholesale_accounts wa ON wa.id=s.account_id AND s.account_type='wholesale'
            LEFT JOIN shop_usage u        ON u.shop_id=s.account_id AND u.month=?
            ORDER BY s.created_at DESC LIMIT 200
        """, (month,)).fetchall()]
    return jsonify({
        "active_subscriptions": active, "mrr": mrr, "arr": round(mrr*12, 2),
        "trials": trials, "converted": converted, "churn": churn,
        "plan_distribution": plan_dist, "revenue_by_month": rev_months,
        "tenants": tenants,
    })


@app.route("/api/saas/tenants", methods=["GET"])
def all_tenants_s15():
    if not session.get("admin_user"):
        return jsonify({"error": "unauthorized"}), 403
    month = datetime.now(timezone.utc).strftime("%Y-%m")
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute("""
            SELECT s.id, s.account_id, s.account_type, s.plan_name, s.status,
                   s.start_date, s.end_date, s.amount_paid, s.created_at,
                   COALESCE(sp.price_monthly,0) AS price_monthly,
                   COALESCE(rs.shop_name, wa.business_name, 'Shop #'||s.account_id) AS shop_name,
                   COALESCE(u.bills_count,0) AS bills_this_month,
                   COALESCE(t.converted,-1) AS trial_status
            FROM subscriptions s
            LEFT JOIN subscription_plans sp ON sp.id=s.plan_id
            LEFT JOIN retail_shops rs      ON rs.id=s.account_id AND s.account_type='retail'
            LEFT JOIN wholesale_accounts wa ON wa.id=s.account_id AND s.account_type='wholesale'
            LEFT JOIN shop_usage u  ON u.shop_id=s.account_id AND u.month=?
            LEFT JOIN saas_trials t ON t.shop_id=s.account_id AND t.account_type=s.account_type
            ORDER BY s.created_at DESC
        """, (month,)).fetchall()]
    return jsonify(rows)


@app.route("/api/saas/admin/upgrade", methods=["POST"])
def admin_upgrade_plan_s15():
    if not session.get("admin_user"):
        return jsonify({"error": "unauthorized"}), 403
    data      = request.get_json(silent=True) or {}
    shop_id   = int(data.get("shop_id", 0))
    plan_id   = int(data.get("plan_id", 1))
    days      = int(data.get("days", 30))
    notes     = data.get("notes", "Admin manual grant")
    acct_type = data.get("account_type", "retail")
    with get_conn() as conn:
        plan = conn.execute("SELECT * FROM subscription_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({"error": "Plan not found"}), 404
        now     = datetime.now(timezone.utc)
        new_end = (now + timedelta(days=days)).isoformat()
        existing = conn.execute(
            "SELECT id, plan_name FROM subscriptions WHERE account_id=? AND account_type=?",
            (shop_id, acct_type)).fetchone()
        if existing:
            old_plan = existing["plan_name"] or "none"
            conn.execute("""UPDATE subscriptions SET plan_id=?, plan_name=?, status='active',
                end_date=?, updated_at=?, notes=? WHERE id=?""",
                (plan["id"], plan["name"], new_end, now.isoformat(), notes, existing["id"]))
        else:
            old_plan = "none"
            conn.execute("""INSERT INTO subscriptions
                (account_type,account_id,plan_id,plan_name,status,start_date,end_date,
                 billing_cycle,notes,created_at,updated_at)
                VALUES (?,?,?,?,'active',?,?,'manual',?,?,?)""",
                (acct_type, shop_id, plan["id"], plan["name"],
                 now.isoformat(), new_end, notes, now.isoformat(), now.isoformat()))
        conn.execute("""INSERT INTO plan_events
            (shop_id,account_type,event_type,from_plan,to_plan,notes,created_at)
            VALUES (?,?,?,?,?,?,?)""",
            (shop_id, acct_type, "admin_upgrade", old_plan, plan["name"], notes, now.isoformat()))
    return jsonify({"success": True, "plan": plan["name"], "end_date": new_end})


@app.route("/api/saas/admin/cancel", methods=["POST"])
def admin_cancel_s15():
    if not session.get("admin_user"):
        return jsonify({"error": "unauthorized"}), 403
    data    = request.get_json(silent=True) or {}
    shop_id = int(data.get("shop_id", 0))
    reason  = data.get("reason", "Admin cancel")
    now     = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        conn.execute("UPDATE subscriptions SET status='cancelled', updated_at=?, notes=? WHERE account_id=?",
                     (now, reason, shop_id))
        conn.execute("""INSERT INTO plan_events
            (shop_id,account_type,event_type,notes,created_at) VALUES (?,?,?,?,?)""",
            (shop_id, "retail", "admin_cancel", reason, now))
    return jsonify({"success": True})


@app.route("/api/saas/my-plan", methods=["GET"])
def my_plan_s15():
    shop_id = session.get("portal_user")
    if not shop_id:
        return jsonify({"error": "unauthorized"}), 403
    month = datetime.now(timezone.utc).strftime("%Y-%m")
    pi    = get_shop_plan(int(shop_id))
    with get_conn() as conn:
        usage = conn.execute(
            "SELECT * FROM shop_usage WHERE shop_id=? AND month=?", (shop_id, month)).fetchone()
        invoices = [dict(r) for r in conn.execute(
            "SELECT * FROM saas_invoices WHERE shop_id=? ORDER BY created_at DESC LIMIT 24",
            (shop_id,)).fetchall()]
        trial = conn.execute(
            "SELECT * FROM saas_trials WHERE shop_id=? AND account_type='retail'", (shop_id,)).fetchone()
    u = dict(usage) if usage else {"bills_count": 0, "ai_calls": 0, "api_calls": 0}
    lim = pi["features"]
    bills_pct = 0
    if lim.get("bills_per_month", -1) > 0 and int(u.get("bills_count", 0)) > 0:
        bills_pct = round(int(u["bills_count"]) / lim["bills_per_month"] * 100, 1)
    days_left = None
    if pi.get("end_date"):
        try:
            exp = datetime.fromisoformat(pi["end_date"].replace("Z", "+00:00"))
            days_left = max(0, (exp - datetime.now(timezone.utc)).days)
        except Exception:
            pass
    return jsonify({
        "plan": pi["plan"], "status": pi["status"],
        "end_date": pi.get("end_date"), "days_left": days_left,
        "features": pi["features"], "feature_labels": FEATURE_LABELS,
        "usage": u, "bills_pct": bills_pct,
        "limits": {"bills_per_month": lim.get("bills_per_month", 300),
                   "max_staff": lim.get("max_staff", 2)},
        "invoices": invoices,
        "trial": dict(trial) if trial else None,
        "plan_order": PLAN_ORDER,
    })


@app.route("/api/saas/start-trial", methods=["POST"])
def start_trial_s15():
    shop_id = session.get("portal_user")
    if not shop_id:
        return jsonify({"error": "unauthorized"}), 403
    data    = request.get_json(silent=True) or {}
    plan_id = int(data.get("plan_id", 2))
    with get_conn() as conn:
        existing_trial = conn.execute(
            "SELECT * FROM saas_trials WHERE shop_id=? AND account_type='retail'", (shop_id,)).fetchone()
        if existing_trial:
            return jsonify({"error": "Trial already used", "converted": bool(existing_trial["converted"])}), 400
        plan = conn.execute("SELECT * FROM subscription_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({"error": "Plan not found"}), 404
        now       = datetime.now(timezone.utc)
        trial_end = (now + timedelta(days=14)).isoformat()
        conn.execute("""INSERT OR REPLACE INTO saas_trials
            (shop_id, account_type, trial_start, trial_end, converted) VALUES (?,?,?,?,0)""",
            (shop_id, "retail", now.isoformat(), trial_end))
        existing_sub = conn.execute(
            "SELECT id FROM subscriptions WHERE account_id=? AND account_type='retail'", (shop_id,)).fetchone()
        if existing_sub:
            conn.execute("""UPDATE subscriptions SET plan_id=?, plan_name=?, status='trial',
                start_date=?, end_date=?, updated_at=? WHERE id=?""",
                (plan["id"], plan["name"], now.isoformat(), trial_end, now.isoformat(), existing_sub["id"]))
        else:
            conn.execute("""INSERT INTO subscriptions
                (account_type,account_id,plan_id,plan_name,status,start_date,end_date,
                 billing_cycle,created_at,updated_at) VALUES ('retail',?,?,?,'trial',?,?,'monthly',?,?)""",
                (shop_id, plan["id"], plan["name"], now.isoformat(), trial_end,
                 now.isoformat(), now.isoformat()))
        conn.execute("""INSERT INTO plan_events
            (shop_id,account_type,event_type,to_plan,notes,created_at) VALUES (?,?,?,?,?,?)""",
            (shop_id, "retail", "trial_start", plan["name"], "14-day free trial", now.isoformat()))
    return jsonify({"success": True, "trial_end": trial_end, "plan": plan["name"], "days": 14})


@app.route("/api/saas/check-limit", methods=["POST"])
def check_usage_limit_s15():
    shop_id = session.get("portal_user")
    if not shop_id:
        return jsonify({"error": "unauthorized"}), 403
    data  = request.get_json(silent=True) or {}
    pi    = get_shop_plan(int(shop_id))
    month = datetime.now(timezone.utc).strftime("%Y-%m")
    with get_conn() as conn:
        usage = conn.execute(
            "SELECT bills_count FROM shop_usage WHERE shop_id=? AND month=?", (shop_id, month)).fetchone()
    current = int(usage["bills_count"] if usage else 0)
    limit   = pi["features"].get("bills_per_month", 300)
    over    = (limit != -1 and current >= limit)
    return jsonify({
        "metric": data.get("metric", "bills"), "current": current, "limit": limit,
        "over_limit": over, "plan": pi["plan"],
        "pct_used": round(current / limit * 100, 1) if limit > 0 else 0,
    })


@app.route("/api/saas/features", methods=["GET"])
def my_features_s15():
    shop_id = session.get("portal_user")
    if not shop_id:
        return jsonify({"error": "unauthorized"}), 403
    pi = get_shop_plan(int(shop_id))
    return jsonify({"plan": pi["plan"], "status": pi["status"],
                    "features": pi["features"], "labels": FEATURE_LABELS})


@app.route("/api/razorpay/webhook", methods=["POST"])
def razorpay_webhook_s15():
    import hmac as _hmac
    secret        = os.environ.get("RAZORPAY_WEBHOOK_SECRET", "")
    payload_bytes = request.get_data()
    sig           = request.headers.get("X-Razorpay-Signature", "")
    if secret:
        expected = _hmac.new(secret.encode(), payload_bytes, hashlib.sha256).hexdigest()
        if not _hmac.compare_digest(expected, sig):
            return jsonify({"error": "bad signature"}), 400
    try:
        ev    = request.get_json(silent=True) or {}
        etype = ev.get("event", "")
        now   = datetime.now(timezone.utc).isoformat()
        with get_conn() as conn:
            if etype == "payment.captured":
                pmt    = ev.get("payload", {}).get("payment", {}).get("entity", {})
                pid    = pmt.get("id", "")
                oid    = pmt.get("order_id", "")
                amount = float(pmt.get("amount", 0)) / 100
                sub    = conn.execute(
                    "SELECT * FROM subscriptions WHERE razorpay_order_id=?", (oid,)).fetchone()
                if sub:
                    cycle   = sub["billing_cycle"] or "monthly"
                    days    = 365 if cycle == "yearly" else 30
                    new_end = (datetime.now(timezone.utc) + timedelta(days=days)).isoformat()
                    conn.execute("""UPDATE subscriptions SET status='active',
                        razorpay_payment_id=?, end_date=?, amount_paid=?, updated_at=? WHERE id=?""",
                        (pid, new_end, amount, now, sub["id"]))
                    inv_num = f"INV-{datetime.now(timezone.utc).strftime('%Y%m')}-{sub['id']:05d}"
                    conn.execute("""INSERT OR IGNORE INTO saas_invoices
                        (shop_id,account_type,subscription_id,plan_name,amount,billing_cycle,
                         period_start,period_end,status,razorpay_payment_id,invoice_number,paid_at,created_at)
                        VALUES (?,?,?,?,?,?,?,?,'paid',?,?,?,?)""",
                        (sub["account_id"], sub["account_type"], sub["id"], sub["plan_name"],
                         amount, cycle, now, new_end, pid, inv_num, now, now))
                    conn.execute("""UPDATE saas_trials SET converted=1, converted_plan=?,
                        converted_at=? WHERE shop_id=? AND account_type=?""",
                        (sub["plan_name"], now, sub["account_id"], sub["account_type"]))
            elif etype == "payment.failed":
                pmt = ev.get("payload", {}).get("payment", {}).get("entity", {})
                oid = pmt.get("order_id", "")
                conn.execute("""UPDATE subscriptions SET status='payment_failed', updated_at=?
                    WHERE razorpay_order_id=?""", (now, oid))
    except Exception:
        pass
    return jsonify({"status": "ok"}), 200


# ════════════════════════════════════════════════════════════════════
# STAGE 16 — Cash Drawer, Denominations, Expenses, Device Sharing
# ════════════════════════════════════════════════════════════════════
def init_stage16_db():
    with get_conn() as conn:
        # Cash denominations — opening + current count per denomination
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cash_drawer (
                denomination INTEGER PRIMARY KEY,  -- 1,2,5,10,20,50,100,200,500,2000
                kind         TEXT DEFAULT 'note',  -- 'note' or 'coin'
                count        INTEGER DEFAULT 0,
                min_alert    INTEGER DEFAULT 5,
                updated_at   TEXT DEFAULT (datetime('now'))
            )
        """)
        # Seed defaults if empty
        cnt = conn.execute("SELECT COUNT(*) FROM cash_drawer").fetchone()[0]
        if cnt == 0:
            defaults = [
                (1, 'coin', 20, 10),
                (2, 'coin', 20, 10),
                (5, 'coin', 30, 10),
                (10, 'note', 20, 10),
                (20, 'note', 25, 10),
                (50, 'note', 20, 8),
                (100, 'note', 30, 10),
                (200, 'note', 15, 5),
                (500, 'note', 20, 5),
                (2000, 'note', 5, 2),
            ]
            for d, k, c, m in defaults:
                conn.execute("INSERT INTO cash_drawer (denomination, kind, count, min_alert) VALUES (?,?,?,?)",
                             (d, k, c, m))

        # Every cash IN / OUT — sale, change, expense, exchange, opening, manual
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cash_movements (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                ts            TEXT DEFAULT (datetime('now')),
                kind          TEXT NOT NULL,    -- sale_in, change_out, expense_out, exchange_in, exchange_out, manual_in, manual_out, opening
                amount        REAL DEFAULT 0,
                breakdown     TEXT DEFAULT '{}',-- JSON {denom: count}
                ref_type      TEXT DEFAULT '',  -- bill / expense / exchange
                ref_id        TEXT DEFAULT '',
                staff_id      INTEGER DEFAULT 0,
                staff_name    TEXT DEFAULT '',
                note          TEXT DEFAULT ''
            )
        """)

        # Expenses (staff advance, tea/milk, supplier payment, credit, etc.)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                ts            TEXT DEFAULT (datetime('now')),
                date          TEXT NOT NULL,
                expense_type  TEXT NOT NULL,    -- staff_advance, tea, milk, supplier, credit, fuel, utility, rent, other (MANDATORY)
                category      TEXT DEFAULT '',  -- sub-category
                payee         TEXT DEFAULT '',  -- staff name, supplier name etc.
                payee_id      INTEGER DEFAULT 0,
                amount        REAL NOT NULL,
                payment_mode  TEXT DEFAULT 'Cash', -- Cash/UPI/Card
                breakdown     TEXT DEFAULT '{}',
                bill_no       TEXT DEFAULT '',  -- supplier bill ref
                notes         TEXT DEFAULT '',
                approved_by   TEXT DEFAULT '',
                staff_id      INTEGER DEFAULT 0,
                staff_name    TEXT DEFAULT '',
                created_at    TEXT DEFAULT (datetime('now'))
            )
        """)

        # Exchange events — customer gave 6x20 for 100 etc.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS exchange_events (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                ts              TEXT DEFAULT (datetime('now')),
                party           TEXT DEFAULT '',  -- 'customer' / 'petrol_bunk' / 'bank' / 'other'
                party_name      TEXT DEFAULT '',
                amount          REAL DEFAULT 0,
                given           TEXT DEFAULT '{}', -- JSON breakdown WE gave
                received        TEXT DEFAULT '{}', -- JSON breakdown WE received
                staff_id        INTEGER DEFAULT 0,
                staff_name      TEXT DEFAULT '',
                notes           TEXT DEFAULT ''
            )
        """)

        # Device registration for phone↔PC sharing
        conn.execute("""
            CREATE TABLE IF NOT EXISTS devices (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                device_id     TEXT UNIQUE NOT NULL,
                device_name   TEXT DEFAULT '',
                device_type   TEXT DEFAULT 'pc',   -- pc / phone / tablet
                staff_id      INTEGER DEFAULT 0,
                staff_name    TEXT DEFAULT '',
                last_seen     TEXT DEFAULT (datetime('now')),
                is_online     INTEGER DEFAULT 1,
                user_agent    TEXT DEFAULT ''
            )
        """)

        # Cross-device messages (face-scan results, bill sync, etc.)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS device_messages (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                ts            TEXT DEFAULT (datetime('now')),
                from_device   TEXT NOT NULL,
                to_device     TEXT NOT NULL,   -- specific id OR 'broadcast'
                msg_type      TEXT NOT NULL,   -- face_scan, bill_payload, cart_push, customer_select
                payload       TEXT DEFAULT '{}',
                status        TEXT DEFAULT 'pending',  -- pending / accepted / ignored / consumed
                consumed_at   TEXT DEFAULT '',
                expires_at    TEXT DEFAULT ''
            )
        """)

        # Wanted list (since /wanted now has a template)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS wanted_list (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                ts           TEXT DEFAULT (datetime('now')),
                medicine     TEXT NOT NULL,
                customer     TEXT DEFAULT '',
                phone        TEXT DEFAULT '',
                qty          INTEGER DEFAULT 1,
                priority     TEXT DEFAULT 'medium',  -- low / medium / high
                status       TEXT DEFAULT 'open',    -- open / fulfilled / closed
                notes        TEXT DEFAULT '',
                fulfilled_at TEXT DEFAULT ''
            )
        """)


# ── CASH DRAWER APIs ────────────────────────────────────────────────
@app.route("/cash-drawer")
def cash_drawer_page():
    return render_template("cash_drawer.html")


@app.route("/api/cash/drawer", methods=["GET"])
def api_cash_drawer():
    """Return current denomination counts + total + alerts."""
    with get_conn() as conn:
        rows = conn.execute("SELECT denomination, kind, count, min_alert, updated_at FROM cash_drawer ORDER BY denomination DESC").fetchall()
    items = []
    total = 0
    alerts = []
    for r in rows:
        d = dict(r)
        d["value"] = d["denomination"] * d["count"]
        total += d["value"]
        if d["count"] <= d["min_alert"]:
            alerts.append({"denomination": d["denomination"], "kind": d["kind"], "count": d["count"], "min": d["min_alert"]})
        items.append(d)
    return jsonify({"items": items, "total_cash": total, "alerts": alerts})


@app.route("/api/cash/drawer/adjust", methods=["POST"])
def api_cash_drawer_adjust():
    """Manual adjustment of a denomination count (with audit log)."""
    d = request.json or {}
    denom = int(d.get("denomination", 0))
    new_count = int(d.get("count", 0))
    note = d.get("note", "Manual adjustment")
    if denom <= 0:
        return jsonify({"error": "denomination required"}), 400
    with get_conn() as conn:
        cur = conn.execute("SELECT count FROM cash_drawer WHERE denomination=?", (denom,)).fetchone()
        if not cur:
            return jsonify({"error": "denomination not configured"}), 404
        old = cur[0]
        diff = new_count - old
        conn.execute("UPDATE cash_drawer SET count=?, updated_at=datetime('now') WHERE denomination=?", (new_count, denom))
        conn.execute("""INSERT INTO cash_movements (kind, amount, breakdown, ref_type, note, staff_name)
                        VALUES (?,?,?,?,?,?)""",
                     ("manual_in" if diff >= 0 else "manual_out",
                      abs(diff * denom),
                      json.dumps({str(denom): abs(diff)}),
                      "adjust", note, session.get("staff_name", "")))
    return jsonify({"status": "ok", "old": old, "new": new_count})


@app.route("/api/cash/drawer/threshold", methods=["POST"])
def api_cash_drawer_threshold():
    """Update alert threshold for a denomination."""
    d = request.json or {}
    denom = int(d.get("denomination", 0))
    min_alert = int(d.get("min_alert", 0))
    with get_conn() as conn:
        conn.execute("UPDATE cash_drawer SET min_alert=? WHERE denomination=?", (min_alert, denom))
    return jsonify({"status": "ok"})


def _suggest_change(amount, drawer):
    """ML-lite: best denomination combo for `amount` given availability."""
    # Greedy: pick largest available first, never use more than available
    result = {}
    remaining = round(amount)
    # drawer is dict {denom: count}
    for denom in sorted(drawer.keys(), reverse=True):
        if remaining <= 0:
            break
        avail = drawer[denom]
        if avail <= 0:
            continue
        need = remaining // denom
        use = min(need, avail)
        if use > 0:
            result[denom] = use
            remaining -= use * denom
    return result, remaining  # remaining > 0 means we can't fully give


@app.route("/api/cash/suggest-change", methods=["POST"])
def api_cash_suggest_change():
    """Given a balance amount, suggest the optimal denomination breakdown."""
    d = request.json or {}
    amount = float(d.get("amount", 0))
    if amount <= 0:
        return jsonify({"breakdown": {}, "shortage": 0})
    with get_conn() as conn:
        rows = conn.execute("SELECT denomination, count FROM cash_drawer").fetchall()
    drawer = {r[0]: r[1] for r in rows}
    breakdown, shortage = _suggest_change(amount, drawer)
    return jsonify({"breakdown": {str(k): v for k, v in breakdown.items()},
                    "shortage": shortage,
                    "amount_requested": amount,
                    "amount_can_give": amount - shortage})


@app.route("/api/cash/record-bill", methods=["POST"])
def api_cash_record_bill():
    """When bill is paid in cash, record received notes + dispensed change."""
    d = request.json or {}
    bill_id = d.get("bill_id", "")
    received = d.get("received", {})    # {denom: count} customer handed over
    given_change = d.get("given_change", {})  # {denom: count} we returned
    if not received:
        return jsonify({"error": "received breakdown required"}), 400
    received_total = sum(int(k) * int(v) for k, v in received.items())
    given_total = sum(int(k) * int(v) for k, v in given_change.items())
    bill_amount = received_total - given_total

    with get_conn() as conn:
        # Add received notes to drawer
        for denom, count in received.items():
            conn.execute("UPDATE cash_drawer SET count = count + ?, updated_at=datetime('now') WHERE denomination=?",
                         (int(count), int(denom)))
        # Remove given change
        for denom, count in given_change.items():
            conn.execute("UPDATE cash_drawer SET count = MAX(0, count - ?), updated_at=datetime('now') WHERE denomination=?",
                         (int(count), int(denom)))
        # Log movements
        conn.execute("""INSERT INTO cash_movements (kind, amount, breakdown, ref_type, ref_id, staff_name)
                        VALUES (?,?,?,?,?,?)""",
                     ("sale_in", received_total, json.dumps(received), "bill", bill_id,
                      session.get("staff_name", "")))
        if given_total > 0:
            conn.execute("""INSERT INTO cash_movements (kind, amount, breakdown, ref_type, ref_id, staff_name)
                            VALUES (?,?,?,?,?,?)""",
                         ("change_out", given_total, json.dumps(given_change), "bill", bill_id,
                          session.get("staff_name", "")))
    return jsonify({"status": "ok", "bill_amount": bill_amount,
                    "received_total": received_total, "change_given": given_total})


@app.route("/api/cash/movements", methods=["GET"])
def api_cash_movements():
    """Recent movements (last 100)."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM cash_movements ORDER BY id DESC LIMIT 100"
        ).fetchall()
    return jsonify({"movements": [dict(r) for r in rows]})


@app.route("/api/cash/exchange", methods=["POST"])
def api_cash_exchange():
    """Record an exchange (customer gave 6×20 for 100, or petrol bunk 1000 → mixed)."""
    d = request.json or {}
    party = d.get("party", "customer")
    party_name = d.get("party_name", "")
    given = d.get("given", {})       # what we gave them
    received = d.get("received", {})  # what we got
    notes = d.get("notes", "")

    given_total = sum(int(k) * int(v) for k, v in given.items())
    received_total = sum(int(k) * int(v) for k, v in received.items())

    with get_conn() as conn:
        # We received - add to drawer
        for denom, count in received.items():
            conn.execute("UPDATE cash_drawer SET count = count + ?, updated_at=datetime('now') WHERE denomination=?",
                         (int(count), int(denom)))
        # We gave - remove
        for denom, count in given.items():
            conn.execute("UPDATE cash_drawer SET count = MAX(0, count - ?), updated_at=datetime('now') WHERE denomination=?",
                         (int(count), int(denom)))
        cur = conn.execute("""INSERT INTO exchange_events (party, party_name, amount, given, received, notes, staff_name)
                              VALUES (?,?,?,?,?,?,?)""",
                           (party, party_name, max(given_total, received_total),
                            json.dumps(given), json.dumps(received), notes,
                            session.get("staff_name", "")))
        eid = cur.lastrowid
        if given_total > 0:
            conn.execute("""INSERT INTO cash_movements (kind, amount, breakdown, ref_type, ref_id, note, staff_name)
                            VALUES (?,?,?,?,?,?,?)""",
                         ("exchange_out", given_total, json.dumps(given), "exchange", str(eid),
                          f"Exchange with {party_name}", session.get("staff_name", "")))
        if received_total > 0:
            conn.execute("""INSERT INTO cash_movements (kind, amount, breakdown, ref_type, ref_id, note, staff_name)
                            VALUES (?,?,?,?,?,?,?)""",
                         ("exchange_in", received_total, json.dumps(received), "exchange", str(eid),
                          f"Exchange with {party_name}", session.get("staff_name", "")))
    return jsonify({"status": "ok", "id": eid,
                    "given_total": given_total, "received_total": received_total})


@app.route("/api/cash/exchanges", methods=["GET"])
def api_cash_exchanges():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM exchange_events ORDER BY id DESC LIMIT 50").fetchall()
    out = []
    for r in rows:
        d = dict(r)
        try: d["given"] = json.loads(d["given"] or "{}")
        except: d["given"] = {}
        try: d["received"] = json.loads(d["received"] or "{}")
        except: d["received"] = {}
        out.append(d)
    return jsonify({"exchanges": out})


# ── EXPENSES APIs ──────────────────────────────────────────────────
@app.route("/expenses")
def expenses_page():
    return render_template("expenses.html")


EXPENSE_TYPES = ["staff_advance", "tea", "milk", "supplier", "credit", "fuel", "utility", "rent", "stationery", "maintenance", "other"]


@app.route("/api/expenses", methods=["GET"])
def api_expenses_list():
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    etype = request.args.get("type", "")
    q = "SELECT * FROM expenses WHERE 1=1"
    p = []
    if date_from:
        q += " AND date >= ?"; p.append(date_from)
    if date_to:
        q += " AND date <= ?"; p.append(date_to)
    if etype:
        q += " AND expense_type = ?"; p.append(etype)
    q += " ORDER BY id DESC LIMIT 500"
    with get_conn() as conn:
        rows = conn.execute(q, p).fetchall()
    expenses = [dict(r) for r in rows]
    total = sum(e["amount"] for e in expenses)
    by_type = {}
    for e in expenses:
        by_type[e["expense_type"]] = by_type.get(e["expense_type"], 0) + e["amount"]
    return jsonify({"expenses": expenses, "total": total, "by_type": by_type,
                    "types": EXPENSE_TYPES})


@app.route("/api/expenses", methods=["POST"])
def api_expenses_create():
    d = request.json or {}
    # MANDATORY: expense_type
    etype = (d.get("expense_type") or "").strip()
    if not etype:
        return jsonify({"error": "expense_type is mandatory", "types": EXPENSE_TYPES}), 400
    if etype not in EXPENSE_TYPES:
        return jsonify({"error": f"invalid expense_type. Allowed: {EXPENSE_TYPES}"}), 400
    amount = float(d.get("amount", 0))
    if amount <= 0:
        return jsonify({"error": "amount required"}), 400

    date = d.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    payee = d.get("payee", "")
    payment_mode = d.get("payment_mode", "Cash")
    notes = d.get("notes", "")
    bill_no = d.get("bill_no", "")
    category = d.get("category", "")
    breakdown = d.get("breakdown", {})  # cash denomination if paid in cash

    with get_conn() as conn:
        cur = conn.execute("""INSERT INTO expenses
            (date, expense_type, category, payee, amount, payment_mode, breakdown, bill_no, notes, staff_id, staff_name)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (date, etype, category, payee, amount, payment_mode, json.dumps(breakdown), bill_no, notes,
             session.get("staff_id", 0), session.get("staff_name", "")))
        eid = cur.lastrowid

        # If Cash + has breakdown → deduct from drawer
        if payment_mode == "Cash" and breakdown:
            for denom, count in breakdown.items():
                conn.execute("UPDATE cash_drawer SET count = MAX(0, count - ?), updated_at=datetime('now') WHERE denomination=?",
                             (int(count), int(denom)))
            conn.execute("""INSERT INTO cash_movements (kind, amount, breakdown, ref_type, ref_id, note, staff_name)
                            VALUES (?,?,?,?,?,?,?)""",
                         ("expense_out", amount, json.dumps(breakdown), "expense", str(eid),
                          f"{etype}: {payee}", session.get("staff_name", "")))
    return jsonify({"status": "ok", "id": eid})


@app.route("/api/expenses/<int:eid>", methods=["DELETE"])
def api_expenses_delete(eid):
    with get_conn() as conn:
        conn.execute("DELETE FROM expenses WHERE id=?", (eid,))
    return jsonify({"status": "ok"})


@app.route("/api/expenses/summary", methods=["GET"])
def api_expenses_summary():
    """Today / this week / this month summary."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    month_start = today[:7] + "-01"
    with get_conn() as conn:
        today_total = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date=?", (today,)).fetchone()[0]
        month_total = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date>=?", (month_start,)).fetchone()[0]
        by_type_month = conn.execute("""SELECT expense_type, COUNT(*), SUM(amount)
                                        FROM expenses WHERE date>=?
                                        GROUP BY expense_type ORDER BY SUM(amount) DESC""", (month_start,)).fetchall()
    return jsonify({
        "today": today_total,
        "month": month_total,
        "by_type_month": [{"type": r[0], "count": r[1], "amount": r[2]} for r in by_type_month]
    })


# ── DEVICE PAIRING + SHARING APIs ──────────────────────────────────
@app.route("/api/devices/register", methods=["POST"])
def api_device_register():
    d = request.json or {}
    device_id = d.get("device_id") or hashlib.md5(f"{time.time()}{random.random()}".encode()).hexdigest()[:12]
    device_name = d.get("device_name", "Unnamed")
    device_type = d.get("device_type", "pc")
    user_agent = d.get("user_agent", request.headers.get("User-Agent", "")[:200])
    with get_conn() as conn:
        conn.execute("""INSERT OR REPLACE INTO devices
            (device_id, device_name, device_type, staff_id, staff_name, last_seen, is_online, user_agent)
            VALUES (?,?,?,?,?, datetime('now'), 1, ?)""",
            (device_id, device_name, device_type,
             session.get("staff_id", 0), session.get("staff_name", ""), user_agent))
    return jsonify({"status": "ok", "device_id": device_id})


@app.route("/api/devices", methods=["GET"])
def api_devices_list():
    with get_conn() as conn:
        # Online if seen in last 60s
        conn.execute("UPDATE devices SET is_online=0 WHERE datetime(last_seen) < datetime('now','-60 seconds')")
        rows = conn.execute("SELECT * FROM devices ORDER BY last_seen DESC LIMIT 50").fetchall()
    return jsonify({"devices": [dict(r) for r in rows]})


@app.route("/api/devices/heartbeat", methods=["POST"])
def api_device_heartbeat():
    d = request.json or {}
    device_id = d.get("device_id", "")
    if not device_id:
        return jsonify({"error": "device_id required"}), 400
    with get_conn() as conn:
        conn.execute("UPDATE devices SET last_seen=datetime('now'), is_online=1 WHERE device_id=?", (device_id,))
    return jsonify({"status": "ok"})


@app.route("/api/devices/send", methods=["POST"])
def api_device_send():
    """Push a message to another device (face_scan, bill_payload, customer_select)."""
    d = request.json or {}
    from_device = d.get("from_device", "")
    to_device = d.get("to_device", "broadcast")
    msg_type = d.get("msg_type", "")
    payload = d.get("payload", {})
    if not msg_type:
        return jsonify({"error": "msg_type required"}), 400
    with get_conn() as conn:
        cur = conn.execute("""INSERT INTO device_messages
            (from_device, to_device, msg_type, payload, expires_at)
            VALUES (?,?,?,?, datetime('now','+30 minutes'))""",
            (from_device, to_device, msg_type, json.dumps(payload)))
    return jsonify({"status": "ok", "id": cur.lastrowid})


@app.route("/api/devices/poll/<device_id>", methods=["GET"])
def api_device_poll(device_id):
    """Long-poll-style — return pending messages for this device."""
    with get_conn() as conn:
        # Touch heartbeat
        conn.execute("UPDATE devices SET last_seen=datetime('now'), is_online=1 WHERE device_id=?", (device_id,))
        rows = conn.execute("""SELECT * FROM device_messages
            WHERE (to_device=? OR to_device='broadcast')
              AND status='pending'
              AND (expires_at='' OR datetime(expires_at) > datetime('now'))
            ORDER BY id ASC LIMIT 20""", (device_id,)).fetchall()
    msgs = []
    for r in rows:
        d = dict(r)
        try: d["payload"] = json.loads(d["payload"] or "{}")
        except: d["payload"] = {}
        msgs.append(d)
    return jsonify({"messages": msgs})


@app.route("/api/devices/ack/<int:msg_id>", methods=["POST"])
def api_device_ack(msg_id):
    """Mark a message accepted / ignored / consumed."""
    d = request.json or {}
    status = d.get("status", "consumed")  # accepted / ignored / consumed
    with get_conn() as conn:
        conn.execute("UPDATE device_messages SET status=?, consumed_at=datetime('now') WHERE id=?",
                     (status, msg_id))
    return jsonify({"status": "ok"})


# ── WANTED LIST APIs (was missing) ─────────────────────────────────
@app.route("/api/wanted", methods=["GET"])
def api_wanted_list():
    status = request.args.get("status", "")
    q = "SELECT * FROM wanted_list"
    p = []
    if status:
        q += " WHERE status=?"; p.append(status)
    q += " ORDER BY id DESC LIMIT 200"
    with get_conn() as conn:
        rows = conn.execute(q, p).fetchall()
    return jsonify({"items": [dict(r) for r in rows]})


@app.route("/api/wanted", methods=["POST"])
def api_wanted_create():
    d = request.json or {}
    if not d.get("medicine"):
        return jsonify({"error": "medicine required"}), 400
    with get_conn() as conn:
        cur = conn.execute("""INSERT INTO wanted_list
            (medicine, customer, phone, qty, priority, notes)
            VALUES (?,?,?,?,?,?)""",
            (d.get("medicine", ""), d.get("customer", ""), d.get("phone", ""),
             int(d.get("qty", 1)), d.get("priority", "medium"), d.get("notes", "")))
    return jsonify({"status": "ok", "id": cur.lastrowid})


@app.route("/api/wanted/<int:wid>", methods=["DELETE"])
def api_wanted_delete(wid):
    with get_conn() as conn:
        conn.execute("DELETE FROM wanted_list WHERE id=?", (wid,))
    return jsonify({"status": "ok"})


@app.route("/api/wanted/<int:wid>/fulfill", methods=["POST"])
def api_wanted_fulfill(wid):
    with get_conn() as conn:
        conn.execute("UPDATE wanted_list SET status='fulfilled', fulfilled_at=datetime('now') WHERE id=?", (wid,))
    return jsonify({"status": "ok"})


# ── BILL RECEIPT PAGE (printable + WhatsApp share) ─────────────────
@app.route("/bill/<bill_id>")
def bill_receipt_page(bill_id):
    with get_conn() as conn:
        r = conn.execute("SELECT * FROM bills WHERE id=?", (bill_id,)).fetchone()
        if not r:
            return render_template("bill_receipt.html", bill=None, items=[])
        bill = dict(r)
        try: items = json.loads(bill.get("items") or "[]")
        except: items = []
    return render_template("bill_receipt.html", bill=bill, items=items)


@app.route("/mobile-bill")
def mobile_billing_page():
    return render_template("mobile_billing.html")


# ════════════════════════════════════════════════════════════════════
# STAGE 17 — Refill Reminders, AI Calls, Pick→Pack Queue, Dosage
# ════════════════════════════════════════════════════════════════════
def init_stage17_db():
    with get_conn() as conn:
        # Reminder schedules (global defaults + per-customer overrides)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS reminder_schedules (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                scope           TEXT NOT NULL,            -- 'global' or 'customer'
                customer_phone  TEXT DEFAULT '',          -- for per-customer override
                category        TEXT DEFAULT '',          -- 'chronic' / 'antibiotic' / 'eye_drops' / 'all'
                days_before     INTEGER DEFAULT 3,        -- call N days before stock-out
                call_time       TEXT DEFAULT '10:00',
                language        TEXT DEFAULT 'ta',        -- ta/en/hi
                channel         TEXT DEFAULT 'both',      -- whatsapp/call/both
                is_active       INTEGER DEFAULT 1,
                greeting        TEXT DEFAULT '',
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)
        # Seed defaults
        cnt = conn.execute("SELECT COUNT(*) FROM reminder_schedules WHERE scope='global'").fetchone()[0]
        if cnt == 0:
            conn.execute("""INSERT INTO reminder_schedules
                (scope, category, days_before, call_time, language, channel, greeting)
                VALUES ('global', 'all', 3, '10:00', 'ta', 'both',
                        'வணக்கம்! Selvam Medicals-ல் இருந்து அழைக்கிறோம். உங்கள் மருந்து 3 நாட்களில் முடிந்துவிடும்.')""")

        # Call queue + history
        conn.execute("""
            CREATE TABLE IF NOT EXISTS call_queue (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_name   TEXT,
                customer_phone  TEXT NOT NULL,
                kind            TEXT DEFAULT 'refill',    -- refill / followup / birthday / outofstock / promo
                medicine        TEXT DEFAULT '',
                expected_stockout TEXT DEFAULT '',
                language        TEXT DEFAULT 'ta',
                scheduled_at    TEXT DEFAULT (datetime('now')),
                status          TEXT DEFAULT 'queued',     -- queued / calling / done / press1 / press2 / failed / skipped
                attempts        INTEGER DEFAULT 0,
                response        TEXT DEFAULT '',           -- press1 / press2 / voicemail
                whatsapp_sent   INTEGER DEFAULT 0,
                whatsapp_text   TEXT DEFAULT '',
                resulting_bill  TEXT DEFAULT '',           -- auto-created bill id on press1
                notes           TEXT DEFAULT '',
                called_at       TEXT DEFAULT '',
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)

        # AI Call billing/subscription
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ai_call_subscriptions (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_name       TEXT NOT NULL,             -- starter / pro / unlimited
                price_per_month REAL DEFAULT 0,
                included_calls  INTEGER DEFAULT 0,
                extra_call_cost REAL DEFAULT 0,
                whatsapp_included INTEGER DEFAULT 0,
                is_active       INTEGER DEFAULT 1,
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)
        # Seed 3 plans if empty
        if conn.execute("SELECT COUNT(*) FROM ai_call_subscriptions").fetchone()[0] == 0:
            for r in [
                ('Starter', 499, 100, 2.50, 500),
                ('Pro', 1499, 500, 1.80, 2000),
                ('Unlimited', 3999, 9999, 0, 9999),
            ]:
                conn.execute("""INSERT INTO ai_call_subscriptions
                    (plan_name, price_per_month, included_calls, extra_call_cost, whatsapp_included)
                    VALUES (?,?,?,?,?)""", r)

        # Pick / Pack queue — bills get a workflow status
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pickpack_queue (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                bill_id         TEXT NOT NULL,
                customer_name   TEXT DEFAULT '',
                customer_phone  TEXT DEFAULT '',
                source          TEXT DEFAULT 'manual',     -- manual / call_press1 / whatsapp_reply / shop_portal
                items_json      TEXT DEFAULT '[]',
                total           REAL DEFAULT 0,
                status          TEXT DEFAULT 'queued',     -- queued / picking / packing / ready / dispatched / cancelled
                priority        TEXT DEFAULT 'normal',     -- urgent / high / normal / low
                assigned_to     TEXT DEFAULT '',
                notes           TEXT DEFAULT '',
                created_at      TEXT DEFAULT (datetime('now')),
                started_at      TEXT DEFAULT '',
                ready_at        TEXT DEFAULT '',
                dispatched_at   TEXT DEFAULT ''
            )
        """)

        # Shop busy mode toggle (system-wide flag)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS shop_state (
                key   TEXT PRIMARY KEY,
                value TEXT,
                updated_at TEXT DEFAULT (datetime('now'))
            )
        """)
        # Default: not busy
        conn.execute("INSERT OR IGNORE INTO shop_state (key, value) VALUES ('busy_mode', '0')")
        conn.execute("INSERT OR IGNORE INTO shop_state (key, value) VALUES ('busy_message', 'Currently busy — bills will be processed shortly')")

        # Dosage knowledge base — common Indian medicines
        conn.execute("""
            CREATE TABLE IF NOT EXISTS dosage_kb (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                medicine_pattern TEXT NOT NULL UNIQUE,    -- DOLO, CROCIN, PAN, BRUFEN
                dosage_en     TEXT,
                dosage_ta     TEXT,
                timing        TEXT,                       -- before/after food, morning/night
                kind          TEXT                        -- tablet/syrup/drops/cream
            )
        """)
        # Seed common dosage patterns
        if conn.execute("SELECT COUNT(*) FROM dosage_kb").fetchone()[0] == 0:
            seeds = [
                ("DOLO 650", "1 tablet after food, max 4/day, gap 6 hrs", "உணவுக்குப் பிறகு 1 மாத்திரை, தினமும் 4-க்கு மேல் வேண்டாம், 6 மணி நேர இடைவெளி", "after_food", "tablet"),
                ("CROCIN", "1 tablet every 6 hrs after food", "ஒவ்வொரு 6 மணி நேரத்திற்கும் 1 மாத்திரை, உணவுக்குப் பிறகு", "after_food", "tablet"),
                ("PAN 40", "1 tablet morning, before food", "காலையில் 1 மாத்திரை, உணவுக்கு முன்", "before_food", "tablet"),
                ("PANTOP 40", "1 tablet morning, before food", "காலையில் 1 மாத்திரை, உணவுக்கு முன்", "before_food", "tablet"),
                ("BRUFEN", "1 tablet 2-3 times after food", "உணவுக்குப் பிறகு தினமும் 2-3 முறை 1 மாத்திரை", "after_food", "tablet"),
                ("AUGMENTIN", "1 tablet every 12 hrs, complete full course", "12 மணி நேரத்திற்கு 1 முறை. முழு course-ஐ முடிக்கவும்", "with_food", "tablet"),
                ("AZITHRAL", "1 tablet daily after food, 3-5 days", "தினமும் 1 மாத்திரை, உணவுக்குப் பிறகு, 3-5 நாட்கள்", "after_food", "tablet"),
                ("METFORMIN", "1 tablet after meals, twice daily", "உணவுக்குப் பிறகு தினமும் 2 முறை", "after_food", "tablet"),
                ("AMLODIPINE", "1 tablet daily morning", "காலையில் தினமும் 1 மாத்திரை", "morning", "tablet"),
                ("ATORVASTATIN", "1 tablet at bedtime", "இரவில் படுக்கைக்கு முன் 1 மாத்திரை", "night", "tablet"),
                ("LEVOCETIRIZINE", "1 tablet at night", "இரவில் 1 மாத்திரை", "night", "tablet"),
                ("CETIRIZINE", "1 tablet at night", "இரவில் 1 மாத்திரை", "night", "tablet"),
                ("EYE DROP", "2 drops 3-4 times daily in affected eye", "பாதிக்கப்பட்ட கண்ணில் தினமும் 3-4 முறை 2 சொட்டுகள்", "as_needed", "drops"),
                ("EAR DROP", "2-3 drops in affected ear 2-3 times daily", "பாதிக்கப்பட்ட காதில் தினமும் 2-3 முறை 2-3 சொட்டுகள்", "as_needed", "drops"),
                ("SYP", "5-10 ml as per age, after food", "வயதுக்கு ஏற்றபடி 5-10 ml, உணவுக்குப் பிறகு", "after_food", "syrup"),
                ("CREAM", "Apply thin layer 2-3 times daily on affected area", "பாதிப்பு உள்ள இடத்தில் தினமும் 2-3 முறை மெல்லிய அடுக்கு பூசவும்", "topical", "cream"),
                ("OINTMENT", "Apply 2-3 times daily on affected area", "பாதிப்பு உள்ள இடத்தில் தினமும் 2-3 முறை பூசவும்", "topical", "cream"),
                ("INSULIN", "Inject as per prescribed units before food", "உணவுக்கு முன் குறிப்பிட்ட அளவில் ஊசி", "before_food", "injection"),
            ]
            for s in seeds:
                conn.execute("""INSERT OR IGNORE INTO dosage_kb (medicine_pattern, dosage_en, dosage_ta, timing, kind)
                                VALUES (?,?,?,?,?)""", s)


# ── DOSAGE LOOKUP API ──────────────────────────────────────────────
@app.route("/api/dosage/lookup", methods=["GET"])
def api_dosage_lookup():
    """Look up dosage instructions for a medicine name."""
    name = (request.args.get("name") or "").upper()
    if not name:
        return jsonify({"dosage_en": "", "dosage_ta": "", "match": ""})
    with get_conn() as conn:
        rows = conn.execute("SELECT medicine_pattern, dosage_en, dosage_ta, timing, kind FROM dosage_kb").fetchall()
    # Find best match
    for r in rows:
        if r["medicine_pattern"] in name:
            return jsonify({
                "match": r["medicine_pattern"],
                "dosage_en": r["dosage_en"],
                "dosage_ta": r["dosage_ta"],
                "timing": r["timing"],
                "kind": r["kind"]
            })
    return jsonify({"match": "", "dosage_en": "", "dosage_ta": ""})


@app.route("/api/dosage/all", methods=["GET"])
def api_dosage_all():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM dosage_kb ORDER BY medicine_pattern").fetchall()
    return jsonify({"items": [dict(r) for r in rows]})


@app.route("/api/dosage", methods=["POST"])
def api_dosage_add():
    d = request.json or {}
    if not d.get("medicine_pattern"):
        return jsonify({"error": "medicine_pattern required"}), 400
    with get_conn() as conn:
        conn.execute("""INSERT OR REPLACE INTO dosage_kb
            (medicine_pattern, dosage_en, dosage_ta, timing, kind)
            VALUES (?,?,?,?,?)""",
            (d.get("medicine_pattern", "").upper(), d.get("dosage_en", ""),
             d.get("dosage_ta", ""), d.get("timing", ""), d.get("kind", "")))
    return jsonify({"status": "ok"})


# ── REFILL REMINDER SYSTEM ──────────────────────────────────────────
@app.route("/refill-reminders")
def refill_reminders_page():
    return render_template("refill_reminders.html")


@app.route("/api/reminders/schedules", methods=["GET"])
def api_reminder_schedules():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM reminder_schedules ORDER BY scope, id").fetchall()
    return jsonify({"schedules": [dict(r) for r in rows]})


@app.route("/api/reminders/schedules", methods=["POST"])
def api_reminder_schedule_create():
    d = request.json or {}
    with get_conn() as conn:
        conn.execute("""INSERT INTO reminder_schedules
            (scope, customer_phone, category, days_before, call_time, language, channel, greeting)
            VALUES (?,?,?,?,?,?,?,?)""",
            (d.get("scope", "customer"), d.get("customer_phone", ""), d.get("category", "all"),
             int(d.get("days_before", 3)), d.get("call_time", "10:00"),
             d.get("language", "ta"), d.get("channel", "both"),
             d.get("greeting", "")))
    return jsonify({"status": "ok"})


@app.route("/api/reminders/schedules/<int:sid>", methods=["DELETE"])
def api_reminder_schedule_delete(sid):
    with get_conn() as conn:
        conn.execute("DELETE FROM reminder_schedules WHERE id=?", (sid,))
    return jsonify({"status": "ok"})


@app.route("/api/reminders/auto-detect", methods=["POST"])
def api_reminders_auto_detect():
    """Analyse last bills to detect customers about to run out of stock; queue calls."""
    with get_conn() as conn:
        # Get global setting
        glob = conn.execute("SELECT days_before, language, channel, greeting FROM reminder_schedules WHERE scope='global' LIMIT 1").fetchone()
        days_before = glob["days_before"] if glob else 3
        # Customers with chronic conditions or repeat bills in last 60 days
        # For each customer, check their last bill date; if items typically last 30 days, alert 27 days in
        rows = conn.execute("""
            SELECT phone, cust, MAX(date) as last_date, COUNT(*) as bill_count, MAX(items) as last_items, SUM(total) as spend
            FROM bills WHERE phone != '' AND phone IS NOT NULL
            GROUP BY phone HAVING bill_count >= 2
            ORDER BY last_date DESC LIMIT 50
        """).fetchall()
        queued = 0
        for r in rows:
            try:
                # Estimate stockout = last_date + 28 days
                last = datetime.strptime(r["last_date"], "%Y-%m-%d") if len(r["last_date"])==10 else datetime.now()
                stockout = last + timedelta(days=28)
                call_date = stockout - timedelta(days=days_before)
                if call_date.date() <= datetime.now().date() + timedelta(days=7):
                    # Skip if already queued recently
                    exists = conn.execute("""SELECT id FROM call_queue WHERE customer_phone=? AND kind='refill'
                                              AND status IN ('queued','calling') AND created_at > datetime('now','-7 days')""",
                                          (r["phone"],)).fetchone()
                    if exists: continue
                    items = json.loads(r["last_items"] or "[]") if r["last_items"] else []
                    med_names = ", ".join([i.get("name", "") for i in items[:3]])
                    conn.execute("""INSERT INTO call_queue
                        (customer_name, customer_phone, kind, medicine, expected_stockout, language, scheduled_at)
                        VALUES (?,?,'refill',?,?,?,?)""",
                        (r["cust"] or "Customer", r["phone"], med_names, stockout.strftime("%Y-%m-%d"),
                         glob["language"] if glob else "ta", call_date.strftime("%Y-%m-%d") + " 10:00"))
                    queued += 1
            except Exception:
                continue
    return jsonify({"status": "ok", "queued": queued})


@app.route("/call-dashboard")
def call_dashboard_page():
    return render_template("call_dashboard.html")


@app.route("/api/calls/queue", methods=["GET"])
def api_call_queue_list():
    status = request.args.get("status", "")
    q = "SELECT * FROM call_queue"
    p = []
    if status:
        q += " WHERE status=?"; p.append(status)
    q += " ORDER BY scheduled_at ASC, id DESC LIMIT 200"
    with get_conn() as conn:
        rows = conn.execute(q, p).fetchall()
        stats = {}
        for s in ['queued', 'calling', 'done', 'press1', 'press2', 'failed']:
            stats[s] = conn.execute("SELECT COUNT(*) FROM call_queue WHERE status=?", (s,)).fetchone()[0]
    return jsonify({"calls": [dict(r) for r in rows], "stats": stats})


@app.route("/api/calls/queue", methods=["POST"])
def api_call_queue_create():
    d = request.json or {}
    with get_conn() as conn:
        cur = conn.execute("""INSERT INTO call_queue
            (customer_name, customer_phone, kind, medicine, language, scheduled_at)
            VALUES (?,?,?,?,?,?)""",
            (d.get("customer_name", ""), d.get("customer_phone", ""),
             d.get("kind", "refill"), d.get("medicine", ""),
             d.get("language", "ta"),
             d.get("scheduled_at", datetime.now().strftime("%Y-%m-%d %H:%M"))))
    return jsonify({"status": "ok", "id": cur.lastrowid})


@app.route("/api/calls/<int:cid>/start", methods=["POST"])
def api_call_start(cid):
    """Mark call as in-progress (simulates outbound call)."""
    with get_conn() as conn:
        conn.execute("UPDATE call_queue SET status='calling', attempts=attempts+1, called_at=datetime('now') WHERE id=?", (cid,))
    return jsonify({"status": "ok"})


@app.route("/api/calls/<int:cid>/response", methods=["POST"])
def api_call_response(cid):
    """Record IVR response: press1 (reorder) / press2 (decline) / voicemail / no_answer."""
    d = request.json or {}
    resp = d.get("response", "")  # press1 / press2 / voicemail
    new_status = {
        "press1": "press1",
        "press2": "press2",
        "voicemail": "done",
        "no_answer": "failed",
        "wrong_number": "failed",
    }.get(resp, "done")

    with get_conn() as conn:
        call = conn.execute("SELECT * FROM call_queue WHERE id=?", (cid,)).fetchone()
        if not call:
            return jsonify({"error": "call not found"}), 404
        conn.execute("UPDATE call_queue SET status=?, response=? WHERE id=?",
                     (new_status, resp, cid))

        # If PRESS 1 → auto-create pickpack queue entry
        bill_id = ""
        if resp == "press1":
            # Find last bill for this customer to use as template
            last_bill = conn.execute(
                "SELECT items, total FROM bills WHERE phone=? ORDER BY ts DESC LIMIT 1",
                (call["customer_phone"],)
            ).fetchone()
            items_json = last_bill["items"] if last_bill else "[]"
            total = last_bill["total"] if last_bill else 0
            bill_id = f"AUTO-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            conn.execute("""INSERT INTO pickpack_queue
                (bill_id, customer_name, customer_phone, source, items_json, total, status, priority, notes)
                VALUES (?,?,?,?,?,?,'queued','high',?)""",
                (bill_id, call["customer_name"], call["customer_phone"],
                 "call_press1", items_json, total,
                 f"Auto-bill from refill call ID #{cid}"))
            conn.execute("UPDATE call_queue SET resulting_bill=? WHERE id=?", (bill_id, cid))

    return jsonify({"status": "ok", "new_status": new_status, "resulting_bill": bill_id})


@app.route("/api/calls/<int:cid>/whatsapp", methods=["POST"])
def api_call_whatsapp_log(cid):
    """Log that WhatsApp was sent (manually via wa.me)."""
    d = request.json or {}
    with get_conn() as conn:
        conn.execute("UPDATE call_queue SET whatsapp_sent=1, whatsapp_text=? WHERE id=?",
                     (d.get("text", ""), cid))
    return jsonify({"status": "ok"})


# ── AI CALL SUBSCRIPTION PLANS ──────────────────────────────────────
@app.route("/api/calls/plans", methods=["GET"])
def api_call_plans():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM ai_call_subscriptions WHERE is_active=1").fetchall()
    return jsonify({"plans": [dict(r) for r in rows]})


# ── PICK / PACK QUEUE ───────────────────────────────────────────────
@app.route("/pickpack")
def pickpack_page():
    return render_template("pickpack.html")


@app.route("/api/pickpack", methods=["GET"])
def api_pickpack_list():
    status = request.args.get("status", "")
    q = "SELECT * FROM pickpack_queue"
    p = []
    if status:
        q += " WHERE status=?"; p.append(status)
    q += " ORDER BY priority DESC, id ASC LIMIT 200"
    with get_conn() as conn:
        rows = conn.execute(q, p).fetchall()
        stats = {}
        for s in ['queued', 'picking', 'packing', 'ready', 'dispatched', 'cancelled']:
            stats[s] = conn.execute("SELECT COUNT(*) FROM pickpack_queue WHERE status=?", (s,)).fetchone()[0]
    items = []
    for r in rows:
        d = dict(r)
        try: d["items"] = json.loads(d["items_json"] or "[]")
        except: d["items"] = []
        items.append(d)
    return jsonify({"items": items, "stats": stats})


@app.route("/api/pickpack", methods=["POST"])
def api_pickpack_create():
    d = request.json or {}
    bill_id = d.get("bill_id") or f"PP-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    with get_conn() as conn:
        cur = conn.execute("""INSERT INTO pickpack_queue
            (bill_id, customer_name, customer_phone, source, items_json, total, priority, notes)
            VALUES (?,?,?,?,?,?,?,?)""",
            (bill_id, d.get("customer_name", ""), d.get("customer_phone", ""),
             d.get("source", "manual"), json.dumps(d.get("items", [])),
             float(d.get("total", 0)), d.get("priority", "normal"), d.get("notes", "")))
    return jsonify({"status": "ok", "id": cur.lastrowid, "bill_id": bill_id})


@app.route("/api/pickpack/<int:pid>/status", methods=["POST"])
def api_pickpack_status(pid):
    d = request.json or {}
    new_status = d.get("status", "")
    if new_status not in ['queued', 'picking', 'packing', 'ready', 'dispatched', 'cancelled']:
        return jsonify({"error": "invalid status"}), 400
    field_map = {
        'picking': 'started_at', 'ready': 'ready_at', 'dispatched': 'dispatched_at'
    }
    with get_conn() as conn:
        conn.execute("UPDATE pickpack_queue SET status=?, assigned_to=COALESCE(?, assigned_to) WHERE id=?",
                     (new_status, d.get("assigned_to"), pid))
        if new_status in field_map:
            conn.execute(f"UPDATE pickpack_queue SET {field_map[new_status]}=datetime('now') WHERE id=?", (pid,))
    return jsonify({"status": "ok"})


# ── SHOP BUSY MODE ──────────────────────────────────────────────────
@app.route("/api/shop/state", methods=["GET"])
def api_shop_state():
    with get_conn() as conn:
        rows = conn.execute("SELECT key, value FROM shop_state").fetchall()
    return jsonify({r["key"]: r["value"] for r in rows})


@app.route("/api/shop/busy", methods=["POST"])
def api_shop_busy_toggle():
    d = request.json or {}
    busy = "1" if d.get("busy") else "0"
    msg = d.get("message", "")
    with get_conn() as conn:
        conn.execute("UPDATE shop_state SET value=?, updated_at=datetime('now') WHERE key='busy_mode'", (busy,))
        if msg:
            conn.execute("UPDATE shop_state SET value=?, updated_at=datetime('now') WHERE key='busy_message'", (msg,))
    return jsonify({"status": "ok", "busy": busy == "1"})


# ════════════════════════════════════════════════════════════════════
# STAGE 18 — Anti-Fraud: Payment Verify + Audio + Camera Attendance + Audit
# ════════════════════════════════════════════════════════════════════
def init_stage18_db():
    with get_conn() as conn:
        # Patch customers — Family Head columns
        cust_cols = table_columns(conn, "customers")
        if "family_head_id" not in cust_cols:
            conn.execute("ALTER TABLE customers ADD COLUMN family_head_id INTEGER DEFAULT NULL")
        if "family_relation" not in cust_cols:
            conn.execute("ALTER TABLE customers ADD COLUMN family_relation TEXT DEFAULT ''")
        if "is_chronic" not in cust_cols:
            conn.execute("ALTER TABLE customers ADD COLUMN is_chronic INTEGER DEFAULT 0")

        # Patch bills — payment verification fields
        bcols = table_columns(conn, "bills")
        for col, decl in [
            ("payment_verified", "INTEGER DEFAULT 0"),
            ("payment_verified_by", "TEXT DEFAULT ''"),
            ("payment_verified_at", "TEXT DEFAULT ''"),
            ("payment_amount_heard", "REAL DEFAULT 0"),
            ("audio_filename", "TEXT DEFAULT ''"),
            ("audio_duration", "REAL DEFAULT 0"),
            ("audio_transcript", "TEXT DEFAULT ''"),
            ("ip_address", "TEXT DEFAULT ''"),
            ("device_id", "TEXT DEFAULT ''"),
        ]:
            if col not in bcols:
                conn.execute(f"ALTER TABLE bills ADD COLUMN {col} {decl}")

        # Payment verification audit
        conn.execute("""
            CREATE TABLE IF NOT EXISTS payment_verifications (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                bill_id         TEXT NOT NULL,
                staff_id        INTEGER DEFAULT 0,
                staff_name      TEXT DEFAULT '',
                status          TEXT NOT NULL,            -- verified / not_received / partial / disputed
                expected_amount REAL DEFAULT 0,
                received_amount REAL DEFAULT 0,
                payment_method  TEXT DEFAULT '',          -- Cash / UPI / Card / GPay / PhonePe / PayTM
                gpay_sound_heard INTEGER DEFAULT 0,       -- 1 if sound box played
                audio_filename  TEXT DEFAULT '',
                audio_duration  REAL DEFAULT 0,
                transcript      TEXT DEFAULT '',
                notes           TEXT DEFAULT '',
                ip_address      TEXT DEFAULT '',
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)

        # Camera attendance — face capture on in/out
        conn.execute("""
            CREATE TABLE IF NOT EXISTS camera_attendance (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id        INTEGER NOT NULL,
                staff_name      TEXT DEFAULT '',
                event           TEXT NOT NULL,            -- in / out
                face_image      TEXT DEFAULT '',          -- base64 data URL or filename
                face_vector     TEXT DEFAULT '',          -- 128-d for cross-match
                match_score     REAL DEFAULT 0,           -- vs registered staff face
                match_status    TEXT DEFAULT '',          -- verified / mismatch / unregistered
                location        TEXT DEFAULT '',
                ip_address      TEXT DEFAULT '',
                device_id       TEXT DEFAULT '',
                latitude        REAL DEFAULT 0,
                longitude       REAL DEFAULT 0,
                ts              TEXT DEFAULT (datetime('now'))
            )
        """)

        # Per-staff registered face (for matching)
        scols = table_columns(conn, "staff")
        if "face_vector" not in scols:
            conn.execute("ALTER TABLE staff ADD COLUMN face_vector TEXT DEFAULT ''")
        if "photo_url" not in scols:
            conn.execute("ALTER TABLE staff ADD COLUMN photo_url TEXT DEFAULT ''")

        # Security audit log — every sensitive action
        conn.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                ts              TEXT DEFAULT (datetime('now')),
                action          TEXT NOT NULL,            -- bill_save / bill_delete / discount / price_change / cash_adjust / login / etc.
                target_type     TEXT DEFAULT '',          -- bill / customer / staff / drawer / expense
                target_id       TEXT DEFAULT '',
                staff_id        INTEGER DEFAULT 0,
                staff_name      TEXT DEFAULT '',
                old_value       TEXT DEFAULT '',
                new_value       TEXT DEFAULT '',
                ip_address      TEXT DEFAULT '',
                user_agent      TEXT DEFAULT '',
                device_id       TEXT DEFAULT '',
                severity        TEXT DEFAULT 'info',      -- info / warn / critical
                notes           TEXT DEFAULT ''
            )
        """)

        # GPay/UPI sound box log — detected QR-payment audio events
        conn.execute("""
            CREATE TABLE IF NOT EXISTS soundbox_events (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                provider        TEXT DEFAULT 'GPay',      -- GPay / PhonePe / PayTM / BharatPe
                amount_detected REAL DEFAULT 0,
                bill_id         TEXT DEFAULT '',
                staff_name      TEXT DEFAULT '',
                audio_filename  TEXT DEFAULT '',
                transcript      TEXT DEFAULT '',
                matched         INTEGER DEFAULT 0,         -- did we match it to a bill?
                ts              TEXT DEFAULT (datetime('now'))
            )
        """)


# ── HELPERS ─────────────────────────────────────────────────────────
def _audit(action, target_type="", target_id="", old="", new="", severity="info", notes=""):
    """Write to audit log. Safe to call from any route."""
    try:
        with get_conn() as conn:
            conn.execute("""INSERT INTO audit_log
                (action, target_type, target_id, staff_id, staff_name, old_value, new_value,
                 ip_address, user_agent, severity, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (action, target_type, str(target_id),
                 session.get("staff_id", 0), session.get("staff_name", ""),
                 str(old)[:500] if old else "", str(new)[:500] if new else "",
                 request.remote_addr if request else "",
                 (request.headers.get("User-Agent", "")[:200]) if request else "",
                 severity, notes))
    except Exception as e:
        print(f"[audit] {e}")


# ── PAYMENT VERIFICATION APIs ───────────────────────────────────────
import os as _os
AUDIO_DIR = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "frontend", "static", "audio")
_os.makedirs(AUDIO_DIR, exist_ok=True)


@app.route("/payment-verify")
def payment_verify_page():
    return render_template("payment_verify.html")


@app.route("/api/payment-verify", methods=["POST"])
def api_payment_verify():
    """Mark a bill payment as verified or not.
    Body: { bill_id, status: 'verified'|'not_received'|'partial'|'disputed',
            expected_amount, received_amount, payment_method, gpay_sound_heard, notes }"""
    d = request.json or {}
    bill_id = d.get("bill_id", "")
    status = d.get("status", "")
    if not bill_id or status not in ("verified", "not_received", "partial", "disputed"):
        return jsonify({"error": "bill_id and valid status required"}), 400

    with get_conn() as conn:
        bill = conn.execute("SELECT id, total, staff_name FROM bills WHERE id=?", (bill_id,)).fetchone()
        if not bill:
            return jsonify({"error": "bill not found"}), 404

        expected = float(d.get("expected_amount", bill["total"] or 0))
        received = float(d.get("received_amount", expected))
        method = d.get("payment_method", "")
        sound = 1 if d.get("gpay_sound_heard") else 0
        notes = d.get("notes", "")

        conn.execute("""INSERT INTO payment_verifications
            (bill_id, staff_id, staff_name, status, expected_amount, received_amount,
             payment_method, gpay_sound_heard, audio_filename, audio_duration, transcript, notes, ip_address)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (bill_id, session.get("staff_id", 0), session.get("staff_name", ""),
             status, expected, received, method, sound,
             d.get("audio_filename", ""), float(d.get("audio_duration", 0)),
             d.get("transcript", ""), notes, request.remote_addr or ""))

        is_verified = 1 if status == "verified" and received >= expected else 0
        conn.execute("""UPDATE bills SET payment_verified=?, payment_verified_by=?,
            payment_verified_at=datetime('now'), payment_amount_heard=?
            WHERE id=?""",
            (is_verified, session.get("staff_name", ""), received, bill_id))

    _audit("payment_verify", "bill", bill_id, new=f"{status} ₹{received}",
           severity="warn" if status != "verified" else "info")
    return jsonify({"status": "ok", "verified": bool(is_verified)})


@app.route("/api/payment-verify/audio", methods=["POST"])
def api_payment_audio_upload():
    """Save audio recording for a bill payment."""
    bill_id = request.form.get("bill_id", "")
    if not bill_id:
        return jsonify({"error": "bill_id required"}), 400
    if "audio" not in request.files:
        return jsonify({"error": "no audio file"}), 400

    f = request.files["audio"]
    ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    safe_bid = "".join(c for c in bill_id if c.isalnum())[:30]
    filename = f"bill_{safe_bid}_{ts}.webm"
    full_path = _os.path.join(AUDIO_DIR, filename)
    f.save(full_path)

    # Pseudo-transcript (real ASR would go here)
    transcript = request.form.get("transcript", "")

    with get_conn() as conn:
        conn.execute("UPDATE bills SET audio_filename=?, audio_transcript=? WHERE id=?",
                     (filename, transcript, bill_id))

    return jsonify({"status": "ok", "filename": filename, "url": f"/static/audio/{filename}"})


@app.route("/api/payment-verify/pending", methods=["GET"])
def api_payment_pending():
    """Bills awaiting verification (created in last 24h, not yet verified)."""
    with get_conn() as conn:
        rows = conn.execute("""SELECT id, date, cust, phone, pay, total, staff_name, audio_filename,
                                      payment_verified, ts
                               FROM bills
                               WHERE payment_verified=0
                                 AND ts > strftime('%s', 'now', '-1 days')
                               ORDER BY ts DESC LIMIT 200""").fetchall()
    return jsonify({"bills": [dict(r) for r in rows]})


@app.route("/api/payment-verify/history", methods=["GET"])
def api_payment_history():
    with get_conn() as conn:
        rows = conn.execute("""SELECT pv.*, b.cust, b.phone
                               FROM payment_verifications pv
                               LEFT JOIN bills b ON pv.bill_id = b.id
                               ORDER BY pv.id DESC LIMIT 200""").fetchall()
    return jsonify({"verifications": [dict(r) for r in rows]})


# ── CAMERA ATTENDANCE ───────────────────────────────────────────────
@app.route("/staff-attendance")
def staff_attendance_page():
    return render_template("staff_attendance.html")


@app.route("/api/staff-attendance", methods=["POST"])
def api_staff_attendance():
    d = request.json or {}
    staff_id = int(d.get("staff_id", 0))
    event = d.get("event", "")  # in / out
    if event not in ("in", "out"):
        return jsonify({"error": "event must be in/out"}), 400

    face_image = d.get("face_image", "")[:200000]  # cap base64
    face_vector = d.get("face_vector", "")
    location = d.get("location", "")
    lat = float(d.get("latitude", 0))
    lng = float(d.get("longitude", 0))

    with get_conn() as conn:
        staff_row = conn.execute("SELECT name, face_vector FROM staff WHERE id=?", (staff_id,)).fetchone()
        staff_name = staff_row["name"] if staff_row else "Unknown"
        # Match against registered staff face
        match_status = "unregistered"
        match_score = 0
        if staff_row and staff_row["face_vector"] and face_vector:
            try:
                stored = json.loads(staff_row["face_vector"])
                current = json.loads(face_vector) if isinstance(face_vector, str) else face_vector
                if isinstance(stored, list) and isinstance(current, list) and len(stored) == len(current):
                    dist = math.sqrt(sum((a-b)**2 for a, b in zip(stored, current)))
                    match_score = round(max(0, 1.0 - dist/0.6) * 100, 1)
                    match_status = "verified" if dist <= 0.5 else "mismatch"
            except Exception:
                pass

        conn.execute("""INSERT INTO camera_attendance
            (staff_id, staff_name, event, face_image, face_vector, match_score, match_status,
             location, ip_address, latitude, longitude)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (staff_id, staff_name, event,
             face_image, json.dumps(face_vector) if not isinstance(face_vector, str) else face_vector,
             match_score, match_status, location, request.remote_addr or "", lat, lng))

    _audit(f"attendance_{event}", "staff", staff_id, new=f"{event} (face={match_status})")
    return jsonify({"status": "ok", "match_status": match_status, "match_score": match_score})


@app.route("/api/staff-attendance", methods=["GET"])
def api_staff_attendance_list():
    date = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    staff_id = request.args.get("staff_id", "")
    q = "SELECT * FROM camera_attendance WHERE substr(ts,1,10)=?"
    p = [date]
    if staff_id:
        q += " AND staff_id=?"; p.append(int(staff_id))
    q += " ORDER BY id DESC LIMIT 500"
    with get_conn() as conn:
        rows = conn.execute(q, p).fetchall()
    return jsonify({"events": [dict(r) for r in rows]})


@app.route("/api/staff/<int:sid>/enroll-face", methods=["POST"])
def api_staff_enroll_face(sid):
    d = request.json or {}
    vec = d.get("face_vector") or d.get("vector") or []
    if not isinstance(vec, list) or len(vec) < 32:
        return jsonify({"error": "valid face_vector required"}), 400
    photo = d.get("photo_url", "")
    with get_conn() as conn:
        conn.execute("UPDATE staff SET face_vector=?, photo_url=? WHERE id=?",
                     (json.dumps(vec), photo, sid))
    _audit("staff_face_enroll", "staff", sid, new=f"vector_len={len(vec)}")
    return jsonify({"status": "ok"})


# ── AUDIT LOG VIEW ──────────────────────────────────────────────────
@app.route("/audit-log")
def audit_log_page():
    return render_template("audit_log.html")


@app.route("/api/audit-log", methods=["GET"])
def api_audit_log():
    severity = request.args.get("severity", "")
    action = request.args.get("action", "")
    staff_id = request.args.get("staff_id", "")
    q = "SELECT * FROM audit_log WHERE 1=1"
    p = []
    if severity:
        q += " AND severity=?"; p.append(severity)
    if action:
        q += " AND action LIKE ?"; p.append(f"%{action}%")
    if staff_id:
        q += " AND staff_id=?"; p.append(int(staff_id))
    q += " ORDER BY id DESC LIMIT 500"
    with get_conn() as conn:
        rows = conn.execute(q, p).fetchall()
    return jsonify({"events": [dict(r) for r in rows]})


# ── GPay / Sound Box Events ─────────────────────────────────────────
@app.route("/api/soundbox/event", methods=["POST"])
def api_soundbox_event():
    """Log a detected GPay/UPI sound-box event."""
    d = request.json or {}
    amount = float(d.get("amount", 0))
    provider = d.get("provider", "GPay")
    bill_id = d.get("bill_id", "")
    transcript = d.get("transcript", "")

    with get_conn() as conn:
        # Auto-match to nearest unverified cash/UPI bill in last 5 min for same amount
        matched = 0
        if not bill_id and amount > 0:
            cand = conn.execute("""SELECT id FROM bills
                                   WHERE total BETWEEN ?-1 AND ?+1
                                     AND payment_verified=0
                                     AND ts > strftime('%s','now','-5 minutes')
                                   ORDER BY ts DESC LIMIT 1""", (amount, amount)).fetchone()
            if cand:
                bill_id = cand[0]
                matched = 1
                # Mark bill as verified via sound box
                conn.execute("""UPDATE bills SET payment_verified=1, payment_verified_by=?,
                                payment_verified_at=datetime('now'), payment_amount_heard=?
                                WHERE id=?""",
                             (f"SoundBox({provider})", amount, bill_id))

        conn.execute("""INSERT INTO soundbox_events
            (provider, amount_detected, bill_id, staff_name, transcript, matched)
            VALUES (?,?,?,?,?,?)""",
            (provider, amount, bill_id, session.get("staff_name", ""), transcript, matched))

    return jsonify({"status": "ok", "matched_bill": bill_id, "auto_verified": bool(matched)})


@app.route("/api/soundbox/events", methods=["GET"])
def api_soundbox_events():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM soundbox_events ORDER BY id DESC LIMIT 100").fetchall()
    return jsonify({"events": [dict(r) for r in rows]})


# ── ANTI-CHEAT DASHBOARD ────────────────────────────────────────────
@app.route("/security-dashboard")
def security_dashboard_page():
    return render_template("security_dashboard.html")


@app.route("/api/security/summary", methods=["GET"])
def api_security_summary():
    today = datetime.now().strftime("%Y-%m-%d")
    with get_conn() as conn:
        # Bills today
        bills_today = conn.execute("SELECT COUNT(*) FROM bills WHERE date=?", (today,)).fetchone()[0]
        # Unverified bills
        unverified = conn.execute("SELECT COUNT(*) FROM bills WHERE payment_verified=0 AND date=?", (today,)).fetchone()[0]
        # Attendance events today
        att_today = conn.execute("SELECT COUNT(*) FROM camera_attendance WHERE substr(ts,1,10)=?", (today,)).fetchone()[0]
        att_mismatches = conn.execute("SELECT COUNT(*) FROM camera_attendance WHERE substr(ts,1,10)=? AND match_status='mismatch'", (today,)).fetchone()[0]
        # Audit log critical/warn
        critical_audits = conn.execute("SELECT COUNT(*) FROM audit_log WHERE severity IN ('critical','warn') AND substr(ts,1,10)=?", (today,)).fetchone()[0]
        # Cash drawer
        cash_total = conn.execute("SELECT SUM(denomination*count) FROM cash_drawer").fetchone()[0] or 0
        # Soundbox events
        sound_events = conn.execute("SELECT COUNT(*) FROM soundbox_events WHERE substr(ts,1,10)=?", (today,)).fetchone()[0]
        # Suspicious: bills with large discounts (>10%)
        suspicious = conn.execute("""SELECT COUNT(*) FROM bills
                                     WHERE date=? AND sub>0 AND (disc*100.0/sub) > 10""", (today,)).fetchone()[0]

    return jsonify({
        "bills_today": bills_today,
        "unverified": unverified,
        "att_today": att_today,
        "att_mismatches": att_mismatches,
        "critical_audits": critical_audits,
        "cash_total": cash_total,
        "soundbox_events": sound_events,
        "suspicious_bills": suspicious,
        "verification_rate": round((bills_today - unverified) / bills_today * 100, 1) if bills_today else 100,
    })


# ════════════════════════════════════════════════════════════════════
# STAGE 19 — REPORTS UNIVERSE (Sales/Purchase/Stock/Finance/AI/Compliance)
# ════════════════════════════════════════════════════════════════════
def init_stage19_db():
    """Stage 19 uses existing tables — no new schema. Just ensures indexes exist."""
    with get_conn() as conn:
        # Speed-up indexes
        try:
            conn.execute("CREATE INDEX IF NOT EXISTS idx_bills_date ON bills(date)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_bills_ts ON bills(ts)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_bills_phone ON bills(phone)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_bills_doctor ON bills(doctor)")
        except Exception as e:
            print(f"[stage19 idx] {e}")


def _parse_date_range(args):
    """Parse from/to args. Returns (from_str, to_str, ts_from, ts_to)."""
    from_s = args.get("from", "")
    to_s = args.get("to", "")
    preset = args.get("preset", "")
    today = datetime.now()

    if preset == "today":
        from_s = to_s = today.strftime("%Y-%m-%d")
    elif preset == "yesterday":
        d = (today - timedelta(days=1)).strftime("%Y-%m-%d")
        from_s = to_s = d
    elif preset == "week":
        from_s = (today - timedelta(days=6)).strftime("%Y-%m-%d")
        to_s = today.strftime("%Y-%m-%d")
    elif preset == "month":
        from_s = today.strftime("%Y-%m-01")
        to_s = today.strftime("%Y-%m-%d")
    elif preset == "last_month":
        first_this = today.replace(day=1)
        last_prev = first_this - timedelta(days=1)
        from_s = last_prev.strftime("%Y-%m-01")
        to_s = last_prev.strftime("%Y-%m-%d")
    elif preset == "year":
        from_s = today.strftime("%Y-01-01")
        to_s = today.strftime("%Y-%m-%d")
    elif preset == "all":
        from_s = "2020-01-01"
        to_s = today.strftime("%Y-%m-%d")

    if not from_s: from_s = today.strftime("%Y-%m-01")
    if not to_s: to_s = today.strftime("%Y-%m-%d")

    try:
        ts_from = int(datetime.strptime(from_s, "%Y-%m-%d").timestamp())
        ts_to = int((datetime.strptime(to_s, "%Y-%m-%d") + timedelta(days=1)).timestamp())
    except Exception:
        ts_from, ts_to = 0, 9999999999
    return from_s, to_s, ts_from, ts_to


# ── REPORTS HUB PAGE ──
@app.route("/reports")
def reports_hub_page():
    return render_template("reports_hub.html")


# ── 1. KPI SUMMARY (all-in-one for the hub header) ─────────────────
@app.route("/api/reports/summary", methods=["GET"])
def rep_summary():
    fs, ts, tf, tt = _parse_date_range(request.args)
    with get_conn() as conn:
        # Sales totals
        sales = conn.execute(
            "SELECT COUNT(*), COALESCE(SUM(total),0), COALESCE(SUM(sub),0), COALESCE(SUM(tax),0), COALESCE(SUM(disc),0), COALESCE(SUM(profit),0) "
            "FROM bills WHERE ts BETWEEN ? AND ?", (tf, tt)).fetchone()
        # Unique customers
        ucusts = conn.execute(
            "SELECT COUNT(DISTINCT phone) FROM bills WHERE ts BETWEEN ? AND ? AND phone != ''",
            (tf, tt)).fetchone()[0]
        # Avg basket
        avg_basket = (sales[1] / sales[0]) if sales[0] else 0
        # Expenses
        exp_total = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date BETWEEN ? AND ?",
            (fs, ts)).fetchone()[0]
        # Purchases (stock receiving values where present)
        pur_total = 0
        try:
            pur_total = conn.execute(
                "SELECT COALESCE(SUM(total_amount),0) FROM receiving_orders WHERE substr(received_date,1,10) BETWEEN ? AND ?",
                (fs, ts)).fetchone()[0] or 0
        except Exception:
            pass

    return jsonify({
        "from": fs, "to": ts,
        "bills": sales[0],
        "revenue": sales[1],
        "subtotal": sales[2],
        "gst": sales[3],
        "discount": sales[4],
        "profit": sales[5],
        "unique_customers": ucusts,
        "avg_basket": round(avg_basket, 2),
        "expenses": exp_total,
        "purchases": pur_total,
        "net": round((sales[1] or 0) - (exp_total or 0), 2),
    })


# ── 2. ITEM-WISE SALES (best sellers, qty, revenue, profit) ────────
@app.route("/api/reports/item-sales", methods=["GET"])
def rep_item_sales():
    fs, ts, tf, tt = _parse_date_range(request.args)
    sort = request.args.get("sort", "revenue")  # revenue / qty / profit
    limit = int(request.args.get("limit", 100))

    with get_conn() as conn:
        rows = conn.execute("SELECT items, total, profit FROM bills WHERE ts BETWEEN ? AND ?", (tf, tt)).fetchall()

    agg = {}  # name -> {qty, revenue, bills, profit_estimate}
    for r in rows:
        try:
            items = json.loads(r["items"] or "[]")
        except Exception:
            continue
        bill_total = r["total"] or 1
        bill_profit = r["profit"] or 0
        for it in items:
            nm = it.get("name", "")
            if not nm: continue
            qty = int(it.get("qty", 1))
            amt = float(it.get("amount", it.get("price", 0) * qty))
            d = agg.setdefault(nm, {"name": nm, "qty": 0, "revenue": 0, "bills": 0, "profit": 0})
            d["qty"] += qty
            d["revenue"] += amt
            d["bills"] += 1
            d["profit"] += bill_profit * (amt / bill_total) if bill_total else 0

    items = list(agg.values())
    key = "qty" if sort == "qty" else ("profit" if sort == "profit" else "revenue")
    items.sort(key=lambda x: x[key], reverse=True)
    items = items[:limit]
    for it in items:
        it["revenue"] = round(it["revenue"], 2)
        it["profit"] = round(it["profit"], 2)

    return jsonify({"from": fs, "to": ts, "items": items, "total_unique": len(agg)})


# ── 3. DAILY SALES TIMELINE ────────────────────────────────────────
@app.route("/api/reports/daily-timeline", methods=["GET"])
def rep_daily_timeline():
    fs, ts, tf, tt = _parse_date_range(request.args)
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT substr(date,1,10) as d, COUNT(*) as bills, SUM(total) as revenue, SUM(profit) as profit "
            "FROM bills WHERE ts BETWEEN ? AND ? GROUP BY d ORDER BY d", (tf, tt)).fetchall()
    days = []
    for r in rows:
        days.append({"date": r["d"], "bills": r["bills"], "revenue": round(r["revenue"] or 0, 2), "profit": round(r["profit"] or 0, 2)})
    return jsonify({"from": fs, "to": ts, "days": days})


# ── 4. HOURLY HEATMAP (when does the shop get busy?) ───────────────
@app.route("/api/reports/hourly-heatmap", methods=["GET"])
def rep_hourly_heatmap():
    fs, ts, tf, tt = _parse_date_range(request.args)
    with get_conn() as conn:
        rows = conn.execute("SELECT ts, total FROM bills WHERE ts BETWEEN ? AND ?", (tf, tt)).fetchall()
    grid = [[0] * 24 for _ in range(7)]   # weekday × hour, revenue
    counts = [[0] * 24 for _ in range(7)]  # bill counts
    for r in rows:
        try:
            d = datetime.fromtimestamp(r["ts"])
            grid[d.weekday()][d.hour] += r["total"] or 0
            counts[d.weekday()][d.hour] += 1
        except Exception:
            continue
    return jsonify({
        "from": fs, "to": ts,
        "revenue_grid": [[round(c, 2) for c in row] for row in grid],
        "bill_counts": counts,
        "weekdays": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
    })


# ── 5. MARKET BASKET — items bought together ───────────────────────
@app.route("/api/reports/market-basket", methods=["GET"])
def rep_market_basket():
    fs, ts, tf, tt = _parse_date_range(request.args)
    min_count = int(request.args.get("min", 3))
    limit = int(request.args.get("limit", 50))

    with get_conn() as conn:
        rows = conn.execute("SELECT items FROM bills WHERE ts BETWEEN ? AND ?", (tf, tt)).fetchall()

    item_count = {}
    pair_count = {}
    total_baskets = 0

    for r in rows:
        try:
            items = json.loads(r["items"] or "[]")
        except Exception:
            continue
        names = list({it.get("name", "") for it in items if it.get("name")})
        if len(names) < 1:
            continue
        total_baskets += 1
        for n in names:
            item_count[n] = item_count.get(n, 0) + 1
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                a, b = sorted([names[i], names[j]])
                key = (a, b)
                pair_count[key] = pair_count.get(key, 0) + 1

    pairs = []
    for (a, b), c in pair_count.items():
        if c < min_count:
            continue
        # Confidence A→B = P(B|A) = count(A,B) / count(A)
        ca = item_count.get(a, 1)
        cb = item_count.get(b, 1)
        conf_ab = c / ca if ca else 0
        conf_ba = c / cb if cb else 0
        # Lift = P(B|A) / P(B) = (c / ca) / (cb / total)
        lift = (c * total_baskets) / (ca * cb) if ca and cb else 0
        pairs.append({
            "a": a, "b": b, "count": c,
            "support": round(c / total_baskets, 4) if total_baskets else 0,
            "conf_ab": round(conf_ab, 3),
            "conf_ba": round(conf_ba, 3),
            "lift": round(lift, 2),
        })
    pairs.sort(key=lambda x: x["lift"], reverse=True)
    return jsonify({"from": fs, "to": ts, "baskets": total_baskets, "pairs": pairs[:limit]})


# ── 6. CUSTOMER COHORT — RFM (recency, frequency, monetary) ────────
@app.route("/api/reports/customer-rfm", methods=["GET"])
def rep_customer_rfm():
    fs, ts, tf, tt = _parse_date_range(request.args)
    with get_conn() as conn:
        rows = conn.execute("""SELECT phone, cust, MAX(ts) as last_ts, COUNT(*) as freq, SUM(total) as monetary, SUM(profit) as profit
                               FROM bills WHERE ts BETWEEN ? AND ? AND phone != ''
                               GROUP BY phone ORDER BY monetary DESC""", (tf, tt)).fetchall()
    now_ts = datetime.now().timestamp()
    cohorts = {"champions": 0, "loyal": 0, "at_risk": 0, "lost": 0, "new": 0}
    customers = []
    for r in rows:
        days = int((now_ts - r["last_ts"]) / 86400)
        freq = r["freq"]
        mon = r["monetary"] or 0

        if days <= 14 and freq >= 5: cohort = "champions"
        elif days <= 30 and freq >= 3: cohort = "loyal"
        elif days <= 30 and freq == 1: cohort = "new"
        elif 30 < days <= 90: cohort = "at_risk"
        else: cohort = "lost"
        cohorts[cohort] += 1

        customers.append({
            "phone": r["phone"], "name": r["cust"] or "-",
            "days_since": days, "freq": freq,
            "monetary": round(mon, 2), "profit": round(r["profit"] or 0, 2),
            "cohort": cohort, "avg_bill": round(mon / freq if freq else 0, 2),
        })
    return jsonify({"from": fs, "to": ts, "cohorts": cohorts, "customers": customers[:200]})


# ── 7. ABANDONED CUSTOMERS (haven't visited recently) ──────────────
@app.route("/api/reports/abandoned-customers", methods=["GET"])
def rep_abandoned():
    days = int(request.args.get("days", 60))
    cutoff = int((datetime.now() - timedelta(days=days)).timestamp())
    with get_conn() as conn:
        rows = conn.execute("""SELECT phone, cust, MAX(ts) as last_ts, COUNT(*) as visits, SUM(total) as spend
                               FROM bills WHERE phone != ''
                               GROUP BY phone HAVING MAX(ts) < ? AND visits >= 2
                               ORDER BY spend DESC LIMIT 200""", (cutoff,)).fetchall()
    out = []
    for r in rows:
        out.append({
            "phone": r["phone"], "name": r["cust"],
            "last_visit": datetime.fromtimestamp(r["last_ts"]).strftime("%Y-%m-%d"),
            "days_away": int((datetime.now().timestamp() - r["last_ts"]) / 86400),
            "visits": r["visits"], "spend": round(r["spend"] or 0, 2),
        })
    return jsonify({"days": days, "customers": out})


# ── 8. DOCTOR-WISE SALES ───────────────────────────────────────────
@app.route("/api/reports/doctor-sales", methods=["GET"])
def rep_doctor_sales():
    fs, ts, tf, tt = _parse_date_range(request.args)
    with get_conn() as conn:
        rows = conn.execute("""SELECT doctor, COUNT(*) as rx, SUM(total) as revenue, SUM(profit) as profit,
                                      COUNT(DISTINCT phone) as patients
                               FROM bills WHERE doctor != '' AND doctor IS NOT NULL AND ts BETWEEN ? AND ?
                               GROUP BY doctor ORDER BY revenue DESC""", (tf, tt)).fetchall()
    return jsonify({"from": fs, "to": ts, "doctors": [
        {"name": r["doctor"], "rx_count": r["rx"], "revenue": round(r["revenue"] or 0, 2),
         "profit": round(r["profit"] or 0, 2), "patients": r["patients"]}
        for r in rows
    ]})


# ── 9. CUSTOMER-WISE SALES ─────────────────────────────────────────
@app.route("/api/reports/customer-sales", methods=["GET"])
def rep_customer_sales():
    fs, ts, tf, tt = _parse_date_range(request.args)
    limit = int(request.args.get("limit", 100))
    with get_conn() as conn:
        rows = conn.execute("""SELECT phone, cust, COUNT(*) as visits, SUM(total) as revenue, SUM(profit) as profit
                               FROM bills WHERE phone != '' AND ts BETWEEN ? AND ?
                               GROUP BY phone ORDER BY revenue DESC LIMIT ?""", (tf, tt, limit)).fetchall()
    return jsonify({"from": fs, "to": ts, "customers": [
        {"phone": r["phone"], "name": r["cust"], "visits": r["visits"],
         "revenue": round(r["revenue"] or 0, 2), "profit": round(r["profit"] or 0, 2)}
        for r in rows
    ]})


# ── 10. EXPIRY REPORT ──────────────────────────────────────────────
@app.route("/api/reports/expiry", methods=["GET"])
def rep_expiry():
    days = int(request.args.get("days", 90))
    cutoff = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    today_s = datetime.now().strftime("%Y-%m-%d")
    with get_conn() as conn:
        rows = conn.execute("""SELECT id, n, c, p, s, batch, expiry FROM medicines
                               WHERE expiry != '' AND expiry IS NOT NULL AND expiry <= ? AND s > 0
                               ORDER BY expiry LIMIT 500""", (cutoff,)).fetchall()
    out = []
    for r in rows:
        try:
            d2 = datetime.strptime(r["expiry"][:10], "%Y-%m-%d")
            days_left = (d2 - datetime.now()).days
            tier = "expired" if days_left < 0 else ("30d" if days_left <= 30 else ("60d" if days_left <= 60 else "90d"))
        except Exception:
            days_left = 999; tier = "unknown"
        out.append({
            "id": r["id"], "name": r["n"], "category": r["c"], "price": r["p"],
            "stock": r["s"], "batch": r["batch"], "expiry": r["expiry"],
            "days_left": days_left, "tier": tier,
            "stock_value": round(r["p"] * r["s"], 2),
        })
    return jsonify({"total_value": round(sum(o["stock_value"] for o in out), 2), "items": out})


# ── 11. SLOW MOVERS (no sale in N days, has stock) ────────────────
@app.route("/api/reports/slow-movers", methods=["GET"])
def rep_slow_movers():
    days = int(request.args.get("days", 60))
    cutoff_ts = int((datetime.now() - timedelta(days=days)).timestamp())
    with get_conn() as conn:
        # Collect all sold item names in window
        rows = conn.execute("SELECT items FROM bills WHERE ts >= ?", (cutoff_ts,)).fetchall()
        sold = set()
        for r in rows:
            try:
                for it in json.loads(r["items"] or "[]"):
                    if it.get("name"): sold.add(it["name"].upper())
            except Exception:
                continue
        # Stock items not sold
        all_items = conn.execute(
            "SELECT id, n, c, p, s FROM medicines WHERE s > 0 ORDER BY p*s DESC LIMIT 5000"
        ).fetchall()
    out = []
    for r in all_items:
        if r["n"].upper() in sold: continue
        out.append({
            "id": r["id"], "name": r["n"], "category": r["c"],
            "price": r["p"], "stock": r["s"],
            "stock_value": round(r["p"] * r["s"], 2),
        })
        if len(out) >= 200: break
    out.sort(key=lambda x: x["stock_value"], reverse=True)
    return jsonify({"days": days, "items": out, "total_value": round(sum(o["stock_value"] for o in out), 2)})


# ── 12. DENOMINATION REPORT (links to existing) ────────────────────
@app.route("/api/reports/denomination", methods=["GET"])
def rep_denomination():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM cash_drawer ORDER BY denomination DESC").fetchall()
    items = [{**dict(r), "value": r["denomination"] * r["count"]} for r in rows]
    return jsonify({"items": items, "total": sum(i["value"] for i in items)})


# ── 13. CUSTOMER BALANCE (credit / outstanding) ────────────────────
@app.route("/api/reports/customer-balance", methods=["GET"])
def rep_customer_balance():
    with get_conn() as conn:
        # Customers where total purchase > sum of received_payment (simplified credit calc)
        # Using bills with pay='Credit'
        rows = conn.execute("""SELECT phone, cust, SUM(CASE WHEN pay='Credit' THEN total ELSE 0 END) as credit,
                                      SUM(total) as total_billed, COUNT(*) as bills
                               FROM bills WHERE phone != ''
                               GROUP BY phone HAVING credit > 0 ORDER BY credit DESC LIMIT 200""").fetchall()
    return jsonify({"customers": [
        {"phone": r["phone"], "name": r["cust"], "credit": round(r["credit"] or 0, 2),
         "total_billed": round(r["total_billed"] or 0, 2), "bills": r["bills"]}
        for r in rows
    ]})


# ── 14. NEW ITEMS REPORT ──────────────────────────────────────────
@app.route("/api/reports/new-items", methods=["GET"])
def rep_new_items():
    """Items added recently (uses rowid as proxy if no created_at)."""
    limit = int(request.args.get("limit", 100))
    with get_conn() as conn:
        rows = conn.execute("SELECT id, n, c, p, s FROM medicines ORDER BY rowid DESC LIMIT ?", (limit,)).fetchall()
    return jsonify({"items": [dict(r) for r in rows]})


# ── 15. SALES RETURNS (negative qty / sales returns) ──────────────
@app.route("/api/reports/sales-returns", methods=["GET"])
def rep_sales_returns():
    fs, ts, tf, tt = _parse_date_range(request.args)
    with get_conn() as conn:
        rows = conn.execute("SELECT id, date, cust, phone, total, items FROM bills WHERE total < 0 AND ts BETWEEN ? AND ?", (tf, tt)).fetchall()
    out = []
    for r in rows:
        out.append({
            "id": r["id"], "date": r["date"], "cust": r["cust"], "phone": r["phone"],
            "amount": r["total"],
        })
    return jsonify({"returns": out, "total_refunded": round(sum(o["amount"] for o in out), 2)})


# ── 16. ITEM LEDGER (batch + complete) ─────────────────────────────
@app.route("/api/reports/item-ledger", methods=["GET"])
def rep_item_ledger():
    name = (request.args.get("name") or "").upper().strip()
    if not name:
        return jsonify({"error": "name parameter required"}), 400
    with get_conn() as conn:
        med = conn.execute("SELECT * FROM medicines WHERE UPPER(n) LIKE ?", (f"%{name}%",)).fetchone()
        if not med:
            return jsonify({"error": "medicine not found"}), 404
        # Find all bills containing this item
        rows = conn.execute("SELECT id, date, ts, cust, items, total FROM bills WHERE items LIKE ? ORDER BY ts DESC LIMIT 500", (f"%{med['n']}%",)).fetchall()
    moves = []
    for r in rows:
        try:
            for it in json.loads(r["items"] or "[]"):
                if it.get("name") == med["n"]:
                    moves.append({
                        "date": r["date"], "bill_id": r["id"], "customer": r["cust"],
                        "qty_out": it.get("qty", 0), "rate": it.get("price", 0),
                        "amount": it.get("amount", 0),
                    })
        except Exception:
            continue
    return jsonify({
        "medicine": dict(med), "moves": moves,
        "total_sold": sum(m["qty_out"] for m in moves),
        "total_revenue": round(sum(m["amount"] for m in moves), 2),
    })


# ── 17. PRICE ELASTICITY (basic: stock vs sale velocity per item) ──
@app.route("/api/reports/price-elasticity", methods=["GET"])
def rep_price_elasticity():
    """Rough: top medicines by velocity, suggest if price could be raised (low stock + high demand) or lowered (high stock + low demand)."""
    fs, ts, tf, tt = _parse_date_range(request.args)
    with get_conn() as conn:
        bill_rows = conn.execute("SELECT items FROM bills WHERE ts BETWEEN ? AND ?", (tf, tt)).fetchall()
        meds = {r["n"]: dict(r) for r in conn.execute("SELECT id, n, c, p, s FROM medicines").fetchall()}
    sold = {}
    for r in bill_rows:
        try:
            for it in json.loads(r["items"] or "[]"):
                nm = it.get("name", "")
                if not nm: continue
                sold[nm] = sold.get(nm, 0) + int(it.get("qty", 1))
        except Exception:
            continue
    days_in_range = max(1, (tt - tf) / 86400)
    insights = []
    for nm, qty in sold.items():
        med = meds.get(nm)
        if not med: continue
        velocity = qty / days_in_range
        stock = med["s"]
        price = med["p"]
        days_to_stockout = stock / velocity if velocity > 0 else 999
        suggestion = ""
        if days_to_stockout < 7 and stock > 0:
            suggestion = "⬆ Raise price 5-10% (high demand, low stock)"
        elif days_to_stockout > 180:
            suggestion = "⬇ Lower 10-15% or offer (overstocked vs demand)"
        elif velocity > 5:
            suggestion = "🔥 Fast mover — keep stocked"
        insights.append({
            "name": nm, "category": med["c"], "price": price,
            "stock": stock, "qty_sold": qty,
            "velocity": round(velocity, 2),
            "days_to_stockout": round(days_to_stockout, 0),
            "suggestion": suggestion,
        })
    insights.sort(key=lambda x: x["velocity"], reverse=True)
    return jsonify({"items": insights[:100]})


# ── 18. PATIENT REMINDER (chronic patients due refill) ─────────────
@app.route("/api/reports/patient-reminder", methods=["GET"])
def rep_patient_reminder():
    """Patients whose last refill was 25-32 days ago and bought chronic meds."""
    today = datetime.now()
    earliest = int((today - timedelta(days=32)).timestamp())
    latest = int((today - timedelta(days=25)).timestamp())
    with get_conn() as conn:
        rows = conn.execute("""SELECT phone, cust, MAX(ts) as last_ts, MAX(items) as items, SUM(total) as spend
                               FROM bills WHERE phone != '' GROUP BY phone
                               HAVING last_ts BETWEEN ? AND ? ORDER BY last_ts DESC""",
                            (earliest, latest)).fetchall()
    out = []
    for r in rows:
        try: items = json.loads(r["items"] or "[]")
        except: items = []
        names = [i.get("name", "") for i in items][:3]
        out.append({
            "phone": r["phone"], "name": r["cust"],
            "last_visit": datetime.fromtimestamp(r["last_ts"]).strftime("%Y-%m-%d"),
            "days_ago": int((today.timestamp() - r["last_ts"]) / 86400),
            "last_items": names, "spend": round(r["spend"] or 0, 2),
        })
    return jsonify({"patients": out})


# ── 19. SALES FORECAST (next 7 / 30 days, naive) ───────────────────
@app.route("/api/reports/forecast", methods=["GET"])
def rep_forecast():
    horizon = int(request.args.get("days", 30))
    cutoff_ts = int((datetime.now() - timedelta(days=90)).timestamp())
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT substr(date,1,10) as d, SUM(total) as rev FROM bills WHERE ts >= ? GROUP BY d ORDER BY d",
            (cutoff_ts,)).fetchall()
    if not rows:
        return jsonify({"days": horizon, "forecast": [], "expected_revenue": 0})
    daily = [r["rev"] or 0 for r in rows]
    avg = sum(daily) / len(daily)
    # Simple weekday weighting
    weekday_factor = [1.0] * 7
    weekday_counts = [0] * 7
    weekday_sums = [0] * 7
    for r in rows:
        try:
            wd = datetime.strptime(r["d"], "%Y-%m-%d").weekday()
            weekday_sums[wd] += r["rev"] or 0
            weekday_counts[wd] += 1
        except Exception:
            continue
    overall = sum(weekday_sums) / max(1, sum(weekday_counts))
    for i in range(7):
        if weekday_counts[i] and overall:
            weekday_factor[i] = (weekday_sums[i] / weekday_counts[i]) / overall

    forecast = []
    expected = 0
    for i in range(1, horizon + 1):
        d = datetime.now() + timedelta(days=i)
        wd = d.weekday()
        f = avg * weekday_factor[wd]
        forecast.append({"date": d.strftime("%Y-%m-%d"), "expected": round(f, 2), "weekday": d.strftime("%a")})
        expected += f
    return jsonify({"days": horizon, "forecast": forecast, "expected_revenue": round(expected, 2), "daily_avg": round(avg, 2)})


# ── 20. STAFF PERFORMANCE ──────────────────────────────────────────
@app.route("/api/reports/staff-performance", methods=["GET"])
def rep_staff_performance():
    fs, ts, tf, tt = _parse_date_range(request.args)
    with get_conn() as conn:
        rows = conn.execute("""SELECT staff_name, COUNT(*) as bills, SUM(total) as revenue, AVG(total) as avg_bill
                               FROM bills WHERE staff_name != '' AND ts BETWEEN ? AND ?
                               GROUP BY staff_name ORDER BY revenue DESC""", (tf, tt)).fetchall()
    return jsonify({"from": fs, "to": ts, "staff": [
        {"name": r["staff_name"], "bills": r["bills"],
         "revenue": round(r["revenue"] or 0, 2),
         "avg_bill": round(r["avg_bill"] or 0, 2)}
        for r in rows
    ]})


# ── 21. EXPORT ANY REPORT AS CSV ───────────────────────────────────
@app.route("/api/reports/export", methods=["GET"])
def rep_export_csv():
    """Generic CSV exporter. Pass report=<name> + filter params."""
    import io, csv as csv_mod
    name = request.args.get("report", "")
    # Re-dispatch into our local APIs
    api_map = {
        "item-sales": rep_item_sales,
        "daily-timeline": rep_daily_timeline,
        "customer-sales": rep_customer_sales,
        "doctor-sales": rep_doctor_sales,
        "customer-rfm": rep_customer_rfm,
        "abandoned-customers": rep_abandoned,
        "expiry": rep_expiry,
        "slow-movers": rep_slow_movers,
        "patient-reminder": rep_patient_reminder,
        "staff-performance": rep_staff_performance,
        "market-basket": rep_market_basket,
        "customer-balance": rep_customer_balance,
        "price-elasticity": rep_price_elasticity,
    }
    fn = api_map.get(name)
    if not fn:
        return jsonify({"error": f"unknown report: {name}", "available": list(api_map.keys())}), 400
    resp = fn()
    data = resp.get_json() if hasattr(resp, "get_json") else resp
    # Pick the main array key
    rows_key = None
    for k in ("items", "customers", "doctors", "days", "patients", "pairs", "staff", "returns"):
        if isinstance(data.get(k), list):
            rows_key = k; break
    if not rows_key or not data.get(rows_key):
        return jsonify({"error": "no rows"}), 200
    out = io.StringIO()
    w = csv_mod.writer(out)
    rows = data[rows_key]
    keys = list(rows[0].keys()) if rows else []
    w.writerow(keys)
    for r in rows:
        w.writerow([r.get(k, "") for k in keys])
    return Response(out.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={name}.csv"})


# ════════════════════════════════════════════════════════════════════
# STAGE 20 — License/Anti-Piracy · Migration · Pre-orders · Staff Tasks
#            · High-Margin Combos · Session Timeout · Auto-Backup
# ════════════════════════════════════════════════════════════════════
import uuid as _uuid, platform as _platform, shutil as _shutil, glob as _glob

def init_stage20_db():
    with get_conn() as conn:
        # License — single row, tied to hardware fingerprint
        conn.execute("""
            CREATE TABLE IF NOT EXISTS app_license (
                id              INTEGER PRIMARY KEY,
                license_key     TEXT NOT NULL,
                shop_name       TEXT DEFAULT '',
                owner_name      TEXT DEFAULT '',
                hw_fingerprint  TEXT NOT NULL,
                machine_code    TEXT DEFAULT '',
                activation_code TEXT DEFAULT '',
                activated_at    TEXT DEFAULT (datetime('now')),
                expires_at      TEXT DEFAULT '2099-12-31',
                tier            TEXT DEFAULT 'production',
                status          TEXT DEFAULT 'active',
                last_check      TEXT DEFAULT (datetime('now')),
                checksum        TEXT DEFAULT ''
            )
        """)

        # Customer pre-orders
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pre_orders (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_name   TEXT DEFAULT '',
                customer_phone  TEXT DEFAULT '',
                medicine        TEXT NOT NULL,
                qty             INTEGER DEFAULT 1,
                requested_date  TEXT DEFAULT (datetime('now')),
                needed_by       TEXT DEFAULT '',
                status          TEXT DEFAULT 'pending',
                priority        TEXT DEFAULT 'normal',
                notes           TEXT DEFAULT '',
                fulfilled_at    TEXT DEFAULT '',
                fulfilled_bill  TEXT DEFAULT '',
                created_by      TEXT DEFAULT ''
            )
        """)

        # Recurring staff tasks
        conn.execute("""
            CREATE TABLE IF NOT EXISTS staff_tasks_def (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                title           TEXT NOT NULL,
                description     TEXT DEFAULT '',
                frequency       TEXT NOT NULL,
                category        TEXT DEFAULT '',
                est_minutes     INTEGER DEFAULT 30,
                priority        TEXT DEFAULT 'normal',
                is_active       INTEGER DEFAULT 1,
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS staff_task_log (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                task_def_id     INTEGER NOT NULL,
                title           TEXT DEFAULT '',
                assigned_to     INTEGER DEFAULT 0,
                assigned_name   TEXT DEFAULT '',
                due_date        TEXT DEFAULT '',
                status          TEXT DEFAULT 'pending',
                started_at      TEXT DEFAULT '',
                completed_at    TEXT DEFAULT '',
                quality_score   REAL DEFAULT 0,
                actual_minutes  INTEGER DEFAULT 0,
                notes           TEXT DEFAULT '',
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)

        # Migration jobs
        conn.execute("""
            CREATE TABLE IF NOT EXISTS migration_jobs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                filename        TEXT NOT NULL,
                file_type       TEXT DEFAULT 'txt',
                date_from       TEXT DEFAULT '',
                date_to         TEXT DEFAULT '',
                rows_parsed     INTEGER DEFAULT 0,
                rows_imported   INTEGER DEFAULT 0,
                rows_skipped    INTEGER DEFAULT 0,
                status          TEXT DEFAULT 'pending',
                error_log       TEXT DEFAULT '',
                created_at      TEXT DEFAULT (datetime('now')),
                completed_at    TEXT DEFAULT ''
            )
        """)

        # Seed default recurring tasks if empty
        n = conn.execute("SELECT COUNT(*) FROM staff_tasks_def").fetchone()[0]
        if n == 0:
            seeds = [
                ("Counter cleanup", "Wipe billing counter + display rack + cash drawer", "daily", "cleaning", 15, "high"),
                ("Floor sweep + mop", "Sweep + mop entire shop floor", "daily", "cleaning", 30, "high"),
                ("Cash tally with drawer", "Match physical cash with system drawer count", "daily", "finance", 20, "high"),
                ("Stock placement check", "Place newly arrived medicines in correct racks", "daily", "stock", 30, "normal"),
                ("Expiry check (front rows)", "Check front-row stock for expiry in next 30 days", "daily", "expiry", 15, "normal"),
                ("WhatsApp follow-up", "Send refill reminders to top 10 chronic customers", "daily", "crm", 20, "normal"),
                ("Glass + showcase clean", "Clean all glass cases and showcases", "weekly", "cleaning", 45, "normal"),
                ("Full stock count - top 100 items", "Audit physical vs system stock for top 100", "weekly", "stock", 90, "high"),
                ("Supplier payment review", "Review pending supplier payments, schedule next week", "weekly", "finance", 30, "high"),
                ("Customer database backup", "Export customer list + bills to USB drive", "weekly", "backup", 15, "high"),
                ("Deep clean storage racks", "Empty racks, clean, reorganize", "monthly", "cleaning", 180, "normal"),
                ("Expiry write-off + return memo", "Generate expiry returns to suppliers", "monthly", "expiry", 60, "high"),
                ("GST filing prep", "Compile GSTR-1 data, send to CA", "monthly", "compliance", 90, "high"),
                ("Salary + advance reconciliation", "Calculate net salary after advances", "monthly", "hr", 45, "high"),
                ("Drug Register audit", "Verify narcotic register entries", "monthly", "compliance", 60, "high"),
                ("Full inventory recount", "Physical count of ENTIRE stock vs system", "bimonthly", "stock", 360, "high"),
                ("AC + cold storage service", "Service AC, check refrigerator temperature logs", "bimonthly", "maintenance", 120, "high"),
                ("Drug Inspector visit prep", "Update all registers, organize licenses", "bimonthly", "compliance", 120, "high"),
                ("Loyalty tier review", "Review customer tiers, send promotional WA", "bimonthly", "crm", 60, "normal"),
            ]
            for t in seeds:
                conn.execute("""INSERT INTO staff_tasks_def
                    (title, description, frequency, category, est_minutes, priority)
                    VALUES (?,?,?,?,?,?)""", t)


# ─── HARDWARE FINGERPRINT (for license binding) ────────────────────
def get_hw_fingerprint():
    """Build a stable hardware fingerprint from MAC + node name + platform."""
    try:
        mac = _uuid.getnode()
        node = _platform.node() or ""
        plat = _platform.platform() or ""
        raw = f"{mac}|{node}|{plat}"
        return hashlib.sha256(raw.encode()).hexdigest()[:32]
    except Exception:
        return "fallback-fingerprint-" + hashlib.sha256(str(time.time()).encode()).hexdigest()[:16]


def get_machine_code():
    """Short machine code shown to user for offline activation."""
    fp = get_hw_fingerprint()
    # Format as XXXX-XXXX-XXXX-XXXX
    short = hashlib.sha256(fp.encode()).hexdigest()[:16].upper()
    return "-".join(short[i:i+4] for i in range(0, 16, 4))


def compute_activation_code(machine_code, owner_secret="SELVAM-MEDIVISION-2026"):
    """Generate activation code for a machine. In production owner uses a CLI tool to generate this."""
    raw = f"{machine_code}|{owner_secret}"
    code = hashlib.sha256(raw.encode()).hexdigest()[:16].upper()
    return "-".join(code[i:i+4] for i in range(0, 16, 4))


def check_license():
    """Returns (is_valid, license_dict or None, reason)."""
    try:
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM app_license WHERE id=1").fetchone()
        if not row:
            return (False, None, "not_activated")
        lic = dict(row)
        # Verify HW fingerprint matches
        cur_fp = get_hw_fingerprint()
        if lic["hw_fingerprint"] != cur_fp:
            return (False, lic, "hardware_mismatch")
        # Check expiry
        try:
            exp = datetime.strptime(lic["expires_at"][:10], "%Y-%m-%d")
            if exp < datetime.now():
                return (False, lic, "expired")
        except Exception:
            pass
        # Verify activation code matches what we'd expect for this machine
        expected = compute_activation_code(get_machine_code())
        if lic["activation_code"] != expected:
            return (False, lic, "tampered")
        return (True, lic, "ok")
    except Exception as e:
        return (False, None, f"error:{e}")


# ─── LICENSE-CHECK MIDDLEWARE ──────────────────────────────────────
# Routes that work even without license (so user can activate)
_LICENSE_FREE_PATHS = {
    "/license", "/api/license/info", "/api/license/activate",
    "/sw.js", "/manifest.json", "/offline", "/static/",
}

@app.before_request
def _license_gate():
    p = request.path or ""
    # Skip license check for safe paths
    if any(p == s or p.startswith(s) for s in _LICENSE_FREE_PATHS):
        return None
    # Check license
    valid, lic, reason = check_license()
    if valid:
        return None
    # Allow grace period of 7 days from first run
    if reason == "not_activated":
        # First-run grace: allow 7 days
        if not hasattr(app, "_first_run_ts"):
            app._first_run_ts = time.time()
        elapsed_days = (time.time() - app._first_run_ts) / 86400
        if elapsed_days < 7:
            return None  # Allow grace
    # Block + redirect to activation page
    if p.startswith("/api/"):
        return jsonify({
            "error": "license_required",
            "reason": reason,
            "machine_code": get_machine_code(),
            "activate_at": "/license"
        }), 403
    return redirect("/license")


# ─── SESSION TIMEOUT ───────────────────────────────────────────────
@app.before_request
def _session_timeout():
    """Auto-logout after 8h idle."""
    if "staff_id" not in session:
        return None
    last = session.get("last_activity", 0)
    now = time.time()
    if last and (now - last) > 8 * 3600:
        session.clear()
        if (request.path or "").startswith("/api/"):
            return jsonify({"error": "session_expired"}), 401
        return redirect("/login")
    session["last_activity"] = now


# ─── DAILY BACKUP ──────────────────────────────────────────────────
def daily_backup():
    """Backup database.db to backups/<date>.db if not already done today."""
    try:
        backup_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "backups")
        _os.makedirs(backup_dir, exist_ok=True)
        today = datetime.now().strftime("%Y%m%d")
        target = _os.path.join(backup_dir, f"database_{today}.db")
        if _os.path.exists(target):
            return False  # already backed up today
        src = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "database.db")
        if _os.path.exists(src):
            _shutil.copy2(src, target)
            # Keep last 30 days
            backups = sorted(_glob.glob(_os.path.join(backup_dir, "database_*.db")))
            for old in backups[:-30]:
                try: _os.remove(old)
                except: pass
            return True
    except Exception as e:
        print(f"[backup] {e}")
    return False


_LAST_BACKUP_CHECK = [0]
@app.before_request
def _backup_check():
    """Trigger daily backup once per day (lazy check)."""
    now = time.time()
    if now - _LAST_BACKUP_CHECK[0] > 3600:  # check once per hour
        _LAST_BACKUP_CHECK[0] = now
        daily_backup()


# ─── LICENSE APIs ──────────────────────────────────────────────────
@app.route("/license")
def license_page():
    return render_template("license.html")


@app.route("/api/license/info", methods=["GET"])
def api_license_info():
    valid, lic, reason = check_license()
    machine_code = get_machine_code()
    out = {
        "valid": valid,
        "reason": reason,
        "machine_code": machine_code,
        "hardware": _platform.node(),
        "system": _platform.platform(),
    }
    if lic:
        out["shop_name"] = lic.get("shop_name", "")
        out["owner_name"] = lic.get("owner_name", "")
        out["activated_at"] = lic.get("activated_at", "")
        out["expires_at"] = lic.get("expires_at", "")
        out["tier"] = lic.get("tier", "")
    # First-run grace info
    if reason == "not_activated":
        if hasattr(app, "_first_run_ts"):
            elapsed = (time.time() - app._first_run_ts) / 86400
            out["grace_days_remaining"] = max(0, round(7 - elapsed, 1))
    return jsonify(out)


@app.route("/api/license/activate", methods=["POST"])
def api_license_activate():
    """Activate with shop info + activation code."""
    d = request.json or {}
    shop_name = d.get("shop_name", "").strip()
    owner_name = d.get("owner_name", "").strip()
    activation_code = d.get("activation_code", "").strip().upper()

    if not shop_name or not activation_code:
        return jsonify({"error": "shop_name and activation_code required"}), 400

    machine_code = get_machine_code()
    expected = compute_activation_code(machine_code)
    if activation_code != expected:
        return jsonify({"error": "invalid activation code for this machine",
                        "machine_code": machine_code}), 400

    hw = get_hw_fingerprint()
    key = f"MV-{machine_code}-LIC"
    with get_conn() as conn:
        conn.execute("DELETE FROM app_license")  # single-row
        conn.execute("""INSERT INTO app_license
            (id, license_key, shop_name, owner_name, hw_fingerprint, machine_code,
             activation_code, tier, status, expires_at)
            VALUES (1,?,?,?,?,?,?,'production','active','2099-12-31')""",
            (key, shop_name, owner_name, hw, machine_code, activation_code))

    _audit("license_activate", "license", "1", new=f"{shop_name} ({owner_name})", severity="critical")
    return jsonify({"status": "ok", "license_key": key, "message": "License activated successfully"})


# ─── MIGRATION APIs ────────────────────────────────────────────────
@app.route("/migrate")
def migrate_page():
    return render_template("migrate.html")


def _parse_txt_bills(text):
    """Try to parse common DOS-billing-software TXT formats.
    Heuristics for typical formats:
      Format A: pipe-delimited: BILLNO|DATE|CUST|PHONE|ITEM|QTY|RATE|AMT|TOTAL
      Format B: fixed-width columns with header
      Format C: tab-separated
    Returns: list of bill dicts with parsed items."""
    bills = {}  # bill_id -> bill dict
    lines = text.split("\n")

    # Detect delimiter
    sample = "\n".join(lines[:50])
    if "|" in sample and sample.count("|") > 10:
        delim = "|"
    elif "\t" in sample:
        delim = "\t"
    else:
        delim = ","

    skipped = 0
    for ln in lines:
        ln = ln.strip()
        if not ln or ln.startswith("#"):
            continue
        # Skip headers (containing "BILL" "DATE" "ITEM" all upper)
        upper = ln.upper()
        if "BILLNO" in upper and "DATE" in upper:
            continue
        parts = [p.strip() for p in ln.split(delim)]
        if len(parts) < 5:
            skipped += 1
            continue
        try:
            # Try to extract: bill_no, date, customer, item_name, qty, rate, amount
            bill_no = parts[0]
            date_str = parts[1] if len(parts) > 1 else ""
            customer = parts[2] if len(parts) > 2 else ""
            phone = ""
            # Look for a 10-digit phone
            for p in parts:
                if p.isdigit() and len(p) == 10:
                    phone = p; break

            # Date parsing — try DD/MM/YY, DD-MM-YYYY, etc.
            date_clean = ""
            for fmt in ("%d/%m/%y", "%d/%m/%Y", "%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%y"):
                try:
                    d = datetime.strptime(date_str, fmt)
                    if d.year < 2000: d = d.replace(year=d.year + 100)
                    date_clean = d.strftime("%Y-%m-%d")
                    break
                except Exception:
                    continue
            if not date_clean:
                skipped += 1
                continue

            # Find numeric values - qty, rate, amount
            nums = []
            for p in parts:
                try:
                    v = float(p.replace(",", "").replace("Rs.", "").replace("₹", ""))
                    nums.append(v)
                except Exception:
                    pass
            if len(nums) < 2:
                skipped += 1
                continue

            # Item name heuristic: longest non-numeric, non-date part
            item_name = ""
            for p in parts[2:]:
                if p and not p.replace(".", "").replace(",", "").isdigit() and len(p) > len(item_name):
                    if "/" not in p and "-" not in p and p != customer:
                        item_name = p
            if not item_name:
                item_name = parts[3] if len(parts) > 3 else "Unknown"

            qty = int(nums[0]) if nums[0] < 1000 else 1
            rate = nums[1] if len(nums) > 1 else 0
            amt = nums[-1] if len(nums) > 2 else (qty * rate)

            if bill_no not in bills:
                bills[bill_no] = {
                    "id": bill_no, "date": date_clean, "cust": customer, "phone": phone,
                    "items": [], "sub": 0, "tax": 0, "disc": 0, "total": 0,
                }
            bills[bill_no]["items"].append({
                "name": item_name, "qty": qty, "price": rate, "amount": amt
            })
            bills[bill_no]["sub"] += amt
        except Exception:
            skipped += 1
            continue

    # Compute tax/total
    for b in bills.values():
        b["sub"] = round(b["sub"], 2)
        b["tax"] = round(b["sub"] * 0.12, 2)
        b["total"] = round(b["sub"] + b["tax"], 2)

    return list(bills.values()), skipped


@app.route("/api/migrate/preview", methods=["POST"])
def api_migrate_preview():
    """Parse uploaded text → return preview without saving."""
    if "file" in request.files:
        text = request.files["file"].read().decode("utf-8", errors="ignore")
    else:
        d = request.json or {}
        text = d.get("text", "")
    if not text:
        return jsonify({"error": "no content"}), 400
    bills, skipped = _parse_txt_bills(text)
    return jsonify({
        "bills_count": len(bills),
        "skipped_lines": skipped,
        "preview": bills[:20],
        "total_value": sum(b["total"] for b in bills),
    })


@app.route("/api/migrate/commit", methods=["POST"])
def api_migrate_commit():
    """Parse + commit to bills table."""
    d = request.json or {}
    text = d.get("text", "")
    filename = d.get("filename", "migration.txt")
    if not text:
        return jsonify({"error": "no content"}), 400
    bills, skipped = _parse_txt_bills(text)

    imported = 0
    errors = []
    with get_conn() as conn:
        # Log job
        cur = conn.execute("""INSERT INTO migration_jobs
            (filename, file_type, rows_parsed, status)
            VALUES (?,?,?,?)""", (filename, "txt", len(bills), "running"))
        job_id = cur.lastrowid

        for b in bills:
            try:
                # Generate unique bill ID with MG prefix
                bid = f"MG-{b['id']}-{int(time.time() % 10000)}"
                ts = int(datetime.strptime(b["date"], "%Y-%m-%d").timestamp())
                conn.execute("""INSERT OR IGNORE INTO bills
                    (id, ts, date, cust, phone, pay, sub, disc, tax, total, items, doctor, rx, prescription,
                     bill_type, customer_type, whatsapp_sent, staff_name, staff_id, profit)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (bid, ts, b["date"], b["cust"], b["phone"], "Cash",
                     b["sub"], 0, b["tax"], b["total"], json.dumps(b["items"]),
                     "", "", "", "retail", "customer", 0, "Migration", 0, b["sub"] * 0.20))
                imported += 1
            except Exception as e:
                errors.append(str(e)[:100])

        conn.execute("""UPDATE migration_jobs SET rows_imported=?, rows_skipped=?,
                       status='completed', completed_at=datetime('now'), error_log=?
                       WHERE id=?""",
                    (imported, skipped + (len(bills) - imported),
                     "\n".join(errors[:10]), job_id))

    _audit("migration", "bills", job_id, new=f"{imported} bills imported", severity="critical")
    return jsonify({
        "status": "ok", "job_id": job_id,
        "imported": imported, "skipped": skipped, "errors": errors[:5]
    })


@app.route("/api/migrate/jobs", methods=["GET"])
def api_migrate_jobs():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM migration_jobs ORDER BY id DESC LIMIT 50").fetchall()
    return jsonify({"jobs": [dict(r) for r in rows]})


# ─── PRE-ORDERS APIs ───────────────────────────────────────────────
@app.route("/preorders")
def preorders_page():
    return render_template("preorders.html")


@app.route("/api/preorders", methods=["GET"])
def api_preorders_list():
    status = request.args.get("status", "")
    q = "SELECT * FROM pre_orders"
    p = []
    if status:
        q += " WHERE status=?"; p.append(status)
    q += " ORDER BY id DESC LIMIT 200"
    with get_conn() as conn:
        rows = conn.execute(q, p).fetchall()
        stats = {}
        for s in ["pending", "ordered", "ready", "fulfilled", "cancelled"]:
            stats[s] = conn.execute("SELECT COUNT(*) FROM pre_orders WHERE status=?", (s,)).fetchone()[0]
    return jsonify({"preorders": [dict(r) for r in rows], "stats": stats})


@app.route("/api/preorders", methods=["POST"])
def api_preorders_create():
    d = request.json or {}
    if not d.get("medicine"):
        return jsonify({"error": "medicine required"}), 400
    with get_conn() as conn:
        cur = conn.execute("""INSERT INTO pre_orders
            (customer_name, customer_phone, medicine, qty, needed_by, priority, notes, created_by)
            VALUES (?,?,?,?,?,?,?,?)""",
            (d.get("customer_name", ""), d.get("customer_phone", ""),
             d.get("medicine", ""), int(d.get("qty", 1)),
             d.get("needed_by", ""), d.get("priority", "normal"),
             d.get("notes", ""), session.get("staff_name", "")))
    return jsonify({"status": "ok", "id": cur.lastrowid})


@app.route("/api/preorders/<int:pid>/status", methods=["POST"])
def api_preorders_status(pid):
    d = request.json or {}
    new_status = d.get("status", "")
    if new_status not in ["pending", "ordered", "ready", "fulfilled", "cancelled"]:
        return jsonify({"error": "invalid status"}), 400
    with get_conn() as conn:
        if new_status == "fulfilled":
            conn.execute("UPDATE pre_orders SET status=?, fulfilled_at=datetime('now'), fulfilled_bill=? WHERE id=?",
                         (new_status, d.get("bill_id", ""), pid))
        else:
            conn.execute("UPDATE pre_orders SET status=? WHERE id=?", (new_status, pid))
    return jsonify({"status": "ok"})


# ─── STAFF TASKS APIs ──────────────────────────────────────────────
@app.route("/staff-tasks")
def staff_tasks_page():
    return render_template("staff_tasks.html")


@app.route("/api/staff-tasks/definitions", methods=["GET"])
def api_task_defs():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM staff_tasks_def WHERE is_active=1 ORDER BY frequency, priority DESC").fetchall()
    return jsonify({"tasks": [dict(r) for r in rows]})


@app.route("/api/staff-tasks/log", methods=["GET"])
def api_task_log():
    date_from = request.args.get("from", (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
    with get_conn() as conn:
        rows = conn.execute("""SELECT * FROM staff_task_log
                               WHERE substr(created_at,1,10) >= ?
                               ORDER BY id DESC LIMIT 500""", (date_from,)).fetchall()
        stats = {}
        for s in ["pending", "in_progress", "completed", "skipped"]:
            stats[s] = conn.execute("SELECT COUNT(*) FROM staff_task_log WHERE status=? AND substr(created_at,1,10)>=?", (s, date_from)).fetchone()[0]
    return jsonify({"log": [dict(r) for r in rows], "stats": stats})


@app.route("/api/staff-tasks/assign", methods=["POST"])
def api_task_assign():
    d = request.json or {}
    task_def_id = int(d.get("task_def_id", 0))
    staff_id = int(d.get("staff_id", 0))
    due = d.get("due_date", datetime.now().strftime("%Y-%m-%d"))
    with get_conn() as conn:
        tdef = conn.execute("SELECT * FROM staff_tasks_def WHERE id=?", (task_def_id,)).fetchone()
        if not tdef:
            return jsonify({"error": "task not found"}), 404
        staff = conn.execute("SELECT name FROM staff WHERE id=?", (staff_id,)).fetchone() if staff_id else None
        cur = conn.execute("""INSERT INTO staff_task_log
            (task_def_id, title, assigned_to, assigned_name, due_date, status)
            VALUES (?,?,?,?,?,'pending')""",
            (task_def_id, tdef["title"], staff_id,
             staff["name"] if staff else "", due))
    return jsonify({"status": "ok", "id": cur.lastrowid})


@app.route("/api/staff-tasks/<int:lid>/complete", methods=["POST"])
def api_task_complete(lid):
    d = request.json or {}
    quality = float(d.get("quality_score", 80))
    minutes = int(d.get("actual_minutes", 0))
    notes = d.get("notes", "")
    with get_conn() as conn:
        conn.execute("""UPDATE staff_task_log
                       SET status='completed', completed_at=datetime('now'),
                           quality_score=?, actual_minutes=?, notes=?
                       WHERE id=?""", (quality, minutes, notes, lid))
    return jsonify({"status": "ok"})


@app.route("/api/staff-tasks/leaderboard", methods=["GET"])
def api_task_leaderboard():
    """ML-light: rank staff by completion rate × avg quality × speed."""
    days = int(request.args.get("days", 30))
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT assigned_to, assigned_name,
                   COUNT(*) as total,
                   SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) as done,
                   AVG(CASE WHEN status='completed' THEN quality_score ELSE 0 END) as avg_q,
                   AVG(CASE WHEN status='completed' THEN actual_minutes ELSE 0 END) as avg_m
            FROM staff_task_log
            WHERE substr(created_at,1,10) >= ? AND assigned_to > 0
            GROUP BY assigned_to ORDER BY done DESC
        """, (cutoff,)).fetchall()

        # Per-category strengths (which staff is best at what)
        cats = conn.execute("""
            SELECT l.assigned_to, l.assigned_name, d.category,
                   AVG(l.quality_score) as q, COUNT(*) as c
            FROM staff_task_log l JOIN staff_tasks_def d ON l.task_def_id = d.id
            WHERE l.status='completed' AND substr(l.created_at,1,10) >= ?
            GROUP BY l.assigned_to, d.category
            HAVING c >= 1 ORDER BY q DESC
        """, (cutoff,)).fetchall()

    staff_list = []
    for r in rows:
        completion_rate = (r["done"] / r["total"] * 100) if r["total"] else 0
        score = completion_rate * 0.5 + (r["avg_q"] or 0) * 0.5
        staff_list.append({
            "staff_id": r["assigned_to"], "name": r["assigned_name"],
            "total": r["total"], "done": r["done"],
            "completion_rate": round(completion_rate, 1),
            "avg_quality": round(r["avg_q"] or 0, 1),
            "avg_minutes": round(r["avg_m"] or 0, 0),
            "score": round(score, 1),
        })

    # Best at category
    best_at = {}
    for r in cats:
        if r["category"] not in best_at or r["q"] > best_at[r["category"]]["q"]:
            best_at[r["category"]] = {"staff": r["assigned_name"], "q": round(r["q"], 1), "tasks_done": r["c"]}

    return jsonify({"staff": staff_list, "best_at_category": best_at})


# ─── HIGH-MARGIN COMBINATION REPORT ────────────────────────────────
@app.route("/combinations")
def combinations_page():
    return render_template("combinations.html")


@app.route("/api/combinations/high-margin", methods=["GET"])
def api_combinations():
    """Find items frequently bought together, ranked by combined margin potential."""
    fs, ts, tf, tt = _parse_date_range(request.args)
    min_count = int(request.args.get("min", 2))
    limit = int(request.args.get("limit", 200))

    with get_conn() as conn:
        rows = conn.execute("SELECT items FROM bills WHERE ts BETWEEN ? AND ?", (tf, tt)).fetchall()
        # Get medicine prices for margin estimation
        meds = {m["n"]: dict(m) for m in conn.execute("SELECT n, p, c FROM medicines").fetchall()}

    pair_count = {}
    pair_value = {}
    item_count = {}
    total_baskets = 0

    for r in rows:
        try: items = json.loads(r["items"] or "[]")
        except: continue
        names = list({it.get("name", "") for it in items if it.get("name")})
        amounts = {it.get("name"): it.get("amount", 0) for it in items if it.get("name")}
        if len(names) < 1: continue
        total_baskets += 1
        for n in names:
            item_count[n] = item_count.get(n, 0) + 1
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                a, b = sorted([names[i], names[j]])
                key = (a, b)
                pair_count[key] = pair_count.get(key, 0) + 1
                pair_value[key] = pair_value.get(key, 0) + amounts.get(a, 0) + amounts.get(b, 0)

    combos = []
    for (a, b), c in pair_count.items():
        if c < min_count: continue
        avg_value = pair_value[(a, b)] / c
        # Estimated margin: ~22% blended
        est_margin = avg_value * 0.22
        # Recommendation strength = count × margin
        score = c * est_margin
        # Categorize by syrup/tablet/etc.
        cat_a = meds.get(a, {}).get("c", "")
        cat_b = meds.get(b, {}).get("c", "")
        kind = f"{cat_a} + {cat_b}"
        combos.append({
            "a": a, "b": b, "count": c,
            "avg_basket_value": round(avg_value, 2),
            "est_margin_per_combo": round(est_margin, 2),
            "total_combo_revenue": round(pair_value[(a, b)], 2),
            "score": round(score, 2),
            "kind": kind,
        })

    combos.sort(key=lambda x: x["score"], reverse=True)
    return jsonify({"combos": combos[:limit], "baskets_analyzed": total_baskets})


@app.route("/api/combinations/syrups", methods=["GET"])
def api_combinations_syrups():
    """All available syrups + their stock and price."""
    with get_conn() as conn:
        rows = conn.execute("""SELECT id, n, c, p, s FROM medicines
                               WHERE (UPPER(c) LIKE '%SYRUP%' OR UPPER(c) LIKE '%SYP%'
                                      OR UPPER(n) LIKE '%SYRUP%' OR UPPER(n) LIKE '%SYP%'
                                      OR UPPER(n) LIKE '%TONIC%' OR UPPER(n) LIKE '%SUSP%'
                                      OR UPPER(n) LIKE '%DROPS%')
                                 AND s > 0
                               ORDER BY s DESC LIMIT 200""").fetchall()
    return jsonify({"syrups": [dict(r) for r in rows], "total": len(rows)})


# ════════════════════════════════════════════════════════════════════
# STAGE 21 — MASTER SAAS CONTROL
#  · Plans (feature-gated, custom pricing)
#  · Clients (retail + wholesale)
#  · GPay QR + manual proof upload + vendor approval
#  · Subscription reminders (3 days before expiry)
#  · Discounts
#  · Installed-app sync via internet
# ════════════════════════════════════════════════════════════════════
def init_stage21_db():
    with get_conn() as conn:
        # Master plans (extends existing subscription_plans if any)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS saas_master_plans (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                code            TEXT UNIQUE NOT NULL,
                name            TEXT NOT NULL,
                tagline         TEXT DEFAULT '',
                price           REAL NOT NULL,
                billing_cycle   TEXT DEFAULT 'monthly',
                features_json   TEXT DEFAULT '[]',
                limits_json     TEXT DEFAULT '{}',
                client_type     TEXT DEFAULT 'retail',
                is_active       INTEGER DEFAULT 1,
                is_featured     INTEGER DEFAULT 0,
                sort_order      INTEGER DEFAULT 0,
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)

        # Master clients (extends existing structures)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS saas_master_clients (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                shop_name       TEXT NOT NULL,
                owner_name      TEXT DEFAULT '',
                phone           TEXT DEFAULT '',
                whatsapp        TEXT DEFAULT '',
                email           TEXT DEFAULT '',
                address         TEXT DEFAULT '',
                client_type     TEXT DEFAULT 'retail',
                machine_code    TEXT DEFAULT '',
                plan_id         INTEGER DEFAULT 0,
                discount_pct    REAL DEFAULT 0,
                subscription_start TEXT DEFAULT '',
                subscription_end   TEXT DEFAULT '',
                status          TEXT DEFAULT 'trial',
                trial_end       TEXT DEFAULT '',
                last_seen       TEXT DEFAULT '',
                total_paid      REAL DEFAULT 0,
                notes           TEXT DEFAULT '',
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)

        # Payment proofs (GPay QR flow)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS saas_payment_proofs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id       INTEGER NOT NULL,
                shop_name       TEXT DEFAULT '',
                plan_id         INTEGER DEFAULT 0,
                amount          REAL NOT NULL,
                billing_cycle   TEXT DEFAULT 'monthly',
                payment_method  TEXT DEFAULT 'gpay',
                transaction_id  TEXT DEFAULT '',
                payer_name      TEXT DEFAULT '',
                payer_phone     TEXT DEFAULT '',
                proof_image     TEXT DEFAULT '',
                proof_url       TEXT DEFAULT '',
                notes           TEXT DEFAULT '',
                status          TEXT DEFAULT 'pending',
                approved_by     TEXT DEFAULT '',
                approved_at     TEXT DEFAULT '',
                rejection_reason TEXT DEFAULT '',
                extends_until   TEXT DEFAULT '',
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)

        # Reminders sent
        conn.execute("""
            CREATE TABLE IF NOT EXISTS saas_reminders_sent (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id       INTEGER NOT NULL,
                reminder_type   TEXT NOT NULL,
                channel         TEXT DEFAULT 'whatsapp',
                message         TEXT DEFAULT '',
                sent_at         TEXT DEFAULT (datetime('now')),
                opened          INTEGER DEFAULT 0,
                responded       INTEGER DEFAULT 0
            )
        """)

        # Vendor settings (GPay UPI ID, QR image path, brand colors etc.)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS saas_vendor_settings (
                key             TEXT PRIMARY KEY,
                value           TEXT DEFAULT ''
            )
        """)
        defaults = [
            ("upi_id", "selvammedicals@oksbi"),
            ("upi_name", "Arvind Selvam SR"),
            ("qr_image_url", ""),
            ("vendor_phone", "+919876543210"),
            ("vendor_email", "arvind@selvammedicals.in"),
            ("brand_name", "MediVision AI"),
            ("central_server_url", "https://medivision.in"),
        ]
        for k, v in defaults:
            conn.execute("INSERT OR IGNORE INTO saas_vendor_settings (key, value) VALUES (?,?)", (k, v))

        # Seed default 3 plans if empty
        cnt = conn.execute("SELECT COUNT(*) FROM saas_master_plans").fetchone()[0]
        if cnt == 0:
            plans = [
                ("STARTER_R", "Starter (Retail)", "For small pharmacies", 499, "monthly",
                 ["billing", "gst", "stock", "customers", "whatsapp_share"],
                 {"max_users": 2, "max_bills_per_month": 500, "ai_calls_per_month": 50, "branches": 1},
                 "retail", 1, 0),
                ("PRO_R", "Pro (Retail)", "All AI tools + ML reports", 1499, "monthly",
                 ["billing", "gst", "stock", "customers", "whatsapp_share", "ai_billing",
                  "voice_billing", "face_recognition", "narcotic_register", "loyalty",
                  "refill_reminders", "market_basket", "rfm_cohorts", "cash_drawer"],
                 {"max_users": 5, "max_bills_per_month": 5000, "ai_calls_per_month": 500, "branches": 2},
                 "retail", 1, 1),
                ("ELITE_R", "Elite (Retail)", "Unlimited everything", 3999, "monthly",
                 ["*"],
                 {"max_users": 999, "max_bills_per_month": 999999, "ai_calls_per_month": 9999, "branches": 10},
                 "retail", 1, 0),
                ("STARTER_W", "Starter (Wholesale)", "For small distributors", 2999, "monthly",
                 ["wholesale", "shop_orders", "invoicing", "receivables", "whatsapp_share"],
                 {"max_users": 3, "max_shops": 10, "max_orders_per_month": 500},
                 "wholesale", 1, 0),
                ("PRO_W", "Pro (Wholesale)", "Full distributor toolkit", 7999, "monthly",
                 ["wholesale", "shop_orders", "invoicing", "receivables", "whatsapp_share",
                  "schemes", "ai_dispatch", "forecasting", "sales_rep_tracking"],
                 {"max_users": 10, "max_shops": 50, "max_orders_per_month": 5000},
                 "wholesale", 1, 1),
            ]
            for p in plans:
                conn.execute("""INSERT INTO saas_master_plans
                    (code, name, tagline, price, billing_cycle, features_json, limits_json, client_type, is_active, is_featured)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (p[0], p[1], p[2], p[3], p[4],
                     json.dumps(p[5]), json.dumps(p[6]),
                     p[7], p[8], p[9]))


# ─── HELPERS ───────────────────────────────────────────────────────
def _vendor_setting(key, default=""):
    with get_conn() as conn:
        r = conn.execute("SELECT value FROM saas_vendor_settings WHERE key=?", (key,)).fetchone()
    return r[0] if r else default


# ─── PUBLIC SUBSCRIBE PAGE ─────────────────────────────────────────
@app.route("/subscribe")
def subscribe_page():
    return render_template("subscribe.html")


@app.route("/api/saas/plans/public", methods=["GET"])
def api_plans_public():
    """Plans visible to clients (filtered by client_type)."""
    ctype = request.args.get("type", "retail")
    with get_conn() as conn:
        rows = conn.execute("""SELECT id, code, name, tagline, price, billing_cycle, features_json, limits_json, is_featured
                               FROM saas_master_plans
                               WHERE is_active=1 AND client_type=?
                               ORDER BY sort_order, price""", (ctype,)).fetchall()
    plans = []
    for r in rows:
        d = dict(r)
        try: d["features"] = json.loads(d.pop("features_json") or "[]")
        except: d["features"] = []
        try: d["limits"] = json.loads(d.pop("limits_json") or "{}")
        except: d["limits"] = {}
        plans.append(d)
    return jsonify({"plans": plans, "vendor": {
        "upi_id": _vendor_setting("upi_id"),
        "upi_name": _vendor_setting("upi_name"),
        "qr_image_url": _vendor_setting("qr_image_url"),
        "phone": _vendor_setting("vendor_phone"),
        "email": _vendor_setting("vendor_email"),
        "brand": _vendor_setting("brand_name"),
    }})


@app.route("/api/saas/proof/submit", methods=["POST"])
def api_proof_submit():
    """Client uploads GPay payment proof for vendor review."""
    # Accept multipart (file) or JSON
    proof_url = ""
    if "proof" in request.files:
        f = request.files["proof"]
        if f.filename:
            ts_str = datetime.now().strftime("%Y%m%d%H%M%S")
            safe = "".join(c for c in f.filename if c.isalnum() or c in "._-")[:40]
            fname = f"proof_{ts_str}_{safe}"
            full = _os.path.join(AUDIO_DIR, "..", "proofs", fname)
            _os.makedirs(_os.path.dirname(full), exist_ok=True)
            f.save(full)
            proof_url = f"/static/proofs/{fname}"

    d = request.form if request.form else (request.json or {})

    shop_name = d.get("shop_name", "").strip()
    plan_id = int(d.get("plan_id", 0))
    amount = float(d.get("amount", 0))
    txn_id = d.get("transaction_id", "").strip()
    payer_phone = d.get("payer_phone", "").strip()
    payer_name = d.get("payer_name", "").strip()
    notes = d.get("notes", "")
    machine_code = d.get("machine_code", "")

    if not shop_name or amount <= 0 or not txn_id:
        return jsonify({"error": "shop_name, amount, transaction_id required"}), 400

    with get_conn() as conn:
        # Find or create client
        client = conn.execute("SELECT * FROM saas_master_clients WHERE machine_code=? OR phone=?",
                              (machine_code, payer_phone)).fetchone()
        if client:
            client_id = client["id"]
        else:
            cur = conn.execute("""INSERT INTO saas_master_clients
                (shop_name, owner_name, phone, whatsapp, machine_code, plan_id, status, created_at)
                VALUES (?,?,?,?,?,?,'pending_approval',datetime('now'))""",
                (shop_name, payer_name, payer_phone, payer_phone, machine_code, plan_id))
            client_id = cur.lastrowid

        cur = conn.execute("""INSERT INTO saas_payment_proofs
            (client_id, shop_name, plan_id, amount, payment_method, transaction_id,
             payer_name, payer_phone, proof_url, notes, status)
            VALUES (?,?,?,?,'gpay',?,?,?,?,?,'pending')""",
            (client_id, shop_name, plan_id, amount, txn_id,
             payer_name, payer_phone, proof_url, notes))

    return jsonify({"status": "ok", "id": cur.lastrowid,
                    "message": "Payment proof submitted. We'll activate within 24 hours.",
                    "whatsapp_vendor": _vendor_setting("vendor_phone")})


# ─── ADMIN DASHBOARD ───────────────────────────────────────────────
@app.route("/admin/saas-master")
def admin_saas_master():
    return render_template("admin_saas_master.html")


# ─── PLAN ADMIN APIs ───────────────────────────────────────────────
@app.route("/api/admin/saas/plans", methods=["GET"])
def api_admin_plans():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM saas_master_plans ORDER BY client_type, sort_order, price").fetchall()
    out = []
    for r in rows:
        d = dict(r)
        try: d["features"] = json.loads(d["features_json"] or "[]")
        except: d["features"] = []
        try: d["limits"] = json.loads(d["limits_json"] or "{}")
        except: d["limits"] = {}
        out.append(d)
    return jsonify({"plans": out})


@app.route("/api/admin/saas/plans", methods=["POST"])
def api_admin_plan_save():
    d = request.json or {}
    pid = d.get("id")
    code = d.get("code", "").strip().upper()
    name = d.get("name", "").strip()
    if not code or not name:
        return jsonify({"error": "code and name required"}), 400
    features = d.get("features", [])
    limits = d.get("limits", {})

    with get_conn() as conn:
        if pid:
            conn.execute("""UPDATE saas_master_plans
                SET code=?, name=?, tagline=?, price=?, billing_cycle=?, features_json=?,
                    limits_json=?, client_type=?, is_active=?, is_featured=?, sort_order=?
                WHERE id=?""",
                (code, name, d.get("tagline", ""), float(d.get("price", 0)),
                 d.get("billing_cycle", "monthly"), json.dumps(features),
                 json.dumps(limits), d.get("client_type", "retail"),
                 1 if d.get("is_active", True) else 0,
                 1 if d.get("is_featured", False) else 0,
                 int(d.get("sort_order", 0)), pid))
        else:
            cur = conn.execute("""INSERT INTO saas_master_plans
                (code, name, tagline, price, billing_cycle, features_json, limits_json,
                 client_type, is_active, is_featured, sort_order)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (code, name, d.get("tagline", ""), float(d.get("price", 0)),
                 d.get("billing_cycle", "monthly"), json.dumps(features),
                 json.dumps(limits), d.get("client_type", "retail"),
                 1 if d.get("is_active", True) else 0,
                 1 if d.get("is_featured", False) else 0,
                 int(d.get("sort_order", 0))))
            pid = cur.lastrowid
    return jsonify({"status": "ok", "id": pid})


@app.route("/api/admin/saas/plans/<int:pid>", methods=["DELETE"])
def api_admin_plan_delete(pid):
    with get_conn() as conn:
        conn.execute("DELETE FROM saas_master_plans WHERE id=?", (pid,))
    return jsonify({"status": "ok"})


# ─── CLIENT ADMIN APIs ─────────────────────────────────────────────
@app.route("/api/admin/saas/clients", methods=["GET"])
def api_admin_clients():
    status = request.args.get("status", "")
    q = """SELECT c.*, p.name as plan_name, p.code as plan_code, p.price as plan_price
           FROM saas_master_clients c
           LEFT JOIN saas_master_plans p ON c.plan_id = p.id"""
    params = []
    if status:
        q += " WHERE c.status=?"; params.append(status)
    q += " ORDER BY c.id DESC LIMIT 500"
    with get_conn() as conn:
        rows = conn.execute(q, params).fetchall()
        stats = {
            "total": conn.execute("SELECT COUNT(*) FROM saas_master_clients").fetchone()[0],
            "active": conn.execute("SELECT COUNT(*) FROM saas_master_clients WHERE status='active'").fetchone()[0],
            "trial": conn.execute("SELECT COUNT(*) FROM saas_master_clients WHERE status='trial'").fetchone()[0],
            "pending": conn.execute("SELECT COUNT(*) FROM saas_master_clients WHERE status='pending_approval'").fetchone()[0],
            "expired": conn.execute("SELECT COUNT(*) FROM saas_master_clients WHERE status='expired'").fetchone()[0],
            "expiring_soon": conn.execute(
                "SELECT COUNT(*) FROM saas_master_clients WHERE status='active' "
                "AND date(subscription_end) BETWEEN date('now') AND date('now','+3 days')").fetchone()[0],
            "mrr": conn.execute(
                "SELECT COALESCE(SUM(p.price * (1 - c.discount_pct/100.0)), 0) "
                "FROM saas_master_clients c JOIN saas_master_plans p ON c.plan_id = p.id "
                "WHERE c.status='active'").fetchone()[0],
        }
    out = []
    for r in rows:
        d = dict(r)
        # Compute days remaining
        if d.get("subscription_end"):
            try:
                end = datetime.strptime(d["subscription_end"][:10], "%Y-%m-%d")
                d["days_remaining"] = (end - datetime.now()).days
            except Exception:
                d["days_remaining"] = None
        out.append(d)
    return jsonify({"clients": out, "stats": stats})


@app.route("/api/admin/saas/clients", methods=["POST"])
def api_admin_client_save():
    d = request.json or {}
    cid = d.get("id")
    if not d.get("shop_name"):
        return jsonify({"error": "shop_name required"}), 400
    with get_conn() as conn:
        if cid:
            conn.execute("""UPDATE saas_master_clients SET shop_name=?, owner_name=?, phone=?,
                whatsapp=?, email=?, address=?, client_type=?, machine_code=?, plan_id=?,
                discount_pct=?, subscription_start=?, subscription_end=?, status=?, notes=?
                WHERE id=?""",
                (d.get("shop_name"), d.get("owner_name", ""), d.get("phone", ""),
                 d.get("whatsapp", d.get("phone", "")), d.get("email", ""), d.get("address", ""),
                 d.get("client_type", "retail"), d.get("machine_code", ""),
                 int(d.get("plan_id", 0)), float(d.get("discount_pct", 0)),
                 d.get("subscription_start", ""), d.get("subscription_end", ""),
                 d.get("status", "active"), d.get("notes", ""), cid))
        else:
            cur = conn.execute("""INSERT INTO saas_master_clients
                (shop_name, owner_name, phone, whatsapp, email, address, client_type,
                 machine_code, plan_id, discount_pct, subscription_start, subscription_end, status, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (d.get("shop_name"), d.get("owner_name", ""), d.get("phone", ""),
                 d.get("whatsapp", d.get("phone", "")), d.get("email", ""), d.get("address", ""),
                 d.get("client_type", "retail"), d.get("machine_code", ""),
                 int(d.get("plan_id", 0)), float(d.get("discount_pct", 0)),
                 d.get("subscription_start", datetime.now().strftime("%Y-%m-%d")),
                 d.get("subscription_end", (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")),
                 d.get("status", "active"), d.get("notes", "")))
            cid = cur.lastrowid
    _audit("client_save", "saas_client", cid, new=d.get("shop_name"), severity="info")
    return jsonify({"status": "ok", "id": cid})


@app.route("/api/admin/saas/clients/<int:cid>", methods=["DELETE"])
def api_admin_client_delete(cid):
    with get_conn() as conn:
        conn.execute("DELETE FROM saas_master_clients WHERE id=?", (cid,))
    _audit("client_delete", "saas_client", cid, severity="warn")
    return jsonify({"status": "ok"})


@app.route("/api/admin/saas/clients/<int:cid>/extend", methods=["POST"])
def api_admin_client_extend(cid):
    """Extend subscription by N days."""
    d = request.json or {}
    days = int(d.get("days", 30))
    with get_conn() as conn:
        client = conn.execute("SELECT subscription_end FROM saas_master_clients WHERE id=?", (cid,)).fetchone()
        if not client:
            return jsonify({"error": "client not found"}), 404
        cur_end = client["subscription_end"]
        try:
            base = datetime.strptime(cur_end[:10], "%Y-%m-%d") if cur_end else datetime.now()
            if base < datetime.now(): base = datetime.now()
        except Exception:
            base = datetime.now()
        new_end = base + timedelta(days=days)
        conn.execute("UPDATE saas_master_clients SET subscription_end=?, status='active' WHERE id=?",
                     (new_end.strftime("%Y-%m-%d"), cid))
    return jsonify({"status": "ok", "new_end": new_end.strftime("%Y-%m-%d")})


# ─── PAYMENT PROOF REVIEW APIs ─────────────────────────────────────
@app.route("/api/admin/saas/proofs", methods=["GET"])
def api_admin_proofs():
    status = request.args.get("status", "")
    q = "SELECT * FROM saas_payment_proofs"
    p = []
    if status:
        q += " WHERE status=?"; p.append(status)
    q += " ORDER BY id DESC LIMIT 200"
    with get_conn() as conn:
        rows = conn.execute(q, p).fetchall()
        stats = {s: conn.execute("SELECT COUNT(*) FROM saas_payment_proofs WHERE status=?", (s,)).fetchone()[0]
                 for s in ["pending", "approved", "rejected"]}
    return jsonify({"proofs": [dict(r) for r in rows], "stats": stats})


@app.route("/api/admin/saas/proofs/<int:pid>/approve", methods=["POST"])
def api_admin_proof_approve(pid):
    d = request.json or {}
    extend_days = int(d.get("extend_days", 30))
    with get_conn() as conn:
        proof = conn.execute("SELECT * FROM saas_payment_proofs WHERE id=?", (pid,)).fetchone()
        if not proof:
            return jsonify({"error": "proof not found"}), 404

        client_id = proof["client_id"]
        client = conn.execute("SELECT subscription_end, total_paid FROM saas_master_clients WHERE id=?", (client_id,)).fetchone()

        # Extend subscription
        cur_end = client["subscription_end"] if client else ""
        try:
            base = datetime.strptime(cur_end[:10], "%Y-%m-%d") if cur_end else datetime.now()
            if base < datetime.now(): base = datetime.now()
        except Exception:
            base = datetime.now()
        new_end = base + timedelta(days=extend_days)

        # Mark proof approved
        conn.execute("""UPDATE saas_payment_proofs
            SET status='approved', approved_by=?, approved_at=datetime('now'),
                extends_until=? WHERE id=?""",
            (session.get("staff_name", "Admin"), new_end.strftime("%Y-%m-%d"), pid))

        # Update client
        new_total = (client["total_paid"] if client else 0) + proof["amount"]
        conn.execute("""UPDATE saas_master_clients
            SET status='active', subscription_end=?, plan_id=?, total_paid=?
            WHERE id=?""",
            (new_end.strftime("%Y-%m-%d"), proof["plan_id"], new_total, client_id))

    _audit("proof_approve", "payment_proof", pid, new=f"₹{proof['amount']} → +{extend_days}d", severity="critical")
    return jsonify({"status": "ok", "new_subscription_end": new_end.strftime("%Y-%m-%d")})


@app.route("/api/admin/saas/proofs/<int:pid>/reject", methods=["POST"])
def api_admin_proof_reject(pid):
    d = request.json or {}
    reason = d.get("reason", "Invalid payment proof")
    with get_conn() as conn:
        conn.execute("""UPDATE saas_payment_proofs
            SET status='rejected', approved_by=?, approved_at=datetime('now'), rejection_reason=?
            WHERE id=?""", (session.get("staff_name", "Admin"), reason, pid))
    _audit("proof_reject", "payment_proof", pid, new=reason, severity="warn")
    return jsonify({"status": "ok"})


# ─── REMINDERS ─────────────────────────────────────────────────────
@app.route("/api/admin/saas/reminders/run", methods=["POST"])
def api_admin_reminders_run():
    """Find clients expiring in 3 days, generate WhatsApp reminder URLs."""
    with get_conn() as conn:
        rows = conn.execute("""SELECT c.*, p.name as plan_name, p.price
            FROM saas_master_clients c
            LEFT JOIN saas_master_plans p ON c.plan_id = p.id
            WHERE c.status='active'
              AND date(c.subscription_end) BETWEEN date('now') AND date('now','+3 days')
            ORDER BY c.subscription_end""").fetchall()

    reminders = []
    for r in rows:
        end = r["subscription_end"][:10] if r["subscription_end"] else ""
        try:
            days_left = (datetime.strptime(end, "%Y-%m-%d") - datetime.now()).days
        except Exception:
            days_left = 0
        # Build WA message
        msg = (f"வணக்கம் {r['owner_name'] or r['shop_name']}!\n"
               f"உங்கள் MediVision AI subscription {days_left} நாட்களில் முடியும்.\n"
               f"Plan: {r['plan_name'] or '-'} · ₹{r['price'] or 0}/month\n"
               f"Renew: open the app → /subscribe\n"
               f"\nMediVision AI · {_vendor_setting('vendor_phone')}")
        phone = (r["whatsapp"] or r["phone"] or "").replace("+", "").replace(" ", "")
        if phone and len(phone) == 10:
            phone = "91" + phone
        wa_url = f"https://wa.me/{phone}?text={urllib.parse.quote(msg)}" if phone else ""
        reminders.append({
            "client_id": r["id"], "shop_name": r["shop_name"],
            "owner": r["owner_name"], "phone": phone,
            "days_left": days_left, "subscription_end": end,
            "plan": r["plan_name"], "amount": r["price"] or 0,
            "wa_url": wa_url, "message": msg,
        })

    # Mark all expired clients as expired
    with get_conn() as conn:
        conn.execute("""UPDATE saas_master_clients
            SET status='expired'
            WHERE status='active' AND date(subscription_end) < date('now')""")

    return jsonify({"reminders": reminders, "count": len(reminders)})


@app.route("/api/admin/saas/reminders/log", methods=["POST"])
def api_admin_reminder_log():
    """Log that a reminder was sent (clicked the WA button)."""
    d = request.json or {}
    with get_conn() as conn:
        conn.execute("""INSERT INTO saas_reminders_sent
            (client_id, reminder_type, channel, message)
            VALUES (?,?,?,?)""",
            (int(d.get("client_id", 0)), d.get("reminder_type", "expiry_3d"),
             d.get("channel", "whatsapp"), d.get("message", "")))
    return jsonify({"status": "ok"})


# ─── VENDOR SETTINGS APIs ──────────────────────────────────────────
@app.route("/api/admin/saas/settings", methods=["GET"])
def api_admin_settings():
    with get_conn() as conn:
        rows = conn.execute("SELECT key, value FROM saas_vendor_settings").fetchall()
    return jsonify({r["key"]: r["value"] for r in rows})


@app.route("/api/admin/saas/settings", methods=["POST"])
def api_admin_settings_save():
    d = request.json or {}
    with get_conn() as conn:
        for k, v in d.items():
            conn.execute("INSERT OR REPLACE INTO saas_vendor_settings (key, value) VALUES (?,?)", (k, str(v)))
    return jsonify({"status": "ok"})


# ─── CENTRAL SYNC API (installed apps call this from internet) ─────
@app.route("/api/saas/sync", methods=["POST"])
def api_saas_sync():
    """Installed apps call this every few hours to refresh subscription status."""
    d = request.json or {}
    machine_code = d.get("machine_code", "")
    if not machine_code:
        return jsonify({"error": "machine_code required"}), 400

    with get_conn() as conn:
        client = conn.execute("""SELECT c.*, p.code as plan_code, p.name as plan_name,
                                       p.features_json, p.limits_json, p.price
            FROM saas_master_clients c
            LEFT JOIN saas_master_plans p ON c.plan_id = p.id
            WHERE c.machine_code=?""", (machine_code,)).fetchone()

        if not client:
            return jsonify({
                "found": False,
                "message": "Machine not registered. Visit /subscribe to set up.",
                "subscribe_url": "/subscribe"
            })

        conn.execute("UPDATE saas_master_clients SET last_seen=datetime('now') WHERE id=?", (client["id"],))

    end = client["subscription_end"][:10] if client["subscription_end"] else ""
    try:
        days_left = (datetime.strptime(end, "%Y-%m-%d") - datetime.now()).days
    except Exception:
        days_left = -1

    return jsonify({
        "found": True,
        "shop_name": client["shop_name"],
        "plan_code": client["plan_code"],
        "plan_name": client["plan_name"],
        "status": client["status"],
        "subscription_end": end,
        "days_left": days_left,
        "features": json.loads(client["features_json"] or "[]"),
        "limits": json.loads(client["limits_json"] or "{}"),
        "needs_renewal": days_left <= 3,
    })


# ════════════════════════════════════════════════════════════════════
#  Initialize database
init_db()
init_compliance_db()
init_stage8_db()
init_stage9_db()
init_wholesale_db()
init_stage11_db()
init_stage12_db()
init_stage13_db()
init_phase3_db()
init_stage14_db()
init_stage15_db()
init_stage16_db()
init_stage17_db()
init_stage18_db()
init_stage19_db()
init_stage20_db()
init_stage21_db()


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=True, port=5001)